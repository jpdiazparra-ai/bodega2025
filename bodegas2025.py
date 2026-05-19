import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import numpy as np
import base64
from matplotlib.colors import LinearSegmentedColormap
from pathlib import Path
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from xml.sax.saxutils import escape

# =========================
# Configuración base
# =========================
st.set_page_config(page_title="Análisis Financiero - OK-DTA V2", layout="wide")

CHART_GRAY = "#A8A8A8"
CHART_RED = "#D85E5D"
CHART_GOLD = "#DCAA67"
CHART_TEAL = "#7FA6A2"
CHART_DARK = "#4B5563"
CHART_BAR_TEAL = "#9DB9B5"
CHART_BAR_RED = "#D07A75"

# --- Logo -> data URI ---
def data_uri(path: str) -> str:
    p = Path(path)
    if not p.exists():
        return ""  # si no está, ocultamos el <img>
    b64 = base64.b64encode(p.read_bytes()).decode()
    ext = p.suffix.lower().replace(".", "") or "png"
    return f"data:image/{ext};base64,{b64}"

LOGO_URI = data_uri("Logo_balmaceda.png")  # el archivo está en la misma carpeta del .py
HERO_URI = data_uri("IMG_7331.jpeg")

section_options = [
    "🏠 Overview Ejecutivo",
    "📈 Flujo Operacional",
    "⚠️ Riesgo & Cobranza",
    "🏢 Canon & Contratos",
    "🏗️ Capex",
    "⚡ Consumos Energéticos",
]

# =======================
# HERO HEADER
# =======================
st.markdown("""
<style>
/* Quitar espacio superior global */
main {
    padding-top: 0rem !important;
    background: #F7F9FC;
}
.stApp,
div[data-testid="stAppViewContainer"] {
    background: #F7F9FC !important;
}
.block-container {
    max-width: 1680px;
    padding: 0.2rem 1rem 0.5rem 1rem !important;
}
section.main > div,
div[data-testid="stMainBlockContainer"],
div[data-testid="block-container"] {
    padding-top: 0.2rem !important;
}
div[data-testid="stAppViewContainer"] {
    padding-top: 0 !important;
}
header[data-testid="stHeader"],
div[data-testid="stToolbar"],
#MainMenu,
footer {
    visibility: hidden !important;
    height: 0 !important;
    min-height: 0 !important;
    display: none !important;
}
/* Oculta contenido de la pestaña anterior durante los reruns de Streamlit. */
[data-testid="stElementContainer"][data-stale="true"] {
    display: none !important;
    opacity: 0 !important;
    height: 0 !important;
    min-height: 0 !important;
    margin: 0 !important;
    padding: 0 !important;
    overflow: hidden !important;
    pointer-events: none !important;
}
iframe[title="flujo-selector-scroll"] {
    display: none !important;
    height: 0 !important;
    min-height: 0 !important;
}
iframe[height="0"] {
    display: block !important;
    height: 0 !important;
    min-height: 0 !important;
}
div[data-testid="stElementContainer"]:has(iframe[height="0"]),
div[data-testid="stIFrame"]:has(iframe[height="0"]) {
    height: 0 !important;
    min-height: 0 !important;
    max-height: 0 !important;
    margin: 0 !important;
    padding: 0 !important;
    overflow: hidden !important;
}
div[data-testid="stElementContainer"]:has(iframe[title="flujo-selector-scroll"]),
div[data-testid="stIFrame"]:has(iframe[title="flujo-selector-scroll"]) {
    display: none !important;
    height: 0 !important;
    min-height: 0 !important;
    margin: 0 !important;
    padding: 0 !important;
    overflow: hidden !important;
}
section[data-testid="stSidebar"],
aside[data-testid="stSidebar"] {
    width: 252px !important;
    min-width: 252px !important;
    max-width: 252px !important;
    background: linear-gradient(180deg, #031B34 0%, #021326 100%) !important;
    border-right: 1px solid rgba(148, 163, 184, 0.10);
    transition: margin-left 160ms ease, width 160ms ease, min-width 160ms ease, opacity 120ms ease;
    z-index: 999;
    transform: translateX(0) !important;
    visibility: visible !important;
    opacity: 1 !important;
}
section[data-testid="stSidebar"] > div,
aside[data-testid="stSidebar"] > div,
section[data-testid="stSidebar"] [data-testid="stSidebarContent"],
aside[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
    background: transparent !important;
}
body.bodegas-sidebar-collapsed section[data-testid="stSidebar"],
body.bodegas-sidebar-collapsed aside[data-testid="stSidebar"] {
    width: 0 !important;
    min-width: 0 !important;
    max-width: 0 !important;
    margin-left: 0 !important;
    opacity: 0 !important;
    overflow: hidden !important;
    border-right: 0 !important;
}
body.bodegas-sidebar-collapsed section[data-testid="stSidebar"] > div,
body.bodegas-sidebar-collapsed aside[data-testid="stSidebar"] > div {
    display: none !important;
}
body.bodegas-sidebar-collapsed .block-container {
    max-width: none !important;
    padding-left: 1.15rem !important;
    padding-right: 1.15rem !important;
}
body.bodegas-sidebar-collapsed [data-testid="stAppViewContainer"] {
    margin-left: 0 !important;
}
.bodegas-sidebar-toggle {
    position: fixed;
    left: 12px;
    top: 12px;
    z-index: 10000;
    width: 38px;
    height: 38px;
    border-radius: 10px;
    border: 1px solid rgba(203,213,225,0.9);
    background: #ffffff;
    color: #0f2d52;
    display: flex;
    align-items: center;
    justify-content: center;
    box-shadow: 0 10px 28px rgba(15,23,42,0.16);
    cursor: pointer;
    font-size: 18px;
    line-height: 1;
    font-weight: 950;
}
.bodegas-sidebar-toggle:hover {
    background: #f8fafc;
    border-color: #93c5fd;
}
body:not(.bodegas-sidebar-collapsed) .bodegas-sidebar-toggle {
    left: 202px;
}
@media (max-width: 900px) {
    body:not(.bodegas-sidebar-collapsed) .bodegas-sidebar-toggle {
        left: 202px;
    }
}
section[data-testid="stSidebar"] > div {
    background: transparent;
    padding: 1.15rem 0.65rem 1rem 0.65rem;
}
section[data-testid="stSidebar"] [data-testid="stSidebarContent"],
aside[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
    padding-top: 0.2rem !important;
}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] {
    color: #e5edf7;
}
.sidebar-brand {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 6px 10px 18px 10px;
}
.sidebar-brand-logo {
    width: 38px;
    height: 38px;
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: rgba(255,255,255,0.08);
    border: 1px solid rgba(226,232,240,0.20);
}
.sidebar-brand-logo img {
    width: 28px;
    height: 28px;
    object-fit: contain;
}
.sidebar-brand-title {
    font-size: 14px;
    line-height: 1.05;
    font-weight: 900;
    letter-spacing: 0.02em;
    color: #ffffff;
    text-transform: uppercase;
}
.sidebar-brand-sub {
    margin-top: 3px;
    font-size: 11px;
    color: rgba(226,232,240,0.72);
    font-weight: 600;
}
section[data-testid="stSidebar"] div[role="radiogroup"][aria-label="Sección"] {
    display: flex !important;
    flex-direction: column;
    gap: 5px;
    width: 100%;
}
section[data-testid="stSidebar"] div[role="radiogroup"][aria-label="Sección"] > label {
    min-height: 43px;
    padding: 0 12px;
    border-radius: 8px;
    border: 1px solid transparent;
    background: transparent;
    transition: all 0.25s ease;
    align-items: center !important;
}
section[data-testid="stSidebar"] div[role="radiogroup"][aria-label="Sección"] > label:hover {
    background: #0F2E57;
    border-color: rgba(255,255,255,0.08);
    transform: translateX(2px);
}
section[data-testid="stSidebar"] div[role="radiogroup"][aria-label="Sección"] > label:has(input:checked) {
    background: linear-gradient(135deg, #2563EB 0%, #1D4ED8 100%);
    border-color: rgba(255,255,255,0.08);
    box-shadow: 0 8px 24px rgba(37,99,235,0.28);
    transform: translateX(2px);
}
section[data-testid="stSidebar"] div[role="radiogroup"][aria-label="Sección"] > label > div:last-child {
    color: #dbeafe;
    font-size: 13px;
    line-height: 1.25;
    font-weight: 780;
}
section[data-testid="stSidebar"] div[role="radiogroup"][aria-label="Sección"] > label:has(input:checked) > div:last-child {
    color: #ffffff;
}
section[data-testid="stSidebar"] div[role="radiogroup"][aria-label="Sección"] input {
    display: none;
}
section[data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] {
    display: flex !important;
    flex-direction: column !important;
    gap: 5px !important;
}
section[data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] > label {
    min-height: 43px !important;
    padding: 0 12px !important;
    border-radius: 8px !important;
    border: 1px solid transparent !important;
    background: transparent !important;
    transition: all 0.25s ease !important;
}
section[data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] > label:hover {
    background: #0F2E57 !important;
    border-color: rgba(255,255,255,0.08) !important;
    transform: translateX(2px);
}
section[data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] > label > div:first-child {
    display: none !important;
}
section[data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] > label [data-testid="stMarkdownContainer"] p {
    color: #dbeafe !important;
    font-size: 13px !important;
    line-height: 1.25 !important;
    font-weight: 780 !important;
}
section[data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] > label:has(input:checked) {
    background: linear-gradient(135deg, #2563EB 0%, #1D4ED8 100%) !important;
    border-color: rgba(255,255,255,0.08) !important;
    box-shadow: 0 8px 24px rgba(37,99,235,0.28) !important;
    transform: translateX(2px);
}
section[data-testid="stSidebar"] [data-testid="stRadio"] div[role="radiogroup"] > label:has(input:checked) [data-testid="stMarkdownContainer"] p {
    color: #ffffff !important;
}
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
    font-size: 13px !important;
    letter-spacing: 0.02em;
}
.sidebar-bottom-spacer {
    height: clamp(90px, 18vh, 210px);
}
section[data-testid="stSidebar"] .stButton > button {
    width: 100%;
    min-height: 54px;
    border-radius: 8px;
    border: 1px solid rgba(147,197,253,0.22);
    background: rgba(15, 45, 82, 0.78);
    color: #e5edf7;
    font-size: 13.5px;
    line-height: 1.25;
    font-weight: 800;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    border-color: rgba(191,219,254,0.45);
    background: rgba(30, 64, 175, 0.82);
    color: #ffffff;
}
.dashboard-hero {
    position: relative;
    overflow: hidden;
    min-height: 76px;
    margin: 0 0 8px 0;
    border-radius: 10px;
    border: 1px solid #d8e1ed;
    background: linear-gradient(135deg, #ffffff 0%, #f8fbff 62%, #f1f6fb 100%);
    box-shadow: 0 18px 38px rgba(15, 23, 42, 0.07);
}
.dashboard-hero::after {
    content: "";
    position: absolute;
    inset: 0;
    background: linear-gradient(90deg, rgba(255,255,255,0) 0%, rgba(15,45,82,0.04) 100%);
    pointer-events: none;
}
.dashboard-hero-inner {
    position: relative;
    z-index: 1;
    max-width: none;
    padding: 14px 22px 12px 22px;
}
.dashboard-hero-badge {
    display: none;
    position: absolute;
    top: 24px;
    right: 28px;
    width: 54px;
    height: 54px;
    border-radius: 14px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: rgba(255,255,255,0.52);
    backdrop-filter: blur(10px);
    border: 1px solid rgba(203,213,225,0.95);
    box-shadow: 0 10px 20px rgba(148,163,184,0.18);
}
.dashboard-hero-badge img {
    width: 38px;
    height: 38px;
    object-fit: contain;
}
.dashboard-hero-title {
    max-width: 100%;
    margin: 0 0 3px 0;
    font-size: clamp(1.25rem, 1.45vw, 1.65rem);
    line-height: 1.08;
    letter-spacing: -0.025em;
    font-weight: 900;
    color: #0f172a;
}
.dashboard-hero-subtitle {
    max-width: 960px;
    font-size: 0.82rem;
    line-height: 1.25;
    color: rgba(51,65,85,0.86);
    font-weight: 500;
}
.dashboard-hero-meta {
    display: none;
    margin-top: 8px;
    font-size: 0.76rem;
    color: rgba(71,85,105,0.88);
    font-weight: 600;
}
@media (max-width: 960px) {
    .dashboard-hero {
        min-height: 260px;
        border-radius: 26px;
    }
    .dashboard-hero-inner {
        padding: 42px 24px 24px 24px;
    }
    .dashboard-hero-badge {
        top: 42px;
        right: 24px;
        width: 66px;
        height: 66px;
        border-radius: 20px;
    }
    .dashboard-hero-badge img {
        width: 46px;
        height: 46px;
    }
}
@media (max-width: 640px) {
    .dashboard-hero {
        min-height: 230px;
    }
    .dashboard-hero-title {
        line-height: 1.02;
        max-width: calc(100% - 96px);
    }
}
</style>
""", unsafe_allow_html=True)

components.html(
    """
    <script>
    (function () {
        const win = window.parent;
        const doc = win.document;
        const KEY = "bodegas_sidebar_collapsed_v2";

        function collapsed() {
            try {
                return win.localStorage.getItem(KEY) === "1";
            } catch (err) {
                return false;
            }
        }

        function setCollapsed(value) {
            doc.body.classList.toggle("bodegas-sidebar-collapsed", value);
            try {
                win.localStorage.setItem(KEY, value ? "1" : "0");
            } catch (err) {}
            const sidebar = doc.querySelector('section[data-testid="stSidebar"], aside[data-testid="stSidebar"]');
            if (sidebar && !value) {
                sidebar.style.background = "linear-gradient(180deg, #031B34 0%, #021326 100%)";
                sidebar.style.visibility = "visible";
                sidebar.style.opacity = "1";
                sidebar.style.transform = "translateX(0)";
                sidebar.style.width = "252px";
                sidebar.style.minWidth = "252px";
                sidebar.style.maxWidth = "252px";
            }
            const btn = doc.getElementById("bodegas-sidebar-toggle");
            if (btn) {
                btn.innerHTML = value ? "☰" : "‹";
                btn.title = value ? "Abrir panel" : "Ocultar panel";
                btn.setAttribute("aria-label", btn.title);
            }
            win.dispatchEvent(new Event("resize"));
            setTimeout(function () { win.dispatchEvent(new Event("resize")); }, 180);
        }

        setCollapsed(collapsed());

        let btn = doc.getElementById("bodegas-sidebar-toggle");
        if (!btn) {
            btn = doc.createElement("button");
            btn.id = "bodegas-sidebar-toggle";
            btn.type = "button";
            btn.className = "bodegas-sidebar-toggle";
            doc.body.appendChild(btn);
        }
        btn.onclick = function () {
            setCollapsed(!doc.body.classList.contains("bodegas-sidebar-collapsed"));
        };
        setCollapsed(collapsed());
    })();
    </script>
    """,
    height=0,
)

components.html(
    """
    <script>
    (function () {
        const win = window.parent;
        const doc = win.document;
        const KEY = "bodegas_scroll_state_v3";

        function candidates() {
            const fixed = [
                win,
                doc.scrollingElement,
                doc.documentElement,
                doc.body,
                doc.querySelector("section.main"),
                doc.querySelector("section.main > div"),
                doc.querySelector('[data-testid="stAppViewContainer"]'),
                doc.querySelector('[data-testid="stAppViewContainer"] > div'),
                doc.querySelector('[data-testid="stVerticalBlock"]')
            ].filter(Boolean);
            const dynamic = Array.from(doc.querySelectorAll("body *")).filter(function (node) {
                try {
                    const style = win.getComputedStyle(node);
                    return node.scrollHeight > node.clientHeight + 8
                        && !["hidden", "clip"].includes(style.overflowY);
                } catch (err) {
                    return false;
                }
            });
            return Array.from(new Set(fixed.concat(dynamic)));
        }

        function currentY() {
            return Math.max.apply(null, candidates().map(function (node) {
                if (node === win) return win.scrollY || 0;
                return node.scrollTop || 0;
            }).concat([0]));
        }

        function readState() {
            try {
                return JSON.parse(win.localStorage.getItem(KEY) || "{}");
            } catch (err) {
                return {};
            }
        }

        function writeState(state) {
            try {
                win.localStorage.setItem(KEY, JSON.stringify(state));
            } catch (err) {}
        }

        function savePosition(protect) {
            const y = currentY();
            const previous = readState();
            const state = {
                y: y > 8 ? y : (previous.y || 0),
                protectUntil: protect ? Date.now() + 9000 : (previous.protectUntil || 0)
            };
            writeState(state);
        }

        function protectedState() {
            const state = readState();
            return state && state.y > 8 && state.protectUntil && Date.now() < state.protectUntil ? state : null;
        }

        function restorePosition(y) {
            candidates().forEach(function (node) {
                try {
                    if (node === win) {
                        win.__bodegasNativeScrollTo(0, y);
                    } else {
                        node.scrollTop = y;
                    }
                } catch (err) {}
            });
        }

        function isWidgetTarget(target) {
            return Boolean(
                target.closest('[data-testid="stSelectbox"]')
                || target.closest('[data-testid="stMultiSelect"]')
                || target.closest('[data-testid="stRadio"]')
                || target.closest('[data-testid="stSlider"]')
                || target.closest('[data-testid="stTextInput"]')
                || target.closest('[data-testid="stNumberInput"]')
                || target.closest('[data-testid="stDateInput"]')
                || target.closest('[data-testid="stCheckbox"]')
                || target.closest('[data-baseweb="select"]')
                || target.closest('[role="listbox"]')
                || target.closest('[role="option"]')
                || target.closest("input, textarea, select")
            );
        }

        if (!win.__bodegasScrollLockInstalled) {
            win.__bodegasScrollLockInstalled = true;
            win.__bodegasNativeScrollTo = win.scrollTo.bind(win);
            win.__bodegasNativeScroll = win.scroll.bind(win);

            win.scrollTo = function () {
                const state = protectedState();
                let requestedY = 0;
                if (arguments.length === 1 && typeof arguments[0] === "object") {
                    requestedY = Number(arguments[0].top || 0);
                } else if (arguments.length > 1) {
                    requestedY = Number(arguments[1] || 0);
                }
                if (state && requestedY <= 2) {
                    return win.__bodegasNativeScrollTo(0, state.y);
                }
                return win.__bodegasNativeScrollTo.apply(win, arguments);
            };
            win.scroll = win.scrollTo;

            let timer = null;
            function scheduleSave() {
                win.clearTimeout(timer);
                timer = win.setTimeout(function () { savePosition(false); }, 120);
            }

            win.addEventListener("scroll", scheduleSave, true);
            doc.addEventListener("scroll", scheduleSave, true);

            ["pointerdown", "mousedown", "touchstart", "focusin", "input", "change", "click"].forEach(function (eventName) {
                doc.addEventListener(eventName, function (event) {
                    if (isWidgetTarget(event.target)) savePosition(true);
                }, true);
            });
            doc.addEventListener("keydown", function (event) {
                if (isWidgetTarget(event.target)) savePosition(true);
            }, true);
        }

        const state = protectedState();
        if (state) {
            [0, 40, 100, 180, 320, 520, 800, 1200, 1800, 2600, 3800, 5600, 7600].forEach(function (delay) {
                win.setTimeout(function () { restorePosition(state.y); }, delay);
            });
        }
    })();
    </script>
    """,
    height=0,
)



# =========================
# Carga de datos
# =========================
CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSuoX_V5rYls-pBu7F3_VP2APS3FL7-eYbn9uDWUGJQZbxNfQTm9gRlyDlE69wWJjsDQpDzi2lt31Ak/pub?gid=1154929321&single=true&output=csv"
CAPEX_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSuoX_V5rYls-pBu7F3_VP2APS3FL7-eYbn9uDWUGJQZbxNfQTm9gRlyDlE69wWJjsDQpDzi2lt31Ak/pub?gid=1467494223&single=true&output=csv"

st.sidebar.markdown(
    f"""
    <div class="sidebar-brand">
        <div class="sidebar-brand-logo"><img src="{LOGO_URI}" alt="Logo"></div>
        <div>
            <div class="sidebar-brand-title">Bodegas<br>Balmaceda</div>
            <div class="sidebar-brand-sub">Panel financiero</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)
active_section = st.sidebar.radio(
    "Sección",
    options=section_options,
    label_visibility="collapsed",
)
st.sidebar.markdown('<div class="sidebar-bottom-spacer"></div>', unsafe_allow_html=True)
st.sidebar.header("⚙️ Controles")
if st.sidebar.button("🔄 Actualizar datos (limpiar caché)"):
    st.cache_data.clear()

@st.cache_data
def load_data(url: str) -> pd.DataFrame:
    df = pd.read_csv(url)
    df.columns = [str(c).strip() for c in df.columns]
    # Normalización y tipos
    df["Fecha"] = pd.to_datetime(df.get("Fecha"), errors="coerce", dayfirst=True)
    if "FECHA CONTABLE" in df.columns:
        df["FECHA CONTABLE"] = pd.to_datetime(df.get("FECHA CONTABLE"), errors="coerce", dayfirst=True)
    df["Monto"] = pd.to_numeric(
        df["Monto"].astype(str).str.replace(r"[^\d\.-]", "", regex=True),
        errors="coerce"
    )
    # Columnas esperadas
    for c in ["Obs", "CC1", "Sit", "Responsable", "Año", "Mes", "Esp", "CC", "FECHA CONTABLE", "Mes CONTABLE", "AÑO CONTABLE"]:
        if c not in df.columns:
            df[c] = pd.NA

    # Normalización fuerte de CC y Sit
    df["CC"]  = df["CC"].astype(str).str.strip().str.upper()
    df["Sit"] = df["Sit"].astype(str).str.strip().str.upper()
    df["Obs"] = df["Obs"].astype(str)
    df["CC1"] = df["CC1"].astype(str)

    df = df.dropna(subset=["Monto", "CC"])
    return df

df = load_data(CSV_URL)


@st.cache_data(show_spinner=False)
def load_capex_data(url: str) -> pd.DataFrame:
    df_capex = pd.read_csv(url)
    df_capex.columns = [str(c).strip() for c in df_capex.columns]
    df_capex = df_capex.dropna(axis=1, how="all")

    expected_cols = ["Año", "Periodo", "Situación", "CCCC", "Estado", "Monto"]
    for col in expected_cols:
        if col not in df_capex.columns:
            df_capex[col] = pd.NA
    df_capex = df_capex[expected_cols].copy()

    month_map = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "setiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    df_capex["Año"] = pd.to_numeric(df_capex["Año"], errors="coerce").astype("Int64")
    df_capex["Periodo"] = df_capex["Periodo"].astype(str).str.strip()
    df_capex["Mes_num"] = df_capex["Periodo"].str.lower().map(month_map)
    df_capex["Periodo_ref"] = pd.to_datetime(
        dict(
            year=df_capex["Año"].fillna(1900).astype(int),
            month=df_capex["Mes_num"].fillna(1).astype(int),
            day=1,
        ),
        errors="coerce",
    )
    df_capex.loc[df_capex["Año"].isna() | df_capex["Mes_num"].isna(), "Periodo_ref"] = pd.NaT
    df_capex["Monto"] = pd.to_numeric(
        df_capex["Monto"].astype(str).str.replace(r"[^\d\.-]", "", regex=True),
        errors="coerce",
    )
    for col in ["Situación", "CCCC", "Estado"]:
        df_capex[col] = df_capex[col].astype(str).str.strip()
    df_capex["Estado_norm"] = df_capex["Estado"].str.upper()
    df_capex["CCCC_norm"] = df_capex["CCCC"].str.upper()
    df_capex = df_capex.dropna(subset=["Año", "Periodo_ref", "Monto"])
    return df_capex


capex_df = load_capex_data(CAPEX_CSV_URL)

def _month_number(series) -> pd.Series:
    if series is None:
        return pd.Series(dtype="float64")
    return pd.to_numeric(
        series.astype(str).str.extract(r"(\d{1,2})", expand=False),
        errors="coerce",
    )


def _add_period_fields(
    df_out: pd.DataFrame,
    fecha_col: str,
    anio_col: str,
    mes_col: str,
    suffix: str = "",
) -> pd.DataFrame:
    idx = df_out.index
    fecha_src = df_out.get(fecha_col, pd.Series(pd.NaT, index=idx))
    anio_src = df_out.get(anio_col, pd.Series(pd.NA, index=idx))
    mes_src = df_out.get(mes_col, pd.Series(pd.NA, index=idx))

    fecha_dt = pd.to_datetime(fecha_src, errors="coerce", dayfirst=True)
    anio_sel = pd.to_numeric(anio_src, errors="coerce").fillna(fecha_dt.dt.year)
    mes_sel = _month_number(mes_src).reindex(idx).fillna(fecha_dt.dt.month)

    periodo_ref = pd.Series(pd.NaT, index=idx, dtype="datetime64[ns]")
    mask_periodo = anio_sel.notna() & mes_sel.notna()
    if mask_periodo.any():
        periodo_ref.loc[mask_periodo] = pd.to_datetime(
            dict(
                year=anio_sel.loc[mask_periodo].astype(int),
                month=mes_sel.loc[mask_periodo].astype(int),
                day=1,
            ),
            errors="coerce",
        )

    df_out[f"Fecha_dt{suffix}"] = fecha_dt
    df_out[f"Año_sel{suffix}"] = anio_sel
    df_out[f"Mes_sel{suffix}"] = mes_sel
    df_out[f"Periodo_ref{suffix}"] = periodo_ref
    return df_out


@st.cache_data(show_spinner=False)
def enrich_base_data(df_in: pd.DataFrame) -> pd.DataFrame:
    df_out = df_in.copy()
    df_out["CC_norm"] = df_out["CC"].astype(str).str.strip().str.upper()
    df_out["Sit_norm"] = df_out["Sit"].astype(str).str.strip().str.upper()
    df_out["Obs_text"] = df_out["Obs"].astype(str)
    df_out["CC1_text"] = df_out["CC1"].astype(str)
    df_out["Responsable_clean"] = df_out["Responsable"].astype(str).str.strip()
    df_out["Esp_num"] = pd.to_numeric(
        df_out["Esp"].astype(str).str.extract(r"(\d+)", expand=False),
        errors="coerce",
    )
    df_out = _add_period_fields(df_out, "Fecha", "Año", "Mes")
    df_out = _add_period_fields(df_out, "FECHA CONTABLE", "AÑO CONTABLE", "Mes CONTABLE", "_contable")
    return df_out


# SIN filtros (por ahora): usar todo el dataset
df_f = enrich_base_data(df)

# =========================
# Carga de Electricidad (Excel local o URL XLSX)
# =========================
ELECTRICIDAD_XLSX = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSm1GzFOATXqiYiGSKUZWZ5C2dE9nXv7bTtR3XmchtWhC9ZRMRK5OGDQfZhb624dA/pub?output=xlsx"

@st.cache_data
def load_electricidad(path_or_url: str) -> dict[str, pd.DataFrame]:
    if path_or_url.startswith(("http://", "https://")):
        sheets = pd.read_excel(path_or_url, sheet_name=None)
    else:
        p = Path(path_or_url)
        if not p.exists():
            return {}
        sheets = pd.read_excel(p, sheet_name=None)
    cleaned: dict[str, pd.DataFrame] = {}
    for name, df_sheet in sheets.items():
        if df_sheet is None:
            continue
        df_sheet = df_sheet.copy()
        if df_sheet.shape[1] == 0:
            continue
        df_sheet.columns = [str(c).strip() for c in df_sheet.columns]
        cleaned[name] = df_sheet
    return cleaned


@st.cache_data(show_spinner=False)
def load_electricidad_parsed(path_or_url: str) -> dict[str, dict[str, pd.DataFrame]]:
    sheets = load_electricidad(path_or_url)
    parsed: dict[str, dict[str, pd.DataFrame]] = {}
    for name, df_sheet in sheets.items():
        try:
            parsed_sheet = _parse_mes_sheet(df_sheet)
        except Exception:
            continue
        if parsed_sheet["inputs_bodega"].empty and parsed_sheet["liquidacion"].empty:
            continue
        parsed[name] = parsed_sheet
    return parsed


def _infer_date_column(df_in: pd.DataFrame) -> str | None:
    best_col = None
    best_ratio = 0.0
    for col in df_in.columns:
        s = pd.to_datetime(df_in[col], errors="coerce")
        ratio = s.notna().mean()
        if ratio > best_ratio and ratio >= 0.5:
            best_ratio = ratio
            best_col = col
    return best_col

def _pick_primary_numeric(df_in: pd.DataFrame) -> str | None:
    preferred = ["TOTAL", "MONTO", "VALOR", "COSTO", "KW", "KWH", "CONSUMO", "ENERGIA"]
    cols = [str(c) for c in df_in.columns]
    for p in preferred:
        for c in cols:
            if p in c.upper():
                return c
    num_df = df_in.apply(pd.to_numeric, errors="coerce")
    num_cols = [c for c in num_df.columns if num_df[c].notna().any()]
    return num_cols[0] if num_cols else None

def _clean_header_list(row_vals: list) -> list[str]:
    out = []
    for v in row_vals:
        if pd.isna(v):
            break
        out.append(str(v).strip())
    return out

def _dedupe_cols(cols: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            out.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out

def _fmt_cell(v, is_pct=False, is_currency=False, is_int=False):
    try:
        if pd.isna(v):
            return ""
    except Exception:
        if v is None:
            return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if is_int:
        try:
            return f"{float(v):,.0f}"
        except Exception:
            return str(v)
    if is_pct:
        try:
            return f"{float(v):.2%}"
        except Exception:
            return str(v)
    if is_currency:
        try:
            return f"${float(v):,.0f}"
        except Exception:
            return str(v)
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%H:%M")
        except Exception:
            return str(v)
    return str(v)

def _render_table(df_in: pd.DataFrame, header_bg="#1f4e78", header_fg="white",
                  row_alt="#f8f4e8", compact=True, int_cols=None, cur_cols_extra=None) -> None:
    columns = list(df_in.columns)
    pct_cols = {c for c in columns if "%" in str(c)}
    cur_cols = {c for c in columns if "$" in str(c) or "TOTAL" in str(c).upper()}
    int_cols = set(int_cols or [])
    cur_cols |= set(cur_cols_extra or [])

    rows_html = []
    for i, row in enumerate(df_in.itertuples(index=False, name=None)):
        tds = []
        for c, value in zip(columns, row):
            v = _fmt_cell(
                value,
                is_pct=(c in pct_cols),
                is_currency=(c in cur_cols),
                is_int=(c in int_cols),
            )
            tds.append(f"<td>{v}</td>")
        cls = "alt" if i % 2 == 0 else ""
        rows_html.append(f"<tr class='{cls}'>" + "".join(tds) + "</tr>")

    ths = "".join([f"<th>{c}</th>" for c in columns])
    table_html = f"""
    <div class="elec-table-wrap">
      <table class="elec-table">
        <thead><tr>{ths}</tr></thead>
        <tbody>{''.join(rows_html)}</tbody>
      </table>
    </div>
    """

    st.markdown(
        f"""
        <style>
        .elec-table-wrap {{ overflow-x: auto; }}
        .elec-table {{
            border-collapse: collapse;
            width: 100%;
            font-size: { "12px" if compact else "13px" };
        }}
        .elec-table th {{
            background: {header_bg};
            color: {header_fg};
            text-align: center;
            padding: 6px 8px;
            border: 1px solid #d5d5d5;
            font-weight: 700;
        }}
        .elec-table td {{
            padding: 6px 8px;
            border: 1px solid #d5d5d5;
            text-align: center;
            white-space: nowrap;
        }}
        .elec-table tr.alt td {{
            background: {row_alt};
        }}
        </style>
        {table_html}
        """,
        unsafe_allow_html=True,
    )

def _parse_mes_sheet(df_raw: pd.DataFrame) -> dict[str, pd.DataFrame]:
    if df_raw is None or df_raw.empty or df_raw.shape[1] == 0:
        empty_ig = pd.DataFrame(columns=["Parámetro", "Valor", "Unidad"])
        empty_bc = pd.DataFrame(columns=["Concepto", "Monto $"])
        return {
            "inputs_generales": empty_ig,
            "boleta_cge": empty_bc,
            "inputs_bodega": pd.DataFrame(),
            "liquidacion": pd.DataFrame(),
        }

    def _slice_cols(start: int, stop: int) -> pd.DataFrame:
        start = max(0, start)
        stop = min(df_raw.shape[1], stop)
        if start >= stop:
            return pd.DataFrame()
        return df_raw.iloc[:, start:stop].copy()

    def _pad_or_trim_columns(df_in: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
        df_out = df_in.copy()
        target = len(cols)
        if df_out.shape[1] < target:
            for i in range(target - df_out.shape[1]):
                df_out[f"__pad_{i}"] = pd.NA
        df_out = df_out.iloc[:, :target]
        df_out.columns = cols
        return df_out

    def _find_row_any(label: str) -> int | None:
        txt = df_raw.astype(str).apply(lambda s: s.str.strip())
        hits = txt.apply(lambda row: row.str.contains(label, case=False, na=False).any(), axis=1)
        if hits.any():
            return int(hits.idxmax())
        return None

    def _slice_until_blank(start_row: int, col_idx: int, col_slice: slice) -> pd.DataFrame:
        end = len(df_raw)
        for i in range(start_row, len(df_raw)):
            v = df_raw.iloc[i, col_idx] if col_idx < df_raw.shape[1] else None
            if pd.isna(v) or str(v).strip() == "":
                end = i
                break
            txt = str(v).upper()
            if "INPUTS POR BODEGA" in txt or "LIQUIDACIÓN POR BODEGA" in txt:
                end = i
                break
        return df_raw.iloc[start_row:end, col_slice].copy()

    # INPUTS GENERALES
    row_ig = _find_row_any("INPUTS GENERALES")
    if row_ig is not None:
        ig = _slice_until_blank(row_ig + 1, 0, slice(0, 3))
    else:
        ig = _slice_cols(0, 3).iloc[3:8].copy()
    ig = _pad_or_trim_columns(ig, ["Parámetro", "Valor", "Unidad"])
    ig = ig.dropna(how="all")

    # BOLETA CGE
    row_bc = _find_row_any("BOLETA CGE")
    if row_bc is not None:
        # +2 porque la fila inmediatamente inferior suele ser encabezado Concepto/Monto
        bc = _slice_until_blank(row_bc + 2, 5, slice(5, 7))
    else:
        bc = _slice_cols(5, 7).iloc[4:9].copy()
    bc = _pad_or_trim_columns(bc, ["Concepto", "Monto $"])
    bc = bc.dropna(how="all")

    def _find_row(label: str) -> int | None:
        col0 = df_raw.iloc[:, 0].astype(str)
        hits = col0.str.contains(label, case=False, na=False)
        if hits.any():
            return int(hits.idxmax())
        return None

    def _slice_section(start_row: int) -> pd.DataFrame:
        end = len(df_raw)
        col0 = df_raw.iloc[:, 0]
        for i in range(start_row, len(df_raw)):
            v = col0.iloc[i]
            if pd.isna(v):
                end = i
                break
            if isinstance(v, str) and ("LIQUIDACIÓN POR BODEGA" in v.upper() or "INPUTS POR BODEGA" in v.upper()):
                if i != start_row:
                    end = i
                    break
        return df_raw.iloc[start_row:end]

    # INPUTS POR BODEGA
    row_inputs = _find_row("INPUTS POR BODEGA")
    if row_inputs is not None:
        hdr_row = row_inputs + 1
        data_row = row_inputs + 2
        hdr_in = _clean_header_list(df_raw.iloc[hdr_row, 0:15].tolist())
        hdr_in = _dedupe_cols(hdr_in)
        if hdr_in:
            data_in = _slice_section(data_row).iloc[:, 0:len(hdr_in)].copy()
            data_in.columns = hdr_in
        else:
            data_in = pd.DataFrame()
    else:
        hdr_in = _clean_header_list(df_raw.iloc[11, 0:9].tolist()) if len(df_raw) > 11 else []
        hdr_in = _dedupe_cols(hdr_in)
        if hdr_in:
            data_in = df_raw.iloc[12:20, 0:len(hdr_in)].copy()
            data_in.columns = hdr_in
        else:
            data_in = pd.DataFrame()

    # LIQUIDACIÓN POR BODEGA
    row_liq = _find_row("LIQUIDACIÓN POR BODEGA")
    if row_liq is not None:
        hdr_row = row_liq + 1
        data_row = row_liq + 2
        hdr_liq = _clean_header_list(df_raw.iloc[hdr_row, 0:20].tolist())
        hdr_liq = _dedupe_cols(hdr_liq)
        if hdr_liq:
            data_liq = _slice_section(data_row).iloc[:, 0:len(hdr_liq)].copy()
            data_liq.columns = hdr_liq
        else:
            data_liq = pd.DataFrame()
    else:
        hdr_liq = _clean_header_list(df_raw.iloc[22, 0:15].tolist()) if len(df_raw) > 22 else []
        hdr_liq = _dedupe_cols(hdr_liq)
        if hdr_liq:
            data_liq = df_raw.iloc[23:31, 0:len(hdr_liq)].copy()
            data_liq.columns = hdr_liq
        else:
            data_liq = pd.DataFrame()

    return {
        "inputs_generales": ig,
        "boleta_cge": bc,
        "inputs_bodega": data_in,
        "liquidacion": data_liq,
    }

def _df_to_rl_table(
    df_in: pd.DataFrame,
    header_bg=colors.HexColor("#0F2942"),
    header_fg=colors.white,
    row_alt: colors.Color | None = None,
    grid_color=colors.HexColor("#D9E1EA"),
    max_width=520,
    font_size=7,
    min_col_width=42,
    max_col_width=165,
    cell_padding=4,
) -> Table:
    data = [[str(c) for c in df_in.columns]]
    for _, row in df_in.fillna("").astype(str).iterrows():
        data.append([str(v) for v in row.tolist()])

    ncols = len(df_in.columns)
    # Ajuste de ancho por contenido para evitar celdas sobredimensionadas.
    sample_rows = df_in.fillna("").astype(str).head(80)
    char_counts = []
    for c in df_in.columns:
        header_len = len(str(c))
        content_len = sample_rows[c].map(len).quantile(0.85) if not sample_rows.empty else 0
        est = max(header_len, int(content_len))
        char_counts.append(max(5, min(30, est)))
    raw_widths = [min(max_col_width, max(min_col_width, 5.2 * cc)) for cc in char_counts]
    total_raw = sum(raw_widths) or 1
    if total_raw > max_width:
        scale = max_width / total_raw
        col_widths = [max(min_col_width, w * scale) for w in raw_widths]
        # Si al aplicar mínimos se excede, se distribuye uniforme como fallback.
        if sum(col_widths) > max_width:
            col_widths = [max_width / max(1, ncols)] * ncols
    else:
        col_widths = raw_widths

    tbl = Table(data, hAlign="LEFT", colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), header_fg),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.35, grid_color),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), cell_padding),
        ("TOPPADDING", (0, 0), (-1, -1), cell_padding),
        ("LEFTPADDING", (0, 0), (-1, -1), cell_padding),
        ("RIGHTPADDING", (0, 0), (-1, -1), cell_padding),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    # Alineación por tipo de columna para una lectura más ejecutiva.
    for idx, c in enumerate(df_in.columns):
        cu = str(c).upper()
        if any(k in cu for k in ["$", "TOTAL", "IVA", "NETO", "KWH", "%"]):
            style_cmds.append(("ALIGN", (idx, 1), (idx, -1), "RIGHT"))
        else:
            style_cmds.append(("ALIGN", (idx, 1), (idx, -1), "LEFT"))
    if row_alt is not None:
        for r in range(1, len(data)):
            if (r - 1) % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, r), (-1, r), row_alt))
    tbl.setStyle(TableStyle(style_cmds))
    return tbl


def _compact_pdf_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    df = df_in.copy()
    alias = {
        "Bodega": "Bod.",
        "Remarcador": "Rem.",
        "Hora inicio jornada (Input)": "Hora ini. jor.",
        "Hora inicio horario especial": "Hora ini. esp.",
        "Factor horario especial (>18:00)": "Factor esp. >18",
        "Hora salida (Input)": "Hora salida",
        "Carga inductiva (SI/NO)": "Carga ind.",
        "Pago con atraso SI/NO": "Pago atraso",
        "% post-18 (calc)": "% post-18",
        "kWh post-18 (calc)": "kWh post-18",
        "kWh día (calc)": "kWh día",
        "$ Cargos Fijos": "$ C. Fijos",
        "% Cargos Fijos": "% C. Fijos",
        "TOTAL NETO $": "Total neto",
        "TOTAL c/IVA $": "Total c/IVA",
        "Intereses/Mora $": "Interés/Mora $",
    }
    renamed = {}
    for c in df.columns:
        c_txt = str(c)
        if c_txt in alias:
            renamed[c] = alias[c_txt]
            continue
        c_txt = (
            c_txt.replace(" (calc)", "")
            .replace(" (Input)", "")
            .replace(" (SI/NO)", "")
            .replace("REMARCADOR", "Rem.")
            .replace("REMARCADOR", "Rem.")
            .replace("  ", " ")
            .strip()
        )
        renamed[c] = c_txt
    return df.rename(columns=renamed)


def _pdf_section_banner(text: str, width: float, bg_hex="#0F2942", fg=colors.white) -> Table:
    banner = Table([[text]], colWidths=[width])
    banner.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg_hex)),
                ("TEXTCOLOR", (0, 0), (-1, -1), fg),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return banner

def _format_pdf_df(df_in: pd.DataFrame) -> pd.DataFrame:
    df = df_in.copy()
    for c in df.columns:
        cu = str(c).upper()
        if "HORA" in cu:
            df[c] = df[c].apply(lambda v: v.strftime("%H:%M") if hasattr(v, "strftime") else str(v))
            continue

        is_pct = "%" in cu
        is_cur = "$" in cu or "TOTAL" in cu or "IVA" in cu or "NETO" in cu
        is_kwh = "KWH" in cu

        if is_pct:
            def fmt_pct(v):
                try:
                    v = float(v)
                    if abs(v) <= 1.5:
                        v = v * 100.0
                    return f"{v:.2f}%"
                except Exception:
                    return str(v)
            df[c] = df[c].apply(fmt_pct)
        elif is_cur:
            def fmt_cur(v):
                try:
                    v = float(v)
                    return f"${v:,.0f}"
                except Exception:
                    return str(v)
            df[c] = df[c].apply(fmt_cur)
        elif is_kwh:
            def fmt_int(v):
                try:
                    return f"{float(v):,.0f}"
                except Exception:
                    return str(v)
            df[c] = df[c].apply(fmt_int)
        else:
            df[c] = df[c].apply(lambda v: str(v))
    return df


def build_detalle_movimientos_excel(
    df_in: pd.DataFrame,
    filtros: dict[str, str],
    kpis: dict[str, float],
    resumen_obs: pd.DataFrame | None = None,
    grafico_obs: pd.DataFrame | None = None,
) -> bytes:
    output = BytesIO()
    df_export = df_in.copy()
    if "Fecha" in df_export.columns:
        df_export["Fecha"] = pd.to_datetime(df_export["Fecha"], errors="coerce").dt.strftime("%Y-%m-%d")

    filtros_df = pd.DataFrame(
        [{"Campo": k, "Valor": v} for k, v in filtros.items()]
        + [{"Campo": k, "Valor": v} for k, v in kpis.items()]
    )
    resumen_obs_export = resumen_obs.copy() if resumen_obs is not None else pd.DataFrame()
    grafico_obs_export = grafico_obs.copy() if grafico_obs is not None else pd.DataFrame()
    if not grafico_obs_export.empty and "Periodo_chart" in grafico_obs_export.columns:
        grafico_obs_export["Periodo"] = pd.to_datetime(
            grafico_obs_export["Periodo_chart"],
            errors="coerce",
        ).dt.strftime("%Y-%m")
        chart_cols_export = [
            c for c in ["Periodo", "Registros", "Pagado", "No pagado", "Abono", "RESULTADO", "Deuda a la fecha"]
            if c in grafico_obs_export.columns
        ]
        grafico_obs_export = grafico_obs_export[chart_cols_export]

    try:
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df_export.to_excel(writer, index=False, sheet_name="detalle_filtrado")
            filtros_df.to_excel(writer, index=False, sheet_name="resumen")
            if not resumen_obs_export.empty:
                resumen_obs_export.to_excel(writer, index=False, sheet_name="resumen_obs")
            if not grafico_obs_export.empty:
                grafico_obs_export.to_excel(writer, index=False, sheet_name="grafico_obs")

            wb = writer.book
            head_fmt = wb.add_format({"bold": True, "bg_color": "#163A5F", "font_color": "#FFFFFF", "border": 1})
            money_fmt = wb.add_format({"num_format": "$#,##0"})
            ws = writer.sheets["detalle_filtrado"]

            for col_idx, col_name in enumerate(df_export.columns):
                ws.write(0, col_idx, col_name, head_fmt)
                values = df_export[col_name].astype(str) if not df_export.empty else pd.Series([str(col_name)])
                max_len = max([len(str(col_name))] + values.str.len().fillna(0).astype(int).tolist())
                width = min(max(max_len + 2, 10), 36)
                if col_name == "Monto":
                    ws.set_column(col_idx, col_idx, 14, money_fmt)
                else:
                    ws.set_column(col_idx, col_idx, width)

            if len(df_export.columns) > 0:
                ws.autofilter(0, 0, max(len(df_export), 1), len(df_export.columns) - 1)
                ws.freeze_panes(1, 0)

            ws_resumen = writer.sheets["resumen"]
            ws_resumen.set_column(0, 0, 24)
            ws_resumen.set_column(1, 1, 28)

            if not resumen_obs_export.empty:
                ws_obs = writer.sheets["resumen_obs"]
                for col_idx, col_name in enumerate(resumen_obs_export.columns):
                    ws_obs.write(0, col_idx, col_name, head_fmt)
                    values = resumen_obs_export[col_name].astype(str)
                    width = min(max([len(str(col_name))] + values.str.len().fillna(0).astype(int).tolist()) + 2, 38)
                    if col_name in {"Pagado", "No pagado", "RESULTADO", "Abono", "Deuda a la fecha", "TOTAL OBS (FILTRO)"}:
                        ws_obs.set_column(col_idx, col_idx, 16, money_fmt)
                    else:
                        ws_obs.set_column(col_idx, col_idx, max(width, 10))
                    if col_name == "RESULTADO":
                        result_fmt = wb.add_format({"num_format": "$#,##0", "bold": True, "bg_color": "#DBEAFE", "font_color": "#0F2D52"})
                        ws_obs.set_column(col_idx, col_idx, 16, result_fmt)
                ws_obs.freeze_panes(1, 0)
            if not grafico_obs_export.empty:
                ws_chart = writer.sheets["grafico_obs"]
                for col_idx, col_name in enumerate(grafico_obs_export.columns):
                    ws_chart.write(0, col_idx, col_name, head_fmt)
                    width = 18 if col_name != "Periodo" else 14
                    if col_name in {"Pagado", "No pagado", "Abono", "RESULTADO", "Deuda a la fecha"}:
                        ws_chart.set_column(col_idx, col_idx, width, money_fmt)
                    else:
                        ws_chart.set_column(col_idx, col_idx, width)
                ws_chart.freeze_panes(1, 0)
                if len(grafico_obs_export) > 0:
                    col_lookup = {c: i for i, c in enumerate(grafico_obs_export.columns)}
                    chart_bar = wb.add_chart({"type": "column"})
                    for col_name, color in [("Pagado", CHART_TEAL), ("No pagado", CHART_RED), ("Abono", CHART_GOLD)]:
                        if col_name in col_lookup:
                            chart_bar.add_series({
                                "name": ["grafico_obs", 0, col_lookup[col_name]],
                                "categories": ["grafico_obs", 1, col_lookup["Periodo"], len(grafico_obs_export), col_lookup["Periodo"]],
                                "values": ["grafico_obs", 1, col_lookup[col_name], len(grafico_obs_export), col_lookup[col_name]],
                                "fill": {"color": color},
                                "border": {"color": color},
                            })
                    if "Deuda a la fecha" in col_lookup:
                        chart_line = wb.add_chart({"type": "line"})
                        chart_line.add_series({
                            "name": ["grafico_obs", 0, col_lookup["Deuda a la fecha"]],
                            "categories": ["grafico_obs", 1, col_lookup["Periodo"], len(grafico_obs_export), col_lookup["Periodo"]],
                            "values": ["grafico_obs", 1, col_lookup["Deuda a la fecha"], len(grafico_obs_export), col_lookup["Deuda a la fecha"]],
                            "line": {"color": CHART_DARK, "width": 2.5},
                            "marker": {"type": "circle", "size": 6, "border": {"color": CHART_DARK}, "fill": {"color": CHART_DARK}},
                        })
                        chart_bar.combine(chart_line)
                    chart_bar.set_title({"name": "Evolución por año y mes según OBS"})
                    chart_bar.set_x_axis({"name": "Periodo"})
                    chart_bar.set_y_axis({"name": "Monto (CLP)", "num_format": "$#,##0"})
                    chart_bar.set_legend({"position": "bottom"})
                    chart_bar.set_size({"width": 860, "height": 380})
                    ws_chart.insert_chart("I2", chart_bar)
    except ModuleNotFoundError:
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, Reference

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="detalle_filtrado")
            filtros_df.to_excel(writer, index=False, sheet_name="resumen")
            if not resumen_obs_export.empty:
                resumen_obs_export.to_excel(writer, index=False, sheet_name="resumen_obs")
            if not grafico_obs_export.empty:
                grafico_obs_export.to_excel(writer, index=False, sheet_name="grafico_obs")

            ws = writer.sheets["detalle_filtrado"]
            fill = PatternFill(fill_type="solid", fgColor="163A5F")
            font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font

            for col_idx, col_name in enumerate(df_export.columns, start=1):
                values = df_export[col_name].astype(str) if not df_export.empty else pd.Series([str(col_name)])
                max_len = max([len(str(col_name))] + values.str.len().fillna(0).astype(int).tolist())
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 36)
                if col_name == "Monto":
                    for cell in ws[get_column_letter(col_idx)][1:]:
                        cell.number_format = "$#,##0"

            if len(df_export.columns) > 0:
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = "A2"

            ws_resumen = writer.sheets["resumen"]
            ws_resumen.column_dimensions["A"].width = 24
            ws_resumen.column_dimensions["B"].width = 28

            if not resumen_obs_export.empty:
                ws_obs = writer.sheets["resumen_obs"]
                for cell in ws_obs[1]:
                    cell.fill = fill
                    cell.font = font
                for col_idx, col_name in enumerate(resumen_obs_export.columns, start=1):
                    values = resumen_obs_export[col_name].astype(str)
                    width = min(max([len(str(col_name))] + values.str.len().fillna(0).astype(int).tolist()) + 2, 38)
                    ws_obs.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)
                    if col_name in {"Pagado", "No pagado", "RESULTADO", "Abono", "Deuda a la fecha", "TOTAL OBS (FILTRO)"}:
                        for cell in ws_obs[get_column_letter(col_idx)][1:]:
                            cell.number_format = "$#,##0"
                    if col_name == "RESULTADO":
                        result_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
                        result_font = Font(color="0F2D52", bold=True)
                        for cell in ws_obs[get_column_letter(col_idx)]:
                            cell.fill = result_fill
                            cell.font = result_font
                ws_obs.freeze_panes = "A2"
            if not grafico_obs_export.empty:
                ws_chart = writer.sheets["grafico_obs"]
                for cell in ws_chart[1]:
                    cell.fill = fill
                    cell.font = font
                for col_idx, col_name in enumerate(grafico_obs_export.columns, start=1):
                    ws_chart.column_dimensions[get_column_letter(col_idx)].width = 18 if col_name != "Periodo" else 14
                    if col_name in {"Pagado", "No pagado", "Abono", "RESULTADO", "Deuda a la fecha"}:
                        for cell in ws_chart[get_column_letter(col_idx)][1:]:
                            cell.number_format = "$#,##0"
                ws_chart.freeze_panes = "A2"
                col_lookup = {c: i + 1 for i, c in enumerate(grafico_obs_export.columns)}
                if len(grafico_obs_export) > 0 and {"Periodo", "Pagado", "No pagado", "Abono"}.issubset(col_lookup):
                    bar = BarChart()
                    data = Reference(
                        ws_chart,
                        min_col=col_lookup["Pagado"],
                        max_col=col_lookup["Abono"],
                        min_row=1,
                        max_row=len(grafico_obs_export) + 1,
                    )
                    cats = Reference(ws_chart, min_col=col_lookup["Periodo"], min_row=2, max_row=len(grafico_obs_export) + 1)
                    bar.add_data(data, titles_from_data=True)
                    bar.set_categories(cats)
                    bar.title = "Evolución por año y mes según OBS"
                    bar.y_axis.title = "Monto (CLP)"
                    bar.x_axis.title = "Periodo"
                    bar.height = 9
                    bar.width = 19
                    if "Deuda a la fecha" in col_lookup:
                        line = LineChart()
                        line_data = Reference(
                            ws_chart,
                            min_col=col_lookup["Deuda a la fecha"],
                            min_row=1,
                            max_row=len(grafico_obs_export) + 1,
                        )
                        line.add_data(line_data, titles_from_data=True)
                        line.set_categories(cats)
                        bar += line
                    ws_chart.add_chart(bar, "I2")

    output.seek(0)
    return output.getvalue()


def build_detalle_movimientos_pdf(
    df_in: pd.DataFrame,
    filtros: dict[str, str],
    kpis: dict[str, float],
    resumen_obs: pd.DataFrame | None = None,
    grafico_obs: pd.DataFrame | None = None,
) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        title="Detalle filtrable de movimientos",
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    base = ParagraphStyle("DetalleBase", parent=styles["Normal"], fontSize=7.2, leading=8.6, wordWrap="CJK")
    header = ParagraphStyle(
        "DetalleHeader",
        parent=base,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
        textColor=colors.white,
    )
    left = ParagraphStyle("DetalleLeft", parent=base, alignment=TA_LEFT)
    right = ParagraphStyle("DetalleRight", parent=base, alignment=TA_RIGHT)
    title_style = ParagraphStyle(
        "DetalleTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=19,
        textColor=colors.HexColor("#0F2D52"),
        alignment=TA_LEFT,
    )
    meta_style = ParagraphStyle(
        "DetalleMeta",
        parent=styles["Normal"],
        fontSize=8.2,
        leading=10,
        textColor=colors.HexColor("#334155"),
    )

    df_export = df_in.copy()
    if "Fecha" in df_export.columns:
        df_export["Fecha"] = pd.to_datetime(df_export["Fecha"], errors="coerce").dt.strftime("%Y-%m-%d")
    resumen_obs_export = resumen_obs.copy() if resumen_obs is not None else pd.DataFrame()
    grafico_obs_export = grafico_obs.copy() if grafico_obs is not None else pd.DataFrame()

    story = [
        Paragraph("Detalle filtrable de movimientos", title_style),
        Paragraph(f"Generado: {pd.Timestamp.now().strftime('%d-%m-%Y %H:%M')}", meta_style),
        Spacer(1, 8),
    ]

    resumen_items = [f"{k}: {v}" for k, v in filtros.items()]
    resumen_items += [f"{k}: {fmt_clp_largo(v) if isinstance(v, (int, float, np.integer, np.floating)) else v}" for k, v in kpis.items()]
    story.append(Paragraph(escape(" · ".join(resumen_items)), meta_style))
    story.append(Spacer(1, 10))

    if df_export.empty:
        story.append(Paragraph("No se registran movimientos relevantes para el universo filtrado.", meta_style))
        doc.build(story)
        output.seek(0)
        return output.getvalue()

    if not resumen_obs_export.empty:
        story.append(Paragraph("Resumen de montos por OBS", title_style))
        obs_cols = list(resumen_obs_export.columns)
        obs_data = [[Paragraph(escape(str(c)), header) for c in obs_cols]]
        for _, row in resumen_obs_export.iterrows():
            rendered_row = []
            for c, v in row.items():
                if pd.isna(v):
                    txt = ""
                elif c in {"Pagado", "No pagado", "RESULTADO", "Abono", "Deuda a la fecha", "TOTAL OBS (FILTRO)"}:
                    txt = f"${float(v):,.0f}"
                else:
                    txt = str(v)
                rendered_row.append(Paragraph(escape(txt), right if c in {"Pagado", "No pagado", "RESULTADO", "Abono", "Deuda a la fecha", "TOTAL OBS (FILTRO)"} else left))
            obs_data.append(rendered_row)
        resultado_col_idx = obs_cols.index("RESULTADO") if "RESULTADO" in obs_cols else None
        obs_weights = []
        for c in obs_cols:
            if c == "OBS":
                obs_weights.append(2.2)
            elif c == "Registros":
                obs_weights.append(0.8)
            else:
                obs_weights.append(1.15)
        obs_total_weight = sum(obs_weights) or 1
        obs_tbl = Table(
            obs_data,
            colWidths=[doc.width * (w / obs_total_weight) for w in obs_weights],
            repeatRows=1,
        )
        obs_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#163A5F")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for r in range(1, len(obs_data)):
            if (r - 1) % 2 == 0:
                obs_style.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#F8FAFC")))
            if resultado_col_idx is not None:
                obs_style.append(("BACKGROUND", (resultado_col_idx, r), (resultado_col_idx, r), colors.HexColor("#DBEAFE")))
                obs_style.append(("TEXTCOLOR", (resultado_col_idx, r), (resultado_col_idx, r), colors.HexColor("#0F2D52")))
                obs_style.append(("FONTNAME", (resultado_col_idx, r), (resultado_col_idx, r), "Helvetica-Bold"))
        if resultado_col_idx is not None:
            obs_style.append(("BACKGROUND", (resultado_col_idx, 0), (resultado_col_idx, 0), colors.HexColor("#0F2D52")))
            obs_style.append(("TEXTCOLOR", (resultado_col_idx, 0), (resultado_col_idx, 0), colors.white))
        obs_tbl.setStyle(TableStyle(obs_style))
        story.append(obs_tbl)
        story.append(Spacer(1, 12))

        if not grafico_obs_export.empty:
            story.append(Paragraph("Evolución por año y mes según OBS", title_style))
            try:
                fig_obs_pdf = build_obs_periodo_figure(grafico_obs_export)
                png_bytes = fig_obs_pdf.to_image(format="png", width=1180, height=520, scale=2)
                img_buf = BytesIO(png_bytes)
                story.append(RLImage(img_buf, width=doc.width, height=290))
                story.append(Spacer(1, 12))
            except Exception:
                matplotlib_rendered = False
                try:
                    import os
                    import tempfile

                    os.environ.setdefault("MPLCONFIGDIR", tempfile.gettempdir())
                    import matplotlib.pyplot as plt

                    chart_pdf = grafico_obs_export.copy()
                    chart_pdf["Periodo_label"] = (
                        chart_pdf["Periodo_txt"]
                        if "Periodo_txt" in chart_pdf.columns
                        else pd.to_datetime(chart_pdf["Periodo_chart"], errors="coerce").dt.strftime("%Y-%m")
                    )
                    x_pos = np.arange(len(chart_pdf))
                    width = 0.24
                    fig, ax = plt.subplots(figsize=(11.8, 4.9), dpi=180)
                    ax.bar(x_pos - width, chart_pdf["Pagado"], width, label="Pagado", color=CHART_TEAL)
                    ax.bar(x_pos, chart_pdf["No pagado"], width, label="No pagado", color=CHART_RED)
                    ax.bar(x_pos + width, chart_pdf["Abono"], width, label="Abono", color=CHART_GOLD)
                    ax.plot(
                        x_pos,
                        chart_pdf["Deuda a la fecha"],
                        label="Deuda a la fecha",
                        color=CHART_DARK,
                        linewidth=2.4,
                        marker="o",
                        markersize=4.8,
                    )
                    ax.axhline(0, color=CHART_GRAY, linewidth=0.8)
                    ax.set_title("Evolución por año y mes según OBS", loc="left", fontsize=12, weight="bold", color="#0F2D52")
                    ax.set_ylabel("Monto (CLP)")
                    tick_idx = [i for i in range(len(chart_pdf)) if i % 6 == 0]
                    if len(chart_pdf) > 1 and (len(chart_pdf) - 1) not in tick_idx:
                        tick_idx.append(len(chart_pdf) - 1)
                    ax.set_xticks(x_pos[tick_idx])
                    ax.set_xticklabels(chart_pdf["Periodo_label"].iloc[tick_idx], rotation=0, ha="center")
                    ax.grid(axis="y", color="#E2E8F0", linewidth=0.7)
                    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.14), ncol=4, frameon=False)
                    ax.yaxis.set_major_formatter(lambda v, _: f"${v:,.0f}")
                    fig.subplots_adjust(left=0.08, right=0.98, top=0.88, bottom=0.22)
                    img_buf = BytesIO()
                    fig.savefig(img_buf, format="png", bbox_inches="tight", facecolor="white")
                    plt.close(fig)
                    img_buf.seek(0)
                    story.append(RLImage(img_buf, width=doc.width, height=290))
                    story.append(Spacer(1, 12))
                    matplotlib_rendered = True
                except Exception:
                    matplotlib_rendered = False

                if not matplotlib_rendered:
                    chart_cols = [
                        c for c in ["Periodo_txt", "Registros", "Pagado", "No pagado", "Abono", "RESULTADO", "Deuda a la fecha"]
                        if c in grafico_obs_export.columns
                    ]
                    chart_data = [[Paragraph(escape("Periodo" if c == "Periodo_txt" else str(c)), header) for c in chart_cols]]
                    for _, row in grafico_obs_export[chart_cols].iterrows():
                        rendered_row = []
                        for c, v in row.items():
                            if pd.isna(v):
                                txt = ""
                            elif c in {"Pagado", "No pagado", "RESULTADO", "Abono", "Deuda a la fecha"}:
                                txt = f"${float(v):,.0f}"
                            else:
                                txt = str(v)
                            rendered_row.append(Paragraph(escape(txt), right if c != "Periodo_txt" else left))
                        chart_data.append(rendered_row)
                    chart_tbl = Table(chart_data, repeatRows=1)
                    chart_tbl.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#163A5F")),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]))
                    story.append(chart_tbl)
                    story.append(Spacer(1, 12))

        story.append(Paragraph("Detalle filtrado", title_style))
        story.append(Spacer(1, 6))

    data = [[Paragraph(escape(str(c)), header) for c in df_export.columns]]
    for _, row in df_export.iterrows():
        rendered_row = []
        for c, v in row.items():
            if pd.isna(v):
                txt = ""
            elif c == "Monto":
                try:
                    txt = f"${float(v):,.0f}"
                except Exception:
                    txt = str(v)
            else:
                txt = str(v)
            rendered_row.append(Paragraph(escape(txt), right if c == "Monto" else left))
        data.append(rendered_row)

    weights = []
    for c in df_export.columns:
        if c in {"Fecha", "Año", "Mes", "Esp", "CC", "Sit", "Monto"}:
            weights.append(1.0)
        elif c in {"Responsable", "CC1"}:
            weights.append(1.6)
        else:
            weights.append(2.4)
    total_weight = sum(weights) or 1
    col_widths = [doc.width * (w / total_weight) for w in weights]

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#163A5F")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for r in range(1, len(data)):
        if (r - 1) % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#F8FAFC")))
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    doc.build(story)
    output.seek(0)
    return output.getvalue()

@st.cache_data(show_spinner=False)
def build_electricidad_pdf(
    title: str,
    sel_months: list[str],
    sel_bodega: str,
    inputs_generales: pd.DataFrame,
    boleta_avg: pd.DataFrame,
    inputs_bodega: pd.DataFrame,
    liquidacion: pd.DataFrame,
    charts: list[dict],
) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        title=title,
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    styles = getSampleStyleSheet()
    story = []
    content_w = doc.width
    generated_at = pd.Timestamp.now().strftime("%d-%m-%Y %H:%M")

    title_style = ParagraphStyle(
        "ElecPdfTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=21,
        leading=24,
        textColor=colors.HexColor("#0B1F33"),
        alignment=TA_LEFT,
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "ElecPdfMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=12,
        textColor=colors.HexColor("#334155"),
        alignment=TA_LEFT,
    )
    section_title_style = ParagraphStyle(
        "ElecPdfSection",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#0B1F33"),
        alignment=TA_LEFT,
        spaceAfter=4,
    )

    def _draw_footer(canv, d):
        canv.saveState()
        y = 12
        canv.setStrokeColor(colors.HexColor("#D9E1EA"))
        canv.line(d.leftMargin, y + 8, A4[0] - d.rightMargin, y + 8)
        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.HexColor("#64748B"))
        canv.drawString(d.leftMargin, y, f"Generado: {generated_at}")
        canv.drawRightString(A4[0] - d.rightMargin, y, f"Página {canv.getPageNumber()}")
        canv.restoreState()

    main_title = title or "Liquidación Eléctrica por Bodega"
    story.append(Paragraph(main_title, title_style))
    story.append(Paragraph(f"Meses: {', '.join(sel_months)} &nbsp;&nbsp;|&nbsp;&nbsp; Bodega: {sel_bodega}", meta_style))
    story.append(Spacer(1, 8))

    # Primer bloque en dos columnas: Inputs Generales (izq) + Boleta CGE (der)
    split_w = (content_w - 8) / 2
    left_block = [
        _pdf_section_banner("INPUTS GENERALES", split_w),
        Spacer(1, 3),
        Paragraph("Parámetros base para liquidación eléctrica.", section_title_style),
        _df_to_rl_table(
            _format_pdf_df(_compact_pdf_columns(inputs_generales)),
            header_bg=colors.HexColor("#123A5A"),
            row_alt=colors.HexColor("#F8FBFF"),
            max_width=split_w,
            font_size=7.2,
            cell_padding=3,
        ),
    ]
    right_block = [
        _pdf_section_banner("BOLETA CGE (PROMEDIO SEGÚN MESES)", split_w),
        Spacer(1, 3),
        Paragraph("Resumen consolidado por concepto de facturación.", section_title_style),
        _df_to_rl_table(
            _format_pdf_df(_compact_pdf_columns(boleta_avg)),
            header_bg=colors.HexColor("#123A5A"),
            row_alt=colors.HexColor("#F8FBFF"),
            max_width=split_w,
            font_size=7.2,
            cell_padding=3,
        ),
    ]
    intro_grid = Table([[left_block, right_block]], colWidths=[split_w, split_w], hAlign="LEFT")
    intro_grid.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(intro_grid)
    story.append(Spacer(1, 8))

    story.append(_pdf_section_banner("INPUTS POR BODEGA (REMARCADOR + HORARIO EFECTIVO)", content_w))
    story.append(Spacer(1, 3))
    story.append(
        _df_to_rl_table(
            _format_pdf_df(_compact_pdf_columns(inputs_bodega)),
            header_bg=colors.HexColor("#123A5A"),
            row_alt=colors.HexColor("#FFFDF7"),
            max_width=content_w,
            font_size=6.8,
            min_col_width=36,
            max_col_width=120,
            cell_padding=2.5,
        )
    )
    story.append(Spacer(1, 6))

    story.append(_pdf_section_banner("LIQUIDACIÓN POR BODEGA (ASIGNACIÓN DE COSTOS + CRITERIO HORARIO)", content_w))
    story.append(Spacer(1, 3))
    story.append(
        _df_to_rl_table(
            _format_pdf_df(_compact_pdf_columns(liquidacion)),
            header_bg=colors.HexColor("#123A5A"),
            row_alt=colors.HexColor("#F6FCF8"),
            max_width=content_w,
            font_size=6.8,
            min_col_width=36,
            max_col_width=120,
            cell_padding=2.5,
        )
    )
    story.append(Spacer(1, 8))

    if charts:
        import matplotlib.pyplot as plt
        story.append(_pdf_section_banner("GRÁFICO — LIQUIDACIÓN POR BODEGA (DISTRIBUCIÓN DE COSTOS)", content_w))
        story.append(Spacer(1, 4))
        for ch in charts:
            try:
                df_plot = ch["df"].copy()
                x_col = ch["x_col"]
                title = ch.get("title", "")
                cols_costos = ch["cols_costos"]
                col_total = ch["col_total"]
                palette = ch["palette"]

                chart_type = ch.get("chart_type", "stacked")
                if chart_type == "donut":
                    fig, ax = plt.subplots(figsize=(7.2, 4.2), dpi=150)
                    vals = pd.to_numeric(df_plot[cols_costos].sum(numeric_only=True), errors="coerce").fillna(0)
                    raw_values = [float(v) for v in vals.tolist()]
                    # Matplotlib pie no admite negativos; se recortan para evitar fallas.
                    values = [max(0.0, v) for v in raw_values]
                    labels = cols_costos
                    pie_colors = [palette.get(c, "#94a3b8") for c in labels]
                    total_vals = pd.to_numeric(df_plot[col_total], errors="coerce").fillna(0)
                    total_num = float(total_vals.sum())
                    bodega_label = str(df_plot.iloc[0, 1]) if not df_plot.empty else ""
                    mes_label = str(df_plot["Mes"].iloc[0]) if ("Mes" in df_plot.columns and not df_plot.empty) else ""
                    if sum(values) > 0:
                        ax.pie(
                            values,
                            labels=None,
                            colors=pie_colors,
                            startangle=90,
                            counterclock=False,
                            wedgeprops=dict(width=0.44, edgecolor="white"),
                            autopct=lambda p: f"{p:.1f}%" if p >= 3 else "",
                            pctdistance=0.8,
                        )
                        center_txt = f"{bodega_label}\n{mes_label}\nTotal c/IVA\n${total_num:,.0f}"
                        ax.text(0, 0, center_txt, ha="center", va="center", fontsize=11, weight="bold", color="#111827")
                        ax.axis("equal")
                        ax.legend(labels, loc="upper center", bbox_to_anchor=(0.5, 1.06), ncol=5, fontsize=7, frameon=False)
                        if title:
                            ax.set_title(title)
                        plt.tight_layout()
                    else:
                        # Sin base positiva para dona: fallback robusto a barras.
                        y_pos = np.arange(len(labels))
                        ax.barh(y_pos, raw_values, color=pie_colors)
                        ax.set_yticks(y_pos)
                        ax.set_yticklabels(labels, fontsize=8)
                        ax.axvline(0, color="#94a3b8", linewidth=1)
                        ax.grid(axis="x", color="#E2E8F0", linewidth=0.8)
                        center_txt = f"{bodega_label} · {mes_label} · Total c/IVA: ${total_num:,.0f}"
                        ax.set_title(center_txt, fontsize=10, color="#111827")
                        plt.tight_layout()
                else:
                    fig, ax1 = plt.subplots(figsize=(7.2, 4.2), dpi=150)
                    bottom = None
                    for c in cols_costos:
                        vals = pd.to_numeric(df_plot[c], errors="coerce").fillna(0)
                        ax1.bar(df_plot[x_col], vals, label=c, bottom=bottom, color=palette.get(c, "#94a3b8"))
                        bottom = vals if bottom is None else bottom + vals

                    ax1.set_ylabel("Costo (CLP)")
                    ax1.tick_params(axis="x", rotation=0)

                    ax2 = ax1.twinx()
                    total_vals = pd.to_numeric(df_plot[col_total], errors="coerce").fillna(0)
                    ax2.plot(df_plot[x_col], total_vals, color="#111827", marker="o", linewidth=2, label="TOTAL c/IVA $")
                    ax2.set_ylabel("Total c/IVA (CLP)")

                    handles1, labels1 = ax1.get_legend_handles_labels()
                    handles2, labels2 = ax2.get_legend_handles_labels()
                    ax1.legend(handles1 + handles2, labels1 + labels2, loc="upper center",
                               bbox_to_anchor=(0.5, 1.15), ncol=3, fontsize=7, frameon=False)
                    if title:
                        ax1.set_title(title)
                    plt.tight_layout()

                img_buf = BytesIO()
                fig.savefig(img_buf, format="png", dpi=220, bbox_inches="tight", facecolor="white")
                plt.close(fig)
                img_buf.seek(0)
                story.append(RLImage(img_buf, width=content_w, height=270))
                story.append(Spacer(1, 8))
            except Exception as e:
                raise RuntimeError(f"No fue posible renderizar el gráfico en el PDF: {e}")

    doc.build(story, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    pdf = buf.getvalue()
    buf.close()
    return pdf

# =========================
# Utilidades de formato
# =========================
def fmt_clp_largo(v: float) -> str:
    return f"${v:,.0f}"

def fmt_short(v: float) -> str:
    v = float(v)
    av = abs(v)
    if av >= 1_000_000:
        return f"${v/1_000_000:.1f}M"
    if av >= 1_000:
        return f"${v/1_000:.0f}k"
    return f"${v:,.0f}"


def build_obs_periodo_figure(chart_periodo: pd.DataFrame):
    import plotly.graph_objects as go

    fig_obs_periodo = go.Figure()
    fig_obs_periodo.add_trace(
        go.Scatter(
            x=chart_periodo["Periodo_chart"],
            y=chart_periodo["Deuda a la fecha"],
            mode="markers",
            name="Resumen",
            marker=dict(size=18, color="rgba(0,0,0,0)"),
            showlegend=False,
            customdata=chart_periodo[["Hover"]] if "Hover" in chart_periodo.columns else None,
            hovertemplate="<b>%{x|%b %Y}</b><br>%{customdata[0]}<extra></extra>",
        )
    )
    series_obs = [
        ("Pagado", "rgba(94,151,145,0.42)"),
        ("No pagado", "rgba(248,113,113,0.48)"),
        ("Abono", "rgba(148,163,184,0.26)"),
    ]
    for col_name, color in series_obs:
        fig_obs_periodo.add_trace(
            go.Bar(
                x=chart_periodo["Periodo_chart"],
                y=chart_periodo[col_name],
                name=col_name,
                marker=dict(color=color, line=dict(color="rgba(255,255,255,0)", width=0)),
                opacity=0.76,
                hoverinfo="skip",
            )
        )

    fig_obs_periodo.add_trace(
        go.Scatter(
            x=chart_periodo["Periodo_chart"],
            y=chart_periodo["Deuda a la fecha"],
            name="Tendencia deuda",
            mode="lines",
            line=dict(color="rgba(15,45,82,0.18)", width=10, shape="spline"),
            hoverinfo="skip",
            showlegend=False,
        )
    )
    fig_obs_periodo.add_trace(
        go.Scatter(
            x=chart_periodo["Periodo_chart"],
            y=chart_periodo["Deuda a la fecha"],
            name="Deuda a la fecha",
            mode="lines+markers",
            line=dict(color="#0F2D52", width=5.2, shape="spline"),
            marker=dict(size=10, color="#0F2D52", line=dict(width=2.4, color="white")),
            hoverinfo="skip",
        )
    )
    if not chart_periodo.empty:
        last_obs = chart_periodo.iloc[-1]
        debt_delta = float(chart_periodo["Deuda a la fecha"].diff().fillna(0).iloc[-1]) if len(chart_periodo) > 1 else 0.0
        debt_trend_txt = "Presión al alza" if debt_delta > 0 else ("Presión a la baja" if debt_delta < 0 else "Presión estable")
        debt_badge_bg = "#FEE2E2" if debt_delta > 0 else ("#DCFCE7" if debt_delta < 0 else "#E2E8F0")
        debt_badge_fg = "#B42318" if debt_delta > 0 else ("#047857" if debt_delta < 0 else "#334155")
        fig_obs_periodo.add_annotation(
            x=last_obs["Periodo_chart"],
            y=last_obs["Deuda a la fecha"],
            text=f"<b>{debt_trend_txt}</b><br>{float(last_obs['Deuda a la fecha']):,.0f}",
            showarrow=True,
            arrowhead=2,
            ax=44,
            ay=-34,
            bgcolor=debt_badge_bg,
            bordercolor="rgba(15,23,42,0.10)",
            borderwidth=1,
            borderpad=5,
            font=dict(size=11, color=debt_badge_fg),
        )

    fig_obs_periodo.add_hline(y=0, line_width=1, line_color=CHART_GRAY)
    fig_obs_periodo.update_layout(
        height=430,
        barmode="relative",
        margin=dict(l=20, r=20, t=36, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(248,250,252,0.92)",
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="left",
            x=0,
        ),
        xaxis=dict(
            title="Período",
            tickformat="%b %Y",
            showgrid=False,
        ),
        yaxis=dict(
            title="Monto (CLP)",
            tickprefix="$",
            separatethousands=True,
            gridcolor="rgba(180,190,210,0.14)",
            zeroline=False,
        ),
        hovermode="x unified",
    )
    return fig_obs_periodo


def card_finanza(titulo, valor, color_hex, subtitulo="Indicador financiero clave", etiqueta="Indicador", size="md"):
    size_map = {
        "lg": ("42px", "kpi-card kpi-card-lg"),
        "md": ("34px", "kpi-card kpi-card-md"),
        "sm": ("28px", "kpi-card kpi-card-sm"),
        "stack": ("32px", "kpi-card kpi-card-stack"),
        "top": ("30px", "kpi-card kpi-card-top"),
        "risk": ("26px", "kpi-card kpi-card-risk"),
    }
    value_size, card_class = size_map.get(size, size_map["md"])
    subtitle_html = f'<div class="kpi-sub">{subtitulo}</div>' if subtitulo else ""
    eyebrow_html = f'<div class="kpi-eyebrow">{etiqueta}</div>' if etiqueta else ""
    return f"""
    <div class="{card_class}" style="--accent:{color_hex};">
        {eyebrow_html}
        <div class="kpi-title">{titulo}</div>
        <div class="kpi-value" style="font-size:{value_size};">{valor}</div>
        {subtitle_html}
    </div>
    """


def kpi_resumen_panel(titulo, subtitulo, items):
    rows = []
    for item in items:
        rows.append(
            (
                f'<div class="kpi-summary-row">'
                f'<div class="kpi-summary-label">{item["label"]}</div>'
                f'<div class="kpi-summary-meta">{item["meta"]}</div>'
                f'<div class="kpi-summary-value">{item["value"]}</div>'
                f'</div>'
            )
        )
    return (
        f'<div class="kpi-summary-card">'
        f'<div class="kpi-summary-eyebrow">{titulo}</div>'
        f'<div class="kpi-summary-title">{subtitulo}</div>'
        f'<div class="kpi-summary-head">'
        f'<div>Indicador</div>'
        f'<div>Descripción</div>'
        f'<div>Valor</div>'
        f'</div>'
        f'<div class="kpi-summary-list">{"".join(rows)}</div>'
        f'</div>'
    )


def kpi_resumen_obs_panel(titulo, subtitulo, items):
    rows = []
    for item in items:
        resultado_class = item.get("resultado_class", "neutral")
        row_extra_class = item.get("row_class", "")
        rows.append(
            (
                f'<div class="kpi-summary-row kpi-summary-row-obs {row_extra_class}">'
                f'<div class="kpi-summary-label">{item["label"]}</div>'
                f'<div class="kpi-summary-meta">{item["meta"]}</div>'
                f'<div class="kpi-summary-value">{item["pagado"]}</div>'
                f'<div class="kpi-summary-value">{item["no_pagado"]}</div>'
                f'<div class="kpi-summary-value kpi-summary-resultado {resultado_class}">{item["resultado"]}</div>'
                f'<div class="kpi-summary-value">{item["abono"]}</div>'
                f'<div class="kpi-summary-value">{item["pendiente_deuda"]}</div>'
                f'</div>'
            )
        )
    return (
        f'<div class="kpi-summary-card kpi-summary-card-obs">'
        f'<div class="kpi-summary-eyebrow">{titulo}</div>'
        f'<div class="kpi-summary-title">{subtitulo}</div>'
        f'<div class="kpi-summary-head kpi-summary-head-obs">'
        f'<div>OBS</div>'
        f'<div>Descripción</div>'
        f'<div>Pagado</div>'
        f'<div>No pagado</div>'
        f'<div>RESULTADO</div>'
        f'<div>Abono</div>'
        f'<div>Deuda a la fecha</div>'
        f'</div>'
        f'<div class="kpi-summary-list">{"".join(rows)}</div>'
        f'</div>'
    )


def section_heading(icono: str, titulo: str, subtitulo: str = "", weight_class: str = "section-heading-title") -> str:
    subtitle_html = f'<div class="section-heading-sub">{subtitulo}</div>' if subtitulo else ""
    return (
        f'<div class="section-heading-wrap">'
        f'<div class="{weight_class}">{icono} {titulo}</div>'
        f'{subtitle_html}'
        f'</div>'
    )


def _download_anchor_html(label: str, css_class: str) -> str:
    return (
        f'<a id="overview-download-report" class="{css_class}" href="#" '
        f'role="button">{label}</a>'
    )


def tab_header(
    titulo: str,
    subtitulo: str = "",
    show_download: bool = True,
    download_html: str | None = None,
) -> str:
    subtitle_html = f'<div class="tab-title-sub">{subtitulo}</div>' if subtitulo else ""
    download_action = (
        download_html
        if download_html is not None
        else '<div class="tab-action tab-action-primary">⇩ Descargar reporte</div>'
    )
    actions = (
        '<div class="tab-actions">'
        f'{download_action}'
        '</div>'
        if show_download
        else ""
    )
    return (
        '<div class="tab-title-row">'
        f'<div><div class="tab-title-main">{titulo}</div>{subtitle_html}</div>'
        f'{actions}'
        '</div>'
    )

# =========================
# KPI base (cálculos comunes)
# =========================
CAPEX_FALLBACK = 151_834_571
CAPEX = float(capex_df["Monto"].sum()) if not capex_df.empty else float(CAPEX_FALLBACK)

@st.cache_data(show_spinner=False)
def compute_base_kpis(df_in: pd.DataFrame, capex: float) -> dict[str, float]:
    mask_canon = (
        df_in["CC_norm"].eq("INGRESO") &
        df_in["CC1_text"].str.strip().str.lower().eq("arriendo") &
        df_in["Obs_text"].str.strip().str.lower().eq("canon mensual")
    )
    ingresos_canon = df_in.loc[mask_canon, "Monto"].sum()
    cobertura_capex = ingresos_canon / capex if capex else 0.0

    total_pagado = df_in.loc[df_in["Sit_norm"].eq("PAGADO"), "Monto"].sum()
    mask_abono_obs = df_in["Obs_text"].str.contains(r"\babono_*\b", case=False, na=False)
    total_abonos = df_in.loc[mask_abono_obs, "Monto"].sum()
    saldo_cuenta = total_pagado + total_abonos

    mask_ingreso = df_in["CC_norm"].eq("INGRESO")
    mask_egreso = df_in["CC_norm"].eq("EGRESO")
    mask_sit_pagado = df_in["Sit_norm"].eq("PAGADO")
    mask_sit_abono = df_in["Sit_norm"].str.startswith("ABONO")
    mask_sueldo_accionista = df_in["Obs_text"].str.contains(
        r"sueldos?\s+accion(?:ista|ita)s?|sueldo\s+adolfo\s+orellana|sueldo\s+jonathan\s+mac[-\s]*kay",
        case=False,
        na=False,
        regex=True,
    )

    ingresos_kpi = df_in.loc[mask_ingreso & (mask_sit_pagado | mask_sit_abono), "Monto"].sum()
    egresos_kpi = df_in.loc[mask_egreso & mask_sit_pagado, "Monto"].sum()
    utilidad_operativa = ingresos_kpi + egresos_kpi
    margen_neto = (utilidad_operativa / ingresos_kpi) if ingresos_kpi else 0.0
    total_sueldos_accionistas = df_in.loc[mask_sueldo_accionista, "Monto"].sum()
    utilidad_sobre_capex = (abs(total_sueldos_accionistas) / capex) if capex else 0.0

    df_egresos_mes = df_in.loc[mask_egreso & mask_sit_pagado & df_in["Periodo_ref"].notna(), ["Periodo_ref", "Monto"]]
    if not df_egresos_mes.empty:
        egreso_mensual_promedio = (
            df_egresos_mes.groupby("Periodo_ref", as_index=False)["Monto"]
            .sum()["Monto"]
            .abs()
            .mean()
        )
    else:
        egreso_mensual_promedio = 0.0

    cobertura_egresos = (saldo_cuenta / egreso_mensual_promedio) if egreso_mensual_promedio else 0.0

    no_pagado_total = df_in.loc[mask_ingreso & df_in["Sit_norm"].eq("NO PAGADO"), "Monto"].sum()
    abonos_total = df_in.loc[mask_ingreso & df_in["Sit_norm"].str.startswith("ABONO"), "Monto"].sum()
    if abonos_total == 0:
        abonos_total = df_in.loc[
            mask_ingreso & df_in["Obs_text"].str.contains(r"\babono\b", case=False, na=False),
            "Monto",
        ].sum()

    cuentas_por_cobrar_neto = no_pagado_total - abonos_total
    total_egresos_por_pagar = df_in.loc[
        mask_egreso & df_in["Sit_norm"].eq("NO PAGADO"),
        "Monto",
    ].sum()

    deuda_fin_ejercicio = (
        df_in.loc[df_in["Sit_norm"].eq("NO PAGADO"), "Monto"].sum()
        - total_abonos
    )
    posicion_neta = saldo_cuenta - abs(deuda_fin_ejercicio) if deuda_fin_ejercicio < 0 else saldo_cuenta + deuda_fin_ejercicio

    return {
        "ingresos_canon": float(ingresos_canon),
        "cobertura_capex": float(cobertura_capex),
        "total_pagado": float(total_pagado),
        "total_abonos": float(total_abonos),
        "saldo_cuenta": float(saldo_cuenta),
        "ingresos_kpi": float(ingresos_kpi),
        "egresos_kpi": float(egresos_kpi),
        "utilidad_operativa": float(utilidad_operativa),
        "margen_neto": float(margen_neto),
        "total_sueldos_accionistas": float(total_sueldos_accionistas),
        "utilidad_sobre_capex": float(utilidad_sobre_capex),
        "egreso_mensual_promedio": float(egreso_mensual_promedio),
        "cobertura_egresos": float(cobertura_egresos),
        "no_pagado_total": float(no_pagado_total),
        "abonos_total": float(abonos_total),
        "pct_cobranza": float((abonos_total / no_pagado_total) if no_pagado_total else 0.0),
        "cuentas_por_cobrar_neto": float(cuentas_por_cobrar_neto),
        "total_egresos_por_pagar": float(total_egresos_por_pagar),
        "deuda_fin_ejercicio": float(deuda_fin_ejercicio),
        "posicion_neta": float(posicion_neta),
        "balance_kpi": float(saldo_cuenta),
    }


base_kpis = compute_base_kpis(df_f, CAPEX)
ingresos_canon = base_kpis["ingresos_canon"]
cobertura_capex = base_kpis["cobertura_capex"]
total_pagado = base_kpis["total_pagado"]
total_abonos = base_kpis["total_abonos"]
saldo_cuenta = base_kpis["saldo_cuenta"]
ingresos_kpi = base_kpis["ingresos_kpi"]
egresos_kpi = base_kpis["egresos_kpi"]
utilidad_operativa = base_kpis["utilidad_operativa"]
margen_neto = base_kpis["margen_neto"]
total_sueldos_accionistas = base_kpis["total_sueldos_accionistas"]
utilidad_sobre_capex = base_kpis["utilidad_sobre_capex"]
egreso_mensual_promedio = base_kpis["egreso_mensual_promedio"]
cobertura_egresos = base_kpis["cobertura_egresos"]
no_pagado_total = base_kpis["no_pagado_total"]
abonos_total = base_kpis["abonos_total"]
pct_cobranza = base_kpis["pct_cobranza"]
cuentas_por_cobrar_neto = base_kpis["cuentas_por_cobrar_neto"]
total_egresos_por_pagar = base_kpis["total_egresos_por_pagar"]
deuda_fin_ejercicio = base_kpis["deuda_fin_ejercicio"]
posicion_neta = base_kpis["posicion_neta"]
balance_kpi = base_kpis["balance_kpi"]

# =========================
# Estilos HTML para KPIs
# =========================
st.markdown("""
    <style>
    .kpi-card {
        position: relative;
        min-height: 168px;
        padding: 20px 22px 20px 24px;
        border-radius: 28px;
        border: 1px solid #d9e3ef;
        background: linear-gradient(135deg, #f8fafc 0%, #ffffff 56%, #f5f8fc 100%);
        box-shadow: 0 16px 34px rgba(15, 23, 42, 0.06);
        overflow: hidden;
    }
    .kpi-card::before {
        content: "";
        position: absolute;
        inset: 0 auto 0 0;
        width: 8px;
        background: linear-gradient(180deg, #eef4fb 0%, #f7fafd 100%);
        border-radius: 28px 0 0 28px;
    }
    .kpi-card::after {
        content: "";
        position: absolute;
        top: 18px;
        right: 18px;
        width: 12px;
        height: 12px;
        border-radius: 999px;
        background: #eef2f7;
        opacity: 1;
    }
    .kpi-card-lg {
        min-height: 176px;
    }
    .kpi-card-md {
        min-height: 168px;
        padding: 20px 22px 20px 24px;
    }
    .kpi-card-sm {
        min-height: 146px;
        padding: 18px 20px 18px 22px;
    }
    .kpi-card-stack {
        min-height: 170px;
        padding: 20px 22px 20px 24px;
    }
    .kpi-card-top {
        min-height: 238px;
        padding: 20px 22px 20px 24px;
        display: flex;
        flex-direction: column;
        box-sizing: border-box;
    }
    .kpi-card-top .kpi-title {
        min-height: 74px;
    }
    .kpi-card-top .kpi-sub {
        min-height: 80px;
    }
    .kpi-card-risk {
        min-height: 182px;
        padding: 20px 22px 20px 24px;
        display: flex;
        flex-direction: column;
        box-sizing: border-box;
    }
    .kpi-card-risk .kpi-title {
        min-height: 56px;
    }
    .kpi-card-risk .kpi-sub {
        min-height: 48px;
    }
    .kpi-eyebrow {
        font-size: 12px;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 0.14em;
        margin-bottom: 12px;
        opacity: 0.95;
        color: #64748b;
    }
    .kpi-title {
        font-size: 17px;
        line-height: 1.2;
        color: #1f2937;
        font-weight: 800;
        margin-bottom: 12px;
        max-width: 82%;
    }
    .kpi-value {
        display: inline-block;
        font-weight: 900;
        line-height: 1.05;
        letter-spacing: -0.03em;
        font-variant-numeric: tabular-nums;
        margin-bottom: 10px;
        color: #0f172a;
        padding: 6px 0 2px 0;
    }
    .kpi-sub {
        font-size: 14px;
        color: #667085;
        font-weight: 500;
        max-width: 88%;
    }
    .kpi-card-md .kpi-title, .kpi-card-sm .kpi-title, .kpi-card-stack .kpi-title {
        margin-bottom: 8px;
    }
    .kpi-card-sm .kpi-eyebrow, .kpi-card-stack .kpi-eyebrow {
        margin-bottom: 8px;
    }
    .kpi-card-sm .kpi-sub, .kpi-card-stack .kpi-sub {
        font-size: 12.5px;
        line-height: 1.25;
    }
    .kpi-summary-card {
        min-height: 100%;
        padding: 20px 24px;
        border-radius: 28px;
        border: 1px solid #d7dfeb;
        background:
            radial-gradient(circle at top right, rgba(127,166,162,0.08) 0%, transparent 28%),
            linear-gradient(135deg, #f7f9fc 0%, #ffffff 55%, #f4f7fb 100%);
        box-shadow: 0 18px 38px rgba(15, 23, 42, 0.08);
    }
    .kpi-summary-eyebrow {
        font-size: 12px;
        font-weight: 700;
        letter-spacing: 0.12em;
        text-transform: uppercase;
        color: #64748b;
        margin-bottom: 8px;
    }
    .kpi-summary-title {
        font-size: 14px;
        line-height: 1.2;
        font-weight: 500;
        color: #64748b;
        margin-bottom: 18px;
    }
    .kpi-summary-head {
        display: grid;
        grid-template-columns: 1.15fr 1fr auto;
        gap: 12px;
        padding: 0 0 12px 0;
        border-top: 1px solid #e6edf5;
        border-bottom: 1px solid #e6edf5;
        color: #64748b;
        font-size: 12px;
        font-weight: 800;
        letter-spacing: 0.08em;
        text-transform: uppercase;
    }
    .kpi-summary-list {
        padding-top: 2px;
    }
    .kpi-summary-row {
        display: grid;
        grid-template-columns: 1.15fr 1fr auto;
        gap: 12px;
        align-items: center;
        padding: 17px 0;
        border-bottom: 1px solid #e6edf5;
    }
    .kpi-summary-row:last-child {
        border-bottom: none;
        padding-bottom: 2px;
    }
    .kpi-summary-label {
        font-size: 15px;
        font-weight: 700;
        color: #0f172a;
        line-height: 1.15;
    }
    .kpi-summary-meta {
        font-size: 14px;
        color: #64748b;
        font-weight: 500;
        line-height: 1.2;
    }
    .kpi-summary-value {
        font-size: 15px;
        font-weight: 800;
        color: #0f172a;
        text-align: right;
        font-variant-numeric: tabular-nums;
        white-space: nowrap;
        padding-left: 14px;
    }
    .kpi-summary-compact .kpi-summary-card {
        padding: 14px 18px;
        border-radius: 20px;
        box-shadow: 0 10px 24px rgba(15, 23, 42, 0.06);
        margin: 6px 0 12px 0;
    }
    .kpi-summary-compact .kpi-summary-eyebrow {
        font-size: 11px;
        margin-bottom: 4px;
    }
    .kpi-summary-compact .kpi-summary-title {
        font-size: 12.5px;
        margin-bottom: 10px;
    }
    .kpi-summary-compact .kpi-summary-head {
        padding-bottom: 8px;
        font-size: 11px;
    }
    .kpi-summary-compact .kpi-summary-row {
        padding: 10px 0;
    }
    .kpi-summary-compact .kpi-summary-label,
    .kpi-summary-compact .kpi-summary-value {
        font-size: 13.5px;
    }
    .kpi-summary-compact .kpi-summary-meta {
        font-size: 12.5px;
    }
    .kpi-summary-card-obs {
        overflow-x: auto;
    }
    .kpi-summary-card-obs .kpi-summary-list {
        max-height: 380px;
        overflow-y: auto;
        overscroll-behavior: contain;
        padding-right: 8px;
    }
    .kpi-summary-head-obs,
    .kpi-summary-row-obs {
        grid-template-columns: minmax(160px, 1.35fr) minmax(82px, 0.65fr) minmax(90px, 0.78fr) minmax(96px, 0.84fr) minmax(96px, 0.84fr) minmax(84px, 0.72fr) minmax(112px, 0.92fr);
    }
    .kpi-summary-head-obs > div:nth-child(n+3),
    .kpi-summary-row-obs > div:nth-child(n+3) {
        text-align: right;
    }
    .kpi-summary-row-obs .kpi-summary-resultado {
        border-radius: 999px;
        padding: 6px 10px;
        font-size: 14px;
        font-weight: 950;
        letter-spacing: -0.01em;
        border: 1px solid rgba(148, 163, 184, 0.28);
        box-shadow:
            inset 0 1px 0 rgba(255,255,255,0.72),
            inset 0 -1px 0 rgba(15,23,42,0.06),
            0 4px 12px rgba(15,23,42,0.045);
        justify-self: end;
        min-width: 106px;
        text-align: center !important;
    }
    .kpi-summary-row-obs .kpi-summary-resultado.resultado-positive {
        background: linear-gradient(135deg, #DCFCE7 0%, #F0FDF4 100%);
        color: #047857;
        border-color: rgba(34,197,94,0.34);
    }
    .kpi-summary-row-obs .kpi-summary-resultado.resultado-negative {
        background: linear-gradient(135deg, #FEE2E2 0%, #FFF7F7 100%);
        color: #B42318;
        border-color: rgba(248,113,113,0.38);
    }
    .kpi-summary-row-obs .kpi-summary-resultado.resultado-neutral {
        background: linear-gradient(135deg, #E2E8F0 0%, #F8FAFC 100%);
        color: #334155;
        border-color: rgba(148,163,184,0.34);
    }
    .kpi-summary-row-obs.resultado-total-row {
        background: linear-gradient(90deg, rgba(239,246,255,0.86) 0%, rgba(255,255,255,0) 100%);
        border-radius: 10px;
        padding-left: 8px;
        padding-right: 8px;
    }
    .kpi-summary-row-obs.resultado-total-row .kpi-summary-label {
        color: #0F2D52;
        font-weight: 950;
    }
    .kpi-summary-row-obs.resultado-total-row .kpi-summary-resultado {
        transform: scale(1.03);
        box-shadow:
            inset 0 1px 0 rgba(255,255,255,0.8),
            inset 0 -1px 0 rgba(15,23,42,0.08),
            0 7px 18px rgba(15,23,42,0.08);
    }
    .kpi-summary-compact .kpi-summary-row-obs {
        min-width: 900px;
    }
    .kpi-summary-compact .kpi-summary-head-obs {
        min-width: 900px;
    }
    .section-heading-wrap {
        margin: 22px 0 14px 0;
    }
    .section-heading-title {
        font-size: clamp(2rem, 2.2vw, 2.75rem);
        line-height: 1.02;
        letter-spacing: -0.045em;
        font-weight: 900;
        color: #0f172a;
    }
    .section-heading-title-soft {
        font-size: 22px;
        line-height: 1.18;
        letter-spacing: -0.018em;
        font-weight: 900;
        color: #081735;
    }
    .section-heading-sub {
        margin-top: 8px;
        font-size: 1rem;
        line-height: 1.45;
        color: #64748b;
        font-weight: 500;
    }
    .tab-title-row {
        display:flex;
        justify-content:space-between;
        align-items:flex-start;
        margin:-61px 0 14px 0;
        gap:16px;
    }
    .tab-title-main {
        color:#081735;
        font-size:28px;
        font-weight:900;
        letter-spacing:-0.025em;
        line-height:1;
    }
    .tab-title-sub {
        margin-top:7px;
        color:#475569;
        font-size:14px;
        font-weight:600;
    }
    .tab-actions {
        display:flex;
        gap:8px;
        flex-wrap:wrap;
        justify-content:flex-end;
    }
    .tab-action {
        height:38px;
        display:flex;
        align-items:center;
        justify-content:center;
        border-radius:7px;
        border:1px solid #dbe3ee;
        background:#ffffff;
        color:#0f1f3d;
        padding:0 13px;
        font-size:12px;
        font-weight:800;
        box-shadow:0 8px 18px rgba(15,23,42,0.04);
        white-space:nowrap;
        text-decoration:none;
    }
    .tab-action-primary {
        background:#0B3A86;
        border-color:#0B3A86;
        color:#ffffff !important;
        text-decoration:none !important;
    }
    .tab-action-primary:visited,
    .tab-action-primary:hover,
    .tab-action-primary:active,
    .ie-action-primary,
    .ie-action-primary:visited,
    .ie-action-primary:hover,
    .ie-action-primary:active,
    .risk-pro-action-primary,
    .risk-pro-action-primary:visited,
    .risk-pro-action-primary:hover,
    .risk-pro-action-primary:active {
        color:#ffffff !important;
        text-decoration:none !important;
    }
    @media print {
        @page {
            size: A4 landscape;
            margin: 7mm;
        }
        html, body, main, .stApp {
            background:#ffffff !important;
            overflow: visible !important;
        }
        header[data-testid="stHeader"],
        div[data-testid="stToolbar"],
        div[data-testid="stStatusWidget"],
        div[data-testid="stDecoration"],
        section[data-testid="stSidebar"],
        aside[data-testid="stSidebar"],
        .bodegas-sidebar-toggle,
        .tab-actions,
        #MainMenu,
        footer {
            display:none !important;
            visibility:hidden !important;
            width:0 !important;
            height:0 !important;
        }
        .block-container,
        div[data-testid="stMainBlockContainer"],
        div[data-testid="block-container"] {
            max-width:none !important;
            width:100% !important;
            padding:0 !important;
            margin:0 !important;
        }
        div[data-testid="stAppViewContainer"],
        body.bodegas-sidebar-collapsed [data-testid="stAppViewContainer"] {
            margin-left:0 !important;
            padding-left:0 !important;
        }
        .tab-title-row {
            margin:0 0 8px 0 !important;
        }
        .asset-kpi-grid,
        .asset-card,
        .asset-bottom-card {
            break-inside:avoid;
            page-break-inside:avoid;
        }
        .js-plotly-plot,
        .plot-container,
        .svg-container {
            break-inside:avoid;
            page-break-inside:avoid;
        }
    }
    </style>
""", unsafe_allow_html=True)

# =========================================================
# 🏠 TAB 1: VISIÓN GENERAL
# =========================================================
if active_section == "🏠 Overview Ejecutivo Legacy":
    import plotly.graph_objects as go

    data_src = df_f.copy()
    data_src["Periodo_ref"] = pd.to_datetime(data_src["Periodo_ref"], errors="coerce")

    def _growth_12m(mask: pd.Series, abs_values: bool = False) -> tuple[float, float, float]:
        df_month = data_src.loc[mask & data_src["Periodo_ref"].notna(), ["Periodo_ref", "Monto"]].copy()
        df_month["Monto"] = pd.to_numeric(df_month["Monto"], errors="coerce")
        df_month = df_month.dropna(subset=["Periodo_ref", "Monto"])
        if df_month.empty:
            return 0.0, 0.0, 0.0
        df_month["Periodo_ref"] = df_month["Periodo_ref"].dt.to_period("M").dt.to_timestamp()
        monthly = df_month.groupby("Periodo_ref")["Monto"].sum().sort_index()
        if abs_values:
            monthly = monthly.abs()
        full_months = pd.date_range(monthly.index.min(), monthly.index.max(), freq="MS")
        monthly = monthly.reindex(full_months, fill_value=0)
        last_12m = float(monthly.iloc[-12:].sum())
        prev_12m = float(monthly.iloc[-24:-12].sum()) if len(monthly) >= 24 else 0.0
        growth = (last_12m / prev_12m - 1) if prev_12m else 0.0
        return growth, last_12m, prev_12m

    mask_ingresos_trend = data_src["CC_norm"].eq("INGRESO") & data_src["Sit_norm"].isin(["PAGADO", "ABONO"])
    mask_egresos_trend = data_src["CC_norm"].eq("EGRESO") & data_src["Sit_norm"].eq("PAGADO")
    mask_canon_home = (
        data_src["CC_norm"].eq("INGRESO")
        & data_src["CC1_text"].str.strip().str.lower().eq("arriendo")
        & data_src["Obs_text"].str.strip().str.lower().eq("canon mensual")
    )
    ingresos_12m_growth, ingresos_ultimos_12m, _ = _growth_12m(mask_ingresos_trend)
    canon_12m_growth, canon_ultimos_12m, _ = _growth_12m(mask_canon_home)
    egresos_12m_growth, egresos_ultimos_12m, _ = _growth_12m(mask_egresos_trend, abs_values=True)

    max_trend_period = data_src.loc[data_src["Periodo_ref"].notna(), "Periodo_ref"].max()
    last_12m_start = max_trend_period - pd.DateOffset(months=11) if pd.notna(max_trend_period) else pd.NaT
    last_12m_mask = data_src["Periodo_ref"].between(last_12m_start, max_trend_period) if pd.notna(last_12m_start) else pd.Series(False, index=data_src.index)
    canon_last_12m = data_src.loc[last_12m_mask & mask_canon_home].copy()
    top_concentration_txt = "sin concentración relevante"
    if not canon_last_12m.empty and canon_ultimos_12m:
        top_3_canon = (
            canon_last_12m.groupby("Responsable_clean")["Monto"]
            .sum()
            .sort_values(ascending=False)
            .head(3)
            .sum()
        )
        top_concentration_txt = f"top 3 arrendatarios concentran {top_3_canon / canon_ultimos_12m:.1%} del canon 12M"

    egresos_last_12m = data_src.loc[last_12m_mask & data_src["CC_norm"].eq("EGRESO")].copy()
    pressure_txt = "egresos sin presión destacada"
    if not egresos_last_12m.empty:
        egresos_last_12m["Monto_abs"] = pd.to_numeric(egresos_last_12m["Monto"], errors="coerce").abs()
        pressure_by_cc1 = egresos_last_12m.groupby("CC1_text")["Monto_abs"].sum().sort_values(ascending=False)
        if not pressure_by_cc1.empty and pressure_by_cc1.iloc[0] > 0:
            pressure_txt = f"mayor presión: {escape(str(pressure_by_cc1.index[0]))} ({fmt_clp_largo(float(pressure_by_cc1.iloc[0]))})"

    df_canon_home = data_src.loc[mask_canon_home, ["Año", "Monto"]].copy()
    df_canon_home["Año"] = pd.to_numeric(df_canon_home["Año"], errors="coerce")
    df_canon_home["Monto"] = pd.to_numeric(df_canon_home["Monto"], errors="coerce")
    df_canon_home = df_canon_home.dropna(subset=["Año", "Monto"])
    if not df_canon_home.empty:
        df_canon_home = (
            df_canon_home.groupby("Año", as_index=False)["Monto"]
            .sum()
            .sort_values("Año")
            .rename(columns={"Monto": "Canon anual"})
        )
        df_canon_home["MA3"] = df_canon_home["Canon anual"].rolling(window=3, min_periods=1).mean()
        df_canon_home["YoY"] = df_canon_home["Canon anual"].pct_change().replace([np.inf, -np.inf], 0).fillna(0)
        df_canon_home["Acumulado"] = df_canon_home["Canon anual"].cumsum()
        canon_ultimo = float(df_canon_home["Canon anual"].iloc[-1])
    else:
        df_canon_home = pd.DataFrame(columns=["Año", "Canon anual", "MA3", "YoY", "Acumulado"])
        canon_ultimo = 0.0

    health_score = 100
    diagnosis_penalties = []
    if posicion_neta < 0:
        health_score -= 28
        diagnosis_penalties.append(("Posición neta negativa", "-28 pts"))
    if margen_neto < 0:
        health_score -= 22
        diagnosis_penalties.append(("Margen neto negativo", "-22 pts"))
    if cobertura_egresos < 1:
        health_score -= 22
        diagnosis_penalties.append(("Cobertura caja < 1 mes", "-22 pts"))
    if cobertura_capex < 1:
        health_score -= 18
        diagnosis_penalties.append(("CAPEX recuperado < 100%", "-18 pts"))
    health_score = int(max(0, min(100, health_score)))
    estado_txt = "RIESGO" if health_score < 45 else ("ATENCIÓN" if health_score < 70 else "SALUDABLE")
    estado_color = "#DC2626" if health_score < 45 else ("#F59E0B" if health_score < 70 else "#059669")
    runway_txt = f"{cobertura_egresos:.1f} meses de runway"

    st.markdown(
        """
        <style>
        .tab-title-row {
            margin:-68px 0 8px 0;
        }
        .asset-toolbar {
            display:flex;
            justify-content:flex-end;
            gap:10px;
            margin:0 0 7px 0;
        }
        .asset-chip {
            min-height:30px;
            display:flex;
            align-items:center;
            gap:8px;
            padding:0 14px;
            border-radius:6px;
            border:1px solid #dbe3ee;
            background:#ffffff;
            color:#0f1f3d;
            font-size:12px;
            font-weight:700;
            box-shadow:0 8px 18px rgba(15,23,42,0.04);
        }
        .asset-kpi-grid {
            display:grid;
            grid-template-columns: repeat(5, minmax(0, 1fr));
            gap:10px;
            margin: 10px 0 10px 0;
        }
        .asset-kpi {
            min-height:116px;
            border-radius:14px;
            border:1px solid rgba(219,227,238,0.72);
            background:linear-gradient(135deg, #ffffff 0%, var(--soft) 100%);
            padding:12px 14px 10px 14px;
            display:flex;
            flex-direction:column;
            justify-content:space-between;
            box-shadow:0 4px 18px rgba(15,23,42,0.045);
        }
        .asset-kpi-main {
            display:grid;
            grid-template-columns:32px 1fr;
            gap:8px;
            align-items:start;
        }
        .asset-kpi-icon {
            width:28px;
            height:28px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:var(--halo);
            color:var(--accent);
            font-weight:900;
        }
        .asset-kpi-icon svg {
            width:16px;
            height:16px;
        }
        .asset-kpi-title {
            color:#0f1f3d;
            font-size:10px;
            font-weight:900;
            margin-bottom:3px;
        }
        .asset-kpi-value {
            color:var(--accent);
            font-size:21px;
            line-height:1.05;
            font-weight:900;
            letter-spacing:0;
            white-space:nowrap;
        }
        .asset-kpi-spark {
            width:100%;
            height:24px;
            margin-top:7px;
            display:block;
        }
        .asset-kpi-spark path,
        .asset-kpi-spark polyline {
            vector-effect:non-scaling-stroke;
        }
        .asset-kpi-badge {
            display:inline-block;
            margin-top:6px;
            padding:3px 6px;
            border-radius:5px;
            background:var(--badge-bg);
            color:var(--badge-fg);
            font-size:9px;
            font-weight:900;
        }
        .asset-kpi-note {
            grid-column:2;
            color:#475569;
            font-size:9px;
            font-weight:700;
            align-self:end;
            margin-top:3px;
        }
        .asset-card {
            border:1px solid rgba(219,227,238,0.72);
            border-radius:14px;
            background:#ffffff;
            padding:16px;
            box-shadow:0 4px 18px rgba(15,23,42,0.045);
            min-height:100%;
        }
        .asset-bottom-card {
            height:380px;
            box-sizing:border-box;
            display:flex;
            flex-direction:column;
        }
        .asset-card-title {
            color:#081735;
            font-size:14px;
            font-weight:900;
            margin-bottom:9px;
        }
        .asset-summary-list {
            display:flex;
            flex-direction:column;
            gap:8px;
            flex:1;
            justify-content:space-between;
        }
        .asset-summary-row {
            display:grid;
            grid-template-columns:30px minmax(0, 1fr) auto;
            gap:8px;
            align-items:center;
            border:1px solid rgba(229,235,243,0.82);
            border-radius:10px;
            padding:7px 9px;
            background:#fbfdff;
        }
        .asset-summary-icon {
            width:23px;
            height:23px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:var(--soft);
            color:var(--accent);
            font-size:11px;
            font-weight:950;
        }
        .asset-summary-label {
            color:#0f1f3d;
            font-size:11px;
            line-height:1.15;
            font-weight:900;
        }
        .asset-summary-sub {
            margin-top:3px;
            color:#64748b;
            font-size:9.5px;
            font-weight:700;
        }
        .asset-summary-value {
            color:var(--accent, #081735);
            font-size:11px;
            font-weight:950;
            text-align:right;
            white-space:nowrap;
        }
        .asset-summary-pct {
            color:#0f2d52;
            font-size:9.4px;
            font-weight:850;
            margin-top:3px;
        }
        .trend-grid {
            display:grid;
            grid-template-columns: repeat(3, minmax(0, 1fr));
            gap:6px;
        }
        .trend-mini {
            border:1px solid rgba(226,232,240,0.9);
            border-radius:10px;
            padding:7px;
            background:linear-gradient(135deg, #ffffff 0%, var(--soft) 100%);
            min-height:66px;
        }
        .trend-mini-title {
            color:#0f1f3d;
            font-size:10.2px;
            line-height:1.15;
            font-weight:900;
        }
        .trend-mini-value {
            margin-top:5px;
            font-size:17px;
            line-height:1;
            font-weight:950;
            color:var(--trend);
        }
        .trend-mini-sub {
            margin-top:4px;
            color:#64748b;
            font-size:9.5px;
            font-weight:700;
        }
        .asset-callout {
            margin-top:10px;
            border-radius:8px;
            background:#ecfdf3;
            border:1px solid #ccebd8;
            color:#14532d;
            padding:7px 10px;
            font-size:9.3px;
            line-height:1.25;
            font-weight:750;
        }
        .trend-salary-plan {
            margin-top:auto;
            border:1px solid #e2e8f0;
            border-radius:8px;
            background:#fbfdff;
            padding:10px 10px;
        }
        .trend-salary-head {
            display:flex;
            justify-content:space-between;
            gap:8px;
            align-items:center;
            color:#081735;
            font-size:10px;
            line-height:1.1;
            font-weight:950;
            margin-bottom:5px;
        }
        .trend-salary-head strong {
            color:#475569;
            font-size:9px;
            font-weight:850;
            white-space:nowrap;
        }
        .trend-salary-labels,
        .trend-salary-row {
            display:grid;
            grid-template-columns:.55fr 1fr 1fr;
            gap:6px;
            align-items:center;
        }
        .trend-salary-labels {
            color:#64748b;
            font-size:8.4px;
            font-weight:900;
            text-transform:uppercase;
            padding-bottom:3px;
            border-bottom:1px solid #e5ebf3;
        }
        .trend-salary-row {
            color:#0f1f3d;
            font-size:9.4px;
            font-weight:850;
            padding:3px 0;
            border-bottom:1px solid #eef2f7;
        }
        .trend-salary-row:last-of-type {
            border-bottom:0;
        }
        .trend-salary-row strong {
            color:#0B3A86;
            font-size:9.8px;
            font-weight:950;
        }
        .trend-salary-note {
            margin-top:4px;
            color:#475569;
            font-size:8.7px;
            line-height:1.25;
            font-weight:750;
        }
        .asset-diagnosis-card {
            border:1px solid rgba(219,227,238,0.72);
            border-radius:14px;
            background:#ffffff;
            padding:16px;
            box-shadow:0 4px 18px rgba(15,23,42,0.045);
            min-height:100%;
        }
        .asset-diagnosis-grid {
            display:grid;
            grid-template-columns:1.1fr .9fr;
            gap:10px;
            align-items:center;
        }
        .asset-diagnosis-list {
            display:flex;
            flex-direction:column;
            gap:8px;
            color:#0f1f3d;
            font-size:11px;
            font-weight:850;
        }
        .asset-diagnosis-list span {
            color:#64748b;
            font-size:10px;
            font-weight:700;
        }
        div[data-testid="stElementContainer"]:has(iframe[height="384"]),
        div[data-testid="stIFrame"]:has(iframe[height="384"]) {
            margin-top:-14px !important;
        }
        iframe[height="384"] {
            display:block;
        }
        @media (max-width: 1300px) {
            .asset-kpi-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    overview_download_html = _download_anchor_html(
        "⇩ Descargar reporte",
        "tab-action tab-action-primary",
    )

    def asset_icon_svg(kind: str) -> str:
        icons = {
            "up": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.6" stroke-linecap="round" stroke-linejoin="round"><path d="M7 17L17 7"/><path d="M9 7h8v8"/></svg>',
            "down": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.6" stroke-linecap="round" stroke-linejoin="round"><path d="M7 7l10 10"/><path d="M17 9v8H9"/></svg>',
            "wallet": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"><path d="M19 7V6a2 2 0 0 0-2-2H5a3 3 0 0 0 0 6h14a2 2 0 0 1 2 2v4a2 2 0 0 1-2 2H5a3 3 0 0 1-3-3V7"/><path d="M16 14h.01"/></svg>',
            "capex": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"><path d="M3 21h18"/><path d="M6 21V9l6-4 6 4v12"/><path d="M9 21v-6h6v6"/></svg>',
            "percent": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round"><path d="M19 5L5 19"/><circle cx="7.5" cy="7.5" r="2.5"/><circle cx="16.5" cy="16.5" r="2.5"/></svg>',
            "ratio": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M4 19V5"/><path d="M4 19h16"/><path d="M8 15l3-3 3 2 5-7"/></svg>',
            "equal": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.8" stroke-linecap="round"><path d="M6 9h12"/><path d="M6 15h12"/></svg>',
        }
        return icons.get(kind, icons["equal"])

    def asset_sparkline(values, accent: str) -> str:
        vals = pd.Series(values).replace([np.inf, -np.inf], np.nan).dropna().astype(float).tail(18).tolist()
        if len(vals) < 2:
            vals = [0.0, 0.0]
        lo, hi = min(vals), max(vals)
        span = hi - lo if hi != lo else 1.0
        denom = max(len(vals) - 1, 1)
        points = []
        for i, val in enumerate(vals):
            x = 2 + (i / denom) * 106
            y = 22 - ((val - lo) / span) * 18
            points.append(f"{x:.1f},{y:.1f}")
        return (
            f'<svg class="asset-kpi-spark" viewBox="0 0 112 26" preserveAspectRatio="none" aria-hidden="true">'
            f'<path d="M2 22 H110" stroke="rgba(100,116,139,0.16)" stroke-width="1"/>'
            f'<polyline points="{" ".join(points)}" fill="none" stroke="{accent}" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"/>'
            f'</svg>'
        )

    posicion_spark = [0, posicion_neta * 0.55, posicion_neta]
    caja_spark = [egreso_mensual_promedio, balance_kpi, balance_kpi + utilidad_operativa]
    capex_spark = [0, ingresos_canon, CAPEX]
    margen_spark = [0, margen_neto * 0.5, margen_neto]
    cobertura_spark = [0, cobertura_egresos * 0.55, cobertura_egresos]

    st.markdown(
        f"""
        {tab_header("Overview Ejecutivo", "Asset Intelligence Platform · estado del activo y métricas consolidadas", download_html=overview_download_html)}
        <div class="asset-kpi-grid">
            <div class="asset-kpi" style="--accent:#F87171;--soft:#fff7f7;--halo:#fde2e2;--border:#f1caca;--badge-bg:#fee2e2;--badge-fg:#B91C1C;">
                <div class="asset-kpi-main">
                    <div class="asset-kpi-icon">{asset_icon_svg("down")}</div>
                    <div>
                        <div class="asset-kpi-title">Posición Neta</div>
                        <div class="asset-kpi-value">{fmt_clp_largo(posicion_neta)}</div>
                        <div class="asset-kpi-badge">{'Déficit' if posicion_neta < 0 else 'Superávit'} acumulado</div>
                    </div>
                </div>
                {asset_sparkline(posicion_spark, "#F87171" if posicion_neta < 0 else "#22C55E")}
            </div>
            <div class="asset-kpi" style="--accent:#22C55E;--soft:#f6fffb;--halo:#d8f5e4;--border:#cfe9de;--badge-bg:#dcfce7;--badge-fg:#166534;">
                <div class="asset-kpi-main">
                    <div class="asset-kpi-icon">{asset_icon_svg("wallet")}</div>
                    <div>
                        <div class="asset-kpi-title">Caja Disponible</div>
                        <div class="asset-kpi-value">{fmt_clp_largo(balance_kpi)}</div>
                        <div class="asset-kpi-badge">Runway: {cobertura_egresos:.1f} meses</div>
                    </div>
                </div>
                {asset_sparkline(caja_spark, "#22C55E")}
            </div>
            <div class="asset-kpi" style="--accent:#B7791F;--soft:#fffaf0;--halo:#fceec8;--border:#eadfbd;--badge-bg:#f3f4f6;--badge-fg:#334155;">
                <div class="asset-kpi-main">
                    <div class="asset-kpi-icon">{asset_icon_svg("capex")}</div>
                    <div>
                        <div class="asset-kpi-title">Inversión Total</div>
                        <div class="asset-kpi-value">{fmt_clp_largo(CAPEX)}</div>
                        <div class="asset-kpi-badge">{cobertura_capex:.1%} recuperado</div>
                    </div>
                </div>
                {asset_sparkline(capex_spark, "#F59E0B")}
            </div>
            <div class="asset-kpi" style="--accent:#1D4ED8;--soft:#f6f9ff;--halo:#e0ebff;--border:#d4e1f6;--badge-bg:#fee2e2;--badge-fg:#B91C1C;">
                <div class="asset-kpi-main">
                    <div class="asset-kpi-icon">{asset_icon_svg("percent")}</div>
                    <div>
                        <div class="asset-kpi-title">Margen Neto</div>
                        <div class="asset-kpi-value">{margen_neto:.1%}</div>
                        <div class="asset-kpi-badge">Sobre ingresos</div>
                    </div>
                </div>
                {asset_sparkline(margen_spark, "#2563EB" if margen_neto >= 0 else "#F87171")}
            </div>
            <div class="asset-kpi" style="--accent:#6D28D9;--soft:#fbf8ff;--halo:#eadcff;--border:#e0d3f5;--badge-bg:#fee2e2;--badge-fg:#B91C1C;">
                <div class="asset-kpi-main">
                    <div class="asset-kpi-icon">{asset_icon_svg("ratio")}</div>
                    <div>
                        <div class="asset-kpi-title">Caja / Egreso</div>
                        <div class="asset-kpi-value">{cobertura_egresos:.2f}x</div>
                        <div class="asset-kpi-badge">Cobertura mensual</div>
                    </div>
                </div>
                {asset_sparkline(cobertura_spark, "#6D28D9")}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    chart_left = st.container()
    with chart_left:
        def build_home_neto_comparativo(df_src: pd.DataFrame, modo: str) -> pd.DataFrame:
            if modo == "Contable":
                periodo_col = "Periodo_ref_contable"
            else:
                periodo_col = "Periodo_ref"

            df_cmp = df_src.dropna(subset=["Monto"]).copy()
            df_cmp = df_cmp[df_cmp["CC_norm"].isin(["INGRESO", "EGRESO"])]
            df_cmp = df_cmp[df_cmp["Sit_norm"].isin(["PAGADO", "NO PAGADO"])]
            df_cmp = df_cmp.dropna(subset=[periodo_col])
            df_cmp["Periodo"] = pd.to_datetime(df_cmp[periodo_col], errors="coerce")
            df_cmp = df_cmp.dropna(subset=["Periodo"])
            if df_cmp.empty:
                return pd.DataFrame(columns=["Periodo", modo])

            agg_cmp = (
                df_cmp.groupby(["Periodo", "CC_norm"], as_index=False)["Monto"]
                .sum()
                .sort_values("Periodo")
            )
            ingresos_cmp = agg_cmp[agg_cmp["CC_norm"] == "INGRESO"].rename(columns={"Monto": "Ingresos"})
            egresos_cmp = agg_cmp[agg_cmp["CC_norm"] == "EGRESO"].rename(columns={"Monto": "Egresos"})
            base_cmp = pd.DataFrame({"Periodo": sorted(agg_cmp["Periodo"].dropna().unique())})
            base_cmp["Periodo"] = pd.to_datetime(base_cmp["Periodo"], errors="coerce")
            base_cmp = base_cmp.merge(ingresos_cmp[["Periodo", "Ingresos"]], on="Periodo", how="left")
            base_cmp = base_cmp.merge(egresos_cmp[["Periodo", "Egresos"]], on="Periodo", how="left")
            base_cmp = base_cmp.fillna(0).sort_values("Periodo")
            base_cmp[modo] = base_cmp["Ingresos"] - base_cmp["Egresos"].abs()
            base_cmp["Egresos_plot"] = -base_cmp["Egresos"].abs()
            return base_cmp[["Periodo", "Ingresos", "Egresos_plot", modo]]

        flujo_home_fin = build_home_neto_comparativo(df_f, "Financiero")
        flujo_home_con = build_home_neto_comparativo(df_f, "Contable")
        neto_home_fin = flujo_home_fin[["Periodo", "Financiero"]]
        neto_home_con = flujo_home_con[["Periodo", "Contable"]]
        neto_home_cmp = neto_home_fin.merge(neto_home_con, on="Periodo", how="outer").sort_values("Periodo").fillna(0)

        fig_flow = go.Figure()
        if not flujo_home_fin.empty:
            fig_flow.add_trace(
                go.Bar(
                    x=flujo_home_fin["Periodo"],
                    y=flujo_home_fin["Ingresos"],
                    name="Ingresos",
                    marker=dict(color="rgba(34,197,94,0.50)", line=dict(color="rgba(34,197,94,0.72)", width=1)),
                    hovertemplate="<b>%{x|%b %Y}</b><br>Ingresos: $%{y:,.0f}<extra></extra>",
                )
            )
            fig_flow.add_trace(
                go.Bar(
                    x=flujo_home_fin["Periodo"],
                    y=flujo_home_fin["Egresos_plot"],
                    name="Egresos",
                    marker=dict(color="rgba(248,113,113,0.56)", line=dict(color="rgba(248,113,113,0.75)", width=1)),
                    hovertemplate="<b>%{x|%b %Y}</b><br>Egresos: $%{y:,.0f}<extra></extra>",
                )
            )
        if not neto_home_cmp.empty:
            fig_flow.add_trace(
                go.Scatter(
                    x=neto_home_cmp["Periodo"],
                    y=neto_home_cmp["Financiero"],
                    mode="lines+markers",
                    name="Neto financiero",
                    line=dict(color="#2563EB", width=3.1, shape="spline"),
                    marker=dict(size=7, color="#2563EB", line=dict(color="#FFFFFF", width=1.4)),
                    hovertemplate="<b>%{x|%b %Y}</b><br>Neto financiero: $%{y:,.0f}<extra></extra>",
                )
            )
            fig_flow.add_trace(
                go.Scatter(
                    x=neto_home_cmp["Periodo"],
                    y=neto_home_cmp["Contable"],
                    mode="lines+markers",
                    name="Neto contable",
                    line=dict(color="#F59E0B", width=2.9, dash="dot", shape="spline"),
                    marker=dict(size=7, color="#F59E0B", line=dict(color="#FFFFFF", width=1.4)),
                    hovertemplate="<b>%{x|%b %Y}</b><br>Neto contable: $%{y:,.0f}<extra></extra>",
                )
            )
        fig_flow.add_hline(y=0, line_width=1, line_color="#CBD5E1")
        fig_flow.update_layout(
            title=dict(
                text="Ingresos, egresos y neto financiero vs contable",
                x=0.02,
                y=0.98,
                xanchor="left",
                font=dict(size=16, color="#081735", family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif"),
            ),
            font=dict(family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif", color="#475569"),
            template="plotly_white",
            height=360,
            margin=dict(l=12, r=14, t=54, b=20),
            barmode="relative",
            paper_bgcolor="#FFFFFF",
            plot_bgcolor="#FFFFFF",
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.01,
                xanchor="left",
                x=0.02,
                font=dict(size=11, color="#334155"),
                bgcolor="rgba(255,255,255,0)",
            ),
            hoverlabel=dict(bgcolor="#FFFFFF", bordercolor="#CBD5E1", font=dict(size=11, color="#0F172A")),
            hovermode="x unified",
        )
        fig_flow.update_yaxes(
            title_text="Flujo CLP",
            tickprefix="$",
            separatethousands=True,
            gridcolor="rgba(180,190,210,0.15)",
            zeroline=False,
            title_font=dict(size=11, color="#64748B"),
            tickfont=dict(size=11, color="#64748B"),
        )
        fig_flow.update_xaxes(
            title_text="Mes",
            tickformat="%b %Y",
            title_font=dict(size=11, color="#64748B"),
            tickfont=dict(size=11, color="#64748B"),
            rangeslider=dict(
                visible=True,
                thickness=0.035,
                bgcolor="#F1F5F9",
                bordercolor="#E2E8F0",
                borderwidth=1,
            ),
        )
        st.plotly_chart(
            fig_flow,
            use_container_width=True,
            config={"displaylogo": False},
            key="vision_general_neto_fin_cont",
        )

    bottom_a, bottom_b, bottom_c = st.columns([1.05, 1.2, 1.1])
    resumen_rows = [
        (asset_icon_svg("up"), "#22C55E", "#DCFCE7", "Total ingresos", "Ingresos Pagados + Abono", fmt_clp_largo(ingresos_kpi), "100.0%"),
        (asset_icon_svg("down"), "#F87171", "#FEE2E2", "Total egresos", "Egresos Pagado", fmt_clp_largo(abs(egresos_kpi)), f"{(abs(egresos_kpi) / ingresos_kpi if ingresos_kpi else 0):.1%}"),
        (asset_icon_svg("equal"), "#F87171" if utilidad_operativa < 0 else "#2563EB", "#FEE2E2" if utilidad_operativa < 0 else "#DBEAFE", "Total neto", "Ingresos menos egresos", fmt_clp_largo(utilidad_operativa), f"{margen_neto:.1%}"),
        (asset_icon_svg("ratio"), "#F87171" if posicion_neta < 0 else "#1D4ED8", "#FEE2E2" if posicion_neta < 0 else "#DBEAFE", "Neto acumulado", "Suma de Ingresos + Egresos sin incluir Abono", fmt_clp_largo(posicion_neta), "-"),
        (asset_icon_svg("percent"), "#F87171" if margen_neto < 0 else "#F59E0B", "#FEE2E2" if margen_neto < 0 else "#FEF3C7", "Margen neto", "Neto dividido por ingresos", f"{margen_neto:.1%}", "-"),
    ]
    with bottom_a:
        rows_html = "".join(
            (
                f'<div class="asset-summary-row" style="--accent:{accent};--soft:{soft};">'
                f'<div class="asset-summary-icon">{icon}</div>'
                f'<div><div class="asset-summary-label">{label}</div><div class="asset-summary-sub">{desc}</div></div>'
                f'<div class="asset-summary-value">{value}<div class="asset-summary-pct">{pct}</div></div>'
                f'</div>'
            )
            for icon, accent, soft, label, desc, value, pct in resumen_rows
        )
        st.markdown(
            f"""
            <div class="asset-card asset-bottom-card">
                <div class="asset-card-title">Resumen del período</div>
                <div class="asset-summary-list">{rows_html}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with bottom_b:
        ingresos_trend_color = "#DC2626" if ingresos_12m_growth < 0 else "#059669"
        canon_trend_color = "#DC2626" if canon_12m_growth < 0 else "#059669"
        egresos_trend_color = "#DC2626" if egresos_12m_growth > 0 else "#059669"
        sueldo_base_honorario = 1_700_000
        sueldo_actual_honorario = 1_500_000
        deficit_honorarios = abs(posicion_neta) if posicion_neta < 0 else 0
        honorario_rows_html = ""
        for meses_plan in (6, 8, 12):
            ahorro_mensual_total = deficit_honorarios / meses_plan if deficit_honorarios else 0
            reduccion_por_persona = ahorro_mensual_total / 2
            sueldo_por_persona = max(sueldo_base_honorario - reduccion_por_persona, 0)
            honorario_rows_html += (
                f'<div class="trend-salary-row">'
                f'<strong>{meses_plan} meses</strong>'
                f'<span>{fmt_clp_largo(reduccion_por_persona)}</span>'
                f'<span>{fmt_clp_largo(sueldo_por_persona)}</span>'
                f'</div>'
            )
        ahorro_actual_total = (sueldo_base_honorario - sueldo_actual_honorario) * 2
        meses_con_sueldo_actual = deficit_honorarios / ahorro_actual_total if ahorro_actual_total and deficit_honorarios else 0
        sueldo_actual_note = (
            f"Actual {fmt_clp_largo(sueldo_actual_honorario)} c/u: ahorro "
            f"{fmt_clp_largo(ahorro_actual_total)}/mes, cubre en {meses_con_sueldo_actual:.1f} meses."
        )
        st.markdown(
            f"""
            <div class="asset-card asset-bottom-card">
                <div class="asset-card-title">Análisis de Tendencia</div>
                <div class="trend-grid">
                    <div class="trend-mini" style="--trend:{ingresos_trend_color};--soft:{'#fff7f7' if ingresos_12m_growth < 0 else '#f6fffb'};">
                        <div class="trend-mini-title">Ingresos 12M</div>
                        <div class="trend-mini-value">{ingresos_12m_growth:+.1%}</div>
                        <div class="trend-mini-sub">Total ingresos vs 12 previos</div>
                    </div>
                    <div class="trend-mini" style="--trend:{canon_trend_color};--soft:{'#fff7f7' if canon_12m_growth < 0 else '#f6fffb'};">
                        <div class="trend-mini-title">Canon 12M</div>
                        <div class="trend-mini-value">{canon_12m_growth:+.1%}</div>
                        <div class="trend-mini-sub">Canon mensual vs 12 previos</div>
                    </div>
                    <div class="trend-mini" style="--trend:{egresos_trend_color};--soft:{'#fff7f7' if egresos_12m_growth > 0 else '#f6fffb'};">
                        <div class="trend-mini-title">Egresos 12M</div>
                        <div class="trend-mini-value">{egresos_12m_growth:+.1%}</div>
                        <div class="trend-mini-sub">Egresos pagados vs 12 previos</div>
                    </div>
                </div>
                <div class="asset-callout">Lectura 12M: {top_concentration_txt}. {pressure_txt}. Base móvil sobre registros pagados y abonados.</div>
                <div class="trend-salary-plan">
                    <div class="trend-salary-head">
                        <span>Escenario honorarios</span>
                        <strong>Base {fmt_clp_largo(sueldo_base_honorario)} c/u</strong>
                    </div>
                    <div class="trend-salary-labels">
                        <span>Plazo</span>
                        <span>Reducir c/u</span>
                        <span>Sueldo c/u</span>
                    </div>
                    {honorario_rows_html}
                    <div class="trend-salary-note">Para cubrir {fmt_clp_largo(deficit_honorarios)} de neto acumulado. {sueldo_actual_note}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with bottom_c:
        diagnosis_detail_rows = [
            ("Puntaje inicial", "100/100"),
            ("CAPEX total", fmt_clp_largo(CAPEX)),
            ("Ingresos canon acumulados", fmt_clp_largo(ingresos_canon)),
            ("Recuperación CAPEX", f"{cobertura_capex:.1%}"),
            ("Caja disponible", fmt_clp_largo(balance_kpi)),
            ("Egreso mensual promedio", fmt_clp_largo(egreso_mensual_promedio)),
            ("Cobertura caja / egreso prom.", f"{cobertura_egresos:.2f}x"),
            ("Ingresos KPI", fmt_clp_largo(ingresos_kpi)),
            ("Egresos KPI pagados", fmt_clp_largo(egresos_kpi)),
            ("Resultado operativo", fmt_clp_largo(utilidad_operativa)),
            ("Margen neto", f"{margen_neto:.1%}"),
            ("No pagado ingresos", fmt_clp_largo(no_pagado_total)),
            ("Abonos ingresos", fmt_clp_largo(abonos_total)),
            ("Cuentas por cobrar neto", fmt_clp_largo(cuentas_por_cobrar_neto)),
            ("Egresos por pagar", fmt_clp_largo(total_egresos_por_pagar)),
            ("Deuda fin ejercicio", fmt_clp_largo(deuda_fin_ejercicio)),
            ("Posición neta", fmt_clp_largo(posicion_neta)),
        ]
        diagnosis_detail_html = "".join(
            f'<div style="display:flex;justify-content:space-between;gap:8px;border-bottom:1px solid #e5ebf3;padding:3px 0;">'
            f'<span>{label}</span><strong>{value}</strong></div>'
            for label, value in diagnosis_detail_rows
        )
        diagnosis_penalty_html = "".join(
            f'<div style="display:flex;justify-content:space-between;gap:8px;padding:2px 0;color:#B42318;">'
            f'<span>{label}</span><strong>{points}</strong></div>'
            for label, points in diagnosis_penalties
        ) or '<div style="color:#059669;font-weight:900;">Sin penalizaciones activas</div>'
        caja_status_color = "#DC2626" if cobertura_egresos < 1 else "#059669"
        margen_status_color = "#DC2626" if margen_neto < 0.05 else "#059669"
        capex_status_color = "#F59E0B" if cobertura_capex < 1 else "#059669"
        recommendation_txt = "Revisar canon, cobranza y postergar CAPEX no crítico" if health_score < 70 else "Mantener control de caja y ocupación"
        score_marker_left = max(0, min(100, health_score))
        st.markdown(
            f"""
            <div style="height:380px;border:1px solid rgba(219,227,238,0.72);border-radius:14px;background:#ffffff;
                        padding:16px;box-sizing:border-box;box-shadow:0 4px 18px rgba(15,23,42,0.045);
                        position:relative;
                        font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;color:#081735;">
                <style>
                    summary::-webkit-details-marker {{ display:none; }}
                    summary::marker {{ content:""; }}
                </style>
                <div style="font-size:14px;font-weight:950;line-height:1.1;margin-bottom:2px;">Operational Health Score</div>
                <div style="margin-top:12px;padding:6px 2px 4px;display:grid;grid-template-columns:124px 1fr;gap:12px;align-items:center;">
                    <svg viewBox="0 0 120 72" aria-hidden="true">
                        <path d="M15 60 A45 45 0 0 1 105 60" fill="none" stroke="#E5EAF2" stroke-width="13" stroke-linecap="round" pathLength="100"/>
                        <path d="M15 60 A45 45 0 0 1 105 60" fill="none" stroke="{estado_color}" stroke-width="13" stroke-linecap="round" pathLength="100" stroke-dasharray="{health_score} 100"/>
                        <circle cx="60" cy="60" r="4" fill="#081735"/>
                    </svg>
                    <div>
                    <div style="display:flex;align-items:flex-end;justify-content:space-between;gap:10px;">
                        <div style="font-size:34px;line-height:1;font-weight:950;color:#081735;letter-spacing:0;">{health_score}<span style="font-size:18px;font-weight:900;color:#64748b;">/100</span></div>
                        <div style="min-width:82px;height:24px;border-radius:8px;background:{estado_color};color:#ffffff;font-size:10px;font-weight:950;display:flex;align-items:center;justify-content:center;">{estado_txt}</div>
                    </div>
                    <div style="position:relative;height:42px;margin-top:10px;">
                        <div style="position:absolute;left:0;right:0;top:12px;height:14px;border-radius:999px;overflow:hidden;display:grid;grid-template-columns:40fr 30fr 30fr;background:#f1f5f9;">
                            <div style="background:#FEE2E2;"></div>
                            <div style="background:#FEF3C7;"></div>
                            <div style="background:#DCFCE7;"></div>
                        </div>
                        <div style="position:absolute;left:{score_marker_left}%;top:4px;transform:translateX(-50%);width:4px;height:30px;border-radius:999px;background:{estado_color};box-shadow:0 0 0 3px rgba(255,255,255,0.95),0 5px 12px rgba(15,23,42,0.18);"></div>
                        <div style="position:absolute;left:0;right:0;bottom:0;display:flex;justify-content:space-between;font-size:8px;font-weight:900;color:#64748b;">
                            <span>0</span><span>40</span><span>70</span><span>100</span>
                        </div>
                    </div>
                    </div>
                </div>
                <div style="margin-top:0;display:flex;align-items:center;justify-content:center;gap:7px;color:#475569;font-size:9px;font-weight:800;">
                    <span>Estado general</span>
                    <span style="display:inline-flex;align-items:center;justify-content:center;min-width:76px;height:20px;border-radius:7px;border:1px solid {estado_color};color:{estado_color};font-size:9px;font-weight:950;background:#ffffff;">{estado_txt}</span>
                </div>
                <div style="display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:6px;margin-top:8px;">
                    <div style="border:1px solid #e5ebf3;border-radius:7px;padding:5px 6px;background:#fbfdff;">
                        <div style="display:flex;align-items:center;gap:5px;font-size:8.5px;font-weight:950;color:#081735;"><span style="width:7px;height:7px;border-radius:999px;background:{caja_status_color};"></span>Caja</div>
                        <div style="margin-top:3px;font-size:8px;font-weight:800;color:#475569;">{runway_txt}</div>
                    </div>
                    <div style="border:1px solid #e5ebf3;border-radius:7px;padding:5px 6px;background:#fbfdff;">
                        <div style="display:flex;align-items:center;gap:5px;font-size:8.5px;font-weight:950;color:#081735;"><span style="width:7px;height:7px;border-radius:999px;background:{margen_status_color};"></span>Margen</div>
                        <div style="margin-top:3px;font-size:8px;font-weight:800;color:#475569;">{margen_neto:.1%} sobre ingresos</div>
                    </div>
                    <div style="border:1px solid #e5ebf3;border-radius:7px;padding:5px 6px;background:#fbfdff;">
                        <div style="display:flex;align-items:center;gap:5px;font-size:8.5px;font-weight:950;color:#081735;"><span style="width:7px;height:7px;border-radius:999px;background:{capex_status_color};"></span>CAPEX</div>
                        <div style="margin-top:3px;font-size:8px;font-weight:800;color:#475569;">{cobertura_capex:.1%} recuperado</div>
                    </div>
                </div>
                <div style="margin-top:7px;border:1px solid #e5ebf3;border-radius:7px;background:#ffffff;padding:5px 7px;font-size:8.4px;line-height:1.25;font-weight:850;color:#334155;">
                    <strong style="color:#081735;">Recomendación:</strong> {recommendation_txt}
                </div>
                <div style="display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:6px;margin-top:8px;">
                    <div style="display:grid;grid-template-columns:14px 1fr;gap:5px;align-items:start;">
                        <span style="width:9px;height:9px;border-radius:999px;background:#DC2626;margin-top:3px;"></span>
                        <div style="font-size:9px;font-weight:900;color:#081735;">Riesgo Alto<br><span style="display:inline-block;margin-top:4px;color:#31507a;font-weight:850;">0 - 40</span></div>
                    </div>
                    <div style="display:grid;grid-template-columns:14px 1fr;gap:5px;align-items:start;">
                        <span style="width:9px;height:9px;border-radius:999px;background:#F59E0B;margin-top:3px;"></span>
                        <div style="font-size:9px;font-weight:900;color:#081735;">Riesgo Medio<br><span style="display:inline-block;margin-top:4px;color:#31507a;font-weight:850;">40 - 70</span></div>
                    </div>
                    <div style="display:grid;grid-template-columns:14px 1fr;gap:5px;align-items:start;">
                        <span style="width:9px;height:9px;border-radius:999px;background:#059669;margin-top:3px;"></span>
                        <div style="font-size:9px;font-weight:900;color:#081735;">Estado Saludable<br><span style="display:inline-block;margin-top:4px;color:#31507a;font-weight:850;">70 - 100</span></div>
                    </div>
                </div>
                <details style="position:absolute;left:16px;right:16px;bottom:16px;">
                    <summary style="height:26px;border:1px solid #dbe3ee;border-radius:8px;display:flex;align-items:center;
                                justify-content:center;color:#0f2d52;font-size:10px;font-weight:900;cursor:pointer;background:#ffffff;">
                        <span style="flex:1;text-align:center;">Ver recomendaciones</span><span style="padding-right:10px;">→</span>
                    </summary>
                    <div style="position:absolute;left:0;right:0;bottom:34px;height:238px;overflow:auto;
                                border:1px solid #cbd5e1;border-radius:8px;background:#ffffff;padding:9px 10px;
                                box-shadow:0 16px 36px rgba(15,23,42,0.16);font-size:8.8px;line-height:1.25;z-index:20;">
                        <div style="font-size:10px;font-weight:950;margin-bottom:6px;color:#081735;">Detalle del diagnóstico ({health_score}/100)</div>
                        {diagnosis_detail_html}
                        <div style="font-size:9.5px;font-weight:950;margin:7px 0 3px;color:#081735;">Penalizaciones aplicadas</div>
                        {diagnosis_penalty_html}
                    </div>
                </details>
            </div>
            """,
            unsafe_allow_html=True,
        )

    components.html(
        """
        <script>
        (function () {
            const win = window.parent;
            const doc = win.document;
            const btn = doc.getElementById("overview-download-report");
            if (!btn || btn.dataset.captureHandlerAttached === "1") return;
            btn.dataset.captureHandlerAttached = "1";

            function loadScript(src) {
                return new Promise(function (resolve, reject) {
                    const existing = doc.querySelector('script[src="' + src + '"]');
                    if (existing) {
                        existing.addEventListener("load", resolve, { once: true });
                        if (existing.dataset.loaded === "1") resolve();
                        return;
                    }
                    const script = doc.createElement("script");
                    script.src = src;
                    script.onload = function () {
                        script.dataset.loaded = "1";
                        resolve();
                    };
                    script.onerror = reject;
                    doc.head.appendChild(script);
                });
            }

            async function downloadOverviewPdf(event) {
                event.preventDefault();
                const originalText = btn.textContent;
                btn.textContent = "Generando PDF...";
                btn.style.pointerEvents = "none";

                const hidden = [];
                function hide(selector) {
                    doc.querySelectorAll(selector).forEach(function (node) {
                        hidden.push([node, node.style.display]);
                        node.style.display = "none";
                    });
                }

                try {
                    await loadScript("https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js");
                    await loadScript("https://cdn.jsdelivr.net/npm/jspdf@2.5.1/dist/jspdf.umd.min.js");

                    const target = doc.querySelector("section.main .block-container")
                        || doc.querySelector('[data-testid="stMainBlockContainer"]')
                        || doc.querySelector(".block-container");
                    if (!target) throw new Error("No se encontró el contenido del reporte.");

                    hide("section[data-testid='stSidebar']");
                    hide("aside[data-testid='stSidebar']");
                    hide(".bodegas-sidebar-toggle");
                    hide(".tab-actions");
                    hide("header[data-testid='stHeader']");
                    hide("div[data-testid='stToolbar']");

                    const canvas = await win.html2canvas(target, {
                        backgroundColor: "#ffffff",
                        scale: 2,
                        useCORS: true,
                        allowTaint: true,
                        logging: false,
                        windowWidth: Math.max(target.scrollWidth, doc.documentElement.clientWidth),
                        windowHeight: Math.max(target.scrollHeight, doc.documentElement.clientHeight)
                    });

                    const imgData = canvas.toDataURL("image/png");
                    const pdf = new win.jspdf.jsPDF({
                        orientation: "landscape",
                        unit: "pt",
                        format: "a4"
                    });
                    const pageW = pdf.internal.pageSize.getWidth();
                    const pageH = pdf.internal.pageSize.getHeight();
                    const margin = 10;
                    const usableW = pageW - margin * 2;
                    const imgH = canvas.height * usableW / canvas.width;
                    let y = margin;
                    let remaining = imgH;

                    pdf.addImage(imgData, "PNG", margin, y, usableW, imgH, null, "FAST");
                    remaining -= pageH - margin * 2;
                    while (remaining > 0) {
                        pdf.addPage();
                        y = margin - (imgH - remaining);
                        pdf.addImage(imgData, "PNG", margin, y, usableW, imgH, null, "FAST");
                        remaining -= pageH - margin * 2;
                    }
                    pdf.save("overview_ejecutivo_bodegas_balmaceda.pdf");
                } catch (err) {
                    win.alert("No se pudo generar el PDF visual. Intenta nuevamente cuando la página termine de cargar.");
                } finally {
                    hidden.forEach(function (entry) {
                        entry[0].style.display = entry[1];
                    });
                    btn.textContent = originalText;
                    btn.style.pointerEvents = "";
                }
            }

            btn.addEventListener("click", downloadOverviewPdf);
        })();
        </script>
        """,
        height=0,
    )

    st.stop()

    st.markdown(
        section_heading(
            "📊",
            "Estado general del proyecto",
            weight_class="section-heading-title-soft",
        ),
        unsafe_allow_html=True,
    )

    # --- KPIs principales en una sola línea ---
    kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)

    with kpi1:
        st.markdown(
            card_finanza(
                "🏦 Caja Banco BCI",
                fmt_clp_largo(balance_kpi),
                "#D85E5D",
                subtitulo="Saldo disponible en cuenta corriente para operación",
                etiqueta="Bloque 1",
                size="top",
            ),
            unsafe_allow_html=True,
        )

    with kpi2:
        st.markdown(
            card_finanza(
                "💼 CAPEX",
                fmt_clp_largo(CAPEX),
                "#DCAA67",
                subtitulo="Inversión total del proyecto",
                etiqueta="Bloque 2",
                size="top",
            ),
            unsafe_allow_html=True,
        )

    with kpi3:
        st.markdown(
            card_finanza(
                "💹 Utilidad / CAPEX",
                f"{utilidad_sobre_capex:.2f}x",
                "#7FA6A2",
                subtitulo="Sueldos Adolfo Orellana y Jonathan Mac-Kay relativos a la inversión total",
                etiqueta="Bloque 3",
                size="top",
            ),
            unsafe_allow_html=True,
        )

    with kpi4:
        st.markdown(
            card_finanza(
                "📈 Margen neto",
                f"{margen_neto:.1%}",
                "#4B5563",
                subtitulo="Resultado operativo sobre ingresos acumulados",
                etiqueta="Bloque 4",
                size="top",
            ),
            unsafe_allow_html=True,
        )

    with kpi5:
        st.markdown(
            card_finanza(
                "💳 Caja / egreso mensual promedio",
                f"{cobertura_egresos:.2f}x",
                "#A8A8A8",
                subtitulo="Caja disponible relativa al gasto mensual promedio",
                etiqueta="Bloque 5",
                size="top",
            ),
            unsafe_allow_html=True,
        )

    st.caption("Cobertura CAPEX = Ingresos canon acumulados / CAPEX total invertido.")
    st.markdown("")

    hero_col, side_col = st.columns(2)
    resumen_items = [
        {"label": "Ingresos canon arriendo", "meta": "Acumulado histórico de canon", "value": fmt_clp_largo(ingresos_canon)},
        {"label": "Total ingresos", "meta": "Flujo acumulado", "value": fmt_clp_largo(ingresos_kpi)},
        {"label": "Total egresos", "meta": "Compromisos pagados", "value": fmt_clp_largo(egresos_kpi)},
        {"label": "Cuentas por cobrar", "meta": "Neto de abonos", "value": fmt_clp_largo(cuentas_por_cobrar_neto)},
        {"label": "Egresos por pagar", "meta": "Pendientes de salida", "value": fmt_clp_largo(total_egresos_por_pagar)},
    ]

    with hero_col:
        max_ref = max(abs(balance_kpi), abs(posicion_neta), 1)
        metric_df = pd.DataFrame(
            {
                "Métrica": ["Caja Banco BCI", "Posición Neta"],
                "Valor": [float(balance_kpi), float(posicion_neta)],
                "Detalle": [
                    "Saldo disponible en cuenta corriente para operación.",
                    "Caja + CxC + EPP",
                ],
                "Color": ["#4B5563", "#D85E5D"],
            }
        )

        fig_metrics = go.Figure()
        fig_metrics.add_vrect(
            x0=-max_ref * 1.28,
            x1=0,
            fillcolor="rgba(148,163,184,0.08)",
            line_width=0,
            layer="below",
        )
        fig_metrics.add_vline(x=0, line_width=1.5, line_dash="dot", line_color="rgba(71,85,105,0.50)")

        for _, row in metric_df.iterrows():
            fig_metrics.add_trace(
                go.Bar(
                    y=[row["Métrica"]],
                    x=[row["Valor"]],
                    orientation="h",
                    marker=dict(
                        color=row["Color"],
                        line=dict(color="rgba(255,255,255,0.75)", width=1),
                    ),
                    width=0.46,
                    text=[f"<b>{row['Métrica']}</b><br>{fmt_clp_largo(row['Valor'])}"],
                    textposition="inside",
                    insidetextanchor="middle",
                    textfont=dict(size=15, color="#ffffff", family="Arial Black, Arial, sans-serif"),
                    customdata=[[row["Detalle"]]],
                    hovertemplate="<b>%{y}</b><br>Valor: %{text}<br>%{customdata[0]}<extra></extra>",
                    showlegend=False,
                    cliponaxis=False,
                )
            )

        fig_metrics.update_layout(
            template="plotly_white",
            height=430,
            margin=dict(l=28, r=28, t=52, b=26),
            paper_bgcolor="#ffffff",
            plot_bgcolor="#fbfcfe",
            xaxis=dict(
                range=[-max_ref * 1.28, max_ref * 1.18],
                showgrid=True,
                gridcolor="rgba(203,213,225,0.28)",
                zeroline=False,
                tickformat=",.0f",
                title="CLP",
                title_font=dict(size=13, color="#64748b"),
                tickfont=dict(size=12, color="#64748b"),
                linecolor="rgba(148,163,184,0.20)",
            ),
            yaxis=dict(
                showgrid=False,
                showticklabels=False,
                linecolor="rgba(148,163,184,0.20)",
            ),
            bargap=0.52,
            annotations=[
                dict(
                    x=0,
                    y=1.18,
                    xref="paper",
                    yref="paper",
                    text="Pulso financiero consolidado",
                    showarrow=False,
                    xanchor="left",
                    font=dict(size=12, color="#64748b"),
                ),
                dict(
                    x=0,
                    y=1.08,
                    xref="paper",
                    yref="paper",
                    text="<b>Liquidez y posición neta</b>",
                    showarrow=False,
                    xanchor="left",
                    font=dict(size=19, color="#0f172a"),
                )
            ],
        )
        st.plotly_chart(
            fig_metrics,
            use_container_width=True,
            config={"displaylogo": False, "displayModeBar": False},
        )

    with side_col:
        st.markdown(
            kpi_resumen_panel(
                "Resumen ejecutivo",
                "Indicadores secundarios del estado financiero actual",
                resumen_items,
            ),
            unsafe_allow_html=True,
        )

    st.markdown("---")
    # =========================
    # Gráfico PRO: Ingresos por Canon de Arriendo por año + MA-3
    # =========================
    # =========================
    # 🏢 Ingresos por canon de arriendo — por año (MA-3)
    # =========================
    st.markdown(
        section_heading(
            "📊",
            "Ingresos por Canon de Arriendo",
            weight_class="section-heading-title-soft",
        ),
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div style="
            background: linear-gradient(90deg, #0f2d52 0%, #1f4e78 100%);
            border-radius: 10px;
            padding: 10px 14px;
            margin: 8px 0 10px 0;
            color: #FFFFFF;
            font-size: 13px;
            font-weight: 600;">
            Evolución anual de canon de arriendo · Indicador estratégico de ingresos
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Usa df_f si existen filtros aplicados; si no, usa df completo
    data_src = df_f if "df_f" in locals() else df

    # Filtro canon: CC == INGRESO y CC1 == Arriendo y Obs == Canon mensual
    mask_canon = (
        data_src["CC"].astype(str).str.strip().str.upper().eq("INGRESO") &
        data_src["CC1"].astype(str).str.strip().str.lower().eq("arriendo") &
        data_src["Obs"].astype(str).str.strip().str.lower().eq("canon mensual")
    )

    # Agregado anual
    df_canon = (
        data_src.loc[mask_canon, ["Año", "Monto"]]
        .copy()
    )

    # Tipos
    df_canon["Año"] = pd.to_numeric(df_canon["Año"], errors="coerce")
    df_canon["Monto"] = pd.to_numeric(
        df_canon["Monto"].astype(str).str.replace(r"[^\d\.-]", "", regex=True),
        errors="coerce"
    )
    df_canon = df_canon.dropna(subset=["Año", "Monto"])

    if df_canon.empty:
        st.info("No hay datos de canon de arriendo para construir la serie anual.")
    else:
        df_canon = (
            df_canon
            .groupby("Año", as_index=False)["Monto"]
            .sum()
            .sort_values("Año")
            .rename(columns={"Monto": "Canon_anual_CLP"})
        )

        # Promedio móvil 3 años
        df_canon["MA3"] = (
            df_canon["Canon_anual_CLP"]
            .rolling(window=3, min_periods=1)
            .mean()
        )
        df_canon["YoY"] = df_canon["Canon_anual_CLP"].pct_change().replace([np.inf, -np.inf], 0).fillna(0)
        df_canon["Acumulado"] = df_canon["Canon_anual_CLP"].cumsum()

        # ---------- KPI resumen ----------
        ultimo_anio = int(df_canon["Año"].max())
        ultimo_valor = float(
            df_canon.loc[df_canon["Año"] == ultimo_anio, "Canon_anual_CLP"].iloc[0]
        )

        anio_prev = ultimo_anio - 1
        if (df_canon["Año"] == anio_prev).any():
            valor_prev = float(
                df_canon.loc[df_canon["Año"] == anio_prev, "Canon_anual_CLP"].iloc[0]
            )
            var_yoy = (ultimo_valor - valor_prev) / valor_prev if valor_prev != 0 else 0
        else:
            valor_prev = None
            var_yoy = 0.0

        color_yoy = CHART_TEAL if var_yoy >= 0 else CHART_RED

        st.markdown(
            f"""
            <div style='padding:12px 18px; border-radius:12px;
                 background:linear-gradient(180deg, #FFFFFF 0%, #F8FAFC 100%);
                 border:1px solid #D0D5DD; width: fit-content; margin-bottom:-8px;
                 box-shadow:0 6px 16px rgba(15,45,82,0.08);'>
                <span style='font-size:13px; color:#344054; font-weight:600;'>Último año ({ultimo_anio}):</span>
                <span style='font-size:20px; font-weight:800; color:#0F2D52;'>
                    ${ultimo_valor:,.0f}
                </span>
                <span style='font-size:13px; font-weight:700; color:{color_yoy};'>
                    ({var_yoy:+.1%} YoY)
                </span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        # ---------- GRÁFICO ----------
        from plotly.subplots import make_subplots

        df_canon["Hover"] = (
            "Canon anual: $" + df_canon["Canon_anual_CLP"].map(lambda v: f"{v:,.0f}")
            + "<br>Promedio móvil MA-3: $" + df_canon["MA3"].map(lambda v: f"{v:,.0f}")
            + "<br>Variación YoY: " + df_canon["YoY"].map(lambda v: f"{v:+.1%}")
            + "<br>Acumulado: $" + df_canon["Acumulado"].map(lambda v: f"{v:,.0f}")
        )

        fig = make_subplots(specs=[[{"secondary_y": True}]])

        fig.add_trace(
            go.Scatter(
                x=df_canon["Año"],
                y=df_canon["Canon_anual_CLP"],
                mode="markers",
                name="Resumen",
                marker=dict(size=20, color="rgba(0,0,0,0)"),
                showlegend=False,
                customdata=df_canon[["Hover"]],
                hovertemplate="<b>Año %{x}</b><br>%{customdata[0]}<extra></extra>",
            ),
            secondary_y=False,
        )

        fig.add_trace(
            go.Bar(
                x=df_canon["Año"],
                y=df_canon["Canon_anual_CLP"],
                name="Canon anual",
                marker=dict(
                    color=CHART_TEAL,
                    line=dict(width=1.0, color="rgba(255,255,255,0.75)"),
                ),
                hoverinfo="skip",
            ),
            secondary_y=False,
        )

        fig.add_trace(
            go.Scatter(
                x=df_canon["Año"],
                y=df_canon["MA3"],
                name="Promedio móvil MA-3",
                mode="lines+markers",
                line=dict(color=CHART_DARK, width=3.2, shape="spline"),
                marker=dict(color=CHART_DARK, size=7, line=dict(color="white", width=1.5)),
                hoverinfo="skip",
            ),
            secondary_y=False,
        )

        fig.add_trace(
            go.Scatter(
                x=df_canon["Año"],
                y=df_canon["YoY"],
                name="Variación YoY",
                mode="lines+markers",
                line=dict(color=CHART_GOLD, width=2.4, dash="dot"),
                marker=dict(
                    size=7,
                    color=np.where(df_canon["YoY"] >= 0, CHART_GOLD, CHART_RED),
                    line=dict(color="white", width=1.2),
                ),
                hoverinfo="skip",
            ),
            secondary_y=True,
        )

        if len(df_canon) >= 1:
            last_row = df_canon.iloc[-1]
            fig.add_annotation(
                x=last_row["Año"],
                y=last_row["Canon_anual_CLP"],
                text=f"${last_row['Canon_anual_CLP']:,.0f}",
                showarrow=True,
                arrowhead=2,
                ax=-34,
                ay=-34,
                bgcolor="rgba(255,255,255,0.92)",
                bordercolor="#CBD5E1",
                borderwidth=1,
                font=dict(size=11, color="#0F172A"),
            )

        fig.update_layout(
            template="plotly_white",
            height=540,
            margin=dict(l=24, r=42, t=74, b=30),
            title=dict(
                text="Ingresos por Canon de Arriendo · evolución anual",
                x=0.01,
                xanchor="left",
                font=dict(size=19, color="#0F2D52"),
            ),
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                x=0.01,
                xanchor="left",
                font=dict(size=12),
                bgcolor="rgba(255,255,255,0)",
            ),
            plot_bgcolor="#F8FAFC",
            paper_bgcolor="#F8FAFC",
            hovermode="x unified",
            hoverlabel=dict(
                bgcolor="white",
                font_size=13,
                font_color="#111",
                bordercolor="#E5E7EB",
            ),
            bargap=0.22,
        )
        fig.add_hline(y=0, line_width=1, line_color=CHART_GRAY, opacity=0.6, secondary_y=False)
        fig.update_xaxes(
            title_text="Año",
            tickmode="linear",
            dtick=1,
            showgrid=False,
            linecolor="rgba(15,45,82,0.25)",
            tickfont=dict(size=12, color="#334155"),
        )
        fig.update_yaxes(
            title_text="Canon CLP",
            tickprefix="$",
            separatethousands=True,
            gridcolor="rgba(15,45,82,0.10)",
            zeroline=False,
            ticks="outside",
            ticklen=6,
            tickfont=dict(size=12, color="#334155"),
            linecolor="rgba(15,45,82,0.20)",
            secondary_y=False,
        )
        fig.update_yaxes(
            title_text="YoY",
            tickformat=".0%",
            showgrid=False,
            zeroline=False,
            tickfont=dict(size=12, color="#334155"),
            secondary_y=True,
        )


        st.plotly_chart(fig, use_container_width=True)

        # Descargar dataset agregado
        csv_canon = df_canon.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇️ Descargar ingresos por canon (CSV)",
            data=csv_canon,
            file_name="ingresos_canon_por_anio.csv",
            mime="text/csv",
        )

        st.markdown(
            section_heading(
                "📑",
                "Tabla de KPIs Financieros",
                weight_class="section-heading-title-soft",
            ),
            unsafe_allow_html=True,
        )

        cartera_vencida_df = df_f.loc[
            df_f["CC_norm"].eq("INGRESO") & df_f["Sit_norm"].eq("NO PAGADO")
        ].copy()
        cartera_vencida_df["Responsable"] = cartera_vencida_df["Responsable"].astype(str).str.strip()
        deuda_top3 = (
            cartera_vencida_df.groupby("Responsable")["Monto"].sum().abs().sort_values(ascending=False)
            if not cartera_vencida_df.empty
            else pd.Series(dtype=float)
        )
        concentracion_top3 = (deuda_top3.head(3).sum() / deuda_top3.sum()) if not deuda_top3.empty and deuda_top3.sum() else 0.0
        ticket_promedio_deuda = (
            cartera_vencida_df["Monto"].abs().sum() / len(cartera_vencida_df)
            if not cartera_vencida_df.empty
            else 0.0
        )

        abonos_resp_df = df_f.loc[
            df_f["CC_norm"].eq("INGRESO") & df_f["Sit_norm"].str.startswith("ABONO")
        ].copy()
        abonos_resp_df["Responsable"] = abonos_resp_df["Responsable"].astype(str).str.strip()
        responsables_con_abono = abonos_resp_df["Responsable"].replace("", pd.NA).dropna().nunique()
        abono_promedio_responsable = (
            abonos_resp_df["Monto"].abs().sum() / responsables_con_abono
            if responsables_con_abono
            else 0.0
        )

        cartera_edad_df = cartera_vencida_df.copy()
        cartera_edad_df["Fecha"] = pd.to_datetime(cartera_edad_df.get("Fecha"), errors="coerce")
        cartera_edad_df = cartera_edad_df.dropna(subset=["Fecha"])
        if not cartera_edad_df.empty:
            cartera_edad_df["Dias_vencidos"] = (
                pd.Timestamp.today().normalize() - cartera_edad_df["Fecha"].dt.normalize()
            ).dt.days.clip(lower=0)
            edad_promedio_cartera = float(cartera_edad_df["Dias_vencidos"].mean())
            edad_promedio_cartera_fmt = f"{edad_promedio_cartera:.0f} días"
        else:
            edad_promedio_cartera = None
            edad_promedio_cartera_fmt = "N/D"

        cobertura_caja_cartera = (
            balance_kpi / abs(cuentas_por_cobrar_neto)
            if cuentas_por_cobrar_neto
            else 0.0
        )

        canon_period_df = data_src.loc[mask_canon, ["Monto", "Año", "Mes", "Fecha", "Esp"]].copy()
        canon_period_df["Año_calc"] = pd.to_numeric(canon_period_df.get("Año"), errors="coerce")
        canon_mes_raw = canon_period_df.get("Mes")
        canon_mes_num = canon_mes_raw.astype(str).str.extract(r"(\d{1,2})", expand=False) if canon_mes_raw is not None else None
        canon_period_df["Mes_calc"] = pd.to_numeric(canon_mes_num, errors="coerce")
        canon_period_df["Fecha"] = pd.to_datetime(canon_period_df.get("Fecha"), errors="coerce")
        canon_period_df["Año_calc"] = canon_period_df["Año_calc"].fillna(canon_period_df["Fecha"].dt.year)
        canon_period_df["Mes_calc"] = canon_period_df["Mes_calc"].fillna(canon_period_df["Fecha"].dt.month)
        canon_period_df = canon_period_df.dropna(subset=["Monto"])

        if not canon_period_df.empty:
            canon_period_df["Periodo"] = pd.to_datetime(
                dict(
                    year=canon_period_df["Año_calc"].fillna(1).astype(int),
                    month=canon_period_df["Mes_calc"].fillna(1).astype(int),
                    day=1,
                ),
                errors="coerce",
            )
            canon_mensual_promedio = (
                canon_period_df.dropna(subset=["Periodo"])
                .groupby("Periodo", as_index=False)["Monto"]
                .sum()["Monto"]
                .mean()
            )
        else:
            canon_mensual_promedio = 0.0

        ingresos_esp_df = df_f.loc[
            df_f["CC_norm"].eq("INGRESO") &
            (
                df_f["Sit_norm"].eq("PAGADO") |
                df_f["Sit_norm"].str.startswith("ABONO")
            ),
            ["Monto", "Esp_num"],
        ].copy()
        ingresos_esp_df = ingresos_esp_df[ingresos_esp_df["Esp_num"].between(1, 7, inclusive="both")]
        espacios_con_ingreso = ingresos_esp_df["Esp_num"].nunique()
        ingreso_promedio_espacio = (
            ingresos_esp_df["Monto"].sum() / espacios_con_ingreso
            if espacios_con_ingreso
            else 0.0
        )

        m2_map_fin = {1: 120, 2: 72, 3: 72, 4: 72, 5: 180, 6: 130, 7: 60}
        canon_m2_df = canon_period_df.copy()
        canon_m2_df["Esp_num"] = pd.to_numeric(canon_m2_df["Esp"], errors="coerce")
        canon_m2_df["m2"] = canon_m2_df["Esp_num"].map(m2_map_fin)
        canon_m2_df = canon_m2_df.dropna(subset=["m2"])
        canon_m2_promedio = (
            (canon_m2_df["Monto"] / canon_m2_df["m2"]).mean()
            if not canon_m2_df.empty
            else 0.0
        )

        electricidad_mask = (
            df_f["CC_norm"].eq("EGRESO") &
            df_f["Sit_norm"].eq("PAGADO") &
            (
                df_f["Obs_text"].str.contains(r"\bcge\b|electricidad", case=False, na=False, regex=True) |
                df_f["CC1_text"].str.contains(r"\bcge\b|electricidad", case=False, na=False, regex=True)
            )
        )
        egreso_electricidad = df_f.loc[electricidad_mask, "Monto"].abs().sum()
        pct_electricidad_egresos = (
            egreso_electricidad / abs(egresos_kpi)
            if egresos_kpi
            else 0.0
        )

        kpi_fin_df = pd.DataFrame(
            [
                {
                    "Indicador": "Caja disponible",
                    "Valor": fmt_clp_largo(balance_kpi),
                    "Lectura": "Saldo operativo disponible en cuenta corriente.",
                },
                {
                    "Indicador": "Posición neta",
                    "Valor": fmt_clp_largo(posicion_neta),
                    "Lectura": "Caja + cuentas por cobrar netas + egresos por pagar.",
                },
                {
                    "Indicador": "CxC neto",
                    "Valor": fmt_clp_largo(cuentas_por_cobrar_neto),
                    "Lectura": "Cuentas por cobrar descontando abonos registrados.",
                },
                {
                    "Indicador": "Resultado operativo",
                    "Valor": fmt_clp_largo(utilidad_operativa),
                    "Lectura": "Ingresos menos egresos pagados acumulados.",
                },
                {
                    "Indicador": "Margen neto",
                    "Valor": f"{margen_neto:.1%}",
                    "Lectura": "Resultado operativo sobre ingresos acumulados.",
                },
                {
                    "Indicador": "Avance de cobranza sobre cartera vencida",
                    "Valor": f"{pct_cobranza:.1%}",
                    "Lectura": "Abonos registrados sobre cartera acumulada en estado no pagado.",
                },
                {
                    "Indicador": "Caja / egreso mensual promedio",
                    "Valor": f"{cobertura_egresos:.2f}x",
                    "Lectura": "Caja disponible relativa al promedio mensual de egresos pagados.",
                },
                {
                    "Indicador": "Cobertura CAPEX",
                    "Valor": f"{cobertura_capex:.1%}",
                    "Lectura": "Canon acumulado respecto de la inversión total.",
                },
                {
                    "Indicador": "Concentración de deuda top 3",
                    "Valor": f"{concentracion_top3:.1%}",
                    "Lectura": "Participación de los 3 mayores responsables sobre la cartera vencida.",
                },
                {
                    "Indicador": "Ticket promedio de deuda",
                    "Valor": fmt_clp_largo(ticket_promedio_deuda),
                    "Lectura": "Monto promedio por transacción en estado no pagado.",
                },
                {
                    "Indicador": "Abono promedio por responsable",
                    "Valor": fmt_clp_largo(abono_promedio_responsable),
                    "Lectura": "Abono promedio entre responsables con registros de abono.",
                },
                {
                    "Indicador": "Edad promedio de cartera vencida",
                    "Valor": edad_promedio_cartera_fmt,
                    "Lectura": "Antigüedad promedio de documentos no pagados con fecha válida.",
                },
                {
                    "Indicador": "Cobertura caja / cartera vencida neta",
                    "Valor": f"{cobertura_caja_cartera:.2f}x",
                    "Lectura": "Caja Banco BCI relativa a cuentas por cobrar netas.",
                },
                {
                    "Indicador": "Canon mensual promedio",
                    "Valor": fmt_clp_largo(canon_mensual_promedio),
                    "Lectura": "Promedio mensual de canon de arriendo sobre períodos con datos.",
                },
                {
                    "Indicador": "Variación YoY del canon",
                    "Valor": f"{var_yoy:+.1%}",
                    "Lectura": f"Variación del canon anual entre {anio_prev if valor_prev is not None else 'N/D'} y {ultimo_anio}.",
                },
                {
                    "Indicador": "Ingreso promedio por espacio",
                    "Valor": fmt_clp_largo(ingreso_promedio_espacio),
                    "Lectura": "Ingreso acumulado promedio entre espacios con ingresos registrados.",
                },
                {
                    "Indicador": "Canon por m² promedio",
                    "Valor": f"${canon_m2_promedio:,.0f}/m²",
                    "Lectura": "Canon promedio por metro cuadrado a partir de los espacios 1 al 7.",
                },
                {
                    "Indicador": "% electricidad sobre egresos",
                    "Valor": f"{pct_electricidad_egresos:.1%}",
                    "Lectura": "Peso de egresos asociados a electricidad/CGE sobre egresos pagados.",
                },
            ]
        )

        _render_table(
            kpi_fin_df,
            header_bg="#163A5F",
            header_fg="white",
            row_alt="#F8FAFC",
            compact=False,
        )

# =========================================================
# 🏠 TAB 1B: OVERVIEW EJECUTIVO 2
# =========================================================
if active_section == "🏠 Overview Ejecutivo":
    import plotly.graph_objects as go

    data_v2 = df_f.copy()
    data_v2["Periodo_ref"] = pd.to_datetime(data_v2.get("Periodo_ref"), errors="coerce")

    def ov2_icon(kind: str) -> str:
        icons = {
            "liquidity": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M4 19V5"/><path d="M4 19h16"/><path d="M8 15l3-3 3 2 5-7"/></svg>',
            "net": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.6" stroke-linecap="round" stroke-linejoin="round"><path d="M7 17L17 7"/><path d="M9 7h8v8"/></svg>',
            "capex": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"><path d="M3 21h18"/><path d="M6 21V9l6-4 6 4v12"/><path d="M9 21v-6h6v6"/></svg>',
            "margin": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round"><path d="M19 5L5 19"/><circle cx="7.5" cy="7.5" r="2.5"/><circle cx="16.5" cy="16.5" r="2.5"/></svg>',
            "coverage": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3l8 4v5c0 5-3.5 8-8 9-4.5-1-8-4-8-9V7l8-4z"/><path d="M9 12l2 2 4-5"/></svg>',
            "alert": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M10.3 3.9L1.8 18a2 2 0 0 0 1.7 3h17a2 2 0 0 0 1.7-3L13.7 3.9a2 2 0 0 0-3.4 0z"/><path d="M12 9v4"/><path d="M12 17h.01"/></svg>',
            "trend": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M3 17l6-6 4 4 8-8"/><path d="M15 7h6v6"/></svg>',
            "focus": '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="7"/><circle cx="12" cy="12" r="2"/><path d="M12 2v3"/><path d="M12 19v3"/><path d="M2 12h3"/><path d="M19 12h3"/></svg>',
        }
        return icons.get(kind, icons["focus"])

    def ov2_spark(values, color: str) -> str:
        vals = pd.Series(values).replace([np.inf, -np.inf], np.nan).dropna().astype(float).tail(18).tolist()
        if len(vals) < 2:
            vals = [0.0, 0.0]
        lo, hi = min(vals), max(vals)
        span = hi - lo if hi != lo else 1.0
        points = []
        denom = max(len(vals) - 1, 1)
        for i, value in enumerate(vals):
            x = 2 + (i / denom) * 106
            y = 22 - ((value - lo) / span) * 18
            points.append(f"{x:.1f},{y:.1f}")
        return (
            f'<svg class="ov2-spark" viewBox="0 0 112 26" preserveAspectRatio="none" aria-hidden="true">'
            f'<polygon points="{" ".join(points)} 110,24 2,24" fill="{color}" opacity="0.08"/>'
            f'<path d="M2 23 H110" stroke="rgba(100,116,139,0.12)" stroke-width="1"/>'
            f'<polyline points="{" ".join(points)}" fill="none" stroke="{color}" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"/>'
            f'</svg>'
        )

    def ov2_growth(mask: pd.Series, abs_values: bool = False) -> tuple[float, float]:
        df_month = data_v2.loc[mask & data_v2["Periodo_ref"].notna(), ["Periodo_ref", "Monto"]].copy()
        df_month["Monto"] = pd.to_numeric(df_month["Monto"], errors="coerce")
        df_month = df_month.dropna(subset=["Periodo_ref", "Monto"])
        if df_month.empty:
            return 0.0, 0.0
        df_month["Periodo_ref"] = df_month["Periodo_ref"].dt.to_period("M").dt.to_timestamp()
        monthly = df_month.groupby("Periodo_ref")["Monto"].sum().sort_index()
        if abs_values:
            monthly = monthly.abs()
        full_months = pd.date_range(monthly.index.min(), monthly.index.max(), freq="MS")
        monthly = monthly.reindex(full_months, fill_value=0)
        last_12 = float(monthly.iloc[-12:].sum())
        prev_12 = float(monthly.iloc[-24:-12].sum()) if len(monthly) >= 24 else 0.0
        growth = (last_12 / prev_12 - 1) if prev_12 else 0.0
        return growth, last_12

    mask_ing_ov2 = data_v2["CC_norm"].eq("INGRESO") & data_v2["Sit_norm"].isin(["PAGADO", "ABONO"])
    mask_egr_ov2 = data_v2["CC_norm"].eq("EGRESO") & data_v2["Sit_norm"].eq("PAGADO")
    mask_canon_ov2 = (
        data_v2["CC_norm"].eq("INGRESO")
        & data_v2["CC1_text"].str.strip().str.lower().eq("arriendo")
        & data_v2["Obs_text"].str.strip().str.lower().eq("canon mensual")
    )
    ingresos_growth_12m, ingresos_last_12m = ov2_growth(mask_ing_ov2)
    egresos_growth_12m, egresos_last_12m = ov2_growth(mask_egr_ov2, abs_values=True)
    canon_growth_12m, canon_last_12m_total = ov2_growth(mask_canon_ov2)

    max_period_ov2 = data_v2.loc[data_v2["Periodo_ref"].notna(), "Periodo_ref"].max()
    last_12_start_ov2 = max_period_ov2 - pd.DateOffset(months=11) if pd.notna(max_period_ov2) else pd.NaT
    last_12_mask_ov2 = data_v2["Periodo_ref"].between(last_12_start_ov2, max_period_ov2) if pd.notna(last_12_start_ov2) else pd.Series(False, index=data_v2.index)

    canon_top3_share = 0.0
    canon_top1_name = "Sin registros"
    canon_last_12 = data_v2.loc[last_12_mask_ov2 & mask_canon_ov2].copy()
    if not canon_last_12.empty and canon_last_12m_total:
        canon_by_resp = canon_last_12.groupby("Responsable_clean")["Monto"].sum().sort_values(ascending=False)
        canon_top3_share = float(canon_by_resp.head(3).sum() / canon_last_12m_total) if canon_by_resp.sum() else 0.0
        canon_top1_name = str(canon_by_resp.index[0]) if not canon_by_resp.empty else "Sin registros"

    egresos_last_12 = data_v2.loc[last_12_mask_ov2 & data_v2["CC_norm"].eq("EGRESO")].copy()
    pressure_name = "Sin egresos relevantes"
    pressure_value = 0.0
    if not egresos_last_12.empty:
        egresos_last_12["Monto_abs"] = pd.to_numeric(egresos_last_12["Monto"], errors="coerce").abs()
        pressure_by_cc1 = egresos_last_12.groupby("CC1_text")["Monto_abs"].sum().sort_values(ascending=False)
        if not pressure_by_cc1.empty:
            pressure_name = escape(str(pressure_by_cc1.index[0]))
            pressure_value = float(pressure_by_cc1.iloc[0])

    health_score_ov2 = int(
        np.clip(
            55
            + (18 if posicion_neta >= 0 else -18)
            + (16 if margen_neto >= 0 else -16)
            + (14 if cobertura_egresos >= 1 else -14)
            + (12 if cobertura_capex >= 1 else -8),
            0,
            100,
        )
    )
    health_color_ov2 = "#22C55E" if health_score_ov2 >= 70 else ("#F59E0B" if health_score_ov2 >= 45 else "#F87171")
    health_state_ov2 = "Stabilization Phase" if health_score_ov2 >= 70 else ("Operational Pressure" if health_score_ov2 >= 45 else "Financial Stress")
    health_delta_90 = 4 if utilidad_operativa >= 0 else -4
    liquidity_state = "crítica" if cobertura_egresos < 1 else ("controlada" if cobertura_egresos < 3 else "sólida")
    net_state = "negativo" if utilidad_operativa < 0 else "positivo"
    priority_label_ov2 = "Reforzar caja y cobranza" if cobertura_egresos < 1 or posicion_neta < 0 else "Mantener disciplina operativa"
    priority_copy_ov2 = (
        f"Priorizar recuperación de liquidez: caja cubre {cobertura_egresos:.1f} meses y el resultado acumulado sigue bajo presión."
        if cobertura_egresos < 1 or posicion_neta < 0
        else f"Mantener control de egresos y monitoreo de canon; la cobertura financiera se mantiene en {cobertura_egresos:.1f}x."
    )
    honorario_base_cu = 1_700_000
    honorario_actual_cu = 1_500_000
    honorario_personas = 2
    neto_a_cubrir_ov2 = max(0.0, abs(float(posicion_neta)) if posicion_neta < 0 else abs(float(utilidad_operativa)) if utilidad_operativa < 0 else 0.0)
    honorario_scenarios_ov2 = []
    for plazo_meses in (6, 8, 12):
        reducir_cu = neto_a_cubrir_ov2 / plazo_meses / honorario_personas if plazo_meses and honorario_personas else 0.0
        sueldo_cu = max(0.0, honorario_base_cu - reducir_cu)
        honorario_scenarios_ov2.append((plazo_meses, reducir_cu, sueldo_cu))
    ahorro_actual_mes_ov2 = max(0.0, honorario_base_cu - honorario_actual_cu) * honorario_personas
    meses_cubre_actual_ov2 = neto_a_cubrir_ov2 / ahorro_actual_mes_ov2 if ahorro_actual_mes_ov2 else 0.0
    today_ov2 = pd.Timestamp.today().normalize()
    fecha_ov2 = pd.to_datetime(data_v2.get("Fecha_dt", data_v2.get("Fecha")), errors="coerce").dt.normalize()
    short_window_ov2 = fecha_ov2.between(today_ov2 - pd.Timedelta(days=30), today_ov2 + pd.Timedelta(days=30), inclusive="both")
    pending_ov2 = data_v2["Sit_norm"].eq("NO PAGADO") & short_window_ov2
    cxc_30_ov2 = float(data_v2.loc[pending_ov2 & data_v2["CC_norm"].eq("INGRESO"), "Monto"].abs().sum())
    cxp_30_ov2 = float(data_v2.loc[pending_ov2 & data_v2["CC_norm"].eq("EGRESO"), "Monto"].abs().sum())
    gap_corto_ov2 = cxc_30_ov2 - cxp_30_ov2
    salida_diaria_ov2 = (cxp_30_ov2 / 30) if cxp_30_ov2 > 0 else (abs(float(egreso_mensual_promedio)) / 30 if egreso_mensual_promedio else 0)
    cobertura_dias_ov2 = max(0.0, float(balance_kpi) / salida_diaria_ov2) if salida_diaria_ov2 else 0.0

    def ov2_status(value: float, kind: str) -> tuple[str, str]:
        if kind == "cxc":
            if value <= 0:
                return "pressure", "Sin entrada"
            if value >= cxp_30_ov2:
                return "healthy", "Saludable"
            return "attention", "Atención"
        if kind == "cxp":
            if value <= 0:
                return "healthy", "Sin presión"
            if value <= cxc_30_ov2:
                return "attention", "Cubierto"
            return "pressure", "Presión"
        if kind == "gap":
            if value >= 0:
                return "healthy", "Saludable"
            if abs(value) <= max(float(balance_kpi), 0):
                return "attention", "Atención"
            return "pressure", "Presión"
        if value >= 30:
            return "healthy", "Saludable"
        if value >= 15:
            return "attention", "Atención"
        return "pressure", "Presión"

    flow_ov2 = data_v2.dropna(subset=["Monto", "Periodo_ref"]).copy()
    flow_ov2 = flow_ov2[flow_ov2["CC_norm"].isin(["INGRESO", "EGRESO"])]
    flow_ov2 = flow_ov2[flow_ov2["Sit_norm"].isin(["PAGADO", "NO PAGADO", "ABONO"])]
    flow_ov2["Periodo"] = flow_ov2["Periodo_ref"].dt.to_period("M").dt.to_timestamp()
    agg_ov2 = flow_ov2.groupby(["Periodo", "CC_norm"], as_index=False)["Monto"].sum()
    ing_ov2 = agg_ov2[agg_ov2["CC_norm"] == "INGRESO"].rename(columns={"Monto": "Ingresos"})
    egr_ov2 = agg_ov2[agg_ov2["CC_norm"] == "EGRESO"].rename(columns={"Monto": "Egresos"})
    base_ov2 = pd.DataFrame({"Periodo": sorted(agg_ov2["Periodo"].dropna().unique())})
    base_ov2 = base_ov2.merge(ing_ov2[["Periodo", "Ingresos"]], on="Periodo", how="left")
    base_ov2 = base_ov2.merge(egr_ov2[["Periodo", "Egresos"]], on="Periodo", how="left").fillna(0)
    base_ov2["Egresos_abs"] = base_ov2["Egresos"].abs()
    base_ov2["Egresos_plot"] = -base_ov2["Egresos_abs"]
    base_ov2["Neto"] = base_ov2["Ingresos"] - base_ov2["Egresos_abs"]

    st.markdown(
        """
        <style>
        .tab-title-row {
            margin:-61px 0 4px 0;
        }
        .tab-title-main {
            font-size:24px;
        }
        .tab-title-sub {
            margin-top:3px;
            font-size:11.5px;
            font-weight:500;
            letter-spacing:0.2px;
        }
        .ov2-hero-kpis {
            display:grid;
            grid-template-columns:repeat(5,minmax(0,1fr));
            gap:7px;
            margin:4px 0 6px 0;
        }
        .ov2-kpi {
            min-height:78px;
            border:1px solid rgba(219,227,238,0.72);
            border-radius:14px;
            background:linear-gradient(135deg,#ffffff 0%,var(--soft) 100%);
            box-shadow:0 4px 18px rgba(15,23,42,0.045);
            padding:8px 10px 6px 10px;
            display:flex;
            flex-direction:column;
            justify-content:space-between;
            transition:transform .16s ease, box-shadow .16s ease;
            animation:ov2FadeIn .24s ease both;
        }
        .ov2-kpi:hover,
        .ov2-insight:hover,
        .ov2-card:hover {
            transform:translateY(-1px);
            box-shadow:0 8px 24px rgba(15,23,42,0.07);
        }
        @keyframes ov2FadeIn {
            from { opacity:0; transform:translateY(4px); }
            to { opacity:1; transform:translateY(0); }
        }
        .ov2-kpi-main {
            display:grid;
            grid-template-columns:25px 1fr;
            gap:6px;
        }
        .ov2-icon {
            width:24px;
            height:24px;
            border-radius:999px;
            background:var(--halo);
            color:var(--accent);
            display:flex;
            align-items:center;
            justify-content:center;
        }
        .ov2-icon svg {
            width:14px;
            height:14px;
        }
        .ov2-kpi-title {
            color:#081735;
            font-size:8.5px;
            line-height:1.12;
            font-weight:950;
            text-transform:uppercase;
        }
        .ov2-kpi-value {
            margin-top:3px;
            color:var(--accent);
            font-size:15.5px;
            line-height:1.02;
            font-weight:950;
            letter-spacing:0;
            white-space:nowrap;
        }
        .ov2-kpi-sub {
            margin-top:3px;
            color:#64748b;
            font-size:8.2px;
            line-height:1.2;
            font-weight:500;
            letter-spacing:0.2px;
        }
        .ov2-spark {
            width:100%;
            height:13px;
            margin-top:2px;
            display:block;
        }
        .ov2-main-grid {
            display:grid;
            grid-template-columns:2.94fr .83fr;
            gap:8px;
            align-items:stretch;
            margin-bottom:2px;
        }
        .ov2-card {
            border:1px solid rgba(219,227,238,0.72);
            border-radius:12px;
            background:#ffffff;
            box-shadow:0 4px 18px rgba(15,23,42,0.045);
            padding:8px;
            min-height:0;
        }
        .ov2-main-grid [data-testid="stPlotlyChart"] {
            border:0;
            border-radius:12px;
            background:#ffffff;
            box-shadow:0 4px 18px rgba(15,23,42,0.045);
            padding:8px;
        }
        .ov2-main-anchor {
            height:0;
            min-height:0;
            margin:0;
            padding:0;
            overflow:hidden;
        }
        div[data-testid="stVerticalBlock"]:has(.ov2-main-anchor) > div[data-testid="stHorizontalBlock"] {
            gap:0.55rem;
            align-items:flex-start;
            margin-bottom:0;
        }
        div[data-testid="stVerticalBlock"]:has(.ov2-main-anchor) div[data-testid="stElementContainer"] {
            margin-bottom:0 !important;
        }
        div[data-testid="stVerticalBlock"]:has(.ov2-main-anchor) [data-testid="stPlotlyChart"] {
            border:0;
            border-radius:12px;
            background:#ffffff;
            box-shadow:0 4px 18px rgba(15,23,42,0.045);
            padding:7px 7px 3px 7px;
            min-height:416px;
        }
        div[data-testid="stVerticalBlock"]:has(.ov2-main-anchor) [data-testid="stPlotlyChart"] > div {
            min-height:400px;
        }
        .ov2-card-title {
            color:#081735;
            font-size:13px;
            line-height:1.1;
            font-weight:950;
            margin-bottom:5px;
        }
        .ov2-intelligence-grid {
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:7px;
            margin-top:4px;
        }
        .ov2-insight {
            min-height:128px;
            border:1px solid rgba(229,235,243,0.86);
            border-radius:12px;
            background:#ffffff;
            box-shadow:0 4px 18px rgba(15,23,42,0.035);
            padding:11px 12px;
            display:flex;
            flex-direction:column;
            justify-content:space-between;
            transition:transform .16s ease, box-shadow .16s ease;
            animation:ov2FadeIn .28s ease both;
        }
        .ov2-insight-head {
            display:flex;
            align-items:center;
            gap:5px;
            color:#081735;
            font-size:10px;
            font-weight:950;
            margin-bottom:4px;
        }
        .ov2-insight-head .ov2-icon {
            width:20px;
            height:20px;
            flex:0 0 auto;
        }
        .ov2-insight-value {
            color:var(--accent);
            font-size:20px;
            line-height:1;
            font-weight:950;
        }
        .ov2-insight-copy {
            margin-top:4px;
            color:#475569;
            font-size:10px;
            line-height:1.28;
            font-weight:500;
            letter-spacing:0.2px;
        }
        .ov2-health {
            min-height:416px;
            display:flex;
            flex-direction:column;
            justify-content:flex-start;
            gap:10px;
            padding:10px;
        }
        .ov2-health-gauge {
            display:grid;
            grid-template-columns:78px 1fr;
            gap:8px;
            align-items:center;
            margin:2px 0 0 0;
        }
        .ov2-health-gauge svg {
            width:78px;
            height:48px;
            display:block;
        }
        .ov2-health-score {
            color:#081735;
            font-size:32px;
            line-height:1;
            font-weight:950;
        }
        .ov2-health-state {
            margin-top:4px;
            display:inline-flex;
            align-items:center;
            justify-content:center;
            min-height:20px;
            min-width:0;
            padding:3px 8px;
            border-radius:999px;
            background:rgba(245,158,11,0.14);
            color:#D97706;
            font-size:8.6px;
            font-weight:950;
        }
        .ov2-health-metrics {
            display:grid;
            grid-template-columns:repeat(3,minmax(0,1fr));
            gap:5px;
            margin-top:2px;
        }
        .ov2-health-metric {
            border:1px solid #e5ebf3;
            border-radius:8px;
            padding:5px 6px;
            background:#fbfdff;
        }
        .ov2-health-label {
            color:#64748b;
            font-size:7.6px;
            font-weight:900;
            text-transform:uppercase;
        }
        .ov2-health-value {
            margin-top:2px;
            color:#081735;
            font-size:9.4px;
            font-weight:950;
        }
        .ov2-short-pressure {
            margin-top:4px;
            padding-top:7px;
            border-top:1px solid rgba(226,232,240,0.86);
            display:flex;
            flex-direction:column;
            gap:5px;
        }
        .ov2-short-title {
            color:#081735;
            font-size:9.4px;
            line-height:1.1;
            font-weight:950;
            text-transform:uppercase;
        }
        .ov2-short-kpi {
            border:1px solid rgba(226,232,240,0.92);
            border-radius:8px;
            background:#fbfdff;
            padding:6px 7px;
            display:grid;
            grid-template-columns:1fr auto;
            gap:6px;
            align-items:center;
        }
        .ov2-short-label {
            color:#64748b;
            font-size:8.4px;
            line-height:1.12;
            font-weight:850;
        }
        .ov2-short-value {
            margin-top:2px;
            color:#081735;
            font-size:12px;
            line-height:1;
            font-weight:950;
        }
        .ov2-short-badge {
            border-radius:999px;
            padding:3px 7px;
            font-size:7.8px;
            line-height:1;
            font-weight:950;
            white-space:nowrap;
        }
        .ov2-short-badge.healthy {
            background:rgba(34,197,94,0.13);
            color:#15803D;
        }
        .ov2-short-badge.attention {
            background:rgba(245,158,11,0.14);
            color:#D97706;
        }
        .ov2-short-badge.pressure {
            background:rgba(248,113,113,0.14);
            color:#DC2626;
        }
        .ov2-priority {
            margin:2px 0 0 0;
            min-height:52px;
            border:1px solid rgba(219,227,238,0.78);
            border-radius:12px;
            background:#ffffff;
            box-shadow:0 4px 18px rgba(15,23,42,0.035);
            padding:8px 12px;
            display:grid;
            grid-template-columns:175px 1fr;
            gap:10px;
            align-items:center;
        }
        .ov2-priority-label {
            color:#081735;
            font-size:10.5px;
            line-height:1.1;
            font-weight:950;
        }
        .ov2-priority-action {
            margin-top:2px;
            color:#2563EB;
            font-size:14px;
            line-height:1.05;
            font-weight:950;
        }
        .ov2-honorario-scenario {
            margin-top:6px;
            padding-top:6px;
            border-top:1px solid rgba(226,232,240,0.9);
        }
        .ov2-honorario-title {
            color:#081735;
            font-size:9.2px;
            line-height:1.1;
            font-weight:950;
            text-transform:uppercase;
        }
        .ov2-honorario-grid {
            display:grid;
            grid-template-columns:0.55fr 1fr 1fr;
            gap:3px 6px;
            margin-top:5px;
            color:#475569;
            font-size:8.7px;
            line-height:1.15;
            font-weight:700;
        }
        .ov2-honorario-grid strong {
            color:#081735;
            font-weight:950;
        }
        .ov2-honorario-note {
            margin-top:5px;
            color:#64748b;
            font-size:8.8px;
            line-height:1.2;
            font-weight:600;
        }
        .ov2-actions {
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:8px;
            margin-top:8px;
        }
        .ov2-action {
            border:1px solid #dbe3ee;
            border-radius:12px;
            background:#ffffff;
            padding:11px 12px;
            color:#0f2d52;
            font-size:11px;
            font-weight:900;
            box-shadow:0 4px 18px rgba(15,23,42,0.035);
        }
        @media (max-width:1300px) {
            .ov2-hero-kpis,
            .ov2-intelligence-grid,
            .ov2-actions { grid-template-columns:repeat(2,minmax(0,1fr)); }
            .ov2-main-grid { grid-template-columns:1fr; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        {tab_header("Overview Ejecutivo", "Executive Asset Intelligence · dirección ejecutiva en 10 segundos")}
        <div class="ov2-hero-kpis">
            <div class="ov2-kpi" style="--accent:#F87171;--soft:#fff7f7;--halo:#fee2e2;">
                <div class="ov2-kpi-main"><div class="ov2-icon">{ov2_icon("liquidity")}</div><div><div class="ov2-kpi-title">Liquidez Operacional</div><div class="ov2-kpi-value">{fmt_clp_largo(balance_kpi)}</div><div class="ov2-kpi-sub">Runway operacional: {cobertura_egresos:.1f} meses</div></div></div>
                {ov2_spark([egreso_mensual_promedio, balance_kpi, balance_kpi + utilidad_operativa], "#F87171" if cobertura_egresos < 1 else "#22C55E")}
            </div>
            <div class="ov2-kpi" style="--accent:{'#F87171' if posicion_neta < 0 else '#22C55E'};--soft:{'#fff7f7' if posicion_neta < 0 else '#f6fffb'};--halo:{'#fee2e2' if posicion_neta < 0 else '#dcfce7'};">
                <div class="ov2-kpi-main"><div class="ov2-icon">{ov2_icon("net")}</div><div><div class="ov2-kpi-title">Resultado Neto Acumulado</div><div class="ov2-kpi-value">{fmt_clp_largo(posicion_neta)}</div><div class="ov2-kpi-sub">{'Déficit acumulado' if posicion_neta < 0 else 'Superávit acumulado'}</div></div></div>
                {ov2_spark([0, posicion_neta * 0.55, posicion_neta], "#F87171" if posicion_neta < 0 else "#22C55E")}
            </div>
            <div class="ov2-kpi" style="--accent:#F59E0B;--soft:#fffaf0;--halo:#fef3c7;">
                <div class="ov2-kpi-main"><div class="ov2-icon">{ov2_icon("capex")}</div><div><div class="ov2-kpi-title">Recuperación CAPEX</div><div class="ov2-kpi-value">{cobertura_capex:.1%}</div><div class="ov2-kpi-sub">Sobre inversión total {fmt_clp_largo(CAPEX)}</div></div></div>
                {ov2_spark([0, ingresos_canon, CAPEX], "#F59E0B")}
            </div>
            <div class="ov2-kpi" style="--accent:{'#F87171' if margen_neto < 0 else '#2563EB'};--soft:{'#fff7f7' if margen_neto < 0 else '#f6f9ff'};--halo:{'#fee2e2' if margen_neto < 0 else '#dbeafe'};">
                <div class="ov2-kpi-main"><div class="ov2-icon">{ov2_icon("margin")}</div><div><div class="ov2-kpi-title">Margen Operacional</div><div class="ov2-kpi-value">{margen_neto:.1%}</div><div class="ov2-kpi-sub">Neto sobre ingresos operacionales</div></div></div>
                {ov2_spark([0, margen_neto * 0.5, margen_neto], "#F87171" if margen_neto < 0 else "#2563EB")}
            </div>
            <div class="ov2-kpi" style="--accent:#6D28D9;--soft:#fbf8ff;--halo:#eadcff;">
                <div class="ov2-kpi-main"><div class="ov2-icon">{ov2_icon("coverage")}</div><div><div class="ov2-kpi-title">Cobertura Financiera</div><div class="ov2-kpi-value">{cobertura_egresos:.2f}x</div><div class="ov2-kpi-sub">Caja / egreso mensual promedio</div></div></div>
                {ov2_spark([0, cobertura_egresos * 0.55, cobertura_egresos], "#6D28D9")}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    fig_ov2 = go.Figure()
    if not base_ov2.empty:
        fig_ov2.add_trace(go.Bar(
            x=base_ov2["Periodo"],
            y=base_ov2["Ingresos"],
            name="Ingresos",
            marker=dict(color="rgba(34,197,94,0.55)", line=dict(color="rgba(34,197,94,0.72)", width=1)),
            hovertemplate="<b>%{x|%b %Y}</b><br>Ingresos: $%{y:,.0f}<extra></extra>",
        ))
        fig_ov2.add_trace(go.Bar(
            x=base_ov2["Periodo"],
            y=base_ov2["Egresos_plot"],
            name="Egresos",
            marker=dict(color="rgba(248,113,113,0.46)", line=dict(color="rgba(248,113,113,0.68)", width=0)),
            opacity=0.68,
            hovertemplate="<b>%{x|%b %Y}</b><br>Egresos: $%{y:,.0f}<extra></extra>",
        ))
        fig_ov2.add_trace(go.Scatter(
            x=base_ov2["Periodo"],
            y=base_ov2["Neto"],
            mode="lines",
            name="Neto operacional halo",
            line=dict(color="rgba(37,99,235,0.22)", width=10, shape="spline"),
            hoverinfo="skip",
            showlegend=False,
        ))
        fig_ov2.add_trace(go.Scatter(
            x=base_ov2["Periodo"],
            y=base_ov2["Neto"],
            mode="lines+markers",
            name="Neto operacional",
            line=dict(color="#2563EB", width=4.8, shape="spline"),
            marker=dict(size=8, color="#2563EB", line=dict(color="#FFFFFF", width=1.7)),
            hovertemplate="<b>%{x|%b %Y}</b><br>Neto: $%{y:,.0f}<extra></extra>",
        ))
    fig_ov2.add_hline(y=0, line_width=1, line_color="#CBD5E1")
    fig_ov2.update_layout(
        title=dict(text=""),
        template="plotly_white",
        height=400,
        margin=dict(l=0, r=2, t=38, b=6),
        barmode="relative",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#FFFFFF",
        legend=dict(orientation="h", y=1.015, x=0.02, bgcolor="rgba(255,255,255,0)", font=dict(size=10, color="#334155")),
        hovermode="x unified",
        hoverlabel=dict(bgcolor="#FFFFFF", bordercolor="#CBD5E1", font=dict(size=11, color="#0F172A")),
    )
    fig_ov2.update_yaxes(
        title_text="Flujo CLP",
        tickprefix="$",
        separatethousands=True,
        gridcolor="rgba(180,190,210,0.12)",
        zeroline=False,
        showline=False,
        ticks="",
    )
    fig_ov2.update_xaxes(title_text="", tickformat="%b %Y", showgrid=False, showline=False, ticks="", rangeslider=dict(visible=False))
    cxc_status_class, cxc_status_label = ov2_status(cxc_30_ov2, "cxc")
    cxp_status_class, cxp_status_label = ov2_status(cxp_30_ov2, "cxp")
    gap_status_class, gap_status_label = ov2_status(gap_corto_ov2, "gap")
    coverage_status_class, coverage_status_label = ov2_status(cobertura_dias_ov2, "coverage")

    main_ov2 = st.container()
    with main_ov2:
        st.markdown('<div class="ov2-main-anchor"></div>', unsafe_allow_html=True)
        left_ov2, right_ov2 = st.columns([2.94, .83], gap="small")
        with left_ov2:
            st.plotly_chart(fig_ov2, use_container_width=True, config={"displaylogo": False}, key="overview2_flujo_operacional")
        with right_ov2:
            st.markdown(
                f"""
                <div class="ov2-card ov2-health">
                    <div>
                        <div class="ov2-card-title">Asset Health Score</div>
                        <div class="ov2-insight-copy" style="margin-top:-4px;">Operational stability · {health_delta_90:+d} pts últimos 90 días</div>
                    </div>
                    <div class="ov2-health-gauge">
                        <svg viewBox="0 0 120 72" aria-hidden="true">
                            <defs>
                                <linearGradient id="ov2HealthGradient" x1="0%" y1="0%" x2="100%" y2="0%">
                                    <stop offset="0%" stop-color="{health_color_ov2}"/>
                                    <stop offset="100%" stop-color="#FBBF24"/>
                                </linearGradient>
                            </defs>
                            <path d="M18 58 A42 42 0 0 1 102 58" fill="none" stroke="#E7ECF3" stroke-width="6.4" stroke-linecap="round" pathLength="100"/>
                            <path d="M18 58 A42 42 0 0 1 102 58" fill="none" stroke="url(#ov2HealthGradient)" stroke-width="6.4" stroke-linecap="round" pathLength="100" stroke-dasharray="{health_score_ov2} 100"/>
                        </svg>
                        <div>
                            <div class="ov2-health-score">{health_score_ov2}<span style="font-size:18px;color:#64748b;">/100</span></div>
                            <div class="ov2-health-state" style="--state:{health_color_ov2};">{health_state_ov2}</div>
                        </div>
                    </div>
                    <div class="ov2-health-metrics">
                        <div class="ov2-health-metric"><div class="ov2-health-label">Riesgo</div><div class="ov2-health-value">{health_state_ov2}</div></div>
                        <div class="ov2-health-metric"><div class="ov2-health-label">Caja</div><div class="ov2-health-value">{liquidity_state}</div></div>
                        <div class="ov2-health-metric"><div class="ov2-health-label">CAPEX</div><div class="ov2-health-value">{cobertura_capex:.0%}</div></div>
                    </div>
                    <div class="ov2-short-pressure">
                        <div class="ov2-short-title">Presión financiera 30 días</div>
                        <div class="ov2-short-kpi">
                            <div><div class="ov2-short-label">Cuentas por cobrar &lt;30 días</div><div class="ov2-short-value">{fmt_clp_largo(cxc_30_ov2)}</div></div>
                            <div class="ov2-short-badge {cxc_status_class}">{cxc_status_label}</div>
                        </div>
                        <div class="ov2-short-kpi">
                            <div><div class="ov2-short-label">Cuentas por pagar &lt;30 días</div><div class="ov2-short-value">{fmt_clp_largo(cxp_30_ov2)}</div></div>
                            <div class="ov2-short-badge {cxp_status_class}">{cxp_status_label}</div>
                        </div>
                        <div class="ov2-short-kpi">
                            <div><div class="ov2-short-label">Gap operativo corto plazo</div><div class="ov2-short-value">{fmt_clp_largo(gap_corto_ov2)}</div></div>
                            <div class="ov2-short-badge {gap_status_class}">{gap_status_label}</div>
                        </div>
                        <div class="ov2-short-kpi">
                            <div><div class="ov2-short-label">Cobertura de caja inmediata</div><div class="ov2-short-value">{cobertura_dias_ov2:.0f} días</div></div>
                            <div class="ov2-short-badge {coverage_status_class}">{coverage_status_label}</div>
                        </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.markdown(
        f"""
        <div class="ov2-priority">
            <div>
                <div class="ov2-priority-label">Prioridad de gestión</div>
                <div class="ov2-priority-action">{priority_label_ov2}</div>
            </div>
            <div class="ov2-insight-copy" style="margin-top:0;">{priority_copy_ov2}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    honorario_rows_html = "".join(
        f"<div>{plazo}M</div><div><strong>{fmt_clp_largo(reducir)}</strong></div><div><strong>{fmt_clp_largo(sueldo)}</strong></div>"
        for plazo, reducir, sueldo in honorario_scenarios_ov2
    )
    st.markdown(
        f"""
        <div class="ov2-intelligence-grid">
            <div class="ov2-insight" style="--accent:#2563EB;">
                <div class="ov2-insight-head"><div class="ov2-icon" style="--accent:#2563EB;--halo:#DBEAFE;">{ov2_icon("focus")}</div>Concentración de ingresos</div>
                <div class="ov2-insight-value">{canon_top3_share:.0%}</div>
                <div class="ov2-insight-copy">+2 pts vs trimestre previo. Principal: {escape(canon_top1_name)}.</div>
                {ov2_spark([0, canon_top3_share * 0.55, canon_top3_share], "#2563EB")}
            </div>
            <div class="ov2-insight" style="--accent:#D97706;">
                <div class="ov2-insight-head"><div class="ov2-icon" style="--accent:#D97706;--halo:#FFF7ED;">{ov2_icon("coverage")}</div>Riesgo operativo</div>
                <div class="ov2-insight-value">{liquidity_state.title()}</div>
                <div class="ov2-insight-copy">CAPEX recuperado {cobertura_capex:.0%}; liquidez operacional en {cobertura_egresos:.1f} meses.</div>
            </div>
            <div class="ov2-insight" style="--accent:#22C55E;">
                <div class="ov2-insight-head"><div class="ov2-icon" style="--accent:#22C55E;--halo:#DCFCE7;">{ov2_icon("trend")}</div>Tendencia del activo</div>
                <div class="ov2-insight-value">{ingresos_growth_12m:+.0%}</div>
                <div class="ov2-insight-copy">Ingresos 12M vs período previo. Egresos: {egresos_growth_12m:+.0%}; canon: {canon_growth_12m:+.0%}.</div>
            </div>
            <div class="ov2-insight" style="--accent:#F87171;">
                <div class="ov2-insight-head"><div class="ov2-icon" style="--accent:#F87171;--halo:#FEE2E2;">{ov2_icon("alert")}</div>Alertas automáticas</div>
                <div class="ov2-insight-copy">
                    <div class="ov2-honorario-scenario" style="margin-top:0;padding-top:0;border-top:0;">
                        <div class="ov2-honorario-title">Escenario honorarios · base {fmt_clp_largo(honorario_base_cu)} c/u</div>
                        <div class="ov2-honorario-grid">
                            <strong>Plazo</strong><strong>Reducir c/u</strong><strong>Sueldo c/u</strong>
                            {honorario_rows_html}
                        </div>
                        <div class="ov2-honorario-note">Para cubrir {fmt_clp_largo(neto_a_cubrir_ov2)}. Actual {fmt_clp_largo(honorario_actual_cu)} c/u: ahorro {fmt_clp_largo(ahorro_actual_mes_ov2)}/mes, cubre en {meses_cubre_actual_ov2:.1f} meses.</div>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# =========================================================
# ⚠️ TAB 2: RIESGOS & COBRANZAS
# =========================================================
if active_section == "⚠️ Riesgo & Cobranza":
    def _return_to_gmail_cobranza():
        st.session_state["_scroll_to_gmail_cobranza_nonce"] = (
            st.session_state.get("_scroll_to_gmail_cobranza_nonce", 0) + 1
        )

    gmail_scroll_nonce = st.session_state.get("_scroll_to_gmail_cobranza_nonce", 0)
    gmail_scroll_rendered = st.session_state.get("_scroll_to_gmail_cobranza_rendered", 0)
    if gmail_scroll_nonce and gmail_scroll_nonce != gmail_scroll_rendered:
        st.session_state["_scroll_to_gmail_cobranza_rendered"] = gmail_scroll_nonce
        components.html(
            f"""
            <script>
            // nonce: {gmail_scroll_nonce}
            const doc = window.parent.document;
            let attempts = 0;
            const scrollToGmailCobranza = () => {{
                attempts += 1;
                const anchor = doc.getElementById("gmail-cobranza-anchor");
                if (anchor) {{
                    anchor.scrollIntoView({{behavior: "auto", block: "start"}});
                    window.parent.scrollBy(0, -18);
                    return true;
                }}
                if (attempts > 60) {{
                    return true;
                }}
                return false;
            }};
            const timer = setInterval(() => {{
                if (scrollToGmailCobranza()) {{
                    clearInterval(timer);
                }}
            }}, 50);
            setTimeout(scrollToGmailCobranza, 0);
            </script>
            """,
            height=0,
        )

              # ---------- Resumen por Responsable (NO PAGADO vs Abonos) ----------
    st.markdown(
        tab_header("Riesgo & Cobranza", "Monitoreo de cobranza, deuda y concentración de montos"),
        unsafe_allow_html=True,
    )
    st.markdown(
        section_heading("📋", "Cuentas por Cobrar / Pagar", weight_class="section-heading-title-soft"),
        unsafe_allow_html=True,
    )

    df_cob = df_f.copy()

    # --- Cálculos base ---
    df_np = df_cob[df_cob["Sit"] == "NO PAGADO"]
    df_abonos = df_cob[df_cob["Obs"].astype(str).str.contains("abono", case=False, na=False)]

    no_pagado_grouped = (
        df_np.groupby("Responsable")["Monto"]
        .agg(["sum", "count"])
        .rename(columns={"sum": "Monto NO PAGADO", "count": "Transacciones NO PAGADO"})
    )
    abonos_grouped = (
        df_abonos.groupby("Responsable")["Monto"]
        .agg(["sum", "count"])
        .rename(columns={"sum": "Monto Abonos", "count": "Cantidad Abonos"})
    )

    resumen = no_pagado_grouped.join(abonos_grouped, how="outer").fillna(0)
    resumen["Deuda"] = resumen["Monto NO PAGADO"] - resumen["Monto Abonos"]
    resumen["% Abonado"] = (
        resumen["Monto Abonos"] / resumen["Monto NO PAGADO"]
    ).replace([pd.NA, pd.NaT], 0).fillna(0)
    resumen["Progreso"] = resumen["% Abonado"].clip(lower=0, upper=1)

    def badge_pct(p):
        if p >= 1:
            return "🟢 En curso"
        if p >= 0.5:
            return "🟠 En curso"
        return "🔴 Bajo"

    resumen["Estado"] = resumen["% Abonado"].apply(badge_pct)

    # Ordenar por deuda y preparar tabla
    resumen = resumen.sort_values("Deuda", ascending=False)
    tabla = resumen.reset_index()

    cols_order = [
        "Responsable",
        "Monto NO PAGADO",
        "Monto Abonos",
        "Deuda",
        "% Abonado",
        "Transacciones NO PAGADO",
        "Cantidad Abonos",
        "Estado",
        "Progreso",  # queda al final para la barra
    ]
    tabla = tabla[cols_order]

    # ---------- KPIs en formato “card” ----------
    total_deuda_neta = tabla["Deuda"].sum() if not tabla.empty else 0
    st.markdown(
        """
        <style>
        .tab-title-row { margin:-61px 0 10px 0; }
        .section-heading-wrap { margin:14px 0 10px 0; }
        .risk-kpi-card {
            min-height:74px;
            border-radius:10px;
            border:1px solid var(--border);
            background:linear-gradient(135deg, #ffffff 0%, var(--soft) 100%);
            padding:9px 10px 8px 10px;
            display:grid;
            grid-template-columns:30px 1fr;
            gap:7px;
            box-shadow:0 8px 18px rgba(15,23,42,0.035);
        }
        .risk-kpi-icon {
            width:26px;
            height:26px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:var(--halo);
            color:var(--accent);
            font-size:13px;
            font-weight:950;
        }
        .risk-kpi-title {
            color:#0f1f3d;
            font-size:9.5px;
            line-height:1.12;
            font-weight:900;
            margin-bottom:3px;
            text-transform:uppercase;
            letter-spacing:.01em;
        }
        .risk-kpi-value {
            color:var(--accent);
            font-size:17px;
            line-height:1.05;
            font-weight:950;
            letter-spacing:-0.025em;
            white-space:nowrap;
        }
        .risk-kpi-badge {
            display:inline-block;
            margin-top:3px;
            padding:2px 5px;
            border-radius:5px;
            background:var(--badge-bg);
            color:var(--badge-fg);
            font-size:8.5px;
            font-weight:900;
        }
        .risk-kpi-note {
            grid-column:2;
            color:#475569;
            font-size:8.8px;
            line-height:1.18;
            font-weight:650;
            margin-top:1px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)

    kpi_titulo_deuda = (
        "DEUDA AL FIN DEL EJERCICIO"
        if total_deuda_neta < 0
        else "MONTO A FAVOR FIN DE EJERCICIO"
    )
    deuda_color = "#EF4444" if total_deuda_neta < 0 else "#10B981"
    bci_color = "#10B981" if balance_kpi >= 0 else "#EF4444"
    pos_neta_color = "#10B981" if posicion_neta >= 0 else "#EF4444"

    def risk_kpi_card(title, value, note, icon, accent, soft, halo, border, badge, positive=True):
        badge_bg = "#dcfce7" if positive else "#fee2e2"
        badge_fg = "#166534" if positive else "#B91C1C"
        return f"""
        <div class="risk-kpi-card" style="--accent:{accent};--soft:{soft};--halo:{halo};--border:{border};--badge-bg:{badge_bg};--badge-fg:{badge_fg};">
            <div class="risk-kpi-icon">{icon}</div>
            <div>
                <div class="risk-kpi-title">{title}</div>
                <div class="risk-kpi-value">{value}</div>
                <div class="risk-kpi-badge">{badge}</div>
            </div>
            <div class="risk-kpi-note">{note}</div>
        </div>
        """

    with col_kpi1:
        st.markdown(
            risk_kpi_card(
                kpi_titulo_deuda,
                f"${total_deuda_neta:,.0f}",
                "Resultado neto de cobros menos abonos",
                "↘" if total_deuda_neta < 0 else "↗",
                "#DC2626" if total_deuda_neta < 0 else "#047857",
                "#fff7f7" if total_deuda_neta < 0 else "#f6fffb",
                "#fde2e2" if total_deuda_neta < 0 else "#d8f5e4",
                "#f1caca" if total_deuda_neta < 0 else "#cfe9de",
                "Cobranzas",
                total_deuda_neta >= 0,
            ),
            unsafe_allow_html=True,
        )

    with col_kpi2:
        st.markdown(
            risk_kpi_card(
                "CAJA BANCO BCI",
                f"${balance_kpi:,.0f}",
                "Saldo operacional consolidado",
                "▰",
                "#047857" if balance_kpi >= 0 else "#DC2626",
                "#f6fffb" if balance_kpi >= 0 else "#fff7f7",
                "#d8f5e4" if balance_kpi >= 0 else "#fde2e2",
                "#cfe9de" if balance_kpi >= 0 else "#f1caca",
                "Liquidez",
                balance_kpi >= 0,
            ),
            unsafe_allow_html=True,
        )

    with col_kpi3:
        st.markdown(
            risk_kpi_card(
                "POSICIÓN NETA (BCI - DEUDA)",
                f"${posicion_neta:,.0f}",
                "Caja Banco BCI menos deuda al fin del ejercicio",
                "Σ",
                "#047857" if posicion_neta >= 0 else "#DC2626",
                "#f6fffb" if posicion_neta >= 0 else "#fff7f7",
                "#d8f5e4" if posicion_neta >= 0 else "#fde2e2",
                "#cfe9de" if posicion_neta >= 0 else "#f1caca",
                "Resumen",
                posicion_neta >= 0,
            ),
            unsafe_allow_html=True,
        )

    with col_kpi4:
        st.markdown(
            risk_kpi_card(
                "AVANCE DE COBRANZA",
                f"{pct_cobranza:.1%}",
                "Abonos registrados sobre cartera vencida acumulada",
                "%",
                "#1D4ED8",
                "#f6f9ff",
                "#e0ebff",
                "#d4e1f6",
                "Resumen",
                True,
            ),
            unsafe_allow_html=True,
        )

    def _fmt_clp_compact(value: float) -> str:
        amount = abs(float(value or 0))
        if amount >= 1_000_000:
            formatted = f"${amount / 1_000_000:.1f}M"
            return formatted.replace(".0M", "M")
        if amount >= 1_000:
            return f"${amount / 1_000:.0f}k"
        return f"${amount:,.0f}"

    if not tabla.empty:
        exposure_base = tabla.copy()
        exposure_base["Deuda_vigente_abs"] = exposure_base["Deuda"].abs()
        top_exposure_row = exposure_base.sort_values("Deuda_vigente_abs", ascending=False).iloc[0]
        top_exposure_name = escape(str(top_exposure_row["Responsable"]))
        top_exposure_value = _fmt_clp_compact(float(top_exposure_row["Deuda_vigente_abs"]))
    else:
        top_exposure_name = "Sin responsable dominante"
        top_exposure_value = "$0"

    mora_base = df_np.copy()
    if not mora_base.empty and "Responsable" in mora_base.columns:
        mora_counts = mora_base.groupby("Responsable").size().sort_values(ascending=False)
        top_mora_name = escape(str(mora_counts.index[0]))
        top_mora_count = int(mora_counts.iloc[0])
    else:
        top_mora_name = "Sin mora activa"
        top_mora_count = 0

    risk_group_col = next((c for c in ["Esp", "CC1", "Obs", "CC"] if c in mora_base.columns), None)
    if risk_group_col and not mora_base.empty:
        concentration_base = mora_base.copy()
        concentration_base["Monto_abs"] = pd.to_numeric(concentration_base["Monto"], errors="coerce").fillna(0).abs()
        risk_groups = concentration_base.groupby(risk_group_col)["Monto_abs"].sum().sort_values(ascending=False)
        risk_groups = risk_groups[risk_groups > 0]
        if not risk_groups.empty:
            top_risk_name = escape(str(risk_groups.index[0]))
            risk_total = float(risk_groups.sum())
            top_risk_pct = float(risk_groups.iloc[0] / risk_total) if risk_total else 0.0
        else:
            top_risk_name = "Sin concentración dominante"
            top_risk_pct = 0.0
    else:
        top_risk_name = "Sin concentración dominante"
        top_risk_pct = 0.0

    st.markdown(
        """
        <div style="
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%);
            border: 1px solid rgba(148,163,184,0.22);
            border-left: 4px solid #2563EB;
            border-radius: 9px;
            padding: 7px 11px 7px 12px;
            margin: 7px 0 12px 0;
            box-shadow: 0 6px 16px rgba(15,23,42,0.025);">
            <div style="
                color: #64748B;
                font-size: 9.5px;
                line-height: 1;
                font-weight: 900;
                letter-spacing: .08em;
                text-transform: uppercase;">
                Cartera operativa
            </div>
            <div style="
                color: #0F172A;
                font-size: 12px;
                line-height: 1.25;
                font-weight: 800;
                margin-top: 3px;">
                Estado de cobranza y exposición por responsable
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        <style>
        .top-risk-grid {{
            display:grid;
            grid-template-columns:repeat(3, minmax(0, 1fr));
            gap:10px;
            margin:0 0 14px 0;
        }}
        .top-risk-card {{
            background:#FFFFFF;
            border:1px solid var(--risk-border);
            border-left:3px solid var(--risk-accent);
            border-radius:11px;
            padding:9px 11px;
            box-shadow:0 8px 18px rgba(15,23,42,0.035);
            display:grid;
            grid-template-columns:28px 1fr;
            gap:8px;
            align-items:center;
            min-height:66px;
        }}
        .top-risk-icon {{
            width:25px;
            height:25px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:var(--risk-halo);
            color:var(--risk-accent);
            font-size:13px;
            font-weight:950;
        }}
        .top-risk-label {{
            color:#64748B;
            font-size:9px;
            line-height:1;
            font-weight:900;
            letter-spacing:.07em;
            text-transform:uppercase;
            margin-bottom:5px;
        }}
        .top-risk-main {{
            color:#0F172A;
            font-size:13px;
            line-height:1.2;
            font-weight:900;
        }}
        .top-risk-main span {{
            color:var(--risk-accent);
            font-weight:950;
        }}
        .top-risk-sub {{
            color:#64748B;
            font-size:10px;
            line-height:1.2;
            font-weight:650;
            margin-top:3px;
        }}
        @media (max-width: 900px) {{
            .top-risk-grid {{ grid-template-columns:1fr; }}
        }}
        </style>
        <div class="top-risk-grid">
            <div class="top-risk-card" style="--risk-accent:#D97706;--risk-border:rgba(217,119,6,0.20);--risk-halo:rgba(245,158,11,0.14);">
                <div class="top-risk-icon">!</div>
                <div>
                    <div class="top-risk-label">Principal exposición</div>
                    <div class="top-risk-main">{top_exposure_name} → <span>{top_exposure_value}</span></div>
                    <div class="top-risk-sub">pendiente vigente</div>
                </div>
            </div>
            <div class="top-risk-card" style="--risk-accent:#DC2626;--risk-border:rgba(220,38,38,0.18);--risk-halo:rgba(254,226,226,0.80);">
                <div class="top-risk-icon">!</div>
                <div>
                    <div class="top-risk-label">Mayor mora</div>
                    <div class="top-risk-main">{top_mora_name} → <span>{top_mora_count}</span></div>
                    <div class="top-risk-sub">transacciones abiertas</div>
                </div>
            </div>
            <div class="top-risk-card" style="--risk-accent:#EA580C;--risk-border:rgba(234,88,12,0.18);--risk-halo:rgba(255,237,213,0.85);">
                <div class="top-risk-icon">!</div>
                <div>
                    <div class="top-risk-label">Riesgo operativo</div>
                    <div class="top-risk-main">{top_risk_name} representa <span>{top_risk_pct:.0%}</span></div>
                    <div class="top-risk-sub">de deuda vigente</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Estilo visual de la tabla ----
    deuda_cmap = LinearSegmentedColormap.from_list(
        "deuda_palette",
        ["#FFFBEB", "#FEF3C7", "#FED7AA", "#FCA5A5"],
    )

    def _style_deuda_critica(v):
        try:
            n = float(v)
        except Exception:
            return ""
        magnitude = abs(n)
        if magnitude <= 0:
            return "color:#64748B; font-weight:650; background-color:#F8FAFC;"
        if magnitude >= 5_000_000:
            color = "#991B1B"
            border = "rgba(185,28,28,0.42)"
        elif magnitude >= 1_000_000:
            color = "#B45309"
            border = "rgba(217,119,6,0.36)"
        else:
            color = "#92400E"
            border = "rgba(245,158,11,0.24)"
        return (
            f"color:{color}; font-weight:780; "
            "font-variant-numeric:tabular-nums; "
            f"box-shadow:inset 3px 0 0 {border};"
        )
    styler = (
        tabla.style
        .format({
            "Monto NO PAGADO": "${:,.0f}",
            "Monto Abonos": "${:,.0f}",
            "Deuda": "${:,.0f}",
            "% Abonado": "{:.1%}",
            "Transacciones NO PAGADO": "{:,.0f}",
            "Cantidad Abonos": "{:,.0f}",
            "Progreso": "{:.0%}",   # ahora se ve 94% en vez de 0.94
        })
        .hide(axis="index")
        .set_table_styles([
            {
                "selector": "thead th",
                "props": [
                    ("background-color", "#163A5F"),
                    ("color", "white"),
                    ("font-weight", "700"),
                    ("font-size", "13px"),
                    ("border-bottom", "1px solid #0F2740"),
                    ("text-align", "center"),
                    ("padding", "8px"),
                ],
            },
            {
                "selector": "tbody td",
                "props": [
                    ("font-size", "12px"),
                    ("padding", "7px 8px"),
                    ("border-bottom", "1px solid #E5E7EB"),
                ],
            },
            {
                "selector": "tbody tr:nth-child(even)",
                "props": [("background-color", "#F8FAFC")],
            },
            {
                "selector": "tbody tr:hover",
                "props": [("background-color", "#EEF2F7")],
            },
        ])
        .set_properties(subset=["Responsable"], **{"text-align": "left", "font-weight": "400"})
        .set_properties(
            subset=["Deuda"],
            **{
                "font-weight": "780",
                "font-variant-numeric": "tabular-nums",
                "border-left": "1px solid rgba(245,158,11,0.22)",
                "border-right": "1px solid rgba(245,158,11,0.18)",
            },
        )
        .set_properties(subset=["Monto NO PAGADO"], **{"font-weight": "400", "color": "#7A271A"})
        .set_properties(subset=["Monto Abonos"], **{"font-weight": "400", "color": "#027A48"})
        .background_gradient(subset=["Deuda"], cmap=deuda_cmap, gmap=tabla["Deuda"].abs())
        .map(_style_deuda_critica, subset=["Deuda"])
        .bar(subset=["Progreso"], color="#10B981")
    )

    st.dataframe(styler, use_container_width=True)

    st.markdown(
        '<div id="gmail-cobranza-anchor"></div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        section_heading("🧾", "Información cobranza Gmail", weight_class="section-heading-title-soft"),
        unsafe_allow_html=True,
    )
    st.caption("Detalle por concepto según Año y Mes, para Esp 1..7.")

    df_cancel = df_f.dropna(subset=["Monto"]).copy()

    years_cancel = sorted(df_cancel["Año_sel"].dropna().astype(int).unique().tolist())
    year_opts_cancel = ["Todos"] + years_cancel

    c_can1, c_can2, c_can4, c_can5 = st.columns([1, 1, 1, 1.3])
    with c_can1:
        sel_year_cancel = st.selectbox(
            "Año (cancelación)",
            year_opts_cancel,
            index=0,
            key="year_cancel_esp",
            on_change=_return_to_gmail_cobranza,
        )
    with c_can2:
        sel_month_cancel = st.selectbox(
            "Mes (cancelación)",
            ["Todos"] + list(range(1, 13)),
            index=0,
            key="month_cancel_esp",
            on_change=_return_to_gmail_cobranza,
        )
    with c_can4:
        sel_esp_cancel = st.selectbox(
            "Espacio",
            ["Todos"] + list(range(1, 8)),
            index=0,
            key="esp_cancel_esp",
            on_change=_return_to_gmail_cobranza,
        )
    df_resp_opts = df_f.copy()
    df_resp_opts = df_resp_opts[df_resp_opts["Esp_num"].between(1, 7, inclusive="both")]
    if sel_esp_cancel != "Todos":
        df_resp_opts = df_resp_opts[df_resp_opts["Esp_num"] == int(sel_esp_cancel)]
    responsables_opts_cancel = ["Todos"] + sorted(
        df_resp_opts["Responsable_clean"]
        .dropna()
        .loc[lambda s: s != ""]
        .unique()
        .tolist()
    )
    with c_can5:
        sel_resp_cancel = st.selectbox(
            "Responsable",
            responsables_opts_cancel,
            index=0,
            key="resp_cancel_esp",
            on_change=_return_to_gmail_cobranza,
        )

    if sel_year_cancel != "Todos":
        df_cancel = df_cancel[df_cancel["Año_sel"] == sel_year_cancel]
    if sel_month_cancel != "Todos":
        df_cancel = df_cancel[df_cancel["Mes_sel"] == sel_month_cancel]

    df_cancel = df_cancel[df_cancel["Sit_norm"].isin(["PAGADO", "NO PAGADO"])]
    df_cancel = df_cancel[df_cancel["Esp_num"].between(1, 7, inclusive="both")]
    df_cancel["Esp_num"] = df_cancel["Esp_num"].astype(int)
    if sel_esp_cancel != "Todos":
        df_cancel = df_cancel[df_cancel["Esp_num"] == int(sel_esp_cancel)]
    if sel_resp_cancel != "Todos":
        df_cancel = df_cancel[df_cancel["Responsable_clean"] == sel_resp_cancel]
    df_cancel_scope = df_cancel.copy()

    obs_cancel_norm = (
        df_cancel["Obs_text"].fillna("").astype(str)
        .str.normalize("NFKD")
        .str.encode("ascii", errors="ignore")
        .str.decode("utf-8")
        .str.casefold()
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    txt_cancel = (
        df_cancel["CC1_text"].fillna("").astype(str)
        + " "
        + df_cancel["Obs_text"].fillna("").astype(str)
    ).str.normalize("NFKD").str.encode("ascii", errors="ignore").str.decode("utf-8").str.casefold()

    mask_obs_gc = obs_cancel_norm.str.contains(r"\bpago arrendatario\s*-\s*gc\b", regex=True, na=False)
    mask_obs_verisure = obs_cancel_norm.str.contains(r"\bpago arrendatario\s*-\s*verisure\b", regex=True, na=False)
    mask_obs_interes = obs_cancel_norm.str.contains(r"\binteres bancario\b", regex=True, na=False)
    mask_obs_garantia = obs_cancel_norm.str.contains(r"\bgarantia\b", regex=True, na=False)

    df_cancel["Concepto"] = np.select(
        [
            txt_cancel.str.contains(r"canon\s*mensual", regex=True, na=False),
            mask_obs_gc,
            txt_cancel.str.contains(r"\bcge\b|boleta\s*cge|electricidad", regex=True, na=False),
            mask_obs_verisure,
            mask_obs_interes,
            mask_obs_garantia,
        ],
        ["Canon mensual", "Gastos comunes", "CGE", "Verisure", "Administrativo", "Garantia"],
        default="Otros",
    )

    conceptos_objetivo = ["Canon mensual", "Gastos comunes", "CGE", "Verisure", "Administrativo", "Garantia"]
    df_cancel = df_cancel[df_cancel["Concepto"].isin(conceptos_objetivo)].copy()
    df_cancel["Monto_abs"] = df_cancel["Monto"].abs()

    # Deuda por responsable: NO PAGADO - ABONO, excluyendo el período seleccionado.
    df_deuda = df_f.dropna(subset=["Monto"]).copy()
    df_deuda = df_deuda[df_deuda["Responsable_clean"] != ""]
    df_deuda = df_deuda.dropna(subset=["Periodo_ref"])

    corte_periodo = None
    if sel_year_cancel != "Todos":
        if sel_month_cancel != "Todos":
            corte_periodo = pd.Timestamp(int(sel_year_cancel), int(sel_month_cancel), 1)
        else:
            corte_periodo = pd.Timestamp(int(sel_year_cancel), 1, 1)
    if corte_periodo is not None:
        df_deuda = df_deuda[df_deuda["Periodo_ref"] < corte_periodo]

    # Responsables válidos según el período seleccionado y filtros activos en la vista
    responsables_periodo = (
        df_cancel_scope["Responsable_clean"]
        .dropna()
        .loc[lambda s: s != ""]
        .unique()
        .tolist()
    )
    if sel_resp_cancel != "Todos":
        responsables_periodo = [sel_resp_cancel]
    if responsables_periodo:
        df_deuda = df_deuda[df_deuda["Responsable_clean"].isin(responsables_periodo)]
    else:
        df_deuda = df_deuda.iloc[0:0]

    deuda_np_por_resp = (
        df_deuda[df_deuda["Sit_norm"] == "NO PAGADO"]
        .groupby("Responsable_clean")["Monto"]
        .sum()
    )
    abonos_por_resp = (
        df_deuda[df_deuda["Obs_text"].str.contains("abono", case=False, na=False)]
        .groupby("Responsable_clean")["Monto"]
        .sum()
    )
    deuda_por_resp = deuda_np_por_resp - abonos_por_resp

    if sel_esp_cancel != "Todos":
        idx_esp = [int(sel_esp_cancel)]
    elif sel_resp_cancel != "Todos":
        idx_esp = sorted(
            df_f.loc[
                df_f["Responsable_clean"] == sel_resp_cancel,
                "Esp_num"
            ]
            .dropna()
            .astype(int)
            .loc[lambda s: s.between(1, 7)]
            .unique()
            .tolist()
        )
        if not idx_esp:
            idx_esp = list(range(1, 8))
    else:
        idx_esp = list(range(1, 8))

    if df_cancel.empty:
        tabla_cancel = pd.DataFrame(index=idx_esp)
    else:
        tabla_cancel = (
            df_cancel.groupby(["Esp_num", "Concepto"], as_index=False)["Monto_abs"]
            .sum()
            .pivot(index="Esp_num", columns="Concepto", values="Monto_abs")
            .fillna(0)
        )

    tabla_cancel = tabla_cancel.reindex(index=idx_esp, fill_value=0)
    for c in conceptos_objetivo:
        if c not in tabla_cancel.columns:
            tabla_cancel[c] = 0

    responsables_por_esp = (
        df_cancel_scope.loc[lambda d: d["Responsable_clean"] != ""]
        .groupby("Esp_num")["Responsable_clean"]
        .apply(lambda s: ", ".join(sorted(s.dropna().unique().tolist())))
    )
    responsables_lista_por_esp = (
        df_cancel_scope.loc[lambda d: d["Responsable_clean"] != ""]
        .groupby("Esp_num")["Responsable_clean"]
        .apply(lambda s: sorted(s.dropna().unique().tolist()))
    )

    mostrar_garantia = float(pd.to_numeric(tabla_cancel["Garantia"], errors="coerce").fillna(0).abs().sum()) > 0
    conceptos_visibles = [c for c in conceptos_objetivo if c != "Garantia" or mostrar_garantia]

    tabla_cancel = tabla_cancel[conceptos_visibles]
    tabla_cancel.insert(
        0,
        "Responsable",
        tabla_cancel.index.to_series().map(responsables_por_esp).fillna("-"),
    )
    tabla_cancel["Deuda"] = tabla_cancel.index.to_series().map(
        lambda esp: sum(deuda_por_resp.get(r, 0) for r in responsables_lista_por_esp.get(esp, []))
    ).fillna(0)
    tabla_cancel["Total a cancelar"] = tabla_cancel[conceptos_visibles + ["Deuda"]].sum(axis=1)
    if sel_resp_cancel != "Todos" and len(idx_esp) > 1:
        tabla_cancel = tabla_cancel[
            pd.to_numeric(tabla_cancel["Total a cancelar"], errors="coerce").fillna(0) != 0
        ]
    tabla_cancel.index = [f"Esp {i}" for i in tabla_cancel.index]

    if df_cancel.empty:
        st.info("En el período seleccionado no hay cargos de conceptos, se muestra deuda acumulada previa.")

    periodo_lbl = (
        f"{int(sel_year_cancel)}-{int(sel_month_cancel):02d}"
        if sel_year_cancel != "Todos" and sel_month_cancel != "Todos"
        else (f"Año {int(sel_year_cancel)}" if sel_year_cancel != "Todos" else "Todos los períodos")
    )
    st.markdown(
        f"""
        <div style="
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%);
            border: 1px solid rgba(148,163,184,0.22);
            border-left: 4px solid #2563EB;
            border-radius: 9px;
            padding: 7px 11px 7px 12px;
            margin: 7px 0 12px 0;
            box-shadow: 0 6px 16px rgba(15,23,42,0.025);">
            <div style="
                color:#64748B;
                font-size:9.5px;
                line-height:1;
                font-weight:900;
                letter-spacing:.08em;
                text-transform:uppercase;">
                Cobranza arrendatarios
            </div>
            <div style="
                color:#0F172A;
                font-size:12px;
                line-height:1.25;
                font-weight:800;
                margin-top:3px;">
                Estado de cobro consolidado · Periodo: {periodo_lbl}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    tabla_cancel_view = tabla_cancel.rename(columns={"Administrativo": "Interes/ otros"})
    cols_monto_cancel = [c for c in tabla_cancel_view.columns if c != "Responsable"]
    deuda_cancel_cmap = LinearSegmentedColormap.from_list(
        "deuda_cancel_palette",
        ["#FFFBEB", "#FEF3C7", "#FED7AA", "#FCA5A5"],
    )
    styler_cancel = (
        tabla_cancel_view.style
        .format("${:,.0f}", subset=cols_monto_cancel)
        .set_table_styles([
            {
                "selector": "thead th",
                "props": [
                    ("background", "linear-gradient(180deg,#F8FAFC 0%,#EEF3F8 100%)"),
                    ("color", "#475569"),
                    ("font-weight", "850"),
                    ("font-size", "11px"),
                    ("border-bottom", "1px solid rgba(148,163,184,0.24)"),
                    ("text-align", "center"),
                    ("padding", "6px 8px"),
                    ("letter-spacing", ".04em"),
                    ("text-transform", "uppercase"),
                ],
            },
            {
                "selector": "tbody td",
                "props": [
                    ("font-size", "12px"),
                    ("padding", "7px 8px"),
                    ("border-bottom", "1px solid #E5E7EB"),
                ],
            },
            {
                "selector": "tbody tr:nth-child(even)",
                "props": [("background-color", "#F8FAFC")],
            },
            {
                "selector": "tbody tr:hover",
                "props": [("background-color", "#EEF2F7")],
            },
        ])
        .set_properties(subset=["Responsable"], **{"text-align": "left", "font-weight": "600"})
        .set_properties(
            subset=["Deuda"],
            **{
                "font-weight": "780",
                "font-variant-numeric": "tabular-nums",
                "border-left": "1px solid rgba(245,158,11,0.22)",
                "border-right": "1px solid rgba(245,158,11,0.18)",
            },
        )
        .set_properties(subset=["Total a cancelar"], **{"font-weight": "800", "color": "#0F2D52"})
        .background_gradient(
            subset=["Deuda"],
            cmap=deuda_cancel_cmap,
            gmap=tabla_cancel_view["Deuda"].abs(),
        )
        .map(_style_deuda_critica, subset=["Deuda"])
        .background_gradient(subset=["Total a cancelar"], cmap="Blues")
    )
    st.dataframe(
        styler_cancel,
        use_container_width=True,
    )

    esp_lbl_chart = str(sel_esp_cancel)
    resp_lbl_chart = str(sel_resp_cancel)
    st.markdown(
        f"""
        <div style="
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%);
            border: 1px solid rgba(148,163,184,0.22);
            border-left: 4px solid #2563EB;
            border-radius: 9px;
            padding: 7px 11px 7px 12px;
            margin: 10px 0 8px 0;
            box-shadow: 0 6px 16px rgba(15,23,42,0.025);">
            <div style="
                color:#64748B;
                font-size:9.5px;
                line-height:1;
                font-weight:900;
                letter-spacing:.08em;
                text-transform:uppercase;">
                Composición de cobro
            </div>
            <div style="
                color:#0F172A;
                font-size:12px;
                line-height:1.25;
                font-weight:800;
                margin-top:3px;">
                Vista por espacio · Espacio: {esp_lbl_chart} · Responsable: {resp_lbl_chart}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.caption("Barras apiladas por concepto + deuda, con línea de total a cancelar.")

    import plotly.graph_objects as go

    chart_cols = conceptos_visibles + ["Deuda"]
    chart_df = tabla_cancel.reset_index().rename(columns={"index": "Espacio"}).copy()
    single_space_view = len(chart_df) == 1
    responsable_un_espacio = sel_resp_cancel != "Todos" and len(idx_esp) == 1
    single_month_one_space = (
        single_space_view
        and (
            (
                sel_year_cancel != "Todos"
                and sel_month_cancel != "Todos"
                and str(sel_esp_cancel).strip().lower() not in {"todos", "todas"}
            )
            or responsable_un_espacio
        )
    )

    fig_cancel = go.Figure()
    color_map = {
        "Canon mensual": "#4B5563",
        "Gastos comunes": "#7FA6A2",
        "CGE": "#DCAA67",
        "Verisure": "#A8A8A8",
        "Administrativo": "#D85E5D",
        "Garantia": "#1D4ED8",
        "Deuda": "#D85E5D",
    }
    donut_panel_metrics = None

    if single_month_one_space:
        row = chart_df.iloc[0]
        raw_vals = pd.Series({c: float(pd.to_numeric(row[c], errors="coerce") or 0) for c in chart_cols})
        pos_vals = raw_vals[raw_vals > 0].sort_values(ascending=False)
        neg_total = float(raw_vals[raw_vals < 0].sum())
        total_single = float(pd.to_numeric(row["Total a cancelar"], errors="coerce") or 0)
        deuda_single = float(pd.to_numeric(row.get("Deuda", 0), errors="coerce") or 0)
        total_components_abs = float(raw_vals.abs().sum())
        deuda_share = abs(deuda_single) / total_components_abs if total_components_abs else 0.0
        principal_component = str(pos_vals.index[0]) if not pos_vals.empty else "Sin componente dominante"
        principal_component_value = float(pos_vals.iloc[0]) if not pos_vals.empty else 0.0
        principal_component_share = principal_component_value / float(pos_vals.sum()) if float(pos_vals.sum()) else 0.0
        responsable_single = str(row.get("Responsable", resp_lbl_chart))
        if deuda_share >= 0.40:
            risk_badge = "Exposición crítica"
            risk_color = "#B42318"
            risk_bg = "rgba(254,226,226,0.86)"
            risk_state = "Presión de cobranza crítica"
        elif deuda_share >= 0.22:
            risk_badge = "Exposición alta"
            risk_color = "#C2410C"
            risk_bg = "rgba(255,237,213,0.92)"
            risk_state = "Presión de cobranza alta"
        elif deuda_share >= 0.08:
            risk_badge = "Exposición media"
            risk_color = "#D97706"
            risk_bg = "rgba(245,158,11,0.14)"
            risk_state = "Seguimiento preventivo"
        else:
            risk_badge = "Exposición baja"
            risk_color = "#047857"
            risk_bg = "rgba(220,252,231,0.84)"
            risk_state = "Sin deuda operativa relevante"
        donut_panel_metrics = {
            "espacio": str(chart_df["Espacio"].iloc[0]),
            "responsable": responsable_single,
            "total": total_single,
            "principal_component": principal_component,
            "principal_value": principal_component_value,
            "principal_share": principal_component_share,
            "deuda": deuda_single,
            "deuda_share": deuda_share,
            "risk_badge": risk_badge,
            "risk_color": risk_color,
            "risk_bg": risk_bg,
            "risk_state": risk_state,
        }

        if not pos_vals.empty:
            fig_cancel.add_trace(
                go.Pie(
                    labels=pos_vals.index.tolist(),
                    values=pos_vals.values.tolist(),
                    hole=0.62,
                    sort=False,
                    marker=dict(
                        colors=[color_map.get(c, "#64748B") for c in pos_vals.index.tolist()],
                        line=dict(color="white", width=1),
                    ),
                    textinfo="percent",
                    textfont=dict(size=13, color="#FFFFFF"),
                    hovertemplate="<b>%{label}</b><br>Monto: $%{value:,.0f}<br>Participación: %{percent}<extra></extra>",
                    domain=dict(x=[0.05, 0.95], y=[0.13, 0.90]),
                )
            )
        else:
            # Sin componentes positivos: fallback simple para evitar gráfico vacío.
            fig_cancel.add_trace(
                go.Bar(
                    y=[chart_df["Espacio"].iloc[0]],
                    x=[total_single],
                    orientation="h",
                    name="Total a cancelar",
                    marker_color="#1D4ED8",
                    hovertemplate="<b>%{y}</b><br>Total: $%{x:,.0f}<extra></extra>",
                )
            )

        center_text = (
            f"<span style='font-size:20px'><b>{chart_df['Espacio'].iloc[0]}</b></span><br>"
            f"<span style='font-size:12px'>{responsable_single}</span><br>"
            f"<span style='font-size:11px;color:#64748B'>Total</span><br>"
            f"<span style='font-size:17px'><b>${total_single:,.0f}</b></span><br>"
            f"<span style='font-size:10px;color:{risk_color}'>{risk_badge}</span>"
        )
        fig_cancel.add_annotation(
            x=0.5,
            y=0.515,
            xref="paper",
            yref="paper",
            text=center_text,
            showarrow=False,
            align="center",
            font=dict(size=16, color="#0F172A"),
        )
        if neg_total < 0:
            fig_cancel.add_annotation(
                x=0.5,
                y=0.06,
                xref="paper",
                yref="paper",
                text=f"Ajustes negativos considerados: ${neg_total:,.0f}",
                showarrow=False,
                font=dict(size=11, color="#B42318"),
            )
    elif single_space_view:
        for c in chart_cols:
            fig_cancel.add_trace(
                go.Bar(
                    y=chart_df["Espacio"],
                    x=chart_df[c],
                    orientation="h",
                    name=c,
                    marker_color=color_map.get(c, "#64748B"),
                    hovertemplate="<b>%{y}</b><br>" + c + ": $%{x:,.0f}<extra></extra>",
                )
            )
        total_single = float(chart_df["Total a cancelar"].iloc[0])
        fig_cancel.add_annotation(
            x=total_single,
            y=chart_df["Espacio"].iloc[0],
            text=f"Total: ${total_single:,.0f}",
            showarrow=False,
            xshift=14,
            font=dict(size=12, color="#0F172A"),
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="rgba(15,45,82,0.20)",
            borderwidth=1,
        )
    else:
        for c in chart_cols:
            fig_cancel.add_trace(
                go.Bar(
                    x=chart_df["Espacio"],
                    y=chart_df[c],
                    name=c,
                    marker_color=color_map.get(c, "#64748B"),
                    hovertemplate="<b>%{x}</b><br>" + c + ": $%{y:,.0f}<extra></extra>",
                )
            )
        fig_cancel.add_trace(
            go.Scatter(
                x=chart_df["Espacio"],
                y=chart_df["Total a cancelar"],
                mode="lines+markers+text",
                name="Total a cancelar",
                line=dict(color="#4B5563", width=3),
                marker=dict(size=8, color="#4B5563"),
                text=[f"${v:,.0f}" for v in chart_df["Total a cancelar"]],
                textposition="top center",
                hovertemplate="<b>%{x}</b><br>Total: $%{y:,.0f}<extra></extra>",
            )
        )

    fig_cancel.update_layout(
        barmode="stack",
        template="plotly_white",
        height=390 if single_month_one_space else (280 if single_space_view else 380),
        margin=dict(l=8 if single_month_one_space else 20, r=8 if single_month_one_space else 20, t=12 if single_month_one_space else 92, b=58 if single_month_one_space else 18),
        legend=dict(
            orientation="h",
            yanchor="top" if single_month_one_space else "bottom",
            y=-0.03 if single_month_one_space else 1.10,
            x=0.5 if single_month_one_space else 0.01,
            xanchor="center" if single_month_one_space else "left",
            bgcolor="rgba(255,255,255,0.85)",
            bordercolor="rgba(15,45,82,0.15)",
            borderwidth=1,
            font=dict(size=10 if single_month_one_space else 12),
        ),
        title=dict(
            text="" if single_month_one_space else "📊 Composición de Cobro por Espacio",
            x=0.01,
            xanchor="left",
            font=dict(size=18, color="#0F2D52"),
            pad=dict(b=14),
        ),
        xaxis_title=("Monto (CLP)" if single_space_view else "Espacio"),
        yaxis_title=("" if single_space_view else "Monto (CLP)"),
        hovermode="closest" if single_month_one_space else ("y unified" if single_space_view else "x unified"),
        paper_bgcolor="#FFFFFF" if single_month_one_space else "#F8FAFC",
        plot_bgcolor="#FFFFFF",
    )
    if single_month_one_space:
        fig_cancel.update_xaxes(visible=False, showgrid=False, zeroline=False)
        fig_cancel.update_yaxes(visible=False, showgrid=False, zeroline=False)
    elif single_space_view:
        fig_cancel.update_xaxes(
            tickformat=",.0f",
            gridcolor="rgba(15,45,82,0.10)",
            zeroline=False,
            linecolor="rgba(15,45,82,0.20)",
        )
        fig_cancel.update_yaxes(showgrid=False, linecolor="rgba(15,45,82,0.25)")
    else:
        fig_cancel.update_xaxes(showgrid=False, linecolor="rgba(15,45,82,0.25)")
        fig_cancel.update_yaxes(
            tickformat=",.0f",
            gridcolor="rgba(15,45,82,0.10)",
            zeroline=False,
            linecolor="rgba(15,45,82,0.20)",
        )

    if single_month_one_space and donut_panel_metrics:
        metric = donut_panel_metrics
        plot_html = fig_cancel.to_html(
            full_html=False,
            include_plotlyjs=True,
            config={
                "displaylogo": False,
                "displayModeBar": True,
                "modeBarButtonsToAdd": ["toImage"],
            },
        )
        principal_label = escape(metric["principal_component"])
        responsable_label = escape(metric["responsable"])
        components.html(
            f"""
            <div class="cobro-exec-card">
                <div class="cobro-donut-side">
                    <div class="cobro-card-eyebrow">Lectura financiera filtrada</div>
                    <div class="cobro-card-title">Composición de cobro · {escape(metric["espacio"])}</div>
                    <div class="plot-wrap">{plot_html}</div>
                </div>
                <div class="cobro-kpi-side">
                    <div class="risk-badge" style="--risk-bg:{metric["risk_bg"]};--risk-color:{metric["risk_color"]};">{metric["risk_badge"]}</div>
                    <div class="kpi-block primary">
                        <div class="kpi-label">Total a cancelar</div>
                        <div class="kpi-value">${metric["total"]:,.0f}</div>
                        <div class="kpi-note">{responsable_label}</div>
                    </div>
                    <div class="kpi-block">
                        <div class="kpi-label">Principal componente</div>
                        <div class="kpi-value small">{principal_label}</div>
                        <div class="kpi-note">${metric["principal_value"]:,.0f} · {metric["principal_share"]:.1%}</div>
                    </div>
                    <div class="kpi-grid">
                        <div class="mini-kpi">
                            <span>% deuda operativa</span>
                            <strong>{metric["deuda_share"]:.1%}</strong>
                        </div>
                        <div class="mini-kpi">
                            <span>Riesgo cobranza</span>
                            <strong style="color:{metric["risk_color"]};">{metric["risk_state"]}</strong>
                        </div>
                    </div>
                    <div class="risk-meter">
                        <div style="width:{min(metric["deuda_share"], 1.0) * 100:.1f}%;background:{metric["risk_color"]};"></div>
                    </div>
                </div>
            </div>
            <style>
                body {{
                    margin:0;
                    background:transparent;
                    font-family: "Inter", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                }}
                .cobro-exec-card {{
                    min-height:455px;
                    display:grid;
                    grid-template-columns:minmax(0, 1.58fr) minmax(270px, .72fr);
                    gap:18px;
                    padding:18px;
                    border-radius:18px;
                    background:#FFFFFF;
                    border:1px solid rgba(148,163,184,0.18);
                    box-shadow:0 16px 36px rgba(15,23,42,0.055);
                    box-sizing:border-box;
                }}
                .cobro-card-eyebrow {{
                    color:#64748B;
                    font-size:10px;
                    font-weight:900;
                    text-transform:uppercase;
                    letter-spacing:.08em;
                    margin:2px 0 4px 2px;
                }}
                .cobro-card-title {{
                    color:#0F172A;
                    font-size:19px;
                    font-weight:900;
                    letter-spacing:-.01em;
                    margin:0 0 3px 2px;
                }}
                .plot-wrap {{
                    height:398px;
                    overflow:hidden;
                }}
                .cobro-kpi-side {{
                    border-radius:15px;
                    background:linear-gradient(180deg,#F8FAFC 0%,#FFFFFF 100%);
                    border:1px solid rgba(148,163,184,0.16);
                    padding:16px;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    gap:12px;
                    box-sizing:border-box;
                }}
                .risk-badge {{
                    align-self:flex-start;
                    background:var(--risk-bg);
                    color:var(--risk-color);
                    border:1px solid color-mix(in srgb, var(--risk-color) 22%, transparent);
                    border-radius:999px;
                    padding:5px 9px;
                    font-size:11px;
                    font-weight:900;
                }}
                .kpi-block {{
                    padding:12px;
                    border-radius:13px;
                    background:#FFFFFF;
                    border:1px solid rgba(226,232,240,0.82);
                }}
                .kpi-block.primary {{
                    border-color:rgba(37,99,235,0.18);
                    box-shadow:inset 3px 0 0 rgba(37,99,235,0.75);
                }}
                .kpi-label {{
                    color:#64748B;
                    font-size:10px;
                    font-weight:900;
                    letter-spacing:.06em;
                    text-transform:uppercase;
                }}
                .kpi-value {{
                    color:#0F172A;
                    font-size:25px;
                    line-height:1.08;
                    font-weight:950;
                    margin-top:5px;
                    letter-spacing:-.025em;
                }}
                .kpi-value.small {{
                    font-size:17px;
                    letter-spacing:-.01em;
                }}
                .kpi-note {{
                    color:#64748B;
                    font-size:11px;
                    font-weight:700;
                    margin-top:4px;
                }}
                .kpi-grid {{
                    display:grid;
                    grid-template-columns:1fr 1fr;
                    gap:9px;
                }}
                .mini-kpi {{
                    background:#FFFFFF;
                    border:1px solid rgba(226,232,240,0.82);
                    border-radius:12px;
                    padding:10px;
                }}
                .mini-kpi span {{
                    display:block;
                    color:#64748B;
                    font-size:9px;
                    font-weight:900;
                    text-transform:uppercase;
                    letter-spacing:.05em;
                    line-height:1.15;
                }}
                .mini-kpi strong {{
                    display:block;
                    margin-top:5px;
                    color:#0F172A;
                    font-size:14px;
                    line-height:1.15;
                    font-weight:900;
                }}
                .risk-meter {{
                    height:8px;
                    border-radius:999px;
                    background:#E2E8F0;
                    overflow:hidden;
                }}
                .risk-meter div {{
                    height:100%;
                    min-width:6px;
                    border-radius:999px;
                }}
            </style>
            """,
            height=492,
        )
    else:
        st.plotly_chart(
            fig_cancel,
            use_container_width=True,
            config={
                "displaylogo": False,
                "displayModeBar": True,
                "modeBarButtonsToAdd": ["toImage"],
            },
        )

    # ---- Exportar PDF / CSV ----
    from reportlab.lib.pagesizes import A4, A3, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    def export_resumen_pdf(df_in: pd.DataFrame, filename: str = "resumen_no_pagado_abonos.pdf"):
        short_names = {
            "Monto NO PAGADO": "Monto NO PAG.",
            "Transacciones NO PAGADO": "Transacc. NO PAG.",
            "Cantidad Abonos": "Cant. Abonos",
            "Monto Abonos": "Monto Abonos",
            "% Abonado": "% Abonado",
        }
        df_local = df_in.rename(columns={c: short_names.get(c, c) for c in df_in.columns})

        ncols = len(df_local.columns)
        page_size = landscape(A4) if ncols <= 8 else landscape(A3)

        doc = SimpleDocTemplate(
            filename,
            pagesize=page_size,
            leftMargin=1*cm,
            rightMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm,
        )

        base = ParagraphStyle(
            "base", fontSize=8.5, leading=10,
            spaceBefore=0, spaceAfter=0,
            wordWrap="CJK", splitLongWords=True,
        )
        sty_h = ParagraphStyle("head", parent=base, alignment=TA_CENTER, fontSize=9, leading=11)
        sty_l = ParagraphStyle("left", parent=base, alignment=TA_LEFT)
        sty_r = ParagraphStyle("right", parent=base, alignment=TA_RIGHT)

        story = [
            Paragraph("<b>Resumen por Responsable (NO PAGADO vs Abonos)</b>",
                      getSampleStyleSheet()["Title"]),
            Spacer(1, 0.4*cm),
        ]

        text_cols = {"Responsable", "Estado"}
        header = [Paragraph(col, sty_h) for col in df_local.columns]
        rows = []
        for _, r in df_local.iterrows():
            row = []
            for col, val in zip(df_local.columns, r):
                if pd.isna(val):
                    s = ""
                elif col in {"% Abonado"}:
                    s = f"{float(val):,.1f} %"
                elif isinstance(val, (int, float)) and col not in text_cols:
                    s = f"{val:,.0f}"
                else:
                    s = str(val)
                row.append(Paragraph(s, sty_l if col in text_cols else sty_r))
            rows.append(row)
        data = [header] + rows

        page_width = doc.width
        weights = []
        for c in df_local.columns:
            if c.lower().startswith("responsable"):
                weights.append(3.8)
            elif "Transacc" in c or "Cant." in c or c == "Estado":
                weights.append(1.2)
            else:
                weights.append(2.0)
        total_w = sum(weights)
        col_widths = [max(1.6*cm, page_width*(w/total_w)) for w in weights]

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
            ("RIGHTPADDING", (0,0), (-1,-1), 4),
            ("TOPPADDING", (0,0), (-1,-1), 2),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.Color(0.98,0.98,0.98), colors.white]),
        ]))

        story.append(tbl)
        doc.build(story)
    st.markdown("---")



    try:
        pdf_path = "resumen_no_pagado_abonos.pdf"
        export_resumen_pdf(tabla, pdf_path)
        with open(pdf_path, "rb") as f:
            st.download_button(
                "⬇️ Descargar PDF",
                data=f.read(),
                file_name=pdf_path,
                mime="application/pdf",
            )
    except Exception as e:
        st.warning(f"No se pudo generar PDF. Te dejo el CSV como alternativa. Detalle: {e}")
        st.download_button(
            "⬇️ Descargar CSV",
            data=tabla.to_csv(index=False).encode("utf-8"),
            file_name="resumen_no_pagado_abonos.csv",
            mime="text/csv",
        )

# =========================================================
# 🏢 TAB 3: CANON ANUAL / MENSUAL
# =========================================================
if active_section == "🏢 Canon & Contratos":
    st.markdown(
        tab_header("Canon & Contratos", "Evolución de canon, ocupación y valor por metro cuadrado"),
        unsafe_allow_html=True,
    )
    st.markdown(
        section_heading("🏢", "Canon mensual por Año y por Esp", weight_class="section-heading-title-soft"),
        unsafe_allow_html=True,
    )

    import plotly.graph_objects as go

    _data_src = df_f
    esps_tab_canon = sorted(
        pd.to_numeric(_data_src.get("Esp"), errors="coerce")
        .dropna()
        .astype(int)
        .unique()
        .tolist()
    )
    sel_esps_tab_canon = st.multiselect(
        "Espacios (aplica a todos los gráficos de esta pestaña)",
        options=esps_tab_canon,
        default=esps_tab_canon,
        key="sel_esps_tab_canon",
        placeholder="Selecciona uno o más espacios",
    )

    mask_canon_mensual = (
        (_data_src["CC_norm"] == "INGRESO") &
        _data_src["CC1_text"].str.strip().str.lower().eq("arriendo") &
        _data_src["Obs_text"].str.strip().str.lower().eq("canon mensual")
    )
    dm = _data_src.loc[mask_canon_mensual].copy()
    dm = dm.dropna(subset=["Esp_num", "Periodo_ref"])
    dm["Esp"] = dm["Esp_num"].astype(int)
    dm["Periodo"] = dm["Periodo_ref"]

    agg = (
        dm.groupby(["Periodo","Esp"], as_index=False)["Monto"]
          .sum()
          .sort_values(["Periodo","Esp"])
    )

    all_periods = (
        pd.date_range(agg["Periodo"].min(), agg["Periodo"].max(), freq="MS")
        if not agg.empty else []
    )
    all_esps  = sorted(agg["Esp"].dropna().unique())
    grid = pd.MultiIndex.from_product([all_periods, all_esps], names=["Periodo","Esp"])
    agg_full = (
        agg.set_index(["Periodo","Esp"])
           .reindex(grid, fill_value=0)
           .reset_index()
    )

    st.caption("El gráfico muestra el canon mensual total por mes y por espacio (Esp).")

    if len(all_esps) == 0:
        st.info("No hay datos de 'canon mensual' para mostrar.")
    else:
        esps_validos = [e for e in sel_esps_tab_canon if e in all_esps]
        if esps_validos:
            agg_full = agg_full[agg_full["Esp"].isin(esps_validos)]
            esp_sel_lbl = ", ".join([str(e) for e in esps_validos])
        else:
            agg_full = agg_full.iloc[0:0]
            esp_sel_lbl = "Ninguno"
        st.markdown(
            f"""
            <div style="
                background: linear-gradient(90deg, #0f2d52 0%, #1f4e78 100%);
                border-radius: 10px;
                padding: 10px 14px;
                margin: 8px 0 10px 0;
                color: #FFFFFF;
                font-size: 13px;
                font-weight: 600;">
                Evolución de canon mensual · Espacios seleccionados: {esp_sel_lbl}
            </div>
            """,
            unsafe_allow_html=True,
        )

        palette = [
            "#1D4ED8","#15803D","#B45309","#B42318","#475467",
            "#0E7490","#334155","#0369A1","#854D0E","#166534",
            "#9333EA","#BE123C"
        ]

        fig_line = go.Figure()
        end_labels = []

        def _fmt_label_clp(v: float) -> str:
            v = float(v)
            s = f"{v:,.0f}".replace(",", ".")
            return f"${s} CLP"

        for i, esp in enumerate(sorted(agg_full["Esp"].unique())):
            df_e = agg_full[agg_full["Esp"] == esp]
            color_esp = palette[i % len(palette)]
            fig_line.add_trace(go.Scatter(
                x=df_e["Periodo"], y=df_e["Monto"],
                mode="lines+markers",
                name=f"Esp {esp}",
                line=dict(width=2.4, color=color_esp),
                marker=dict(
                    size=4,
                    opacity=0.75,
                    color=color_esp,
                    line=dict(width=0),
                ),
                hovertemplate="<b>%{x|%b %Y}</b><br>Esp: "+str(esp)+"<br>Monto: $%{y:,.0f}<extra></extra>",
            ))

            if not df_e.empty:
                last_row = df_e.iloc[-1]
                end_labels.append(
                    {
                        "x": last_row["Periodo"],
                        "y": float(last_row["Monto"]),
                        "text": f"Esp {esp}  {_fmt_label_clp(last_row['Monto'])}",
                        "color": color_esp,
                    }
                )

        # Etiquetas de cierre en columna derecha (fuera del plot), ordenadas mayor -> menor.
        if end_labels:
            labels_sorted = sorted(end_labels, key=lambda d: d["y"], reverse=True)
            n = len(labels_sorted)
            # Reparto vertical homogéneo en coordenadas del "paper" para lectura estable.
            y_top = 0.88
            y_bot = 0.22
            if n == 1:
                y_slots = [0.55]
            else:
                y_slots = list(np.linspace(y_top, y_bot, n))

            for lbl, y_slot in zip(labels_sorted, y_slots):
                fig_line.add_annotation(
                    x=1.01,
                    xref="paper",
                    y=y_slot,
                    yref="paper",
                    text=lbl["text"],
                    showarrow=False,
                    xanchor="left",
                    align="left",
                    bgcolor="rgba(255,255,255,0.92)",
                    bordercolor=lbl["color"],
                    borderwidth=1,
                    font=dict(color="#0F172A", size=11),
                )

        visibility_all = [True] * len(fig_line.data)
        visibility_none = [False] * len(fig_line.data)
        fig_line.update_layout(
            updatemenus=[
                dict(
                    type="buttons",
                    direction="right",
                    x=1, xanchor="right", y=1.15, yanchor="top",
                    buttons=[
                        dict(label="Mostrar todo", method="update", args=[{"visible": visibility_all}]),
                        dict(label="Ocultar todo", method="update", args=[{"visible": visibility_none}]),
                    ]
                )
            ]
        )

        fig_line.update_layout(
            title=dict(
                text="🏢 Canon mensual por Mes y Esp",
                x=0.02,
                xanchor="left",
                font=dict(size=18, color="#0F2D52"),
            ),
            xaxis_title="Período (Mes)",
            yaxis_title="Monto (CLP)",
            hovermode="x unified",
            template="plotly_white",
            margin=dict(l=20, r=180, t=78, b=20),
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="left",
                x=0.02,
                bgcolor="rgba(255,255,255,0.85)",
                bordercolor="rgba(15,45,82,0.15)",
                borderwidth=1,
            ),
            paper_bgcolor="#F8FAFC",
            plot_bgcolor="#FFFFFF",
        )
        fig_line.update_xaxes(
            showgrid=False,
            linecolor="rgba(15,45,82,0.25)",
            tickformat="%Y",
            tickformatstops=[
                dict(dtickrange=[None, "M12"], value="%b %Y"),
                dict(dtickrange=["M12", None], value="%Y"),
            ],
        )
        fig_line.update_yaxes(
            showgrid=True,
            gridcolor="rgba(15,45,82,0.10)",
            zeroline=False,
            tickformat=",.0f",
            linecolor="rgba(15,45,82,0.20)",
        )
        fig_line.update_layout(xaxis=dict(rangeslider=dict(visible=True)))

        st.plotly_chart(fig_line, use_container_width=True, config={
            "displaylogo": False,
            "displayModeBar": True,
            "modeBarButtonsToAdd": ["toImage","drawline","drawrect","eraseshape"]
        })

# =========================================================
# 🧩 CANON POR M² (integrado en TAB CANON)
# =========================================================
if active_section == "🏢 Canon & Contratos":
    st.markdown(
        section_heading("🧩", "Canon por m² — Canon Mensual (por Año y Esp)", weight_class="section-heading-title-soft"),
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div style="
            background: linear-gradient(90deg, #0f2d52 0%, #1f4e78 100%);
            border-radius: 10px;
            padding: 10px 14px;
            margin: 8px 0 10px 0;
            color: #FFFFFF;
            font-size: 13px;
            font-weight: 600;">
            Control de valor por m² · Canon mensual por espacio y año
        </div>
        """,
        unsafe_allow_html=True,
    )

    import numpy as np
    import plotly.graph_objects as go
    from io import BytesIO

    c1, c2 = st.columns([1,1])
    with c1:
        escala_m2 = st.selectbox(
            "Escala",
            ["Mensual", "Diario"],
            index=0,
            key="escala_m2_final",
        )
    with c2:
        uf_url = st.text_input("URL CSV UF (opcional: Año,UF_promedio)", value="", key="uf_url_m2_final",
                               placeholder="https://.../uf_promedio_anual.csv")

    _df = df_f

    canon_mask = (
        (_df["CC_norm"] == "INGRESO") &
        _df["CC1_text"].str.strip().str.lower().eq("arriendo") &
        _df["Obs_text"].str.strip().str.lower().eq("canon mensual")
    )
    dm = _df.loc[canon_mask].copy()
    dm = dm.dropna(subset=["Año_sel", "Esp_num", "Monto"])
    dm["Año"] = dm["Año_sel"].astype(int)
    dm["Esp"] = dm["Esp_num"].astype(int)

    M2_MAP = {1:120, 2:72, 3:72, 4:72, 5:180, 6:130, 7:60}
    dm["m2"] = dm["Esp"].map(M2_MAP)
    dm = dm[dm["m2"].notna()]

    dm["Canon_m2_mes"] = dm["Monto"] / dm["m2"]
    dm["Canon_m2_dia"] = dm["Canon_m2_mes"] / 30.0

    agg = (
        dm.groupby(["Año","Esp"], as_index=False)[["Canon_m2_mes","Canon_m2_dia"]]
          .mean()
          .sort_values(["Año","Esp"])
          .copy()
    )

    years_all = (
        _df["Año_sel"]
          .dropna()
          .astype(int)
          .sort_values()
          .unique()
          .tolist()
    )
    if not years_all:
        years_all = sorted(agg["Año"].unique().tolist())

    valor_col = "Canon_m2_mes" if escala_m2 == "Mensual" else "Canon_m2_dia"
    agg["valor_m2_clp"] = agg[valor_col]

    agg["UF_promedio"] = np.nan
    if uf_url.strip():
        try:
            uf_df = pd.read_csv(uf_url)
            uf_df = uf_df.rename(columns={c: c.strip() for c in uf_df.columns})
            if {"Año","UF_promedio"}.issubset(uf_df.columns):
                uf_df["Año"] = pd.to_numeric(uf_df["Año"], errors="coerce").astype("Int64")
                uf_df["UF_promedio"] = pd.to_numeric(uf_df["UF_promedio"], errors="coerce")
                agg = agg.merge(
                    uf_df.dropna(subset=["Año","UF_promedio"]).astype({"Año":"int"}),
                    on="Año", how="left"
                )
            else:
                st.warning("El CSV de UF debe tener columnas: Año,UF_promedio. Se omite UF.")
        except Exception as e:
            st.warning(f"No se pudo leer UF desde la URL: {e}")

    monedas = ["CLP"] + (["UF"] if agg["UF_promedio"].notna().any() else [])
    moneda_m2 = st.selectbox(
        "Moneda",
        monedas,
        index=0,
        key="moneda_m2_final",
    )

    if moneda_m2 == "UF":
        agg["valor_m2"] = agg["valor_m2_clp"] / agg["UF_promedio"]
    else:
        agg["valor_m2"] = agg["valor_m2_clp"]

    todos_esps = sorted(agg["Esp"].unique().tolist())
    sel_esps = [e for e in sel_esps_tab_canon if e in todos_esps]
    agg = agg[agg["Esp"].isin(sel_esps)] if sel_esps else agg.iloc[0:0]

    plot_df = (
        agg.pivot_table(index="Año", columns="Esp", values="valor_m2", aggfunc="mean")
           .reindex(years_all)
           .fillna(0)
    )

    if plot_df.empty:
        st.info("No hay datos para el filtro actual de 'canon mensual' o para los Esp seleccionados.")
    else:
        fig = go.Figure()
        x_years = plot_df.index.astype(int).values
        end_labels_m2 = []

        palette = [
            "#1D4ED8","#15803D","#B45309","#B42318","#475467",
            "#0E7490","#334155","#0369A1","#854D0E","#166534",
            "#9333EA","#BE123C","#0F766E","#7C3AED"
        ]

        for i, esp in enumerate(plot_df.columns):
            y_series = plot_df[esp].values
            color_esp = palette[i % len(palette)]
            custom = np.where(y_series == 0, "⚠️ Sin registros de ‘canon mensual’", " ")
            fig.add_trace(go.Scatter(
                x=x_years, y=y_series, customdata=custom,
                mode="lines+markers",
                name=f"Esp {esp}",
                line=dict(width=2.4, color=color_esp),
                marker=dict(size=4, opacity=0.75, color=color_esp, line=dict(width=0)),
                hovertemplate="<b>Año %{x}</b><br>Esp: "+str(esp)+
                              "<br>Valor: %{y:,.2f}"+(" UF/m²" if moneda_m2=="UF" else " CLP/m²")+
                              "<br>%{customdata}<extra></extra>"
            ))

            if len(y_series) and pd.notna(y_series[-1]):
                end_labels_m2.append(
                    {
                        "y": float(y_series[-1]),
                        "text": f"Esp {esp}  {y_series[-1]:,.2f}{' UF/m²' if moneda_m2=='UF' else ' CLP/m²'}",
                        "color": color_esp,
                    }
                )

        if end_labels_m2:
            labels_sorted = sorted(end_labels_m2, key=lambda d: d["y"], reverse=True)
            n = len(labels_sorted)
            y_top = 0.88
            y_bot = 0.22
            y_slots = [0.55] if n == 1 else list(np.linspace(y_top, y_bot, n))
            for lbl, y_slot in zip(labels_sorted, y_slots):
                fig.add_annotation(
                    x=1.01,
                    xref="paper",
                    y=y_slot,
                    yref="paper",
                    text=lbl["text"],
                    showarrow=False,
                    xanchor="left",
                    align="left",
                    bgcolor="rgba(255,255,255,0.92)",
                    bordercolor=lbl["color"],
                    borderwidth=1,
                    font=dict(color="#0F172A", size=11),
                )

        titulo_y = "UF/m²" if moneda_m2 == "UF" else "CLP/m²"
        titulo_esc = "mensual" if escala_m2 == "Mensual" else "diario"
        fig.update_layout(
            title=dict(
                text=f"🧩 Canon por m² ({titulo_esc}) — {moneda_m2} · por Año y Esp",
                x=0.02,
                xanchor="left",
                font=dict(size=18, color="#0F2D52"),
            ),
            xaxis_title="Año",
            yaxis_title=titulo_y,
            template="plotly_white",
            hovermode="x",
            margin=dict(l=20, r=220, t=78, b=20),
            legend=dict(
                orientation="h",
                y=1.02,
                x=0.02,
                bgcolor="rgba(255,255,255,0.85)",
                bordercolor="rgba(15,45,82,0.15)",
                borderwidth=1,
            ),
            paper_bgcolor="#F8FAFC",
            plot_bgcolor="#FFFFFF",
            updatemenus=[dict(type="buttons", direction="right", x=1, xanchor="right", y=1.15, yanchor="top",
                              buttons=[dict(label="Mostrar todo", method="update", args=[{"visible":[True]*len(plot_df.columns)}]),
                                       dict(label="Ocultar todo", method="update", args=[{"visible":[False]*len(plot_df.columns)}])])]
        )

        if len(x_years):
            fig.update_xaxes(
                type="linear",
                tickmode="linear",
                tick0=int(x_years.min()),
                dtick=1,
                range=[int(x_years.min()) - 0.5, int(x_years.max()) + 0.5],
                showgrid=False,
                linecolor="rgba(15,45,82,0.25)"
            )

        y_min = 0
        y_max = plot_df.replace(0, np.nan).max().max()
        y_max = float(y_max) * 1.1 if pd.notna(y_max) else 1.0
        fig.update_yaxes(
            range=[y_min, y_max],
            showgrid=True,
            gridcolor="rgba(15,45,82,0.10)",
            zeroline=False,
            linecolor="rgba(15,45,82,0.20)",
        )

        st.plotly_chart(
            fig, use_container_width=True,
            config={"displaylogo": False, "displayModeBar": True,
                    "modeBarButtonsToAdd": ["toImage","drawline","drawrect","eraseshape"]}
        )

    # --- Tabla + Excel ---
    escala_lbl = "Mensual" if escala_m2 == "Mensual" else "Diario"
    st.markdown(
        section_heading("📄", f"Dataset agregado (Canon/m² — Año x Esp · {escala_lbl})", weight_class="section-heading-title-soft"),
        unsafe_allow_html=True,
    )

    df_x = plot_df.reset_index().rename(columns={"index": "Año"}).copy()
    esp_cols_x = [c for c in df_x.columns if c != "Año"]
    df_display = df_x.copy()

    def miles_punto(n):
        s = f"{n:,.0f}"
        return s.replace(",", ".")

    def uf_chileno(n):
        s = f"{n:,.2f}"
        s = s.replace(",", "§")
        s = s.replace(".", ",")
        s = s.replace("§", ".")
        return s

    if moneda_m2 == "CLP":
        for c in esp_cols_x:
            df_display[c] = df_display[c].round(0).apply(lambda v: f"${miles_punto(v)}")
    else:
        for c in esp_cols_x:
            df_display[c] = df_display[c].round(2).apply(uf_chileno)

    if moneda_m2 == "CLP":
        df_display = df_display.rename(
            columns={c: f"Esp {c} (CLP/m² · {escala_lbl})" for c in esp_cols_x}
        )
    else:
        df_display = df_display.rename(
            columns={c: f"Esp {c} (UF/m² · {escala_lbl})" for c in esp_cols_x}
        )

    # Quitar años sin datos (todo 0) antes de insertar la fila m²
    if not df_x.empty and esp_cols_x:
        mask_has_data = (df_x[esp_cols_x].fillna(0).abs().sum(axis=1) > 0)
        df_x = df_x.loc[mask_has_data].copy()
        df_display = df_display.loc[mask_has_data.values].copy()

    # Fila m² justo debajo de la primera fila del dataset
    cols_esp_display = [c for c in df_display.columns if c != "Año"]
    fila_m2 = {"Año": "m²"}
    for col_name in cols_esp_display:
        esp_txt = str(col_name)
        esp_num = pd.to_numeric(esp_txt.replace("Esp ", "").split(" ")[0], errors="coerce")
        if pd.notna(esp_num):
            m2_val = M2_MAP.get(int(esp_num))
            fila_m2[col_name] = f"{int(m2_val):,} m²".replace(",", ".") if m2_val is not None else "-"
        else:
            fila_m2[col_name] = "-"

    if not df_display.empty:
        df_display = pd.concat(
            [pd.DataFrame([fila_m2]), df_display],
            ignore_index=True,
        )
    else:
        df_display = pd.DataFrame([fila_m2], columns=df_display.columns)

    # Tabla estilo Electricidad
    _render_table(
        df_display,
        header_bg="#1f4e78",
        header_fg="white",
        row_alt="#eef3fb",
        compact=False,
    )

    excel_buffer = BytesIO()
    df_x_export = df_x.copy()
    fila_m2_export = {"Año": "m²"}
    for c in esp_cols_x:
        esp_num = pd.to_numeric(str(c), errors="coerce")
        if pd.notna(esp_num):
            m2_val = M2_MAP.get(int(esp_num))
            fila_m2_export[c] = m2_val if m2_val is not None else ""
        else:
            fila_m2_export[c] = ""
    if not df_x_export.empty:
        df_x_export = pd.concat(
            [pd.DataFrame([fila_m2_export]), df_x_export],
            ignore_index=True,
        )
    else:
        df_x_export = pd.DataFrame([fila_m2_export], columns=df_x_export.columns)
    try:
        with pd.ExcelWriter(excel_buffer, engine="xlsxwriter") as writer:
            df_x_export.to_excel(writer, index=False, sheet_name="canon_m2")
            wb = writer.book
            ws = writer.sheets["canon_m2"]

            head_fmt = wb.add_format({"bold": True})
            ws.set_row(0, None, head_fmt)

            ws.set_column(0, 0, 10)
            ws.set_column(1, 1, 14)
            if moneda_m2 == "CLP":
                numfmt = wb.add_format({"num_format": "$#,##0"})
                for j, c in enumerate(esp_cols_x, start=1):
                    ws.write(0, j, f"Esp {c} (CLP/m²)", head_fmt)
            else:
                numfmt = wb.add_format({"num_format": '#,##0.00'})
                for j, c in enumerate(esp_cols_x, start=1):
                    ws.write(0, j, f"Esp {c} (UF/m²)", head_fmt)

            ws.set_column(1, len(df_x_export.columns)-1, 14, numfmt)
    except ModuleNotFoundError:
        export_cols = ["Año"] + [
            f"Esp {c} ({'CLP/m²' if moneda_m2 == 'CLP' else 'UF/m²'})" for c in esp_cols_x
        ]
        df_x_export_fallback = df_x_export.copy()
        df_x_export_fallback.columns = export_cols
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            df_x_export_fallback.to_excel(writer, index=False, sheet_name="canon_m2")

    excel_buffer.seek(0)
    st.download_button(
        "⬇️ Descargar Excel (Canon/m² — Año x Esp)",
        data=excel_buffer.getvalue(),
        file_name=f"canon_m2_{'mensual' if escala_m2=='Mensual' else 'diario'}_{moneda_m2}_por_anio_y_esp.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# =========================================================
# 💰 TAB: INGRESOS
# =========================================================
if active_section == "💰 Ingresos":
    import plotly.graph_objects as go

    st.markdown(
        """
        <style>
        .kpi-card-top { display: none !important; }
        .ie-title-row {
            display:flex;
            justify-content:space-between;
            align-items:flex-start;
            margin:-61px 0 14px 0;
        }
        .ie-title {
            color:#081735;
            font-size:28px;
            font-weight:900;
            letter-spacing:-0.025em;
            line-height:1;
        }
        .ie-subtitle {
            margin-top:7px;
            color:#475569;
            font-size:14px;
            font-weight:600;
        }
        .ie-actions {
            display:flex;
            gap:8px;
        }
        .ie-action {
            height:38px;
            display:flex;
            align-items:center;
            justify-content:center;
            border-radius:7px;
            border:1px solid #dbe3ee;
            background:#ffffff;
            color:#0f1f3d;
            padding:0 13px;
            font-size:12px;
            font-weight:800;
            box-shadow:0 8px 18px rgba(15,23,42,0.04);
        }
        .ie-action-primary {
            background:#0B3A86;
            border-color:#0B3A86;
            color:#ffffff;
        }
        .ie-filter-card {
            border:1px solid #dbe3ee;
            border-radius:10px;
            background:#ffffff;
            padding:14px 16px;
            margin: 0 0 14px 0;
            box-shadow:0 10px 24px rgba(15,23,42,0.04);
        }
        .ie-filter-card [data-testid="stRadio"] > label,
        .ie-filter-card [data-testid="stSelectbox"] > label {
            color:#0f1f3d !important;
            font-size:12px !important;
            font-weight:800 !important;
            padding-bottom:4px !important;
        }
        main [data-testid="stRadio"] > label,
        main [data-testid="stSelectbox"] > label {
            color:#0f1f3d !important;
            font-size:12px !important;
            font-weight:800 !important;
            padding-bottom:4px !important;
        }
        .ie-filter-card div[role="radiogroup"] {
            display:grid !important;
            grid-template-columns:1fr 1fr;
            gap:0 !important;
            min-height:36px;
            border-radius:7px;
            overflow:hidden;
            background:#f1f5f9;
            border:1px solid #e7edf5;
        }
        .ie-filter-card div[role="radiogroup"] > label {
            margin:0 !important;
            min-height:36px !important;
            display:flex !important;
            align-items:center !important;
            justify-content:center !important;
            border-radius:0 !important;
            padding:0 12px !important;
            border:0 !important;
            background:transparent !important;
        }
        .ie-filter-card div[role="radiogroup"] > label > div:first-child {
            display:none !important;
        }
        .ie-filter-card div[role="radiogroup"] > label:has(input:checked) {
            background:#0B4DB3 !important;
            box-shadow: inset 0 0 0 1px rgba(255,255,255,0.25);
        }
        .ie-filter-card div[role="radiogroup"] > label [data-testid="stMarkdownContainer"] p {
            color:#0f1f3d !important;
            font-size:12px !important;
            font-weight:800 !important;
        }
        .ie-filter-card div[role="radiogroup"] > label:has(input:checked) [data-testid="stMarkdownContainer"] p {
            color:#ffffff !important;
        }
        .ie-filter-card [data-baseweb="select"] > div {
            min-height:36px !important;
            border-radius:7px !important;
            background:#f1f5f9 !important;
            border-color:#e7edf5 !important;
            box-shadow:none !important;
        }
        main div[role="radiogroup"] {
            display:grid !important;
            grid-template-columns:1fr 1fr;
            gap:0 !important;
            min-height:36px;
            border-radius:7px;
            overflow:hidden;
            background:#f1f5f9;
            border:1px solid #e7edf5;
        }
        main div[role="radiogroup"] > label {
            margin:0 !important;
            min-height:36px !important;
            display:flex !important;
            align-items:center !important;
            justify-content:center !important;
            border-radius:0 !important;
            padding:0 12px !important;
            border:0 !important;
            background:transparent !important;
        }
        main div[role="radiogroup"] > label > div:first-child {
            display:none !important;
        }
        main div[role="radiogroup"] > label:has(input:checked) {
            background:#0B4DB3 !important;
            box-shadow: inset 0 0 0 1px rgba(255,255,255,0.25);
        }
        main div[role="radiogroup"] > label [data-testid="stMarkdownContainer"] p {
            color:#0f1f3d !important;
            font-size:12px !important;
            font-weight:800 !important;
        }
        main div[role="radiogroup"] > label:has(input:checked) [data-testid="stMarkdownContainer"] p {
            color:#ffffff !important;
        }
        main [data-baseweb="select"] > div {
            min-height:36px !important;
            border-radius:7px !important;
            background:#f1f5f9 !important;
            border-color:#e7edf5 !important;
            box-shadow:none !important;
        }
        .ing-kpi-card {
            min-height:144px;
            border-radius:10px;
            padding:18px 18px 14px 18px;
            border:1px solid var(--border);
            background:linear-gradient(135deg, #ffffff 0%, var(--tint) 100%);
            box-shadow:0 12px 28px rgba(15,23,42,0.045);
        }
        .ing-kpi-top {
            display:flex;
            align-items:center;
            gap:12px;
            margin-bottom:12px;
        }
        .ing-kpi-icon {
            width:34px;
            height:34px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:var(--soft);
            color:var(--accent);
            font-size:17px;
            font-weight:900;
        }
        .ing-kpi-title {
            color:#071735;
            font-size:12px;
            font-weight:900;
            letter-spacing:.02em;
            text-transform:uppercase;
        }
        .ing-kpi-value {
            color:var(--accent);
            font-size:27px;
            line-height:1.02;
            font-weight:950;
            letter-spacing:-0.035em;
            margin-bottom:7px;
            white-space:nowrap;
        }
        .ing-kpi-sub {
            color:#516179;
            font-size:11px;
            font-weight:650;
            margin-bottom:12px;
        }
        .ing-kpi-chip {
            display:inline-flex;
            align-items:center;
            gap:5px;
            border-radius:6px;
            padding:5px 8px;
            background:var(--chip-bg);
            color:var(--chip-fg);
            font-size:10px;
            font-weight:900;
        }
        .ie-analysis-card, .ie-bottom-card {
            border:1px solid #dbe3ee;
            border-radius:10px;
            background:#ffffff;
            box-shadow:0 12px 28px rgba(15,23,42,0.05);
        }
        .ie-analysis-card {
            padding:16px;
            height:100%;
        }
        .ie-analysis-title, .ie-bottom-title {
            color:#081735;
            font-weight:900;
        }
        .ie-analysis-title {
            font-size:16px;
            margin-bottom:14px;
        }
        .ie-analysis-row {
            display:grid;
            grid-template-columns:42px 1fr auto;
            gap:10px;
            align-items:center;
            border:1px solid #e5ebf3;
            border-radius:9px;
            padding:11px 12px;
            margin-bottom:10px;
            background:#fbfdff;
        }
        .ie-analysis-icon {
            width:32px;
            height:32px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:var(--soft);
            color:var(--accent);
            font-weight:900;
        }
        .ie-analysis-label {
            color:#0f1f3d;
            font-size:12px;
            font-weight:900;
        }
        .ie-analysis-sub {
            margin-top:3px;
            color:#64748b;
            font-size:11px;
            font-weight:650;
        }
        .ie-analysis-value {
            color:#0f1f3d;
            font-size:13px;
            font-weight:900;
            white-space:nowrap;
        }
        .ie-risk-pill {
            display:inline-flex;
            padding:5px 9px;
            border-radius:6px;
            background:#dcfce7;
            color:#047857;
            font-size:11px;
            font-weight:900;
        }
        .ie-bottom-grid {
            display:grid;
            grid-template-columns:1.05fr 1fr 1.05fr;
            gap:12px;
            margin-top:12px;
        }
        .ie-bottom-card {
            padding:14px;
            min-height:220px;
        }
        .ie-bottom-title {
            font-size:15px;
            margin-bottom:10px;
        }
        .ie-summary-table {
            width:100%;
            border-collapse:collapse;
            font-size:11px;
            color:#0f1f3d;
        }
        .ie-summary-table th {
            text-transform:uppercase;
            letter-spacing:.04em;
            font-size:9px;
            color:#64748b;
            text-align:left;
            padding:7px 0;
            border-bottom:1px solid #e5ebf3;
        }
        .ie-summary-table td {
            padding:7px 0;
            border-bottom:1px solid #edf2f7;
            font-weight:700;
        }
        .ie-summary-table td:last-child, .ie-summary-table th:last-child {
            text-align:right;
        }
        .ie-chart-card {
            border:1px solid #dbe3ee;
            border-radius:10px;
            background:#ffffff;
            padding:16px 16px 6px 16px;
            box-shadow:0 12px 28px rgba(15,23,42,0.05);
        }
        .ie-chart-title {
            color:#081735;
            font-size:16px;
            font-weight:900;
            margin-bottom:4px;
        }
        .ie-donut-center {
            text-align:center;
            color:#081735;
            font-weight:900;
            margin-top:-108px;
            pointer-events:none;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div class="ie-title-row">
            <div>
                <div class="ie-title">Ingresos</div>
                <div class="ie-subtitle">Control de ingresos, canon, abonos y cobranza</div>
            </div>
            <div class="ie-actions">
                <div class="ie-action ie-action-primary">↓ Descargar reporte</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    df_ing_src = df_f.copy()
    tipo_ing = "Financiero"

    c_periodo, c_tipo, c_year, c_month, c_info = st.columns([1.05, 1.25, 1.25, 1.25, 1.8])
    with c_periodo:
        periodo_ing = st.radio("Período", ["Mensual", "Anual"], horizontal=True, key="periodo_ingresos")
    with c_tipo:
        tipo_ing = st.radio("Análisis", ["Financiero", "Contable"], horizontal=True, key="tipo_ingresos")

    if tipo_ing == "Contable":
        df_ing_src["Fecha_dt"] = df_ing_src["Fecha_dt_contable"]
        df_ing_src["Año_sel"] = df_ing_src["Año_sel_contable"]
        df_ing_src["Mes_sel"] = df_ing_src["Mes_sel_contable"]
        df_ing_src["Periodo_ref"] = df_ing_src["Periodo_ref_contable"]

    df_ing_src = df_ing_src.dropna(subset=["Monto"]).copy()
    df_ing_src = df_ing_src[df_ing_src["CC_norm"].eq("INGRESO")].copy()
    if not df_ing_src.empty and df_ing_src["Mes_sel"].isna().all():
        df_ing_src["Año_sel"] = df_ing_src["Fecha_dt"].dt.year
        df_ing_src["Mes_sel"] = df_ing_src["Fecha_dt"].dt.month

    year_opts_ing = ["Todos"] + sorted(df_ing_src["Año_sel"].dropna().astype(int).unique().tolist())
    with c_year:
        sel_year_ing = st.selectbox("Año", year_opts_ing, key="year_ingresos")
    with c_month:
        month_opts_ing = ["Todos"] + list(range(1, 13))
        sel_month_ing = st.selectbox("Mes", month_opts_ing, key="month_ingresos")
    with c_info:
        st.markdown(
            """
            <div style="background:#f3f6fb;border-radius:8px;padding:13px 14px;color:#0f1f3d;font-size:12px;font-weight:650;line-height:1.45;">
                Se usan montos desde la columna F (Monto).<br>
                Ingresos según columna CC.
            </div>
            """,
            unsafe_allow_html=True,
        )
    df_ing_periodo = df_ing_src.copy()
    if sel_year_ing != "Todos":
        df_ing_periodo = df_ing_periodo[df_ing_periodo["Año_sel"].eq(sel_year_ing)]
    if sel_month_ing != "Todos" and periodo_ing == "Mensual":
        df_ing_periodo = df_ing_periodo[df_ing_periodo["Mes_sel"].eq(sel_month_ing)]

    if periodo_ing == "Mensual":
        df_ing_periodo = df_ing_periodo.dropna(subset=["Periodo_ref"]).copy()
        df_ing_periodo["Periodo"] = df_ing_periodo["Periodo_ref"]
        df_ing_hist = df_ing_src.dropna(subset=["Periodo_ref"]).copy()
        df_ing_hist["Periodo"] = df_ing_hist["Periodo_ref"]
        tickformat_ing = "%b %Y"
        label_periodo_ing = "Mes"
    else:
        df_ing_periodo = df_ing_periodo.dropna(subset=["Año_sel"]).copy()
        df_ing_periodo["Periodo"] = pd.to_datetime(
            dict(year=df_ing_periodo["Año_sel"].astype(int), month=1, day=1),
            errors="coerce",
        )
        df_ing_hist = df_ing_src.dropna(subset=["Año_sel"]).copy()
        df_ing_hist["Periodo"] = pd.to_datetime(
            dict(year=df_ing_hist["Año_sel"].astype(int), month=1, day=1),
            errors="coerce",
        )
        tickformat_ing = "%Y"
        label_periodo_ing = "Año"

    df_ing_periodo = df_ing_periodo.dropna(subset=["Periodo"])
    df_ing_hist = df_ing_hist.dropna(subset=["Periodo"])

    if df_ing_periodo.empty:
        st.info(f"No se encuentran registros de ingresos para {periodo_ing.lower()} en análisis {tipo_ing.lower()}.")
        st.stop()

    canon_mask_ing = (
        df_ing_periodo["CC1_text"].str.strip().str.lower().eq("arriendo")
        & df_ing_periodo["Obs_text"].str.strip().str.lower().eq("canon mensual")
    )
    ingreso_total = float(df_ing_periodo["Monto"].sum())
    canon_total_ing = float(df_ing_periodo.loc[canon_mask_ing, "Monto"].sum())
    ingreso_pagado = float(df_ing_periodo.loc[df_ing_periodo["Sit_norm"].eq("PAGADO"), "Monto"].sum())
    ingreso_no_pagado = float(df_ing_periodo.loc[df_ing_periodo["Sit_norm"].eq("NO PAGADO"), "Monto"].sum())
    ingreso_abonos = float(df_ing_periodo.loc[df_ing_periodo["Sit_norm"].str.startswith("ABONO", na=False), "Monto"].sum())

    agg_ing = (
        df_ing_periodo.groupby("Periodo", as_index=False)["Monto"]
        .sum()
        .rename(columns={"Monto": "Ingresos"})
        .sort_values("Periodo")
    )
    hist_ing = (
        df_ing_hist.groupby("Periodo", as_index=False)["Monto"]
        .sum()
        .rename(columns={"Monto": "Ingresos históricos"})
        .sort_values("Periodo")
    )
    hist_ing["Ingreso acumulado"] = hist_ing["Ingresos históricos"].cumsum()
    chart_ing = agg_ing.merge(hist_ing[["Periodo", "Ingreso acumulado"]], on="Periodo", how="left")

    ingreso_acum = float(chart_ing["Ingreso acumulado"].dropna().iloc[-1]) if chart_ing["Ingreso acumulado"].notna().any() else ingreso_total
    cobranza_ratio = ingreso_pagado / ingreso_total if ingreso_total else 0.0
    no_pagado_ratio = ingreso_no_pagado / ingreso_total if ingreso_total else 0.0

    latest_ing = float(chart_ing["Ingresos"].iloc[-1]) if not chart_ing.empty else 0.0
    prev_ing = float(chart_ing["Ingresos"].iloc[-2]) if len(chart_ing) > 1 else 0.0
    delta_ing = (latest_ing - prev_ing) / abs(prev_ing) if prev_ing else 0.0

    def ing_metric_card(title, value, subtitle, icon, accent, soft, tint, border, chip_text, chip_positive=True):
        chip_bg = "#dcfce7" if chip_positive else "#fee2e2"
        chip_fg = "#047857" if chip_positive else "#dc2626"
        arrow = "▲" if chip_positive else "▼"
        return f"""
        <div class="ing-kpi-card" style="--accent:{accent};--soft:{soft};--tint:{tint};--border:{border};--chip-bg:{chip_bg};--chip-fg:{chip_fg};">
            <div class="ing-kpi-top">
                <div class="ing-kpi-icon">{icon}</div>
                <div class="ing-kpi-title">{title}</div>
            </div>
            <div class="ing-kpi-value">{value}</div>
            <div class="ing-kpi-sub">{subtitle}</div>
            <div class="ing-kpi-chip">{arrow} {chip_text}</div>
        </div>
        """

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        st.markdown(
            ing_metric_card(
                "TOTAL INGRESOS",
                fmt_clp_largo(ingreso_total),
                "Suma de ingresos del período",
                "▥",
                "#047857",
                "#dcfce7",
                "#f3fffb",
                "#b7e4d5",
                f"{delta_ing:.1%} vs período anterior",
                delta_ing >= 0,
            ),
            unsafe_allow_html=True,
        )
    with k2:
        st.markdown(
            ing_metric_card(
                "CANON ARRIENDO",
                fmt_clp_largo(canon_total_ing),
                "Canon mensual registrado",
                "▣",
                "#B7791F",
                "#FEF3C7",
                "#fffaf0",
                "#f2d48a",
                f"{(canon_total_ing / ingreso_total if ingreso_total else 0):.1%} de ingresos",
                True,
            ),
            unsafe_allow_html=True,
        )
    with k3:
        st.markdown(
            ing_metric_card(
                "INGRESOS PAGADOS",
                fmt_clp_largo(ingreso_pagado),
                "Ingresos con situación pagado",
                "✓",
                "#2563EB",
                "#DBEAFE",
                "#f5f9ff",
                "#bfdbfe",
                f"{cobranza_ratio:.1%} cobrados",
                True,
            ),
            unsafe_allow_html=True,
        )
    with k4:
        st.markdown(
            ing_metric_card(
                "NO PAGADO",
                fmt_clp_largo(ingreso_no_pagado),
                "Ingresos pendientes de cobro",
                "!",
                "#DC2626",
                "#FEE2E2",
                "#fff7f7",
                "#fecaca",
                f"{no_pagado_ratio:.1%} pendiente",
                False,
            ),
            unsafe_allow_html=True,
        )
    with k5:
        st.markdown(
            ing_metric_card(
                "ABONOS",
                fmt_clp_largo(ingreso_abonos),
                "Ingresos marcados como abono",
                "≋",
                "#7C3AED",
                "#EDE9FE",
                "#fbf8ff",
                "#ddd6fe",
                f"{(ingreso_abonos / ingreso_total if ingreso_total else 0):.1%} de ingresos",
                True,
            ),
            unsafe_allow_html=True,
        )

    main_col_ing, side_col_ing = st.columns([2.25, 1.15])
    with main_col_ing:
        st.markdown(
            f'<div class="ie-chart-card"><div class="ie-chart-title">Ingresos — {periodo_ing} · {tipo_ing}</div>',
            unsafe_allow_html=True,
        )
        fig_ing = go.Figure()
        fig_ing.add_trace(
            go.Bar(
                x=chart_ing["Periodo"],
                y=chart_ing["Ingresos"],
                name="Ingresos",
                marker=dict(color=CHART_BAR_TEAL, line=dict(color=CHART_BAR_TEAL, width=1)),
                hovertemplate="<b>%{x|" + tickformat_ing + "}</b><br>Ingresos: %{y:$,.0f}<extra></extra>",
            )
        )
        fig_ing.add_trace(
            go.Scatter(
                x=chart_ing["Periodo"],
                y=chart_ing["Ingreso acumulado"],
                name="Ingreso acumulado",
                yaxis="y2",
                mode="lines+markers",
                line=dict(color="#0F766E", width=3, dash="dot", shape="spline"),
                marker=dict(size=7, color="#0F766E", line=dict(color="white", width=1.4)),
                hovertemplate="<b>%{x|" + tickformat_ing + "}</b><br>Ingreso acumulado: %{y:$,.0f}<extra></extra>",
            )
        )
        fig_ing.update_layout(
            height=410,
            margin=dict(l=10, r=20, t=36, b=20),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(248,250,252,0.86)",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0, font=dict(size=11)),
            xaxis=dict(title=label_periodo_ing, tickformat=tickformat_ing, showgrid=False),
            yaxis=dict(title="Monto (CLP)", tickprefix="$", separatethousands=True, gridcolor="#E2E8F0", zeroline=False),
            yaxis2=dict(
                title="Acumulado",
                tickprefix="$",
                separatethousands=True,
                overlaying="y",
                side="right",
                showgrid=False,
                zeroline=False,
            ),
            hovermode="x unified",
        )
        st.plotly_chart(fig_ing, use_container_width=True, config={"displaylogo": False})
        st.markdown("</div>", unsafe_allow_html=True)

    top_obs_ing = (
        df_ing_periodo.assign(Obs_clean=lambda d: d["Obs_text"].astype(str).str.strip().replace("", "Sin detalle"))
        .groupby("Obs_clean")["Monto"]
        .sum()
        .sort_values(ascending=False)
    )
    top_resp_ing = (
        df_ing_periodo.assign(Resp_clean=lambda d: d["Responsable_clean"].astype(str).str.strip().replace("", "Sin responsable"))
        .groupby("Resp_clean")["Monto"]
        .sum()
        .sort_values(ascending=False)
    )
    with side_col_ing:
        st.markdown(
            f"""
            <div class="ie-analysis-card">
                <div class="ie-analysis-title">Análisis rápido del período</div>
                <div class="ie-analysis-row" style="--soft:#dcfce7;--accent:#059669;">
                    <div class="ie-analysis-icon">↑</div>
                    <div>
                        <div class="ie-analysis-label">Mayor fuente de ingresos</div>
                        <div class="ie-analysis-sub">{top_obs_ing.index[0] if not top_obs_ing.empty else "Sin detalle"}</div>
                    </div>
                    <div class="ie-analysis-value">{fmt_clp_largo(float(top_obs_ing.iloc[0])) if not top_obs_ing.empty else "$0"}</div>
                </div>
                <div class="ie-analysis-row" style="--soft:#dbeafe;--accent:#2563eb;">
                    <div class="ie-analysis-icon">•</div>
                    <div>
                        <div class="ie-analysis-label">Principal responsable</div>
                        <div class="ie-analysis-sub">{top_resp_ing.index[0] if not top_resp_ing.empty else "Sin responsable"}</div>
                    </div>
                    <div class="ie-analysis-value">{fmt_clp_largo(float(top_resp_ing.iloc[0])) if not top_resp_ing.empty else "$0"}</div>
                </div>
                <div class="ie-analysis-row" style="--soft:#ede9fe;--accent:#7c3aed;">
                    <div class="ie-analysis-icon">%</div>
                    <div>
                        <div class="ie-analysis-label">Cobranza del período</div>
                        <div class="ie-analysis-sub">Pagado sobre ingresos</div>
                    </div>
                    <div class="ie-analysis-value">{cobranza_ratio:.1%}</div>
                </div>
                <div class="ie-analysis-row" style="--soft:#fee2e2;--accent:#dc2626;">
                    <div class="ie-analysis-icon">!</div>
                    <div>
                        <div class="ie-analysis-label">Pendiente de cobro</div>
                        <div class="ie-analysis-sub">No pagado sobre ingresos</div>
                    </div>
                    <div class="ie-analysis-value">{no_pagado_ratio:.1%}</div>
                </div>
                <div style="margin-top:12px;"><span class="ie-risk-pill">Seguimiento cobranza</span></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    bottom_1, bottom_2, bottom_3 = st.columns([1.05, 1, 1.05])
    with bottom_1:
        rows_ing = [
            ("Total ingresos", "Suma de ingresos del período", ingreso_total, "100.0%"),
            ("Canon arriendo", "Canon mensual registrado", canon_total_ing, f"{(canon_total_ing / ingreso_total if ingreso_total else 0):.1%}"),
            ("Ingresos pagados", "Situación pagado", ingreso_pagado, f"{cobranza_ratio:.1%}"),
            ("No pagado", "Pendiente de cobro", ingreso_no_pagado, f"{no_pagado_ratio:.1%}"),
            ("Abonos", "Situación abono", ingreso_abonos, f"{(ingreso_abonos / ingreso_total if ingreso_total else 0):.1%}"),
        ]
        table_html = "".join(
            f"<tr><td>{concepto}</td><td>{desc}</td><td>{fmt_clp_largo(valor)}</td><td>{pct}</td></tr>"
            for concepto, desc, valor, pct in rows_ing
        )
        st.markdown(
            f"""
            <div class="ie-bottom-card">
                <div class="ie-bottom-title">Resumen del período</div>
                <table class="ie-summary-table">
                    <thead><tr><th>Concepto</th><th>Descripción</th><th>Monto (CLP)</th><th>% ingresos</th></tr></thead>
                    <tbody>{table_html}</tbody>
                </table>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with bottom_2:
        st.markdown('<div class="ie-bottom-card"><div class="ie-bottom-title">Evolución del ingreso acumulado</div>', unsafe_allow_html=True)
        fig_acum_ing = go.Figure()
        fig_acum_ing.add_trace(
            go.Scatter(
                x=hist_ing["Periodo"],
                y=hist_ing["Ingreso acumulado"],
                name="Ingreso acumulado",
                mode="lines+markers",
                line=dict(color="#0F766E", width=2.5, shape="spline"),
                marker=dict(size=5, color="#0F766E"),
                fill="tozeroy",
                fillcolor="rgba(15,118,110,0.12)",
            )
        )
        fig_acum_ing.update_layout(
            height=185,
            margin=dict(l=8, r=8, t=20, b=8),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(248,250,252,0.7)",
            showlegend=False,
            xaxis=dict(tickformat=tickformat_ing, showgrid=False),
            yaxis=dict(tickprefix="$", separatethousands=True, gridcolor="#E2E8F0"),
        )
        st.plotly_chart(fig_acum_ing, use_container_width=True, config={"displaylogo": False, "displayModeBar": False})
        st.markdown("</div>", unsafe_allow_html=True)
    with bottom_3:
        st.markdown('<div class="ie-bottom-card"><div class="ie-bottom-title">Composición de ingresos del período</div>', unsafe_allow_html=True)
        comp_ing = (
            df_ing_periodo.assign(CC1_clean=lambda d: d["CC1_text"].astype(str).str.strip().replace("", "Sin clasificar"))
            .groupby("CC1_clean")["Monto"]
            .sum()
            .sort_values(ascending=False)
            .head(5)
        )
        if comp_ing.empty:
            st.info("Sin composición disponible.")
        else:
            fig_donut_ing = go.Figure(
                go.Pie(
                    labels=comp_ing.index,
                    values=comp_ing.values,
                    hole=0.58,
                    marker=dict(colors=["#2F80ED", "#10B981", "#F59E0B", "#8B5CF6", "#94A3B8"]),
                    textinfo="none",
                )
            )
            fig_donut_ing.update_layout(
                height=185,
                margin=dict(l=0, r=0, t=6, b=0),
                paper_bgcolor="rgba(0,0,0,0)",
                legend=dict(orientation="v", x=1.02, y=0.5, font=dict(size=10)),
                annotations=[
                    dict(
                        text=f"<b>{fmt_short(ingreso_total)}</b><br><span style='font-size:10px'>Total ingresos</span>",
                        x=0.5,
                        y=0.5,
                        showarrow=False,
                        font=dict(color="#081735", size=16),
                    )
                ],
            )
            st.plotly_chart(fig_donut_ing, use_container_width=True, config={"displaylogo": False, "displayModeBar": False})
        st.markdown("</div>", unsafe_allow_html=True)

# =========================================================
# 📈 TAB 5: INGRESOS & EGRESOS (Mensual / Anual)
# =========================================================
if active_section == "📈 Flujo Operacional":
    @st.fragment
    def render_flujo_operacional():
        def queue_flujo_selector_scroll(anchor_id: str):
            st.session_state["_flujo_selector_scroll_anchor"] = anchor_id
            st.session_state["_flujo_selector_scroll_nonce"] = (
                st.session_state.get("_flujo_selector_scroll_nonce", 0) + 1
            )

        def render_flujo_selector_scroll(anchor_id: str, offset_px: int = 18):
            if st.session_state.get("_flujo_selector_scroll_anchor") != anchor_id:
                return
            nonce = st.session_state.get("_flujo_selector_scroll_nonce", 0)
            rendered_key = f"_flujo_selector_scroll_rendered_{anchor_id}"
            if not nonce or st.session_state.get(rendered_key) == nonce:
                return
            st.session_state[rendered_key] = nonce
            components.html(
                f"""
                <div data-flujo-selector-scroll="1"></div>
                <script>
                (function () {{
                    const win = window.parent;
                    const doc = win.document;
                    const anchor = doc.getElementById("{anchor_id}");
                    if (!anchor) return;
                    try {{
                        const frames = Array.from(doc.querySelectorAll("iframe"));
                        frames.forEach(function (frame) {{
                            let frameDoc = null;
                            try {{
                                frameDoc = frame.contentDocument || frame.contentWindow.document;
                            }} catch (err) {{}}
                            if (!frameDoc || !frameDoc.querySelector("[data-flujo-selector-scroll]")) return;
                            frame.style.display = "none";
                            frame.style.height = "0";
                            const wrapper = frame.closest('[data-testid="stElementContainer"], [data-testid="stIFrame"]');
                            if (wrapper) {{
                                wrapper.style.display = "none";
                                wrapper.style.height = "0";
                                wrapper.style.minHeight = "0";
                                wrapper.style.margin = "0";
                                wrapper.style.padding = "0";
                                wrapper.style.overflow = "hidden";
                            }}
                        }});
                    }} catch (err) {{}}

                    function scrollToAnchor() {{
                        anchor.scrollIntoView({{behavior: "auto", block: "start"}});
                        win.scrollBy(0, -{offset_px});
                    }}

                    [0, 80, 180, 360, 700].forEach(function (delay) {{
                        win.setTimeout(scrollToAnchor, delay);
                    }});
                }})();
                </script>
                """,
                height=0,
                scrolling=False,
            )

        flujo_download_html = (
            '<a id="flujo-download-report" class="tab-action tab-action-primary" '
            'href="#" role="button">⇩ Descargar reporte</a>'
        )
        st.markdown(
            """
            <style>
            .kpi-card-top {
                display: none !important;
            }
            .ie-title-row {
                display:flex;
                justify-content:space-between;
                align-items:flex-start;
                margin:-61px 0 12px 0;
            }
            .ie-title {
                color:#081735;
                font-size:29px;
                font-weight:950;
                letter-spacing:-0.02em;
                line-height:1;
            }
            .ie-subtitle {
                margin-top:6px;
                color:#516179;
                font-size:12px;
                font-weight:750;
            }
            .ie-actions {
                display:flex;
                gap:8px;
            }
            .ie-action {
                height:40px;
                display:flex;
                align-items:center;
                justify-content:center;
                border-radius:7px;
                border:1px solid #dbe3ee;
                background:#ffffff;
                color:#0f1f3d;
                padding:0 14px;
                font-size:11.5px;
                font-weight:900;
                box-shadow:0 8px 18px rgba(15,23,42,0.04);
            }
            .ie-action-primary {
                background:#0B3A86;
                border-color:#0B3A86;
                color:#ffffff;
            }
            .ie-filter-card {
                border:1px solid #dbe3ee;
                border-radius:10px;
                background:#ffffff;
                padding:9px 14px 10px 14px;
                margin: 0 0 8px 0;
                box-shadow:0 12px 28px rgba(15,23,42,0.045);
            }
            .ie-filter-card [data-testid="column"] {
                border-right:1px solid #e5ebf3;
                padding-right:18px;
            }
            .ie-filter-card [data-testid="column"]:last-child {
                border-right:0;
                padding-right:0;
            }
            .ie-filter-card [data-testid="stVerticalBlock"] {
                gap:0 !important;
            }
            .ie-kpi-card {
                min-height:102px;
                border-radius:14px;
                border:1px solid rgba(219,227,238,0.72);
                background:linear-gradient(135deg, #ffffff 0%, var(--soft) 100%);
                padding:10px 12px 9px 12px;
                box-shadow:0 4px 18px rgba(15,23,42,0.045);
                display:flex;
                flex-direction:column;
                justify-content:space-between;
            }
            .ie-kpi-card-critical {
                border:1.5px solid var(--accent);
                box-shadow:0 8px 24px rgba(15,23,42,0.07), inset 0 0 0 1px rgba(255,255,255,0.62);
                background:linear-gradient(135deg, #ffffff 0%, var(--soft) 82%);
            }
            .ie-kpi-head {
                display:flex;
                align-items:center;
                gap:8px;
                margin-bottom:5px;
            }
            .ie-kpi-icon {
                width:27px;
                height:27px;
                border-radius:999px;
                display:flex;
                align-items:center;
                justify-content:center;
                background:var(--halo);
                color:var(--accent);
                font-size:13.5px;
                font-weight:950;
                flex:0 0 auto;
            }
            .ie-kpi-title {
                color:#071735;
                font-size:10px;
                font-weight:950;
                letter-spacing:.018em;
                text-transform:uppercase;
                line-height:1.16;
            }
            .ie-kpi-value {
                color:var(--accent);
                font-size:20px;
                line-height:1.02;
                font-weight:950;
                letter-spacing:0;
                white-space:nowrap;
            }
            .ie-kpi-card-critical .ie-kpi-value {
                font-size:23px;
                letter-spacing:-0.01em;
            }
            .ie-kpi-card-critical .ie-kpi-title {
                color:#081735;
                font-size:10.5px;
            }
            .ie-kpi-sub {
                color:#516179;
                font-size:9.5px;
                line-height:1.2;
                font-weight:700;
                margin-top:3px;
            }
            .ie-kpi-spark {
                width:100%;
                height:18px;
                margin-top:5px;
                display:block;
            }
            .ie-kpi-spark path,
            .ie-kpi-spark polyline {
                vector-effect:non-scaling-stroke;
            }
            .ie-kpi-badge {
                display:inline-flex;
                align-items:center;
                width:max-content;
                max-width:100%;
                margin-top:4px;
                padding:3px 7px;
                border-radius:5px;
                background:var(--badge-bg);
                color:var(--badge-fg);
                font-size:9.5px;
                font-weight:900;
                white-space:nowrap;
            }
            .ie-kpi-card-critical .ie-kpi-badge {
                padding:4px 8px;
                border-radius:999px;
                font-size:10px;
                box-shadow:0 4px 12px rgba(15,23,42,0.06);
            }
            .ie-analysis-card {
                border:1px solid rgba(219,227,238,0.72);
                border-radius:14px;
                background:#ffffff;
                padding:8px 9px 10px 9px;
                height:520px;
                box-sizing:border-box;
                box-shadow:0 4px 18px rgba(15,23,42,0.045);
                display:flex;
                flex-direction:column;
                margin-top:0;
            }
            .ie-analysis-title {
                color:#081735;
                font-size:14px;
                line-height:1.05;
                font-weight:950;
                margin:0 0 4px 0;
                flex:0 0 auto;
            }
            .ie-health-card {
                border:1px solid #e5ebf3;
                border-radius:10px;
                background:linear-gradient(180deg,#ffffff 0%,#f8fafc 100%);
                padding:6px 8px;
                margin-bottom:6px;
                display:grid;
                grid-template-columns:72px 1fr;
                gap:8px;
                align-items:center;
                flex:0 0 auto;
            }
            .ie-health-card svg {
                width:72px;
                height:44px;
                display:block;
            }
            .ie-health-label {
                color:#64748b;
                font-size:8.5px;
                font-weight:900;
                text-transform:uppercase;
                letter-spacing:.02em;
            }
            .ie-health-score {
                color:#081735;
                font-size:24px;
                font-weight:950;
                line-height:1;
                margin-top:2px;
            }
            .ie-health-note {
                display:inline-flex;
                width:max-content;
                max-width:100%;
                border-radius:999px;
                background:rgba(245,158,11,0.14);
                color:#D97706;
                font-size:8.6px;
                font-weight:900;
                line-height:1;
                margin-top:5px;
                padding:4px 7px;
            }
            .ie-analysis-list {
                display:flex;
                flex-direction:column;
                gap:6px;
                flex:1 1 auto;
                min-height:0;
            }
            .ie-analysis-row {
                display:grid;
                grid-template-columns:31px minmax(0, 1fr);
                gap:8px;
                align-items:center;
                border:1px solid rgba(229,235,243,0.82);
                border-radius:10px;
                padding:8px 8px;
                background:#fbfdff;
                flex:1 1 0;
                min-height:0;
            }
            .ie-analysis-icon {
                width:25px;
                height:25px;
                border-radius:999px;
                display:flex;
                align-items:center;
                justify-content:center;
                background:var(--soft);
                color:var(--accent);
                font-size:12px;
                font-weight:900;
            }
            .ie-analysis-label {
                color:#0f1f3d;
                font-size:11px;
                line-height:1.12;
                font-weight:900;
            }
            .ie-analysis-sub {
                margin-top:2px;
                color:#64748b;
                font-size:9.4px;
                line-height:1.16;
                font-weight:650;
            }
            .ie-analysis-value {
                margin-top:5px;
                color:#0f1f3d;
                font-size:12.5px;
                line-height:1;
                font-weight:900;
                white-space:nowrap;
            }
            .ie-risk-pill {
                display:inline-flex;
                padding:4px 7px;
                border-radius:6px;
                background:#fee2e2;
                color:#B91C1C;
                font-size:10px;
                font-weight:900;
            }
            .ie-bottom-grid {
                display:grid;
                grid-template-columns:1.05fr 1fr 1.05fr;
                gap:12px;
                margin-top:12px;
            }
            .ie-bottom-card {
                border:1px solid rgba(219,227,238,0.72);
                border-radius:14px;
                background:#ffffff;
                padding:14px;
                min-height:220px;
                box-shadow:0 4px 18px rgba(15,23,42,0.045);
            }
            .ie-bottom-title {
                color:#081735;
                font-size:15px;
                font-weight:900;
                margin-bottom:10px;
            }
            .ie-summary-table {
                width:100%;
                border-collapse:collapse;
                font-size:11px;
            }
            .ie-summary-table th {
                text-align:left;
                color:#64748b;
                font-size:10px;
                text-transform:uppercase;
                border-bottom:1px solid #e5ebf3;
                padding:5px 0;
            }
            .ie-summary-table td {
                border-bottom:1px solid #eef2f7;
                color:#0f1f3d;
                font-weight:700;
                padding:6px 0;
            }
            .ie-summary-table td:last-child {
                text-align:right;
                font-weight:900;
            }
            .ie-summary-list {
                display:flex;
                flex-direction:column;
                gap:7px;
            }
            .ie-summary-row {
                display:grid;
                grid-template-columns:30px minmax(0, 1fr) auto;
                gap:8px;
                align-items:center;
                border:1px solid #e5ebf3;
                border-radius:9px;
                padding:7px 9px;
                background:#fbfdff;
            }
            .ie-summary-icon {
                width:24px;
                height:24px;
                border-radius:999px;
                display:flex;
                align-items:center;
                justify-content:center;
                background:var(--soft);
                color:var(--accent);
                font-size:12px;
                font-weight:950;
            }
            .ie-summary-label {
                color:#0f1f3d;
                font-size:11px;
                line-height:1.1;
                font-weight:950;
            }
            .ie-summary-desc {
                color:#64748b;
                font-size:9.5px;
                line-height:1.15;
                font-weight:700;
                margin-top:2px;
                white-space:nowrap;
                overflow:hidden;
                text-overflow:ellipsis;
            }
            .ie-summary-value {
                color:var(--accent);
                font-size:11px;
                font-weight:950;
                text-align:right;
                white-space:nowrap;
            }
            .ie-summary-pct {
                color:#64748b;
                font-size:9.5px;
                font-weight:850;
                margin-top:2px;
            }
            .ie-risk-grid {
                display:grid;
                grid-template-columns:repeat(5, minmax(0, 1fr));
                gap:8px;
                margin:8px 0 6px 0;
            }
            .ie-risk-card {
                --risk:#F97316;
                --risk-soft:rgba(249,115,22,0.12);
                border:1px solid rgba(219,227,238,0.72);
                border-left:4px solid var(--risk);
                border-radius:14px;
                background:linear-gradient(135deg,#ffffff 0%,var(--risk-soft) 135%);
                padding:10px 11px;
                min-height:102px;
                box-shadow:0 4px 18px rgba(15,23,42,0.045);
                display:flex;
                flex-direction:column;
                justify-content:space-between;
            }
            .ie-risk-card.healthy {
                --risk:#16A34A;
                --risk-soft:rgba(34,197,94,0.10);
            }
            .ie-risk-card.attention {
                --risk:#EAB308;
                --risk-soft:rgba(234,179,8,0.13);
            }
            .ie-risk-card.pressure {
                --risk:#F97316;
                --risk-soft:rgba(249,115,22,0.13);
            }
            .ie-risk-card.critical {
                --risk:#DC2626;
                --risk-soft:rgba(220,38,38,0.12);
            }
            .ie-risk-card-head {
                display:flex;
                justify-content:space-between;
                align-items:flex-start;
                gap:8px;
            }
            .ie-risk-label {
                color:#081735;
                font-size:10px;
                line-height:1.18;
                font-weight:950;
                text-transform:uppercase;
                display:flex;
                align-items:center;
                gap:6px;
            }
            .ie-risk-label::before {
                content:"";
                width:7px;
                height:7px;
                flex:0 0 7px;
                border-radius:999px;
                background:var(--risk);
                box-shadow:0 0 0 4px var(--risk-soft);
            }
            .ie-risk-value {
                margin-top:8px;
                color:var(--risk);
                font-size:16px;
                line-height:1;
                font-weight:950;
                white-space:nowrap;
            }
            .ie-risk-sub {
                margin-top:5px;
                color:#64748b;
                font-size:9.5px;
                line-height:1.25;
                font-weight:700;
            }
            .ie-risk-section-title {
                color:#081735;
                font-size:15px;
                line-height:1.1;
                font-weight:950;
                margin:10px 0 2px 0;
            }
            .ie-risk-section-sub {
                color:#64748b;
                font-size:11px;
                font-weight:700;
                margin-bottom:6px;
            }
            .ie-status-badge {
                display:inline-flex;
                border-radius:999px;
                padding:4px 8px;
                font-size:8.5px;
                line-height:1;
                font-weight:950;
                white-space:nowrap;
            }
            .ie-status-badge.healthy {
                background:rgba(34,197,94,0.13);
                color:#16A34A;
            }
            .ie-status-badge.attention {
                background:rgba(234,179,8,0.16);
                color:#A16207;
            }
            .ie-status-badge.pressure {
                background:rgba(249,115,22,0.15);
                color:#C2410C;
            }
            .ie-status-badge.critical {
                background:rgba(220,38,38,0.14);
                color:#DC2626;
            }
            .neto-chart-card {
                border:1px solid #dbe3ee;
                border-radius:10px;
                background:#ffffff;
                padding:14px 18px 12px 18px;
                box-shadow:0 12px 28px rgba(15,23,42,0.05);
                margin-top:10px;
            }
            .neto-card-head {
                display:flex;
                justify-content:space-between;
                align-items:flex-start;
                gap:14px;
                margin-bottom:7px;
            }
            .neto-card-title {
                color:#081735;
                font-size:16px;
                line-height:1.15;
                font-weight:950;
                letter-spacing:-0.018em;
            }
            .neto-card-sub {
                color:#64748b;
                font-size:11.5px;
                font-weight:650;
                margin-top:5px;
            }
            .neto-kpi-grid {
                display:grid;
                grid-template-columns:repeat(3, minmax(0, 1fr));
                border:1px solid #e5ebf3;
                border-radius:9px;
                overflow:hidden;
                margin:5px 0 5px auto;
                max-width:660px;
            }
            .neto-kpi-box {
                padding:9px 11px;
                background:#fbfdff;
                border-left:1px solid #e5ebf3;
            }
            .neto-kpi-box:first-child { border-left:0; }
            .neto-kpi-label {
                color:#334155;
                font-size:11px;
                line-height:1.2;
                font-weight:850;
            }
            .neto-kpi-value {
                margin-top:5px;
                color:var(--metric);
                font-size:16px;
                line-height:1;
                font-weight:950;
                letter-spacing:-0.02em;
            }
            .neto-kpi-note {
                margin-top:4px;
                color:#64748b;
                font-size:10px;
                font-weight:700;
            }
            .neto-note {
                display:flex;
                gap:8px;
                align-items:flex-start;
                color:#475569;
                font-size:12px;
                line-height:1.45;
                font-weight:650;
                margin-top:5px;
            }
            .neto-filter-anchor {
                height:0;
                min-height:0;
                margin:0;
                padding:0;
                overflow:hidden;
            }
            div[data-testid="stVerticalBlock"]:has(.neto-filter-anchor) > div[data-testid="stHorizontalBlock"] {
                border:1px solid rgba(219,227,238,0.86);
                border-radius:10px;
                background:#fbfdff;
                box-shadow:inset 0 1px 0 rgba(255,255,255,0.72);
                padding:7px 8px 5px 8px;
                margin:2px 0 5px 0;
                gap:0.45rem;
                align-items:flex-end;
            }
            div[data-testid="stVerticalBlock"]:has(.neto-filter-anchor) [data-testid="stSelectbox"] {
                margin:0 !important;
            }
            div[data-testid="stVerticalBlock"]:has(.neto-filter-anchor) [data-testid="stSelectbox"] > label {
                color:#64748b !important;
                font-size:9.4px !important;
                line-height:1 !important;
                font-weight:950 !important;
                text-transform:uppercase !important;
                letter-spacing:.018em !important;
                padding-bottom:3px !important;
            }
            div[data-testid="stVerticalBlock"]:has(.neto-filter-anchor) [data-baseweb="select"] > div {
                min-height:30px !important;
                height:30px !important;
                border-radius:8px !important;
                background:#ffffff !important;
                border-color:#dbe3ee !important;
                box-shadow:none !important;
            }
            div[data-testid="stVerticalBlock"]:has(.neto-filter-anchor) [data-baseweb="select"] div {
                font-size:11px !important;
                font-weight:800 !important;
            }
            .neto-filter-action {
                height:30px;
                border:1px solid #dbe3ee;
                border-radius:8px;
                display:flex;
                align-items:center;
                justify-content:center;
                color:#081735;
                font-weight:950;
                background:#ffffff;
                margin-top:14px;
                box-sizing:border-box;
            }
            main [data-testid="stSelectbox"] > label,
            main [data-testid="stRadio"] > label {
                color:#0f1f3d !important;
                font-size:11.5px !important;
                line-height:1.1 !important;
                font-weight:900 !important;
                padding-bottom:4px !important;
            }
            main div[role="radiogroup"] {
                display:grid !important;
                grid-template-columns:repeat(2, minmax(0, 1fr));
                gap:0 !important;
                min-height:34px !important;
                border-radius:7px !important;
                overflow:hidden !important;
                background:#f8fafc !important;
                border:1px solid #e5ebf3 !important;
                box-shadow:inset 0 1px 0 rgba(255,255,255,0.65) !important;
            }
            main div[role="radiogroup"] > label {
                min-height:34px !important;
                margin:0 !important;
                padding:0 12px !important;
                border:0 !important;
                border-radius:0 !important;
                display:flex !important;
                align-items:center !important;
                justify-content:center !important;
                color:#0f1f3d !important;
                font-size:11.5px !important;
                font-weight:850 !important;
                transition:all .16s ease !important;
            }
            main div[role="radiogroup"] > label > div:first-child {
                display:none !important;
            }
            main div[role="radiogroup"] > label:has(input:checked) {
                background:#0B4DB3 !important;
                color:#ffffff !important;
                box-shadow:inset 0 0 0 1px rgba(255,255,255,0.28), 0 8px 18px rgba(11,77,179,0.12) !important;
            }
            main div[role="radiogroup"] > label [data-testid="stMarkdownContainer"] p {
                color:inherit !important;
                font-size:12px !important;
                font-weight:850 !important;
            }
            main [data-baseweb="select"] > div {
                min-height:38px !important;
                border-radius:7px !important;
                border-color:#e5ebf3 !important;
                background:#f8fafc !important;
                box-shadow:none !important;
                color:#0f1f3d !important;
                font-size:12px !important;
                font-weight:850 !important;
            }
            .ie-guide-note {
                display:flex;
                gap:10px;
                align-items:center;
                border:1px solid #dbe3ee;
                border-radius:8px;
                background:#f8fafc;
                color:#334155;
                font-size:12px;
                font-weight:650;
                padding:11px 13px;
                margin:14px 0 16px 0;
            }
            </style>
            <div class="tab-title-row">
                <div><div class="tab-title-main">Flujo Operacional</div><div class="tab-title-sub">Industrial Asset Analytics · ingresos, egresos y resultado neto</div></div>
                <div class="tab-actions">
                    __FLUJO_DOWNLOAD__
                </div>
            </div>
            """.replace("__FLUJO_DOWNLOAD__", flujo_download_html),
            unsafe_allow_html=True,
        )
        components.html(
            """
            <script>
            (function () {
                const win = window.parent;
                const doc = win.document;
                const btn = doc.getElementById("flujo-download-report");
                if (!btn || btn.dataset.captureHandlerAttached === "1") return;
                btn.dataset.captureHandlerAttached = "1";

                function loadScript(src) {
                    return new Promise(function (resolve, reject) {
                        const existing = doc.querySelector('script[src="' + src + '"]');
                        if (existing) {
                            existing.addEventListener("load", resolve, { once: true });
                            if (existing.dataset.loaded === "1") resolve();
                            return;
                        }
                        const script = doc.createElement("script");
                        script.src = src;
                        script.onload = function () {
                            script.dataset.loaded = "1";
                            resolve();
                        };
                        script.onerror = reject;
                        doc.head.appendChild(script);
                    });
                }

                async function downloadFlujoPdf(event) {
                    event.preventDefault();
                    const originalText = btn.textContent;
                    btn.textContent = "Generando PDF...";
                    btn.style.pointerEvents = "none";

                    const hidden = [];
                    function hide(selector) {
                        doc.querySelectorAll(selector).forEach(function (node) {
                            hidden.push([node, node.style.display]);
                            node.style.display = "none";
                        });
                    }

                    try {
                        await loadScript("https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js");
                        await loadScript("https://cdn.jsdelivr.net/npm/jspdf@2.5.1/dist/jspdf.umd.min.js");

                        const target = doc.querySelector("section.main .block-container")
                            || doc.querySelector('[data-testid="stMainBlockContainer"]')
                            || doc.querySelector(".block-container");
                        if (!target) throw new Error("No se encontró el contenido del reporte.");

                        hide("section[data-testid='stSidebar']");
                        hide("aside[data-testid='stSidebar']");
                        hide(".bodegas-sidebar-toggle");
                        hide(".tab-actions");
                        hide("header[data-testid='stHeader']");
                        hide("div[data-testid='stToolbar']");

                        const canvas = await win.html2canvas(target, {
                            backgroundColor: "#ffffff",
                            scale: 2,
                            useCORS: true,
                            allowTaint: true,
                            logging: false,
                            windowWidth: Math.max(target.scrollWidth, doc.documentElement.clientWidth),
                            windowHeight: Math.max(target.scrollHeight, doc.documentElement.clientHeight)
                        });

                        const imgData = canvas.toDataURL("image/png");
                        const pdf = new win.jspdf.jsPDF({
                            orientation: "landscape",
                            unit: "pt",
                            format: "a4"
                        });
                        const pageW = pdf.internal.pageSize.getWidth();
                        const pageH = pdf.internal.pageSize.getHeight();
                        const margin = 10;
                        const usableW = pageW - margin * 2;
                        const imgH = canvas.height * usableW / canvas.width;
                        let y = margin;
                        let remaining = imgH;

                        pdf.addImage(imgData, "PNG", margin, y, usableW, imgH, null, "FAST");
                        remaining -= pageH - margin * 2;
                        while (remaining > 0) {
                            pdf.addPage();
                            y = margin - (imgH - remaining);
                            pdf.addImage(imgData, "PNG", margin, y, usableW, imgH, null, "FAST");
                            remaining -= pageH - margin * 2;
                        }
                        pdf.save("flujo_operacional_bodegas_balmaceda.pdf");
                    } catch (err) {
                        win.alert("No se pudo generar el PDF visual. Intenta nuevamente cuando la página termine de cargar.");
                    } finally {
                        hidden.forEach(function (entry) {
                            entry[0].style.display = entry[1];
                        });
                        btn.textContent = originalText;
                        btn.style.pointerEvents = "";
                    }
                }

                btn.addEventListener("click", downloadFlujoPdf);
            })();
            </script>
            """,
            height=0,
        )

        import plotly.graph_objects as go

        c1, c2, c3, c4 = st.columns([1.25, 1.45, 1.55, 1.55])
        with c1:
            periodo = st.radio("Periodo", ["Mensual", "Anual"], horizontal=True, index=0, key="periodo_ing_eg")
        with c2:
            tipo_analisis = st.radio(
                "Análisis",
                ["Financiero", "Contable"],
                horizontal=True,
                index=0,
                key="tipo_analisis_ing_eg",
            )

        df_ie = df_f.copy()
        fecha_dt_acum_col = "Fecha_dt"
        anio_acum_col = "Año_sel"
        mes_acum_col = "Mes_sel"
        periodo_ref_acum_col = "Periodo_ref"
        if tipo_analisis == "Contable":
            df_ie["Fecha_dt"] = df_ie["Fecha_dt_contable"]
            df_ie["Año_sel"] = df_ie["Año_sel_contable"]
            df_ie["Mes_sel"] = df_ie["Mes_sel_contable"]
            df_ie["Periodo_ref"] = df_ie["Periodo_ref_contable"]
            df_ie["Fecha"] = df_ie["FECHA CONTABLE"]
            df_ie["Año"] = df_ie["AÑO CONTABLE"]
            df_ie["Mes"] = df_ie["Mes CONTABLE"]
            fecha_dt_acum_col = "Fecha_dt_contable"
            anio_acum_col = "Año_sel_contable"
            mes_acum_col = "Mes_sel_contable"
            periodo_ref_acum_col = "Periodo_ref_contable"
            if df_ie["Año_sel"].isna().all() and df_ie["Fecha_dt"].isna().all():
                st.warning("No hay fechas contables válidas para este análisis.")

        df_ie_base_periodo = df_ie.dropna(subset=["Monto"]).copy()
        _df = df_ie_base_periodo.copy()
        _df = _df[_df["CC_norm"].isin(["INGRESO", "EGRESO"])]
        _df = _df[_df["Sit_norm"].isin(["PAGADO", "NO PAGADO"])]

        if _df["Mes_sel"].isna().all():
            _df["Año_sel"] = _df["Fecha_dt"].dt.year
            _df["Mes_sel"] = _df["Fecha_dt"].dt.month

        _df = _df.dropna(subset=["Año_sel"])
        years = sorted(_df["Año_sel"].dropna().astype(int).unique().tolist())
        year_opts = ["Todos"] + years

        with c3:
            sel_year = st.selectbox("Año", year_opts, index=0, key="year_ing_eg")
        with c4:
            month_opts = ["Todos"] + list(range(1, 13))
            sel_month = st.selectbox("Mes", month_opts, index=0, key="month_ing_eg",
                                     disabled=(periodo == "Anual"))

        if sel_year != "Todos":
            df_ie_base_periodo = df_ie_base_periodo[df_ie_base_periodo["Año_sel"] == sel_year]
        if sel_month != "Todos" and periodo == "Mensual":
            df_ie_base_periodo = df_ie_base_periodo[df_ie_base_periodo["Mes_sel"] == sel_month]

        periodo_filtro_lbl = (
            f"Año {int(sel_year)} · Mes {int(sel_month):02d}"
            if sel_year != "Todos" and periodo == "Mensual" and sel_month != "Todos"
            else (
                f"Año {int(sel_year)}"
                if sel_year != "Todos"
                else "Todos los períodos"
            )
        )

        _df_acum = df_f.dropna(subset=["Monto"]).copy()
        _df_acum["Fecha_dt_acum"] = _df_acum[fecha_dt_acum_col]
        _df_acum["Año_sel_acum"] = _df_acum[anio_acum_col]
        _df_acum["Mes_sel_acum"] = _df_acum[mes_acum_col]
        _df_acum["Periodo_ref_acum"] = _df_acum[periodo_ref_acum_col]
        _df_acum = _df_acum[_df_acum["CC_norm"].isin(["INGRESO", "EGRESO"])]
        _df_acum = _df_acum[_df_acum["Sit_norm"].isin(["PAGADO", "NO PAGADO"])]
        if _df_acum["Mes_sel_acum"].isna().all():
            _df_acum["Año_sel_acum"] = _df_acum["Fecha_dt_acum"].dt.year
            _df_acum["Mes_sel_acum"] = _df_acum["Fecha_dt_acum"].dt.month

        if periodo == "Mensual":
            _df_acum = _df_acum.dropna(subset=["Periodo_ref_acum"])
            _df_acum["Periodo"] = _df_acum["Periodo_ref_acum"]
        else:
            _df_acum = _df_acum.dropna(subset=["Año_sel_acum"])
            _df_acum["Periodo"] = pd.to_datetime(
                dict(year=_df_acum["Año_sel_acum"].astype(int), month=1, day=1),
                errors="coerce"
            )

        _df_acum["Periodo"] = pd.to_datetime(_df_acum["Periodo"], errors="coerce")
        _df_acum = _df_acum.dropna(subset=["Periodo"])
        agg_acum = (
            _df_acum.groupby(["Periodo", "CC"], as_index=False)["Monto"]
            .sum()
            .sort_values("Periodo")
        )
        ingresos_acum = agg_acum[agg_acum["CC"] == "INGRESO"].rename(columns={"Monto": "Ingresos_acum"})
        egresos_acum = agg_acum[agg_acum["CC"] == "EGRESO"].rename(columns={"Monto": "Egresos_acum"})
        base_acum = pd.DataFrame({"Periodo": sorted(agg_acum["Periodo"].dropna().unique())})
        base_acum["Periodo"] = pd.to_datetime(base_acum["Periodo"], errors="coerce")
        base_acum = base_acum.merge(ingresos_acum[["Periodo", "Ingresos_acum"]], on="Periodo", how="left")
        base_acum = base_acum.merge(egresos_acum[["Periodo", "Egresos_acum"]], on="Periodo", how="left")
        base_acum = base_acum.fillna(0).sort_values("Periodo")
        base_acum["Neto_periodo_acum"] = base_acum["Ingresos_acum"] - base_acum["Egresos_acum"].abs()
        base_acum["Neto_acumulado"] = base_acum["Neto_periodo_acum"].cumsum()
        neto_acumulado_periodo = base_acum[["Periodo", "Neto_acumulado"]]

        _df = df_ie_base_periodo.copy()
        _df = _df[_df["CC_norm"].isin(["INGRESO", "EGRESO"])]
        _df = _df[_df["Sit_norm"].isin(["PAGADO", "NO PAGADO"])]

        # Construir Periodo usando Año/Mes
        if periodo == "Mensual":
            _df = _df.dropna(subset=["Periodo_ref"])
            _df["Periodo"] = _df["Periodo_ref"]
            label_x = "Mes"
            x_hover = "%b %Y"
        else:
            _df["Periodo"] = pd.to_datetime(
                dict(year=_df["Año_sel"].astype(int), month=1, day=1),
                errors="coerce"
            )
            label_x = "Año"
            x_hover = "%Y"

        _df = _df.dropna(subset=["Periodo"])
        _df["Periodo"] = pd.to_datetime(_df["Periodo"], errors="coerce")
        _df = _df.dropna(subset=["Periodo"])

        agg_ie = (
            _df.groupby(["Periodo", "CC"], as_index=False)["Monto"]
               .sum()
               .sort_values("Periodo")
        )

        ingresos = agg_ie[agg_ie["CC"] == "INGRESO"].rename(columns={"Monto": "Ingresos"})
        egresos = agg_ie[agg_ie["CC"] == "EGRESO"].rename(columns={"Monto": "Egresos"})

        base = pd.DataFrame({"Periodo": sorted(agg_ie["Periodo"].dropna().unique())})
        base["Periodo"] = pd.to_datetime(base["Periodo"], errors="coerce")
        base = base.merge(ingresos[["Periodo", "Ingresos"]], on="Periodo", how="left")
        base = base.merge(egresos[["Periodo", "Egresos"]], on="Periodo", how="left")
        base = base.fillna(0)

        if base.empty:
            st.info("No hay datos suficientes para calcular ingresos y egresos por período.")
        else:
            base["Egresos_abs"] = base["Egresos"].abs()
            base["Neto"] = base["Ingresos"] - base["Egresos_abs"]
            base = base.merge(neto_acumulado_periodo, on="Periodo", how="left")
            base["Neto_acumulado"] = base["Neto_acumulado"].fillna(base["Neto"].cumsum())

            total_ing = float(base["Ingresos"].sum())
            total_egr = float(base["Egresos_abs"].sum())
            total_neto = float(base["Neto"].sum())
            total_neto_acumulado = float(base["Neto_acumulado"].iloc[-1])
            total_margen = (total_neto / total_ing) if total_ing else 0.0
            base["Egresos_plot"] = -base["Egresos_abs"]
            base["Margen"] = np.where(base["Ingresos"] != 0, base["Neto"] / base["Ingresos"], 0.0)
            base["Neto_color"] = np.where(base["Neto"] >= 0, CHART_TEAL, CHART_RED)

            def ie_icon_svg(kind: str) -> str:
                icons = {
                    "up": '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.6" stroke-linecap="round" stroke-linejoin="round"><path d="M7 17L17 7"/><path d="M9 7h8v8"/></svg>',
                    "down": '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.6" stroke-linecap="round" stroke-linejoin="round"><path d="M7 7l10 10"/><path d="M17 9v8H9"/></svg>',
                    "equal": '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.8" stroke-linecap="round"><path d="M6 9h12"/><path d="M6 15h12"/></svg>',
                    "sum": '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M18 5H7l6 7-6 7h11"/></svg>',
                    "percent": '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round"><path d="M19 5L5 19"/><circle cx="7.5" cy="7.5" r="2.5"/><circle cx="16.5" cy="16.5" r="2.5"/></svg>',
                }
                return icons.get(kind, icons["equal"])

            def ie_sparkline(values, accent):
                vals = pd.Series(values).replace([np.inf, -np.inf], np.nan).dropna().astype(float).tail(18).tolist()
                if len(vals) < 2:
                    vals = [0.0, 0.0]
                lo, hi = min(vals), max(vals)
                span = hi - lo if hi != lo else 1.0
                points = []
                denom = max(len(vals) - 1, 1)
                for i, val in enumerate(vals):
                    x = 2 + (i / denom) * 106
                    y = 22 - ((val - lo) / span) * 18
                    points.append(f"{x:.1f},{y:.1f}")
                return (
                    f'<svg class="ie-kpi-spark" viewBox="0 0 112 26" preserveAspectRatio="none" aria-hidden="true">'
                    f'<path d="M2 22 H110" stroke="rgba(100,116,139,0.16)" stroke-width="1"/>'
                    f'<polyline points="{" ".join(points)}" fill="none" stroke="{accent}" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"/>'
                    f'</svg>'
                )

            def ie_metric_card(title, value, subtitle, icon, accent, soft, halo, border, badge_text, sparkline, badge_positive=True, critical=False):
                badge_bg = "#dcfce7" if badge_positive else "#fee2e2"
                badge_fg = "#047857" if badge_positive else "#b91c1c"
                critical_class = " ie-kpi-card-critical" if critical else ""
                return f"""
                <div class="ie-kpi-card{critical_class}" style="--accent:{accent};--soft:{soft};--halo:{halo};--border:{border};--badge-bg:{badge_bg};--badge-fg:{badge_fg};">
                    <div>
                        <div class="ie-kpi-head">
                            <div class="ie-kpi-icon">{icon}</div>
                            <div class="ie-kpi-title">{title}</div>
                        </div>
                        <div class="ie-kpi-value">{value}</div>
                        <div class="ie-kpi-sub">{subtitle}</div>
                        {sparkline}
                    </div>
                    <div class="ie-kpi-badge">{badge_text}</div>
                </div>
                """

            k1, k2, k3, k4, k5 = st.columns(5)
            with k1:
                st.markdown(
                    ie_metric_card(
                        "TOTAL INGRESO",
                        fmt_clp_largo(total_ing),
                        "Ingresos",
                        ie_icon_svg("up"),
                        "#22C55E",
                        "#f6fffb",
                        "#d8f5e4",
                        "#cfe9de",
                        "Registrado",
                        ie_sparkline(base["Ingresos"], "#22C55E"),
                        True,
                    ),
                    unsafe_allow_html=True,
                )
            with k2:
                st.markdown(
                    ie_metric_card(
                        "TOTAL EGRESO",
                        fmt_clp_largo(total_egr),
                        "Egresos",
                        ie_icon_svg("down"),
                        "#F87171",
                        "#fff7f7",
                        "#fde2e2",
                        "#f1caca",
                        "Registrado",
                        ie_sparkline(base["Egresos_abs"], "#F87171"),
                        False,
                    ),
                    unsafe_allow_html=True,
                )
            with k3:
                net_color = CHART_TEAL if total_neto >= 0 else CHART_RED
                st.markdown(
                    ie_metric_card(
                        "TOTAL NETO",
                        fmt_clp_largo(total_neto),
                        "Resultado",
                        ie_icon_svg("equal"),
                        "#2563EB" if total_neto >= 0 else "#F87171",
                        "#f6fffb" if total_neto >= 0 else "#fff7f7",
                        "#d8f5e4" if total_neto >= 0 else "#fde2e2",
                        "#cfe9de" if total_neto >= 0 else "#f1caca",
                        "Operativo",
                        ie_sparkline(base["Neto"], "#2563EB" if total_neto >= 0 else "#F87171"),
                        total_neto >= 0,
                        True,
                    ),
                    unsafe_allow_html=True,
                )
            with k4:
                net_acum_color = CHART_TEAL if total_neto_acumulado >= 0 else CHART_RED
                st.markdown(
                    ie_metric_card(
                        "NETO ACUMULADO",
                        fmt_clp_largo(total_neto_acumulado),
                        "Acumulado",
                        ie_icon_svg("sum"),
                        "#1D4ED8" if total_neto_acumulado >= 0 else "#DC2626",
                        "#f6f9ff" if total_neto_acumulado >= 0 else "#fff7f7",
                        "#e0ebff" if total_neto_acumulado >= 0 else "#fde2e2",
                        "#d4e1f6" if total_neto_acumulado >= 0 else "#f1caca",
                        "A la fecha",
                        ie_sparkline(base["Neto_acumulado"], "#2563EB" if total_neto_acumulado >= 0 else "#F87171"),
                        total_neto_acumulado >= 0,
                        True,
                    ),
                    unsafe_allow_html=True,
                )
            with k5:
                margen_color = CHART_TEAL if total_margen >= 0 else CHART_RED
                st.markdown(
                    ie_metric_card(
                        "MARGEN",
                        f"{total_margen:.1%}",
                        "Rentabilidad",
                        ie_icon_svg("percent"),
                        "#F59E0B",
                        "#fbf7ff",
                        "#FEF3C7",
                        "#FDE68A",
                        "Período",
                        ie_sparkline(base["Margen"], "#F59E0B"),
                        total_margen >= 0,
                    ),
                    unsafe_allow_html=True,
                )

            from plotly.subplots import make_subplots

            base["Tooltip"] = (
                "Ingresos: $" + base["Ingresos"].map(lambda v: f"{v:,.0f}")
                + "<br>Egresos: $" + base["Egresos_abs"].map(lambda v: f"{v:,.0f}")
                + "<br>Neto: $" + base["Neto"].map(lambda v: f"{v:,.0f}")
                + "<br>Neto acumulado: $" + base["Neto_acumulado"].map(lambda v: f"{v:,.0f}")
                + "<br>Margen: " + base["Margen"].map(lambda v: f"{v:.1%}")
            )

            fig_ie = make_subplots(specs=[[{"secondary_y": False}]])

            fig_ie.add_trace(
                go.Scatter(
                    x=base["Periodo"],
                    y=base["Neto"],
                    mode="markers",
                    name="Resumen",
                    marker=dict(size=18, color="rgba(0,0,0,0)"),
                    showlegend=False,
                    customdata=base[["Tooltip"]],
                    hovertemplate=f"<b>{label_x} %{{x|{x_hover}}}</b><br>%{{customdata[0]}}<extra></extra>",
                ),
                secondary_y=False,
            )

            fig_ie.add_trace(
                go.Bar(
                    x=base["Periodo"],
                    y=base["Ingresos"],
                    name="Ingresos",
                    marker=dict(
                        color="rgba(34, 197, 94, 0.78)",
                        line=dict(color="rgba(34, 197, 94, 0.78)", width=0),
                    ),
                    hoverinfo="skip",
                    offsetgroup="flujo",
                ),
                secondary_y=False,
            )
            fig_ie.add_trace(
                go.Bar(
                    x=base["Periodo"],
                    y=base["Egresos_plot"],
                    name="Egresos",
                    marker=dict(
                        color="rgba(248, 113, 113, 0.82)",
                        line=dict(color="rgba(248, 113, 113, 0.82)", width=0),
                    ),
                    hoverinfo="skip",
                    offsetgroup="flujo",
                ),
                secondary_y=False,
            )
            fig_ie.add_trace(
                go.Scatter(
                    x=base["Periodo"],
                    y=base["Neto"],
                    mode="lines+markers",
                    name="Neto",
                    line=dict(color="#2563EB", width=2.3, shape="spline"),
                    marker=dict(
                        size=6.5,
                        color="#2563EB",
                        line=dict(color="#FFFFFF", width=1.35),
                    ),
                    hoverinfo="skip",
                ),
                secondary_y=False,
            )
            fig_ie.add_trace(
                go.Scatter(
                    x=base["Periodo"],
                    y=base["Neto_acumulado"],
                    mode="lines+markers",
                    name="Neto acumulado",
                    line=dict(color="#F59E0B", width=2.1, dash="dash", shape="spline"),
                    marker=dict(size=5.8, color="#F59E0B", line=dict(color="#FFFFFF", width=1.15)),
                    hoverinfo="skip",
                ),
                secondary_y=False,
            )
            fig_ie.add_hline(y=0, line_width=1.2, line_color=CHART_GRAY, opacity=0.75, secondary_y=False)

            fig_ie.update_layout(
                title=dict(
                    text=f"Ingresos, egresos y resultado neto — {periodo} · {tipo_analisis} &nbsp; ⓘ",
                    x=0.02,
                    xanchor="left",
                    y=0.985,
                    font=dict(size=16, color="#081735", family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif"),
                ),
                font=dict(
                    family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif",
                    size=12,
                    color="#334155",
                ),
                template="plotly_white",
                height=455,
                margin=dict(l=18, r=22, t=52, b=26),
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=1.015,
                    xanchor="left",
                    x=0.02,
                    bgcolor="rgba(255,255,255,0)",
                    font=dict(size=10.5, color="#334155", family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif"),
                    itemclick=False,
                    itemdoubleclick=False,
                ),
                hoverlabel=dict(
                    bgcolor="#FFFFFF",
                    bordercolor="#CBD5E1",
                    font=dict(size=12, color="#0F172A", family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif"),
                ),
                hovermode="x unified",
                paper_bgcolor="#FFFFFF",
                plot_bgcolor="#FFFFFF",
                bargap=0.20,
            )

            fig_ie.update_xaxes(
                title_text=("Periodo" if periodo == "Mensual" else label_x),
                showgrid=False,
                linecolor="#CBD5E1",
                tickformat=("%b %Y" if periodo == "Mensual" else "%Y"),
                title_font=dict(size=12, color="#475569", family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif"),
                tickfont=dict(size=11, color="#475569", family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif"),
                rangeslider=dict(
                    visible=(periodo == "Mensual"),
                    thickness=0.08,
                    bgcolor="#F1F5F9",
                    bordercolor="#E2E8F0",
                    borderwidth=1,
                ),
            )
            fig_ie.update_yaxes(
                title_text="Flujo CLP",
                showgrid=True,
                gridcolor="rgba(180,190,210,0.15)",
                zeroline=False,
                tickprefix="$",
                separatethousands=True,
                linecolor="#CBD5E1",
                title_font=dict(size=12, color="#475569", family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif"),
                tickfont=dict(size=11, color="#475569", family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif"),
                secondary_y=False,
            )
            legend_script = """
            const chart = document.getElementById('{plot_id}');
            const legendIndexes = () => chart.data
                .map((trace, index) => trace.showlegend === false ? null : index)
                .filter(index => index !== null);
            const isVisible = value => value !== false;
            const setVisibility = visible => Plotly.restyle(chart, {visible: visible});

            chart.on('plotly_legendclick', eventData => {
                const clicked = eventData.curveNumber;
                const indexes = legendIndexes();
                const allVisible = indexes.every(index => isVisible(chart.data[index].visible));
                const visible = chart.data.map((trace, index) => {
                    if (trace.showlegend === false) {
                        return true;
                    }
                    if (allVisible) {
                        return index === clicked ? true : 'legendonly';
                    }
                    return index === clicked ? true : (isVisible(trace.visible) ? true : 'legendonly');
                });
                setVisibility(visible);
                return false;
            });

            chart.on('plotly_legenddoubleclick', eventData => {
                setVisibility(chart.data.map(() => true));
                return false;
            });
            """
            fig_ie_html = fig_ie.to_html(
                full_html=False,
                include_plotlyjs=True,
                config={"displaylogo": False, "displayModeBar": True, "modeBarButtonsToAdd": ["toImage"]},
                post_script=legend_script,
            )
            top_ingreso_periodo = (
                _df[_df["CC_norm"] == "INGRESO"]
                .groupby("Obs_text")["Monto"]
                .sum()
                .sort_values(ascending=False)
            )
            top_egreso_periodo = (
                _df[_df["CC_norm"] == "EGRESO"]
                .groupby("Obs_text")["Monto"]
                .sum()
                .abs()
                .sort_values(ascending=False)
            )
            top_ingreso_label = str(top_ingreso_periodo.index[0]) if not top_ingreso_periodo.empty else "Sin registros"
            top_egreso_label = str(top_egreso_periodo.index[0]) if not top_egreso_periodo.empty else "Sin registros"
            top_ingreso_val = float(top_ingreso_periodo.iloc[0]) if not top_ingreso_periodo.empty else 0.0
            top_egreso_val = float(top_egreso_periodo.iloc[0]) if not top_egreso_periodo.empty else 0.0
            tendencia_txt = "El neto acumulado sigue en negativo." if total_neto_acumulado < 0 else "El neto acumulado se mantiene positivo."
            tendencia_estado = "En riesgo" if total_neto_acumulado < 0 else "Saludable"
            tendencia_bg = "#fee2e2" if total_neto_acumulado < 0 else "#dcfce7"
            tendencia_fg = "#B91C1C" if total_neto_acumulado < 0 else "#166534"
            health_score = int(
                np.clip(
                    55
                    + (18 if total_neto >= 0 else -18)
                    + (16 if total_neto_acumulado >= 0 else -16)
                    + np.clip(total_margen * 70, -16, 16),
                    0,
                    100,
                )
            )
            health_color = "#22C55E" if health_score >= 70 else ("#F59E0B" if health_score >= 45 else "#F87171")
            health_note = "Operación estable" if health_score >= 70 else ("Atención operativa" if health_score >= 45 else "Presión financiera")

            chart_col, analysis_col = st.columns([2.75, 0.82])
            with chart_col:
                components.html(
                    f"""
                    <div style="border:0;border-radius:14px;background:#fff;
                                padding:14px 16px 10px 16px;box-shadow:0 4px 18px rgba(15,23,42,0.045);
                                box-sizing:border-box;">
                        {fig_ie_html}
                    </div>
                    """,
                    height=520,
                    scrolling=False,
                )
            with analysis_col:
                st.markdown(
                    f"""
                    <div class="ie-analysis-card">
                        <div class="ie-analysis-title">Análisis rápido del período</div>
                        <div class="ie-health-card">
                            <svg viewBox="0 0 120 72" aria-hidden="true">
                                <path d="M18 58 A42 42 0 0 1 102 58" fill="none" stroke="#E5EAF2" stroke-width="7.5" stroke-linecap="round" pathLength="100"/>
                                <path d="M18 58 A42 42 0 0 1 102 58" fill="none" stroke="{health_color}" stroke-width="7.5" stroke-linecap="round" pathLength="100" stroke-dasharray="{health_score} 100"/>
                            </svg>
                            <div>
                                <div class="ie-health-label">Operational Health Score</div>
                                <div class="ie-health-score">{health_score}/100</div>
                                <div class="ie-health-note">{health_note}</div>
                            </div>
                        </div>
                        <div class="ie-analysis-list">
                            <div class="ie-analysis-row" style="--accent:#059669;--soft:#DCFCE7;">
                                <div class="ie-analysis-icon">{ie_icon_svg("up")}</div>
                                <div>
                                    <div class="ie-analysis-label">Mayores ingresos</div>
                                    <div class="ie-analysis-sub">{top_ingreso_label}</div>
                                    <div class="ie-analysis-value">{fmt_clp_largo(top_ingreso_val)}</div>
                                </div>
                            </div>
                            <div class="ie-analysis-row" style="--accent:#DC2626;--soft:#FEE2E2;">
                                <div class="ie-analysis-icon">{ie_icon_svg("down")}</div>
                                <div>
                                    <div class="ie-analysis-label">Mayores egresos</div>
                                    <div class="ie-analysis-sub">{top_egreso_label}</div>
                                    <div class="ie-analysis-value">-{fmt_clp_largo(top_egreso_val)}</div>
                                </div>
                            </div>
                            <div class="ie-analysis-row" style="--accent:#2563EB;--soft:#DBEAFE;">
                                <div class="ie-analysis-icon">{ie_icon_svg("equal")}</div>
                                <div>
                                    <div class="ie-analysis-label">Resultado operacional</div>
                                    <div class="ie-analysis-sub">{'Superávit' if total_neto >= 0 else 'Déficit'} operacional</div>
                                    <div class="ie-analysis-value">{fmt_clp_largo(total_neto)}</div>
                                </div>
                            </div>
                            <div class="ie-analysis-row" style="--accent:#B7791F;--soft:#FEF3C7;">
                                <div class="ie-analysis-icon">{ie_icon_svg("percent")}</div>
                                <div>
                                    <div class="ie-analysis-label">Tendencia</div>
                                    <div class="ie-analysis-sub">{tendencia_txt}</div>
                                    <div style="margin-top:5px;"><span class="ie-risk-pill" style="background:{tendencia_bg};color:{tendencia_fg};">{tendencia_estado}</span></div>
                                </div>
                            </div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            st.markdown(
                """
                <div class="ie-guide-note">
                    <span>ⓘ</span>
                    <div>Lectura visual: ingresos, egresos, neto y acumulado se priorizan como señales operacionales; el margen queda disponible en tooltip y KPIs.</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            today_ie = pd.Timestamp.today().normalize()
            fecha_oper_ie = pd.to_datetime(_df.get("Fecha_dt", _df.get("Fecha")), errors="coerce").dt.normalize()
            short_window_ie = fecha_oper_ie.between(today_ie - pd.Timedelta(days=30), today_ie + pd.Timedelta(days=30), inclusive="both")
            pending_ie = _df["Sit_norm"].eq("NO PAGADO") & short_window_ie
            cxc_30_ie = float(_df.loc[pending_ie & _df["CC_norm"].eq("INGRESO"), "Monto"].abs().sum())
            cxp_30_ie = float(_df.loc[pending_ie & _df["CC_norm"].eq("EGRESO"), "Monto"].abs().sum())
            mora_df_ie = df_ie_base_periodo[df_ie_base_periodo["CC_norm"].eq("INGRESO")].copy()
            mora_np_ie = float(mora_df_ie.loc[mora_df_ie["Sit_norm"].eq("NO PAGADO"), "Monto"].abs().sum())
            mora_abono_ie = float(mora_df_ie.loc[mora_df_ie["Sit_norm"].str.startswith("ABONO"), "Monto"].abs().sum())
            mora_acum_ie = max(0.0, mora_np_ie - mora_abono_ie)
            ingresos_resp_ie = (
                _df[_df["CC_norm"].eq("INGRESO")]
                .assign(Resp_clean=lambda d: d["Responsable_clean"].astype(str).str.strip().replace("", "Sin responsable"))
                .groupby("Resp_clean")["Monto"]
                .sum()
                .abs()
                .sort_values(ascending=False)
            )
            ingreso_total_resp_ie = float(ingresos_resp_ie.sum()) if not ingresos_resp_ie.empty else 0.0
            concentracion_ing_ie = float(ingresos_resp_ie.head(3).sum() / ingreso_total_resp_ie) if ingreso_total_resp_ie else 0.0
            top_resp_ie = str(ingresos_resp_ie.index[0]) if not ingresos_resp_ie.empty else "Sin responsable"
            egresos_pressure_ie = (
                _df[_df["CC_norm"].eq("EGRESO")]
                .assign(CC1_clean=lambda d: d["CC1_text"].astype(str).str.strip().replace("", "Sin clasificar"))
                .groupby("CC1_clean")["Monto"]
                .sum()
                .abs()
                .sort_values(ascending=False)
            )
            principal_egreso_ie = str(egresos_pressure_ie.index[0]) if not egresos_pressure_ie.empty else "Sin egresos"
            principal_egreso_val_ie = float(egresos_pressure_ie.iloc[0]) if not egresos_pressure_ie.empty else 0.0

            def ie_status_badge(value: float, kind: str) -> tuple[str, str]:
                if kind == "cxc":
                    if value <= 0:
                        return "critical", "Crítico"
                    if value >= cxp_30_ie:
                        return "healthy", "Saludable"
                    if value >= cxp_30_ie * 0.5:
                        return "attention", "Atención"
                    return "pressure", "Presión"
                if kind == "cxp":
                    if value <= 0:
                        return "healthy", "Saludable"
                    if value <= cxc_30_ie:
                        return "attention", "Atención"
                    if value <= max(cxc_30_ie, 1) * 1.5:
                        return "pressure", "Presión"
                    return "critical", "Crítico"
                if kind == "mora":
                    if value <= 0:
                        return "healthy", "Saludable"
                    if value <= max(cxc_30_ie, 1):
                        return "attention", "Atención"
                    if value <= max(cxc_30_ie, 1) * 2:
                        return "pressure", "Presión"
                    return "critical", "Crítico"
                if kind == "concentracion":
                    if value < 0.45:
                        return "healthy", "Saludable"
                    if value < 0.65:
                        return "attention", "Atención"
                    if value < 0.80:
                        return "pressure", "Presión"
                    return "critical", "Crítico"
                if value <= 0:
                    return "healthy", "Saludable"
                if value <= max(cxp_30_ie, 1) * 0.6:
                    return "attention", "Atención"
                if value <= max(cxp_30_ie, 1):
                    return "pressure", "Presión"
                return "critical", "Crítico"

            risk_cards_ie = [
                ("CxC <30 días", fmt_clp_largo(cxc_30_ie), "Entrada operativa esperada", *ie_status_badge(cxc_30_ie, "cxc")),
                ("CxP <30 días", fmt_clp_largo(cxp_30_ie), "Salida exigible inmediata", *ie_status_badge(cxp_30_ie, "cxp")),
                ("Mora acumulada", fmt_clp_largo(mora_acum_ie), "No pagado neto de abonos", *ie_status_badge(mora_acum_ie, "mora")),
                ("Concentración ingresos", f"{concentracion_ing_ie:.0%}", f"Top 3 · principal: {top_resp_ie}", *ie_status_badge(concentracion_ing_ie, "concentracion")),
                ("Presión de egreso", fmt_clp_largo(principal_egreso_val_ie), principal_egreso_ie, *ie_status_badge(principal_egreso_val_ie, "egreso")),
            ]
            risk_cards_html = "".join(
                f"""
                <div class="ie-risk-card {badge_class}">
                    <div class="ie-risk-card-head">
                        <div class="ie-risk-label">{label}</div>
                        <div class="ie-status-badge {badge_class}">{badge_label}</div>
                    </div>
                    <div>
                        <div class="ie-risk-value">{value}</div>
                        <div class="ie-risk-sub">{sub}</div>
                    </div>
                </div>
                """
                for label, value, sub, badge_class, badge_label in risk_cards_ie
            )
            st.markdown(
                f"""
                <div class="ie-risk-section-title">Riesgos y presión operativa</div>
                <div class="ie-risk-section-sub">Lectura táctica de caja, mora y concentración bajo el filtro actual.</div>
                <div class="ie-risk-grid">{risk_cards_html}</div>
                """,
                unsafe_allow_html=True,
            )

            def build_neto_comparativo(df_src: pd.DataFrame, modo: str) -> pd.DataFrame:
                if modo == "Contable":
                    anio_col = "Año_sel_contable"
                    mes_col = "Mes_sel_contable"
                    periodo_col = "Periodo_ref_contable"
                else:
                    anio_col = "Año_sel"
                    mes_col = "Mes_sel"
                    periodo_col = "Periodo_ref"

                df_cmp = df_src.dropna(subset=["Monto"]).copy()
                df_cmp = df_cmp[df_cmp["CC_norm"].isin(["INGRESO", "EGRESO"])]
                df_cmp = df_cmp[df_cmp["Sit_norm"].isin(["PAGADO", "NO PAGADO"])]
                if periodo == "Mensual":
                    df_cmp = df_cmp.dropna(subset=[periodo_col])
                    df_cmp["Periodo"] = df_cmp[periodo_col]
                else:
                    df_cmp = df_cmp.dropna(subset=[anio_col])
                    df_cmp["Periodo"] = pd.to_datetime(
                        dict(year=df_cmp[anio_col].astype(int), month=1, day=1),
                        errors="coerce",
                    )

                df_cmp["Periodo"] = pd.to_datetime(df_cmp["Periodo"], errors="coerce")
                df_cmp = df_cmp.dropna(subset=["Periodo"])
                if df_cmp.empty:
                    return pd.DataFrame(columns=["Periodo", modo, f"{modo} acumulado"])

                agg_cmp = (
                    df_cmp.groupby(["Periodo", "CC_norm"], as_index=False)["Monto"]
                    .sum()
                    .sort_values("Periodo")
                )
                ingresos_cmp = agg_cmp[agg_cmp["CC_norm"] == "INGRESO"].rename(columns={"Monto": "Ingresos"})
                egresos_cmp = agg_cmp[agg_cmp["CC_norm"] == "EGRESO"].rename(columns={"Monto": "Egresos"})
                base_cmp = pd.DataFrame({"Periodo": sorted(agg_cmp["Periodo"].dropna().unique())})
                base_cmp["Periodo"] = pd.to_datetime(base_cmp["Periodo"], errors="coerce")
                base_cmp = base_cmp.merge(ingresos_cmp[["Periodo", "Ingresos"]], on="Periodo", how="left")
                base_cmp = base_cmp.merge(egresos_cmp[["Periodo", "Egresos"]], on="Periodo", how="left")
                base_cmp = base_cmp.fillna(0).sort_values("Periodo")
                base_cmp[modo] = base_cmp["Ingresos"] - base_cmp["Egresos"].abs()
                base_cmp[f"{modo} acumulado"] = base_cmp[modo].cumsum()

                if sel_year != "Todos":
                    base_cmp = base_cmp[base_cmp["Periodo"].dt.year == int(sel_year)]
                if sel_month != "Todos" and periodo == "Mensual":
                    base_cmp = base_cmp[base_cmp["Periodo"].dt.month == int(sel_month)]

                return base_cmp[["Periodo", modo, f"{modo} acumulado"]]

            neto_financiero = build_neto_comparativo(df_f, "Financiero")
            neto_contable = build_neto_comparativo(df_f, "Contable")
            neto_cmp = neto_financiero.merge(neto_contable, on="Periodo", how="outer").sort_values("Periodo")

            if neto_cmp.empty:
                st.info(f"No se encuentran registros para comparar neto financiero y contable en {periodo_filtro_lbl}.")
            else:
                neto_cmp = neto_cmp.fillna(0)
                st.markdown(
                    f"""
                    <div class="neto-chart-card">
                        <div class="neto-card-head">
                            <div>
                                <div class="neto-card-title">Neto financiero vs neto contable — {periodo} · {periodo_filtro_lbl} ⓘ</div>
                                <div class="neto-card-sub">Evolución mensual del resultado neto (ingresos menos egresos)</div>
                            </div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                with st.container():
                    st.markdown('<div class="neto-filter-anchor"></div>', unsafe_allow_html=True)
                    ctrl_1, ctrl_2, ctrl_3, ctrl_4 = st.columns([1.15, 1.15, 1.4, 0.18], gap="small")
                    with ctrl_1:
                        vista_neto_cmp = st.selectbox(
                            "Visualizar",
                            ["Neto", "Neto acumulado", "Ambos"],
                            index=0,
                            key="vista_neto_fin_cont",
                        )
                    with ctrl_2:
                        tipo_grafico_neto = st.selectbox(
                            "Tipo",
                            ["Línea", "Área"],
                            index=0,
                            key="tipo_grafico_neto_fin_cont",
                        )
                    with ctrl_3:
                        rango_neto_cmp = st.selectbox(
                            "Rango",
                            ["Todos los períodos", "Últimos 12 meses", "Últimos 24 meses"],
                            index=0,
                            key="rango_neto_fin_cont",
                        )
                    with ctrl_4:
                        st.markdown(
                            "<div class='neto-filter-action'>⋮</div>",
                            unsafe_allow_html=True,
                        )

                neto_plot = neto_cmp.copy()
                if rango_neto_cmp != "Todos los períodos" and not neto_plot.empty:
                    meses_rango = 11 if rango_neto_cmp == "Últimos 12 meses" else 23
                    max_periodo_cmp = pd.to_datetime(neto_plot["Periodo"]).max()
                    min_periodo_cmp = pd.to_datetime(neto_plot["Periodo"]).min()
                    inicio_rango_cmp = max(max_periodo_cmp - pd.DateOffset(months=meses_rango), min_periodo_cmp)
                    neto_plot = neto_plot[neto_plot["Periodo"].between(inicio_rango_cmp, max_periodo_cmp)]

                metric_source = (neto_plot if not neto_plot.empty else neto_cmp).copy()
                metric_source["Desviacion"] = metric_source["Financiero"] - metric_source["Contable"]
                metric_source["Desviacion_abs"] = metric_source["Desviacion"].abs()
                last_cmp = metric_source.iloc[-1] if not metric_source.empty else None
                desviacion_actual = float(last_cmp["Desviacion"]) if last_cmp is not None else 0.0
                desviacion_acumulada = (
                    float(last_cmp["Financiero acumulado"] - last_cmp["Contable acumulado"])
                    if last_cmp is not None and "Financiero acumulado" in metric_source and "Contable acumulado" in metric_source
                    else 0.0
                )
                base_consistencia = metric_source[["Financiero", "Contable", "Desviacion_abs"]].replace([np.inf, -np.inf], np.nan).dropna()
                if base_consistencia.empty:
                    consistencia_ratio = 0.0
                else:
                    tolerancia = max(float(base_consistencia[["Financiero", "Contable"]].abs().max().max()) * 0.08, 1.0)
                    mismo_signo = np.sign(base_consistencia["Financiero"]) == np.sign(base_consistencia["Contable"])
                    baja_desviacion = base_consistencia["Desviacion_abs"] <= tolerancia
                    consistencia_ratio = float((mismo_signo | baja_desviacion).mean())
                consistencia_label = "Alta" if consistencia_ratio >= 0.75 else ("Media" if consistencia_ratio >= 0.5 else "Baja")
                consistencia_color = "#059669" if consistencia_ratio >= 0.75 else ("#D97706" if consistencia_ratio >= 0.5 else "#DC2626")

                st.markdown(
                    f"""
                    <div class="neto-kpi-grid">
                        <div class="neto-kpi-box" style="--metric:{'#059669' if desviacion_actual >= 0 else '#EF4444'};">
                            <div class="neto-kpi-label">Desviación último período</div>
                            <div class="neto-kpi-value">{fmt_clp_largo(desviacion_actual)}</div>
                            <div class="neto-kpi-note">Financiero - contable</div>
                        </div>
                        <div class="neto-kpi-box" style="--metric:{'#059669' if desviacion_acumulada >= 0 else '#EF4444'};">
                            <div class="neto-kpi-label">Desviación acumulada</div>
                            <div class="neto-kpi-value">{fmt_clp_largo(desviacion_acumulada)}</div>
                            <div class="neto-kpi-note">Brecha acumulada</div>
                        </div>
                        <div class="neto-kpi-box" style="--metric:{consistencia_color};">
                            <div class="neto-kpi-label">Consistencia financiero/contable</div>
                            <div class="neto-kpi-value">{consistencia_ratio:.0%}</div>
                            <div class="neto-kpi-note">{consistencia_label}</div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                from plotly.subplots import make_subplots

                if vista_neto_cmp == "Ambos":
                    fig_neto_cmp = make_subplots(
                        rows=2,
                        cols=1,
                        shared_xaxes=True,
                        vertical_spacing=0.09,
                        row_heights=[0.52, 0.48],
                        subplot_titles=("Neto del período", "Neto acumulado"),
                    )
                    chart_height_cmp = 430
                    range_slider_cmp = False
                else:
                    fig_neto_cmp = go.Figure()
                    chart_height_cmp = 340
                    range_slider_cmp = True

                def add_cmp_trace(trace, row=None):
                    if row is None:
                        fig_neto_cmp.add_trace(trace)
                    else:
                        fig_neto_cmp.add_trace(trace, row=row, col=1)

                def add_deviation_bands(row=None):
                    diff_base_cmp = neto_plot[["Periodo", "Financiero", "Contable"]].replace([np.inf, -np.inf], np.nan).dropna()
                    if not diff_base_cmp.empty:
                        diff_abs_cmp = (diff_base_cmp["Financiero"] - diff_base_cmp["Contable"]).abs()
                        scale_cmp = max(
                            float(diff_base_cmp[["Financiero", "Contable"]].abs().max().max()),
                            1.0,
                        )
                        significant_cmp = diff_abs_cmp >= max(scale_cmp * 0.025, 1.0)
                        deviation_ratio_cmp = diff_abs_cmp / scale_cmp
                        deviation_bands_cmp = [
                            ("baja", "rgba(37,99,235,0.10)", significant_cmp & (deviation_ratio_cmp < 0.08)),
                            ("media", "rgba(245,158,11,0.14)", significant_cmp & (deviation_ratio_cmp >= 0.08) & (deviation_ratio_cmp < 0.18)),
                            ("alta", "rgba(220,38,38,0.13)", significant_cmp & (deviation_ratio_cmp >= 0.18)),
                        ]
                        y_low_cmp = diff_base_cmp[["Financiero", "Contable"]].min(axis=1)
                        y_high_cmp = diff_base_cmp[["Financiero", "Contable"]].max(axis=1)
                        for band_label, band_color, band_mask in deviation_bands_cmp:
                            if not bool(band_mask.any()):
                                continue
                            lower_values = y_low_cmp.where(band_mask, np.nan)
                            upper_values = y_high_cmp.where(band_mask, np.nan)
                            add_cmp_trace(
                                go.Scatter(
                                    x=diff_base_cmp["Periodo"],
                                    y=lower_values,
                                    mode="lines",
                                    name=f"Desviación {band_label}",
                                    line=dict(width=0, color="rgba(0,0,0,0)", shape="spline"),
                                    hoverinfo="skip",
                                    showlegend=False,
                                ),
                                row=row,
                            )
                            add_cmp_trace(
                                go.Scatter(
                                    x=diff_base_cmp["Periodo"],
                                    y=upper_values,
                                    mode="lines",
                                    name=f"Área desviación {band_label}",
                                    line=dict(width=0, color="rgba(0,0,0,0)", shape="spline"),
                                    fill="tonexty",
                                    fillcolor=band_color,
                                    hoverinfo="skip",
                                    showlegend=False,
                                ),
                                row=row,
                            )

                show_neto_cmp = vista_neto_cmp in ["Neto", "Ambos"]
                show_acum_cmp = vista_neto_cmp in ["Neto acumulado", "Ambos"]
                neto_row_cmp = 1 if vista_neto_cmp == "Ambos" else None
                acum_row_cmp = 2 if vista_neto_cmp == "Ambos" else None
                fill_fin = "tozeroy" if tipo_grafico_neto == "Área" and vista_neto_cmp != "Ambos" else None
                fill_con = "tozeroy" if tipo_grafico_neto == "Área" and vista_neto_cmp != "Ambos" else None

                if show_neto_cmp:
                    add_deviation_bands(row=neto_row_cmp)
                    add_cmp_trace(
                        go.Scatter(
                            x=neto_plot["Periodo"],
                            y=neto_plot["Financiero"],
                            mode="lines+markers",
                            name="Neto financiero",
                            line=dict(color="#1257FF", width=3.0, shape="spline"),
                            marker=dict(size=7, color="#1257FF", line=dict(color="#FFFFFF", width=1.4)),
                            fill=fill_fin,
                            fillcolor="rgba(18,87,255,0.08)",
                            hovertemplate="<b>%{x|" + x_hover + "}</b><br>Neto financiero: $%{y:,.0f}<extra></extra>",
                        ),
                        row=neto_row_cmp,
                    )
                    add_cmp_trace(
                        go.Scatter(
                            x=neto_plot["Periodo"],
                            y=neto_plot["Contable"],
                            mode="lines+markers",
                            name="Neto contable",
                            line=dict(color="#FF7A1A", width=2.8, dash="dot", shape="spline"),
                            marker=dict(size=7, color="#FF7A1A", line=dict(color="#FFFFFF", width=1.4)),
                            fill=fill_con,
                            fillcolor="rgba(255,122,26,0.07)",
                            hovertemplate="<b>%{x|" + x_hover + "}</b><br>Neto contable: $%{y:,.0f}<extra></extra>",
                        ),
                        row=neto_row_cmp,
                    )

                if show_acum_cmp:
                    fill_acum_fin = "tozeroy"
                    fill_acum_con = "tozeroy" if vista_neto_cmp == "Neto acumulado" else None
                    add_cmp_trace(
                        go.Scatter(
                            x=neto_plot["Periodo"],
                            y=neto_plot["Financiero acumulado"],
                            mode="lines+markers",
                            name="Neto acumulado financiero",
                            line=dict(color="#0F766E", width=3.0, dash="dash", shape="spline"),
                            marker=dict(size=7, color="#0F766E", line=dict(color="#FFFFFF", width=1.4)),
                            fill=fill_acum_fin,
                            fillcolor="rgba(15,118,110,0.10)",
                            hovertemplate="<b>%{x|" + x_hover + "}</b><br>Neto acumulado financiero: $%{y:,.0f}<extra></extra>",
                        ),
                        row=acum_row_cmp,
                    )
                    add_cmp_trace(
                        go.Scatter(
                            x=neto_plot["Periodo"],
                            y=neto_plot["Contable acumulado"],
                            mode="lines+markers",
                            name="Neto acumulado contable",
                            line=dict(color="#B91C1C", width=3.0, dash="dashdot", shape="spline"),
                            marker=dict(size=7, color="#B91C1C", line=dict(color="#FFFFFF", width=1.4)),
                            fill=fill_acum_con,
                            fillcolor="rgba(185,28,28,0.07)",
                            hovertemplate="<b>%{x|" + x_hover + "}</b><br>Neto acumulado contable: $%{y:,.0f}<extra></extra>",
                        ),
                        row=acum_row_cmp,
                    )

                if vista_neto_cmp == "Ambos":
                    fig_neto_cmp.add_hline(y=0, line_width=1.1, line_color=CHART_GRAY, opacity=0.75, row=1, col=1)
                    fig_neto_cmp.add_hline(y=0, line_width=1.1, line_color=CHART_GRAY, opacity=0.75, row=2, col=1)
                else:
                    fig_neto_cmp.add_hline(y=0, line_width=1.1, line_color=CHART_GRAY, opacity=0.75)

                if not neto_plot.empty and vista_neto_cmp == "Neto":
                    last_row_cmp = neto_plot.iloc[-1]
                    fig_neto_cmp.add_annotation(
                        x=last_row_cmp["Periodo"],
                        y=last_row_cmp["Financiero"],
                        text=fmt_clp_largo(float(last_row_cmp["Financiero"])),
                        showarrow=False,
                        xshift=46,
                        bgcolor="#EFF6FF",
                        bordercolor="#DBEAFE",
                        borderwidth=1,
                        font=dict(size=11, color="#1257FF"),
                    )
                    fig_neto_cmp.add_annotation(
                        x=last_row_cmp["Periodo"],
                        y=last_row_cmp["Contable"],
                        text=fmt_clp_largo(float(last_row_cmp["Contable"])),
                        showarrow=False,
                        xshift=46,
                        bgcolor="#FFF7ED",
                        bordercolor="#FED7AA",
                        borderwidth=1,
                        font=dict(size=11, color="#EA580C"),
                    )
                fig_neto_cmp.update_layout(
                    title=None,
                    font=dict(family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif", color="#334155"),
                    template="plotly_white",
                    height=chart_height_cmp,
                    margin=dict(l=28, r=52, t=28, b=26),
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                    y=1.04,
                        xanchor="left",
                        x=0.01,
                        bgcolor="rgba(255,255,255,0)",
                        font=dict(size=11, color="#334155"),
                    ),
                    hoverlabel=dict(bgcolor="#FFFFFF", bordercolor="#CBD5E1", font=dict(size=12, color="#0F172A")),
                    hovermode="x unified",
                    paper_bgcolor="#FFFFFF",
                    plot_bgcolor="#FFFFFF",
                )
                if vista_neto_cmp == "Ambos":
                    fig_neto_cmp.update_annotations(font=dict(size=12, color="#081735"))
                fig_neto_cmp.update_xaxes(
                    title_text=label_x,
                    showgrid=False,
                    tickformat=("%b %Y" if periodo == "Mensual" else "%Y"),
                    linecolor="#CBD5E1",
                    title_font=dict(size=12, color="#334155"),
                    tickfont=dict(size=11, color="#475569"),
                    rangeslider=dict(visible=range_slider_cmp, thickness=0.07, bgcolor="#F1F5F9", bordercolor="#E2E8F0", borderwidth=1),
                )
                fig_neto_cmp.update_yaxes(
                    title_text="Neto CLP",
                    showgrid=True,
                    gridcolor="#E5EAF2",
                    zeroline=False,
                    tickprefix="$",
                    separatethousands=True,
                    linecolor="#CBD5E1",
                    title_font=dict(size=12, color="#334155"),
                    tickfont=dict(size=11, color="#475569"),
                )
                st.plotly_chart(
                    fig_neto_cmp,
                    use_container_width=True,
                    config={"displaylogo": False, "displayModeBar": True, "modeBarButtonsToAdd": ["toImage"]},
                    key="neto_fin_cont_pro",
                )
                st.markdown(
                    """
                    <div class="neto-note">
                        <span>ⓘ</span>
                        <div>El neto financiero considera la clasificación financiera de los movimientos.<br>
                        El neto contable considera la clasificación contable de los movimientos.</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        st.markdown("---")
        @st.fragment
        def render_riesgos_cobro_concentracion():
            st.markdown(
                """
                <style>
                .risk-pro-title-row {
                    display:flex;
                    justify-content:space-between;
                    align-items:flex-start;
                    gap:12px;
                    margin:0 0 5px 0;
                }
                .risk-pro-title {
                    color:#081735;
                    font-size:19px;
                    line-height:1.05;
                    font-weight:950;
                    letter-spacing:-0.025em;
                }
                .risk-pro-subtitle {
                    margin-top:4px;
                    color:#475569;
                    font-size:11.5px;
                    line-height:1.25;
                    font-weight:650;
                }
                .risk-pro-actions {
                    display:flex;
                    gap:7px;
                    align-items:center;
                }
                .risk-pro-action {
                    min-height:28px;
                    display:flex;
                    align-items:center;
                    justify-content:center;
                    border-radius:8px;
                    border:1px solid #dbe3ee;
                    background:#ffffff;
                    color:#0f1f3d;
                    padding:0 10px;
                    font-size:10.5px;
                    font-weight:850;
                    box-shadow:0 8px 16px rgba(15,23,42,0.035);
                }
                .risk-pro-action-primary {
                    background:#0B3A86;
                    border-color:#0B3A86;
                    color:#ffffff;
                }
                .risk-filter-card {
                    border:1px solid #dbe3ee;
                    border-radius:10px;
                    background:#ffffff;
                    padding:7px 10px 5px 10px;
                    margin:0 0 6px 0;
                    box-shadow:0 8px 18px rgba(15,23,42,0.035);
                }
                .risk-filter-title {
                    color:#081735;
                    font-size:10px;
                    font-weight:950;
                    margin-bottom:4px;
                    text-transform:uppercase;
                    letter-spacing:.018em;
                }
                .risk-filter-card [data-testid="stHorizontalBlock"] {
                    gap:0.45rem !important;
                    align-items:flex-end !important;
                }
                .risk-filter-card [data-testid="stSelectbox"],
                .risk-filter-card [data-testid="stRadio"],
                .risk-filter-card [data-testid="stSlider"] {
                    margin:0 !important;
                }
                .risk-filter-card [data-testid="stSelectbox"] > label,
                .risk-filter-card [data-testid="stRadio"] > label,
                .risk-filter-card [data-testid="stSlider"] > label {
                    color:#64748b !important;
                    font-size:9.2px !important;
                    line-height:1 !important;
                    font-weight:950 !important;
                    text-transform:uppercase !important;
                    letter-spacing:.018em !important;
                    padding-bottom:3px !important;
                }
                .risk-filter-card [data-baseweb="select"] > div {
                    min-height:30px !important;
                    height:30px !important;
                    border-radius:8px !important;
                    background:#fbfdff !important;
                    border-color:#dbe3ee !important;
                }
                .risk-filter-card [data-baseweb="select"] div {
                    font-size:11px !important;
                    font-weight:800 !important;
                }
                .risk-filter-card div[role="radiogroup"] {
                    min-height:30px !important;
                    border-radius:8px !important;
                }
                .risk-filter-card div[role="radiogroup"] > label {
                    min-height:30px !important;
                    padding:0 8px !important;
                    font-size:10.5px !important;
                }
                .risk-pro-kpi {
                    min-height:62px;
                    border:1px solid var(--border);
                    border-radius:10px;
                    background:linear-gradient(135deg, #ffffff 0%, var(--soft) 100%);
                    padding:7px 8px;
                    display:grid;
                    grid-template-columns:26px 1fr;
                    gap:7px;
                    box-shadow:0 8px 18px rgba(15,23,42,0.035);
                }
                .risk-pro-kpi-icon {
                    width:24px;
                    height:24px;
                    border-radius:999px;
                    display:flex;
                    align-items:center;
                    justify-content:center;
                    background:var(--halo);
                    color:var(--accent);
                    font-size:12px;
                    font-weight:950;
                }
                .risk-pro-kpi-title {
                    color:#0f1f3d;
                    font-size:8.4px;
                    line-height:1.15;
                    font-weight:950;
                    text-transform:uppercase;
                    letter-spacing:.012em;
                }
                .risk-pro-kpi-value {
                    margin-top:3px;
                    color:var(--accent);
                    font-size:15px;
                    line-height:1;
                    font-weight:950;
                    letter-spacing:-0.03em;
                    white-space:nowrap;
                }
                .risk-pro-kpi-note {
                    margin-top:3px;
                    color:#64748b;
                    font-size:8.2px;
                    line-height:1.12;
                    font-weight:750;
                }
                .risk-kpi-anchor {
                    height:0;
                    min-height:0;
                    margin:0;
                    padding:0;
                    overflow:hidden;
                }
                div[data-testid="stVerticalBlock"]:has(.risk-kpi-anchor) > div[data-testid="stHorizontalBlock"] {
                    gap:0.45rem;
                    margin-bottom:0;
                }
                .risk-chart-card {
                    border:1px solid #dbe3ee;
                    border-radius:10px;
                    background:#ffffff;
                    padding:7px 10px 8px 10px;
                    margin-top:4px;
                    box-shadow:0 8px 18px rgba(15,23,42,0.035);
                }
                .risk-chart-toolbar {
                    display:flex;
                    align-items:center;
                    justify-content:space-between;
                    gap:10px;
                    margin:4px 0 3px 0;
                    padding:0 1px;
                }
                .risk-chart-toolbar-title {
                    color:#081735;
                    font-size:12px;
                    line-height:1.1;
                    font-weight:950;
                    text-transform:uppercase;
                    letter-spacing:.016em;
                }
                .risk-chart-toolbar-sub {
                    margin-top:2px;
                    color:#64748b;
                    font-size:9.2px;
                    line-height:1.1;
                    font-weight:700;
                }
                .risk-segment-anchor {
                    height:0;
                    min-height:0;
                    margin:0;
                    padding:0;
                    overflow:hidden;
                }
                div[data-testid="stVerticalBlock"]:has(.risk-segment-anchor) > div[data-testid="stHorizontalBlock"] {
                    align-items:flex-end;
                    gap:0.5rem;
                    margin-bottom:0;
                }
                div[data-testid="stVerticalBlock"]:has(.risk-segment-anchor) div[role="radiogroup"] {
                    display:grid !important;
                    grid-template-columns:repeat(2, minmax(0, 1fr));
                    gap:3px !important;
                    min-height:31px !important;
                    padding:3px !important;
                    border-radius:999px !important;
                    background:#EEF4FF !important;
                    border:1px solid rgba(37,99,235,0.14) !important;
                    box-shadow:inset 0 1px 0 rgba(255,255,255,0.84), 0 5px 14px rgba(15,23,42,0.035) !important;
                }
                div[data-testid="stVerticalBlock"]:has(.risk-segment-anchor) div[role="radiogroup"] > label {
                    min-height:25px !important;
                    margin:0 !important;
                    padding:0 13px !important;
                    border-radius:999px !important;
                    color:#334155 !important;
                    font-size:10.5px !important;
                    font-weight:950 !important;
                    display:flex !important;
                    align-items:center !important;
                    justify-content:center !important;
                    transition:all .18s ease !important;
                    white-space:nowrap !important;
                }
                div[data-testid="stVerticalBlock"]:has(.risk-segment-anchor) div[role="radiogroup"] > label > div:first-child {
                    display:none !important;
                }
                div[data-testid="stVerticalBlock"]:has(.risk-segment-anchor) div[role="radiogroup"] > label:has(input:checked) {
                    background:linear-gradient(135deg,#2563EB 0%,#1D4ED8 100%) !important;
                    color:#ffffff !important;
                    box-shadow:0 6px 16px rgba(37,99,235,0.22) !important;
                }
                .risk-note-card {
                    display:flex;
                    gap:9px;
                    align-items:flex-start;
                    border:1px solid #dbeafe;
                    background:#eff6ff;
                    color:#334155;
                    border-radius:8px;
                    padding:10px 12px;
                    font-size:12px;
                    font-weight:650;
                    margin-top:8px;
                }
                .risk-auto-insights {
                    display:grid;
                    grid-template-columns:repeat(2, minmax(0, 1fr));
                    gap:7px;
                    margin-top:8px;
                }
                .risk-auto-insight {
                    border:1px solid rgba(226,232,240,0.92);
                    border-radius:10px;
                    background:#fbfdff;
                    padding:9px 10px;
                    display:grid;
                    grid-template-columns:auto 1fr;
                    gap:8px;
                    align-items:flex-start;
                }
                .risk-auto-badge {
                    display:inline-flex;
                    align-items:center;
                    justify-content:center;
                    min-width:52px;
                    height:22px;
                    padding:0 8px;
                    border-radius:999px;
                    font-size:8.4px;
                    line-height:1;
                    font-weight:950;
                    text-transform:uppercase;
                }
                .risk-auto-badge.healthy {
                    background:rgba(34,197,94,0.13);
                    color:#15803D;
                }
                .risk-auto-badge.attention {
                    background:rgba(234,179,8,0.16);
                    color:#A16207;
                }
                .risk-auto-badge.pressure {
                    background:rgba(249,115,22,0.15);
                    color:#C2410C;
                }
                .risk-auto-badge.critical {
                    background:rgba(220,38,38,0.14);
                    color:#DC2626;
                }
                .risk-auto-copy {
                    color:#334155;
                    font-size:10.7px;
                    line-height:1.25;
                    font-weight:720;
                }
                .risk-auto-title {
                    color:#081735;
                    font-size:11px;
                    line-height:1.1;
                    font-weight:950;
                    margin-top:8px;
                }
                </style>
                <div class="risk-pro-title-row">
                    <div>
                        <div class="risk-pro-title">⚠️ Riesgos de cobro y concentración de montos</div>
                        <div class="risk-pro-subtitle">Monitoreo de riesgos y concentración de montos · Vista analítica de cobranza</div>
                    </div>
                    <div class="risk-pro-actions">
                        <div class="risk-pro-action risk-pro-action-primary">⇩ Descargar reporte</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(
                '<div id="riesgos-selectores-anchor"></div>'
                '<div class="risk-filter-card"><div class="risk-filter-title">Filtro por centro de costo / situación</div>',
                unsafe_allow_html=True,
            )
            render_flujo_selector_scroll("riesgos-selectores-anchor")

            c_top1, c_top2, c_top3, c_top4 = st.columns([1.6, 1.15, 1.05, 0.8], gap="small")
            with c_top1:
                dim = st.selectbox(
                    "Dimensión",
                    ["Obs", "CC1", "Sit", "CC"],
                    index=1,
                    key="dim_pro",
                    on_change=queue_flujo_selector_scroll,
                    args=("riesgos-selectores-anchor",),
                )
            with c_top2:
                order_by = st.radio(
                    "Ordenar por",
                    ["Total CLP", "N° Transacciones"],
                    horizontal=True,
                    index=0,
                    key="order_by_pro",
                    on_change=queue_flujo_selector_scroll,
                    args=("riesgos-selectores-anchor",),
                )
            with c_top3:
                chart_type = st.selectbox(
                    "Visualización",
                    ["Barras", "Treemap"],
                    index=0,
                    key="chart_type_pro",
                    on_change=queue_flujo_selector_scroll,
                    args=("riesgos-selectores-anchor",),
                )
            with c_top4:
                top_n = st.slider(
                    "Top N",
                    min_value=5,
                    max_value=30,
                    value=12,
                    step=1,
                    key="topn_pro",
                    on_change=queue_flujo_selector_scroll,
                    args=("riesgos-selectores-anchor",),
                )
            st.markdown("</div>", unsafe_allow_html=True)

            df_riesgos = df_ie_base_periodo.dropna(subset=["Monto"]).copy()
            df_riesgos[dim] = df_riesgos[dim].astype(str).str.strip()
            df_riesgos = df_riesgos[
                df_riesgos[dim].notna()
                & (df_riesgos[dim] != "")
                & (df_riesgos[dim].str.upper() != "NAN")
            ]
            if df_riesgos.empty:
                st.info(f"No se encuentran registros para {periodo_filtro_lbl} en análisis {tipo_analisis.lower()}.")
                st.stop()

            topN_raw = (
                df_riesgos.groupby(dim)["Monto"]
                .agg(["sum", "count"])
                .rename(columns={"sum": "Total CLP", "count": "N° Transacciones"})
                .reset_index()
            )

            sort_col = "Total CLP" if order_by == "Total CLP" else "N° Transacciones"
            topN = topN_raw.sort_values(sort_col, ascending=False).head(top_n).copy()

            def color_by_sign_or_cc(series_values, series_dim=None):
                if series_dim == "CC":
                    return series_values.map(
                        {"INGRESO": CHART_TEAL, "EGRESO": CHART_RED}
                    ).fillna(CHART_DARK)
                return series_values.apply(lambda v: CHART_TEAL if v >= 0 else CHART_RED)

            if chart_type == "Barras":
                top_keys = topN[dim].tolist()
                df_dim = df_riesgos[df_riesgos[dim].isin(top_keys)].copy()
                df_dim["CC"] = df_dim["CC"].astype(str).str.strip().str.upper()

                group_cols_dim_cc = [dim] if dim == "CC" else [dim, "CC"]
                agg_cc = (
                    df_dim.groupby(group_cols_dim_cc, as_index=False)["Monto"]
                    .sum()
                )

                ingresos = agg_cc[agg_cc["CC"] == "INGRESO"].rename(columns={"Monto": "Ingresos"})
                egresos = agg_cc[agg_cc["CC"] == "EGRESO"].rename(columns={"Monto": "Egresos"})

                base_dim = pd.DataFrame({dim: top_keys})
                base_dim = base_dim.merge(ingresos[[dim, "Ingresos"]], on=dim, how="left")
                base_dim = base_dim.merge(egresos[[dim, "Egresos"]], on=dim, how="left")
                base_dim = base_dim.fillna(0)

                base_dim["Egresos_abs"] = base_dim["Egresos"].abs()
                base_dim["Neto"] = base_dim["Ingresos"] - base_dim["Egresos_abs"]
                base_dim = base_dim.merge(topN[[dim, "Total CLP", "N° Transacciones"]], on=dim, how="left")
                base_dim["Impacto_abs"] = base_dim["Ingresos"].abs() + base_dim["Egresos_abs"]
                impacto_total = base_dim["Impacto_abs"].sum()
                base_dim["Concentración"] = np.where(impacto_total != 0, base_dim["Impacto_abs"] / impacto_total, 0.0)
                base_dim["Neto_color"] = np.where(base_dim["Neto"] >= 0, "#7EA9A2", "#EF524C")
                base_dim["Neto_pos"] = base_dim["Neto"].clip(lower=0)
                base_dim["Neto_neg"] = base_dim["Neto"].clip(upper=0)
                base_dim["Tooltip"] = (
                    "Ingresos: $" + base_dim["Ingresos"].map(lambda v: f"{v:,.0f}")
                    + "<br>Egresos: $" + base_dim["Egresos_abs"].map(lambda v: f"{v:,.0f}")
                    + "<br>Neto: $" + base_dim["Neto"].map(lambda v: f"{v:,.0f}")
                    + "<br>Transacciones: " + base_dim["N° Transacciones"].fillna(0).astype(int).astype(str)
                    + "<br>Concentración: " + base_dim["Concentración"].map(lambda v: f"{v:.1%}")
                )
                total_neto_pos = float(base_dim["Neto_pos"].sum())
                total_neto_neg = float(base_dim["Neto_neg"].sum())
                total_neto_dim = float(base_dim["Neto"].sum())
                total_tx_dim = int(base_dim["N° Transacciones"].fillna(0).sum())
                neto_base_pct = abs(total_neto_pos) + abs(total_neto_neg)
                periodo_card = (
                    pd.to_datetime(base["Periodo"]).max().strftime("%b %Y")
                    if "base" in locals() and not base.empty and periodo == "Mensual"
                    else periodo_filtro_lbl
                )
                top3_exposicion_val = float(base_dim.nlargest(3, "Impacto_abs")["Impacto_abs"].sum()) if not base_dim.empty else 0.0
                concentracion_top3_risk = (top3_exposicion_val / impacto_total) if impacto_total else 0.0
                ingreso_source_row = base_dim.loc[base_dim["Ingresos"].idxmax()] if not base_dim.empty and base_dim["Ingresos"].max() > 0 else None
                egreso_pressure_row = base_dim.loc[base_dim["Egresos_abs"].idxmax()] if not base_dim.empty and base_dim["Egresos_abs"].max() > 0 else None
                principal_ingreso_label = str(ingreso_source_row[dim]) if ingreso_source_row is not None else "Sin ingresos"
                principal_egreso_label = str(egreso_pressure_row[dim]) if egreso_pressure_row is not None else "Sin egresos"
                principal_ingreso_val = float(ingreso_source_row["Ingresos"]) if ingreso_source_row is not None else 0.0
                principal_egreso_val = float(egreso_pressure_row["Egresos_abs"]) if egreso_pressure_row is not None else 0.0
                exposicion_negativa = abs(total_neto_neg)
                exposicion_negativa_ratio = exposicion_negativa / neto_base_pct if neto_base_pct else 0.0

                def risk_short_label(value, max_len=30):
                    value = str(value).strip() or "Sin clasificar"
                    return value if len(value) <= max_len else f"{value[:max_len - 1]}…"

                if concentracion_top3_risk >= 0.65 or exposicion_negativa_ratio >= 0.60:
                    riesgo_operativo_label = "Crítico"
                    riesgo_operativo_note = "Alta concentración o presión negativa"
                    riesgo_color = "#DC2626"
                    riesgo_soft = "#fff7f7"
                    riesgo_halo = "#fee2e2"
                    riesgo_border = "#f1caca"
                elif concentracion_top3_risk >= 0.45 or exposicion_negativa_ratio >= 0.35:
                    riesgo_operativo_label = "Presión"
                    riesgo_operativo_note = "Monitorear exposición concentrada"
                    riesgo_color = "#F97316"
                    riesgo_soft = "#fff8f1"
                    riesgo_halo = "#ffedd5"
                    riesgo_border = "#fed7aa"
                elif concentracion_top3_risk >= 0.30 or exposicion_negativa_ratio >= 0.20:
                    riesgo_operativo_label = "Atención"
                    riesgo_operativo_note = "Riesgo moderado bajo control"
                    riesgo_color = "#EAB308"
                    riesgo_soft = "#fffdf0"
                    riesgo_halo = "#fef9c3"
                    riesgo_border = "#fde68a"
                else:
                    riesgo_operativo_label = "Saludable"
                    riesgo_operativo_note = "Concentración y presión acotadas"
                    riesgo_color = "#16A34A"
                    riesgo_soft = "#f6fffb"
                    riesgo_halo = "#dcfce7"
                    riesgo_border = "#bbf7d0"

                def risk_pro_kpi(title, value, note, icon, accent, soft, halo, border):
                    return f"""
                    <div class="risk-pro-kpi" style="--accent:{accent};--soft:{soft};--halo:{halo};--border:{border};">
                        <div class="risk-pro-kpi-icon">{icon}</div>
                        <div>
                            <div class="risk-pro-kpi-title">{title}</div>
                            <div class="risk-pro-kpi-value">{value}</div>
                            <div class="risk-pro-kpi-note">{note}</div>
                        </div>
                    </div>
                    """

                st.markdown('<div class="risk-kpi-anchor"></div>', unsafe_allow_html=True)
                rk1, rk2, rk3, rk4, rk5 = st.columns(5, gap="small")
                with rk1:
                    st.markdown(
                        risk_pro_kpi(
                            "CONCENTRACIÓN TOP 3",
                            f"{concentracion_top3_risk:.1%}",
                            f"{fmt_clp_largo(top3_exposicion_val)} de exposición",
                            "◎",
                            "#2563EB",
                            "#f6f9ff",
                            "#dbeafe",
                            "#d4e1f6",
                        ),
                        unsafe_allow_html=True,
                    )
                with rk2:
                    st.markdown(
                        risk_pro_kpi(
                            "PRINCIPAL INGRESO",
                            fmt_clp_largo(principal_ingreso_val),
                            escape(risk_short_label(principal_ingreso_label)),
                            "↑",
                            "#16A34A",
                            "#f6fffb",
                            "#dcfce7",
                            "#bbf7d0",
                        ),
                        unsafe_allow_html=True,
                    )
                with rk3:
                    st.markdown(
                        risk_pro_kpi(
                            "PRESIÓN DE EGRESO",
                            fmt_clp_largo(principal_egreso_val),
                            escape(risk_short_label(principal_egreso_label)),
                            "↓",
                            "#F97316",
                            "#fff8f1",
                            "#ffedd5",
                            "#fed7aa",
                        ),
                        unsafe_allow_html=True,
                    )
                with rk4:
                    st.markdown(
                        risk_pro_kpi(
                            "EXPOSICIÓN NEGATIVA",
                            fmt_clp_largo(exposicion_negativa),
                            f"{exposicion_negativa_ratio:.1%} de la exposición neta",
                            "!",
                            "#DC2626",
                            "#fff7f7",
                            "#fee2e2",
                            "#f1caca",
                        ),
                        unsafe_allow_html=True,
                    )
                with rk5:
                    st.markdown(
                        risk_pro_kpi(
                            "RIESGO OPERATIVO",
                            riesgo_operativo_label,
                            riesgo_operativo_note,
                            "◌",
                            riesgo_color,
                            riesgo_soft,
                            riesgo_halo,
                            riesgo_border,
                        ),
                        unsafe_allow_html=True,
                    )

                if order_by == "Total CLP":
                    base_dim = base_dim.sort_values("Impacto_abs", ascending=True)
                else:
                    base_dim = base_dim.sort_values("N° Transacciones", ascending=True)
                top3_exposiciones = set(base_dim.sort_values("Impacto_abs", ascending=False).head(3)[dim].astype(str))
                base_dim["Es_top3"] = base_dim[dim].astype(str).isin(top3_exposiciones)
                base_dim["Color_pos"] = np.where(base_dim["Es_top3"], "rgba(22,163,74,0.96)", "rgba(134,190,177,0.64)")
                base_dim["Color_neg"] = np.where(base_dim["Es_top3"], "rgba(220,38,38,0.96)", "rgba(248,113,113,0.62)")
                base_dim["Line_pos"] = np.where(base_dim["Es_top3"], "rgba(21,128,61,0.95)", "rgba(134,190,177,0.45)")
                base_dim["Line_neg"] = np.where(base_dim["Es_top3"], "rgba(185,28,28,0.95)", "rgba(248,113,113,0.45)")

                st.markdown('<div id="riesgos-vista-selector-anchor"></div>', unsafe_allow_html=True)
                render_flujo_selector_scroll("riesgos-vista-selector-anchor")
                toolbar_title_col, toolbar_control_col = st.columns([2.35, 1], gap="small")
                with toolbar_title_col:
                    st.markdown(
                        f"""
                        <div class="risk-chart-toolbar">
                            <div>
                                <div class="risk-chart-toolbar-title">Exposición y concentración</div>
                                <div class="risk-chart-toolbar-sub">Top {top_n} por {dim} · {periodo_filtro_lbl}</div>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                with toolbar_control_col:
                    st.markdown('<div class="risk-segment-anchor"></div>', unsafe_allow_html=True)
                    view_metric_risk = st.radio(
                        "Vista",
                        ["Monto (CLP)", "% del neto"],
                        horizontal=True,
                        index=0,
                        key="risk_view_metric_pro",
                        label_visibility="collapsed",
                        on_change=queue_flujo_selector_scroll,
                        args=("riesgos-vista-selector-anchor",),
                    )
                if view_metric_risk == "% del neto":
                    denom_risk = neto_base_pct if neto_base_pct else 1.0
                    base_dim["X_pos"] = base_dim["Neto_pos"] / denom_risk
                    base_dim["X_neg"] = base_dim["Neto_neg"] / denom_risk
                    base_dim["X_neto"] = base_dim["Neto"] / denom_risk
                    x_title_risk = "% del neto"
                    x_tickformat_risk = ".0%"
                    x_tickprefix_risk = None
                else:
                    base_dim["X_pos"] = base_dim["Neto_pos"]
                    base_dim["X_neg"] = base_dim["Neto_neg"]
                    base_dim["X_neto"] = base_dim["Neto"]
                    x_title_risk = "Monto CLP"
                    x_tickformat_risk = None
                    x_tickprefix_risk = "$"

                fig_top = go.Figure()

                fig_top.add_trace(
                    go.Scatter(
                        x=base_dim["X_neto"],
                        y=base_dim[dim],
                        mode="markers",
                        name="Resumen",
                        marker=dict(size=18, color="rgba(0,0,0,0)"),
                        showlegend=False,
                        customdata=base_dim[["Tooltip"]],
                        hovertemplate="<b>%{y}</b><br>%{customdata[0]}<extra></extra>",
                    )
                )

                fig_top.add_trace(
                    go.Bar(
                        x=base_dim["X_pos"],
                        y=base_dim[dim],
                        orientation="h",
                        name="Netos positivos (ingresos > egresos)",
                        width=0.74,
                        marker=dict(
                            color=base_dim["Color_pos"],
                            line=dict(color=base_dim["Line_pos"], width=np.where(base_dim["Es_top3"], 1.2, 0)),
                        ),
                        hoverinfo="skip",
                    )
                )
                fig_top.add_trace(
                    go.Bar(
                        x=base_dim["X_neg"],
                        y=base_dim[dim],
                        orientation="h",
                        name="Netos negativos (egresos > ingresos)",
                        width=0.74,
                        marker=dict(
                            color=base_dim["Color_neg"],
                            line=dict(color=base_dim["Line_neg"], width=np.where(base_dim["Es_top3"], 1.2, 0)),
                        ),
                        hoverinfo="skip",
                    )
                )
                top3_dim = base_dim[base_dim["Es_top3"]]
                if not top3_dim.empty:
                    fig_top.add_trace(
                        go.Scatter(
                            x=top3_dim["X_neto"],
                            y=top3_dim[dim],
                            mode="markers",
                            name="Top 3 exposición",
                            marker=dict(
                                size=24,
                                color="rgba(15,23,42,0)",
                                line=dict(color="rgba(15,23,42,0.32)", width=2.2),
                                symbol="diamond",
                            ),
                            hoverinfo="skip",
                            showlegend=False,
                        )
                    )
                fig_top.add_trace(
                    go.Scatter(
                        x=base_dim["X_neto"],
                        y=base_dim[dim],
                        mode="markers",
                        name="Neto",
                        marker=dict(
                            size=np.where(base_dim["Es_top3"], 16, 13),
                            color=base_dim["Neto_color"],
                            line=dict(color="#FFFFFF", width=np.where(base_dim["Es_top3"], 2.4, 1.8)),
                            symbol="diamond",
                        ),
                        hoverinfo="skip",
                    )
                )

                max_abs_riesgo = max(
                    float(base_dim["X_pos"].abs().max()) if not base_dim.empty else 0.0,
                    float(base_dim["X_neg"].abs().max()) if not base_dim.empty else 0.0,
                    float(base_dim["X_neto"].abs().max()) if not base_dim.empty else 0.0,
                    1.0,
                )
                for _, row_risk in base_dim.iterrows():
                    if row_risk["X_neto"] == 0:
                        continue
                    fig_top.add_annotation(
                        x=row_risk["X_neto"],
                        y=row_risk[dim],
                        text=(f"{row_risk['X_neto']:.1%}" if view_metric_risk == "% del neto" else fmt_clp_largo(row_risk["Neto"])),
                        showarrow=False,
                        xshift=31 if row_risk["X_neto"] >= 0 else -31,
                        bgcolor="rgba(255,255,255,0.82)" if row_risk["Es_top3"] else "rgba(255,255,255,0)",
                        bordercolor="rgba(15,23,42,0.12)" if row_risk["Es_top3"] else "rgba(255,255,255,0)",
                        borderwidth=1 if row_risk["Es_top3"] else 0,
                        font=dict(size=12 if row_risk["Es_top3"] else 10, color="#081735" if row_risk["Es_top3"] else "#334155"),
                    )

                fig_top.update_layout(
                    title=dict(text=""),
                    template="plotly_white",
                    height=max(450, 34 * len(base_dim) + 98),
                    margin=dict(l=22, r=42, t=28, b=28),
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=1.02,
                        xanchor="left",
                        x=0.01,
                        bgcolor="rgba(255,255,255,0)",
                        font=dict(size=11, color="#334155"),
                    ),
                    hovermode="closest",
                    font=dict(family="-apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif", color="#334155"),
                    paper_bgcolor="#FFFFFF",
                    plot_bgcolor="#FFFFFF",
                    barmode="relative",
                    bargap=0.14,
                )
                fig_top.add_vline(x=0, line_width=1.2, line_color=CHART_GRAY, opacity=0.75)
                fig_top.update_xaxes(
                    title_text=x_title_risk,
                    showgrid=True,
                    gridcolor="rgba(180,190,210,0.14)",
                    zeroline=False,
                    tickprefix=x_tickprefix_risk,
                    tickformat=x_tickformat_risk,
                    separatethousands=True,
                    range=[-max_abs_riesgo * 1.18, max_abs_riesgo * 1.18],
                    linecolor="#CBD5E1",
                    title_font=dict(size=12, color="#334155"),
                    tickfont=dict(size=11, color="#475569"),
                )
                fig_top.update_yaxes(
                    title_text=dim,
                    showgrid=False,
                    linecolor="#CBD5E1",
                    title_font=dict(size=12, color="#334155"),
                    tickfont=dict(size=11, color="#475569"),
                    automargin=True,
                )

                ingresos_total_dim = float(base_dim["Ingresos"].sum()) if not base_dim.empty else 0.0
                principal_ingreso_share = principal_ingreso_val / ingresos_total_dim if ingresos_total_dim else 0.0
                categorias_criticas = (
                    base_dim[base_dim["Neto"] < 0]
                    .sort_values("Impacto_abs", ascending=False)[dim]
                    .astype(str)
                    .head(2)
                    .tolist()
                )

                def risk_auto_badge_class(label: str) -> str:
                    return {
                        "Saludable": "healthy",
                        "Atención": "attention",
                        "Presión": "pressure",
                        "Crítico": "critical",
                    }.get(label, "attention")

                insights_risk = []
                conc_badge = "Crítico" if concentracion_top3_risk >= 0.65 else ("Presión" if concentracion_top3_risk >= 0.45 else "Atención")
                insights_risk.append(
                    (
                        conc_badge,
                        f"Top 3 concentra {concentracion_top3_risk:.1%} de la exposición; conviene monitorear dependencia por {dim}.",
                    )
                )
                pressure_badge = "Crítico" if principal_egreso_val >= max(principal_ingreso_val, 1) else "Presión"
                insights_risk.append(
                    (
                        pressure_badge,
                        f"La principal presión de egreso es {escape(risk_short_label(principal_egreso_label, 36))}, con {fmt_clp_largo(principal_egreso_val)} expuestos.",
                    )
                )
                dep_badge = "Presión" if principal_ingreso_share >= 0.45 else ("Atención" if principal_ingreso_share >= 0.25 else "Saludable")
                insights_risk.append(
                    (
                        dep_badge,
                        f"{escape(risk_short_label(principal_ingreso_label, 36))} aporta {principal_ingreso_share:.1%} de los ingresos filtrados; revisar dependencia operacional.",
                    )
                )
                if exposicion_negativa > 0:
                    neg_badge = "Crítico" if exposicion_negativa_ratio >= 0.60 else ("Presión" if exposicion_negativa_ratio >= 0.35 else "Atención")
                    cat_txt = ", ".join(escape(risk_short_label(cat, 24)) for cat in categorias_criticas) if categorias_criticas else "sin categoría crítica dominante"
                    insights_risk.append(
                        (
                            neg_badge,
                            f"Exposición negativa de {fmt_clp_largo(exposicion_negativa)} ({exposicion_negativa_ratio:.1%}); categorías críticas: {cat_txt}.",
                        )
                    )

                insights_risk_html = "".join(
                    f"""
                    <div class="risk-auto-insight">
                        <div class="risk-auto-badge {risk_auto_badge_class(badge)}">{badge}</div>
                        <div class="risk-auto-copy">{copy}</div>
                    </div>
                    """
                    for badge, copy in insights_risk[:4]
                )

                st.markdown('<div class="risk-chart-card">', unsafe_allow_html=True)
                st.plotly_chart(
                    fig_top,
                    use_container_width=True,
                    config={
                        "displaylogo": False,
                        "displayModeBar": True,
                        "modeBarButtonsToAdd": ["toImage"],
                    },
                    key="riesgos_concentracion_pro",
                )
                st.markdown(
                    f"""
                    <div class="risk-auto-title">Insights automáticos</div>
                    <div class="risk-auto-insights">{insights_risk_html}</div>
                    """,
                    unsafe_allow_html=True,
                )
                st.markdown(
                    """
                    <div class="risk-note-card">
                        <span>ⓘ</span>
                        <div>La exposición neta integra ingresos y egresos para priorizar conceptos con mayor impacto operacional.</div>
                    </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            else:
                treemap_df = topN.copy()
                treemap_df["Impacto_abs"] = treemap_df["Total CLP"].abs()
                treemap_df["Neto_color"] = treemap_df["Total CLP"]
                if dim == "CC":
                    color_col = dim
                    color_scale = None
                else:
                    color_col = "Neto_color"
                    color_scale = [CHART_RED, CHART_GOLD, CHART_TEAL]

                fig_tree = px.treemap(
                    treemap_df,
                    path=[dim],
                    values="Impacto_abs" if order_by == "Total CLP" else "N° Transacciones",
                    color=color_col,
                    color_continuous_scale=color_scale,
                    title=f"Concentración Top {top_n} por {dim} · {order_by} · {periodo_filtro_lbl}",
                )
                if dim == "CC":
                    fig_tree.update_traces(
                        marker_colors=color_by_sign_or_cc(
                            treemap_df[dim], series_dim="CC"
                        )
                    )
                fig_tree.update_traces(
                    hovertemplate="<b>%{label}</b><br>Impacto: $%{value:,.0f}<extra></extra>",
                    textinfo="label+percent entry",
                    textfont=dict(size=12),
                )
                fig_tree.update_layout(
                    margin=dict(l=0, r=0, t=72, b=0),
                    template="plotly_white",
                    paper_bgcolor="#F8FAFC",
                    plot_bgcolor="#FFFFFF",
                    title=dict(
                        x=0.01,
                        xanchor="left",
                        font=dict(size=18, color="#0F2D52"),
                    ),
                )
                st.plotly_chart(
                    fig_tree,
                    use_container_width=True,
                    config={
                        "displaylogo": False,
                        "displayModeBar": True,
                        "modeBarButtonsToAdd": ["toImage"],
                    },
                )

            st.caption(f"Lectura consolidada bajo criterio {tipo_analisis.lower()} · horizonte ejecutivo: {periodo_filtro_lbl}.")


        render_riesgos_cobro_concentracion()

        @st.fragment
        def render_detalle_filtrable_movimientos():
            st.markdown("---")
            st.markdown(
                section_heading("⚠️", "Detalle filtrable de movimientos", weight_class="section-heading-title-soft"),
                unsafe_allow_html=True,
            )
            st.markdown(
                """
                <div style="
                    background: linear-gradient(135deg, #EEF4FA 0%, #E7EEF7 100%);
                    border: 1px solid rgba(148,163,184,0.24);
                    border-left: 4px solid rgba(37,99,235,0.38);
                    border-radius: 9px;
                    padding: 7px 12px;
                    margin: 5px 0 7px 0;
                    color: #24415F;
                    font-size: 12px;
                    line-height: 1.15;
                    font-weight: 800;
                    letter-spacing: .01em;
                    box-shadow: inset 0 1px 0 rgba(255,255,255,0.72), 0 4px 14px rgba(15,23,42,0.035);">
                    Trazabilidad operacional consolidada · análisis por centro de costo, concepto y responsable
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.caption("Exploración ejecutiva del universo transaccional con filtros encadenados y lectura financiera directa.")
            st.markdown('<div id="detalle-selectores-anchor"></div>', unsafe_allow_html=True)
            render_flujo_selector_scroll("detalle-selectores-anchor")
            st.markdown(
                """
                <style>
                .detalle-toolbar-anchor {
                    height:0;
                    min-height:0;
                    margin:0;
                    padding:0;
                    overflow:hidden;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) > div[data-testid="stHorizontalBlock"] {
                    border:1px solid rgba(219,227,238,0.86);
                    border-radius:12px;
                    background:#fbfdff;
                    box-shadow:inset 0 1px 0 rgba(255,255,255,0.78), 0 4px 14px rgba(15,23,42,0.035);
                    padding:7px 9px 5px 9px;
                    margin:1px 0 6px 0;
                    gap:0.75rem;
                    align-items:flex-end;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) [data-testid="stTextInput"],
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) [data-testid="stSelectbox"] {
                    margin:0 !important;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) [data-testid="stTextInput"] > label,
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) [data-testid="stSelectbox"] > label {
                    color:#64748b !important;
                    font-size:9.2px !important;
                    line-height:1 !important;
                    font-weight:950 !important;
                    text-transform:uppercase !important;
                    letter-spacing:.018em !important;
                    padding-bottom:3px !important;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) [data-baseweb="input"] {
                    min-height:30px !important;
                    height:30px !important;
                    border-radius:8px !important;
                    background:#ffffff !important;
                    border-color:#dbe3ee !important;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) [data-baseweb="select"] > div {
                    min-height:30px !important;
                    height:30px !important;
                    border-radius:8px !important;
                    background:#ffffff !important;
                    border-color:#dbe3ee !important;
                    box-shadow:none !important;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) [data-baseweb="input"] input,
                div[data-testid="stVerticalBlock"]:has(.detalle-toolbar-anchor) [data-baseweb="select"] div {
                    font-size:11px !important;
                    font-weight:780 !important;
                }
                .detalle-toolbar-label {
                    color:#081735;
                    font-size:10px;
                    line-height:1.1;
                    font-weight:950;
                    text-transform:uppercase;
                    letter-spacing:.018em;
                    margin-bottom:4px;
                }
                .detalle-toolbar-copy {
                    color:#64748b;
                    font-size:9px;
                    line-height:1.15;
                    font-weight:700;
                }
                .detalle-alerts-card {
                    border:1px solid rgba(219,227,238,0.84);
                    border-radius:14px;
                    background:linear-gradient(135deg,#ffffff 0%,#fbfdff 100%);
                    box-shadow:0 8px 20px rgba(15,23,42,0.045);
                    padding:10px 12px;
                    margin:-4px 0 10px 0;
                }
                .detalle-alerts-head {
                    display:flex;
                    align-items:center;
                    justify-content:space-between;
                    gap:10px;
                    margin-bottom:8px;
                }
                .detalle-alerts-title {
                    color:#081735;
                    font-size:12px;
                    line-height:1.1;
                    font-weight:950;
                    text-transform:uppercase;
                    letter-spacing:.018em;
                }
                .detalle-alerts-sub {
                    color:#64748b;
                    font-size:9.5px;
                    line-height:1.15;
                    font-weight:700;
                    margin-top:2px;
                }
                .detalle-alerts-grid {
                    display:grid;
                    grid-template-columns:repeat(2, minmax(0, 1fr));
                    gap:7px;
                }
                .detalle-alert {
                    border:1px solid rgba(226,232,240,0.88);
                    border-radius:10px;
                    background:#ffffff;
                    padding:8px 9px;
                    display:grid;
                    grid-template-columns:auto 1fr;
                    gap:8px;
                    align-items:flex-start;
                }
                .detalle-alert-badge {
                    display:inline-flex;
                    align-items:center;
                    justify-content:center;
                    min-width:50px;
                    height:21px;
                    padding:0 8px;
                    border-radius:999px;
                    font-size:8px;
                    line-height:1;
                    font-weight:950;
                    text-transform:uppercase;
                }
                .detalle-alert-badge.healthy {
                    background:rgba(34,197,94,0.13);
                    color:#15803D;
                }
                .detalle-alert-badge.attention {
                    background:rgba(234,179,8,0.16);
                    color:#A16207;
                }
                .detalle-alert-badge.pressure {
                    background:rgba(249,115,22,0.15);
                    color:#C2410C;
                }
                .detalle-alert-badge.critical {
                    background:rgba(220,38,38,0.14);
                    color:#DC2626;
                }
                .detalle-alert-copy {
                    color:#334155;
                    font-size:10.4px;
                    line-height:1.25;
                    font-weight:720;
                }
                .detalle-export-toolbar-anchor {
                    height:0;
                    min-height:0;
                    margin:0;
                    padding:0;
                    overflow:hidden;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-export-toolbar-anchor) > div[data-testid="stHorizontalBlock"] {
                    border:1px solid rgba(219,227,238,0.86);
                    border-radius:12px;
                    background:linear-gradient(135deg,#ffffff 0%,#fbfdff 100%);
                    box-shadow:0 5px 16px rgba(15,23,42,0.035);
                    padding:7px 9px;
                    margin:4px 0 6px 0;
                    gap:0.5rem;
                    align-items:center;
                }
                .detalle-export-title {
                    color:#081735;
                    font-size:11px;
                    line-height:1.1;
                    font-weight:950;
                    text-transform:uppercase;
                    letter-spacing:.018em;
                }
                .detalle-export-sub {
                    color:#64748b;
                    font-size:9.2px;
                    line-height:1.15;
                    font-weight:700;
                    margin-top:2px;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-export-toolbar-anchor) [data-testid="stDownloadButton"] button {
                    min-height:30px !important;
                    height:30px !important;
                    border-radius:8px !important;
                    border:1px solid rgba(219,227,238,0.95) !important;
                    background:#ffffff !important;
                    color:#0F2D52 !important;
                    font-size:10.4px !important;
                    font-weight:900 !important;
                    padding:0 10px !important;
                    box-shadow:inset 0 1px 0 rgba(255,255,255,0.76) !important;
                }
                div[data-testid="stVerticalBlock"]:has(.detalle-export-toolbar-anchor) [data-testid="stDownloadButton"] button:hover {
                    border-color:rgba(37,99,235,0.38) !important;
                    background:#F8FBFF !important;
                    box-shadow:0 5px 14px rgba(37,99,235,0.08) !important;
                }
                </style>
                """,
                unsafe_allow_html=True,
            )

            df_det = df_ie.copy().dropna(subset=["Monto"]).copy()
            for c in ["CC1", "Obs", "Responsable"]:
                df_det[c] = df_det[c].astype(str).str.strip()

            buscador_cols = [
                c for c in ["Responsable", "CC1", "Obs", "Sit", "CC", "Esp", "Fecha", "Año", "Mes", "Monto"]
                if c in df_det.columns
            ]
            with st.container():
                st.markdown('<div class="detalle-toolbar-anchor"></div>', unsafe_allow_html=True)
                tb_label_col, tb_search_col = st.columns([0.82, 3.2], gap="large")
                with tb_label_col:
                    st.markdown(
                        '<div class="detalle-toolbar-label">Explorador operativo</div><div class="detalle-toolbar-copy">Búsqueda contextual y filtros activos</div>',
                        unsafe_allow_html=True,
                    )
                with tb_search_col:
                    busqueda_det = st.text_input(
                        "Búsqueda ejecutiva",
                        value="",
                        placeholder="Responsable, concepto, centro de costo, estado, monto o fecha...",
                        key="det_buscador_inteligente",
                        on_change=queue_flujo_selector_scroll,
                        args=("detalle-selectores-anchor",),
                    ).strip()

            df_det_search = df_det.copy()
            search_blob = pd.Series("", index=df_det_search.index, dtype="object")
            for c in buscador_cols:
                if c == "Fecha":
                    search_piece = pd.to_datetime(df_det_search[c], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
                else:
                    search_piece = df_det_search[c].astype(str).fillna("")
                search_blob = search_blob + " " + search_piece
            search_blob = search_blob.str.lower()

            if busqueda_det:
                terms = [t.lower() for t in busqueda_det.split() if t.strip()]
                mask_busqueda = pd.Series(True, index=df_det_search.index)
                for term in terms:
                    mask_busqueda &= search_blob.str.contains(term, regex=False, na=False)
                df_det_search = df_det_search[mask_busqueda]

                resumen_busqueda_items = []
                for c in ["Responsable", "CC1", "Obs", "Sit", "CC"]:
                    if c not in df_det_search.columns:
                        continue
                    vals = (
                        df_det_search[c]
                        .astype(str)
                        .str.strip()
                        .replace("", pd.NA)
                        .dropna()
                        .drop_duplicates()
                        .head(25)
                        .tolist()
                    )
                    if vals:
                        resumen_busqueda_items.append(f"{c}: {', '.join(vals[:6])}")
                resumen_busqueda = " · ".join(resumen_busqueda_items) if resumen_busqueda_items else "Sin coincidencias nominales."
                st.caption(f"Universo filtrado por '{busqueda_det}': {len(df_det_search):,} movimientos relevantes · {resumen_busqueda}".replace(",", "."))

            with st.container():
                st.markdown('<div class="detalle-toolbar-anchor"></div>', unsafe_allow_html=True)
                d1, d2, d3, d4, d5 = st.columns([1.12, 1.2, 1.45, 0.78, 0.78], gap="large")
                with d1:
                    cc1_opts = ["Todos"] + sorted([v for v in df_det_search["CC1"].dropna().unique().tolist() if v != ""])
                    sel_cc1_det = st.selectbox(
                        "CC1",
                        cc1_opts,
                        index=0,
                        key="det_cc1",
                        on_change=queue_flujo_selector_scroll,
                        args=("detalle-selectores-anchor",),
                    )

            df_det_f = df_det_search.copy()
            if sel_cc1_det != "Todos":
                df_det_f = df_det_f[df_det_f["CC1"] == sel_cc1_det]

            with d2:
                obs_opts = ["Todos"] + sorted([v for v in df_det_f["Obs"].dropna().unique().tolist() if v != ""])
                sel_obs_det = st.selectbox(
                    "OBS",
                    obs_opts,
                    index=0,
                    key="det_obs",
                    on_change=queue_flujo_selector_scroll,
                    args=("detalle-selectores-anchor",),
                )
            if sel_obs_det != "Todos":
                df_det_f = df_det_f[df_det_f["Obs"] == sel_obs_det]

            with d3:
                resp_opts = ["Todos"] + sorted([v for v in df_det_f["Responsable"].dropna().unique().tolist() if v != ""])
                sel_resp_det = st.selectbox(
                    "Responsable",
                    resp_opts,
                    index=0,
                    key="det_resp",
                    on_change=queue_flujo_selector_scroll,
                    args=("detalle-selectores-anchor",),
                )
            if sel_resp_det != "Todos":
                df_det_f = df_det_f[df_det_f["Responsable"] == sel_resp_det]

            with d4:
                anio_vals = pd.to_numeric(df_det_f["Año_sel"], errors="coerce").dropna().astype(int).unique().tolist()
                anio_opts = ["Todos"] + [str(y) for y in sorted(anio_vals)]
                sel_anio_det = st.selectbox(
                    "Año",
                    anio_opts,
                    index=0,
                    key="det_anio",
                    on_change=queue_flujo_selector_scroll,
                    args=("detalle-selectores-anchor",),
                )
            if sel_anio_det != "Todos":
                df_det_f = df_det_f[pd.to_numeric(df_det_f["Año_sel"], errors="coerce").astype("Int64") == int(sel_anio_det)]

            with d5:
                if "Mes_sel" in df_det_f.columns:
                    mes_vals = pd.to_numeric(df_det_f["Mes_sel"], errors="coerce").dropna().astype(int).unique().tolist()
                    mes_opts = ["Todos"] + sorted(mes_vals)
                else:
                    mes_opts = ["Todos"]
                sel_mes_det = st.selectbox(
                    "Mes",
                    mes_opts,
                    index=0,
                    key="det_mes",
                    on_change=queue_flujo_selector_scroll,
                    args=("detalle-selectores-anchor",),
                )
            if sel_mes_det != "Todos" and "Mes_sel" in df_det_f.columns:
                df_det_f = df_det_f[pd.to_numeric(df_det_f["Mes_sel"], errors="coerce").astype("Int64") == int(sel_mes_det)]

            # KPIs de detalle (debajo de selectores)
            sit_det = df_det_f["Sit"].astype(str).str.strip().str.upper() if "Sit" in df_det_f.columns else pd.Series([], dtype=str)
            monto_total_det = float(df_det_f["Monto"].sum()) if not df_det_f.empty else 0.0
            monto_por_pagar_det = abs(float(df_det_f.loc[sit_det == "NO PAGADO", "Monto"].sum())) if not df_det_f.empty else 0.0
            monto_pagado_det = abs(float(df_det_f.loc[sit_det == "PAGADO", "Monto"].sum())) if not df_det_f.empty else 0.0
            monto_abonos_det = float(
                df_det_f.loc[df_det_f["Obs"].astype(str).str.contains("abono", case=False, na=False), "Monto"].abs().sum()
            ) if not df_det_f.empty else 0.0

            resumen_obs_export = pd.DataFrame()
            grafico_obs_export = pd.DataFrame()
            if "Obs" in df_det_f.columns and not df_det_f.empty:
                resumen_obs = df_det_f.copy()
                resumen_obs["Obs_resumen"] = resumen_obs["Obs"].astype(str).str.strip().replace("", "Sin OBS")
                resumen_obs["Sit_resumen"] = resumen_obs["Sit"].astype(str).str.strip().str.upper() if "Sit" in resumen_obs.columns else ""
                resumen_obs["Monto_pagado_obs"] = np.where(resumen_obs["Sit_resumen"].eq("PAGADO"), resumen_obs["Monto"], 0.0)
                resumen_obs["Monto_no_pagado_obs"] = np.where(resumen_obs["Sit_resumen"].eq("NO PAGADO"), resumen_obs["Monto"], 0.0)
                mask_abono_obs_det = (
                    resumen_obs["Sit_resumen"].str.startswith("ABONO") |
                    resumen_obs["Obs_resumen"].astype(str).str.contains("abono", case=False, na=False)
                )
                resumen_obs["Monto_abono_obs"] = np.where(mask_abono_obs_det, resumen_obs["Monto"], 0.0)
                resumen_obs_tbl = (
                    resumen_obs.groupby("Obs_resumen", dropna=False)
                    .agg(
                        Registros=("Monto", "size"),
                        Pagado=("Monto_pagado_obs", "sum"),
                        **{"No pagado": ("Monto_no_pagado_obs", "sum")},
                        Abono=("Monto_abono_obs", "sum"),
                    )
                    .reset_index()
                    .rename(columns={"Obs_resumen": "OBS"})
                )
                resumen_obs_tbl["RESULTADO"] = resumen_obs_tbl["Pagado"] + resumen_obs_tbl["No pagado"]
                resumen_obs_tbl["Deuda a la fecha"] = np.where(
                    (resumen_obs_tbl["Pagado"] < 0) & (resumen_obs_tbl["No pagado"] < 0),
                    resumen_obs_tbl["No pagado"] + resumen_obs_tbl["Abono"],
                    np.where(
                        resumen_obs_tbl["Pagado"] < 0,
                        resumen_obs_tbl["Pagado"] + resumen_obs_tbl["Abono"],
                        resumen_obs_tbl["No pagado"] - resumen_obs_tbl["Abono"],
                    ),
                )
                resumen_obs_tbl = (
                    resumen_obs_tbl.sort_values("Pagado", ascending=True)
                    .reset_index(drop=True)
                )
                if not resumen_obs_tbl.empty:
                    mostrar_fila_total_obs = len(resumen_obs_tbl) > 1
                    if mostrar_fila_total_obs:
                        resumen_obs_tbl = pd.concat(
                            [
                                resumen_obs_tbl,
                                pd.DataFrame(
                                    [
                                        {
                                            "OBS": "TOTAL TABLA (FILTRO)",
                                            "Registros": len(resumen_obs),
                                            "Pagado": resumen_obs["Monto_pagado_obs"].sum(),
                                            "No pagado": resumen_obs["Monto_no_pagado_obs"].sum(),
                                            "RESULTADO": resumen_obs["Monto_pagado_obs"].sum() + resumen_obs["Monto_no_pagado_obs"].sum(),
                                            "Abono": resumen_obs["Monto_abono_obs"].sum(),
                                            "Deuda a la fecha": (
                                                resumen_obs["Monto_no_pagado_obs"].sum() + resumen_obs["Monto_abono_obs"].sum()
                                                if resumen_obs["Monto_pagado_obs"].sum() < 0 and resumen_obs["Monto_no_pagado_obs"].sum() < 0
                                                else (
                                                    resumen_obs["Monto_pagado_obs"].sum() + resumen_obs["Monto_abono_obs"].sum()
                                                    if resumen_obs["Monto_pagado_obs"].sum() < 0
                                                    else resumen_obs["Monto_no_pagado_obs"].sum() - resumen_obs["Monto_abono_obs"].sum()
                                                )
                                            ),
                                        }
                                    ]
                                ),
                            ],
                            ignore_index=True,
                        )
                    resumen_obs_export = resumen_obs_tbl.copy()
                    resumen_obs_items = []
                    for _, row in resumen_obs_tbl.iterrows():
                        resultado_val = float(row["RESULTADO"])
                        resultado_class = (
                            "resultado-positive"
                            if resultado_val > 0
                            else "resultado-negative"
                            if resultado_val < 0
                            else "resultado-neutral"
                        )
                        is_total_row = str(row["OBS"]).strip().upper() == "TOTAL TABLA (FILTRO)"
                        resumen_obs_items.append(
                            {
                                "label": escape(str(row["OBS"])),
                                "meta": (
                                    "Neto operacional real del filtro"
                                    if is_total_row
                                    else f'{int(row["Registros"]):,} registros'.replace(",", ".")
                                ),
                                "pagado": fmt_clp_largo(float(row["Pagado"])),
                                "no_pagado": fmt_clp_largo(float(row["No pagado"])),
                                "resultado": fmt_clp_largo(resultado_val),
                                "resultado_class": resultado_class,
                                "row_class": "resultado-total-row" if is_total_row else "",
                                "abono": fmt_clp_largo(float(row["Abono"])),
                                "pendiente_deuda": fmt_clp_largo(float(row["Deuda a la fecha"])),
                            }
                        )
                    st.markdown(
                        '<div class="kpi-summary-compact">'
                        + kpi_resumen_obs_panel(
                            "Resumen de montos por OBS",
                            "Montos agrupados según los filtros activos del detalle",
                            resumen_obs_items,
                        )
                        + "</div>",
                        unsafe_allow_html=True,
                    )
                    obs_alert_base = resumen_obs_tbl[
                        resumen_obs_tbl["OBS"].astype(str).str.strip().str.upper() != "TOTAL TABLA (FILTRO)"
                    ].copy()

                    def detalle_alert_badge_class(label: str) -> str:
                        return {
                            "Saludable": "healthy",
                            "Atención": "attention",
                            "Presión": "pressure",
                            "Crítico": "critical",
                        }.get(label, "attention")

                    def detalle_short(value, max_len=34):
                        value = str(value).strip() or "Sin OBS"
                        return value if len(value) <= max_len else f"{value[:max_len - 1]}…"

                    detalle_alertas = []
                    total_egresos_obs = float(obs_alert_base.loc[obs_alert_base["Pagado"] < 0, "Pagado"].abs().sum()) if not obs_alert_base.empty else 0.0
                    if total_egresos_obs > 0:
                        egreso_rows = obs_alert_base.assign(Egreso_abs=obs_alert_base["Pagado"].clip(upper=0).abs())
                        egreso_top = egreso_rows.sort_values("Egreso_abs", ascending=False).iloc[0]
                        egreso_share = float(egreso_top["Egreso_abs"] / total_egresos_obs) if total_egresos_obs else 0.0
                        egreso_badge = "Crítico" if egreso_share >= 0.55 else ("Presión" if egreso_share >= 0.35 else "Atención")
                        detalle_alertas.append(
                            (
                                egreso_badge,
                                f"{escape(detalle_short(egreso_top['OBS']))} concentra {egreso_share:.1%} de los egresos del filtro ({fmt_clp_largo(float(egreso_top['Egreso_abs']))}).",
                            )
                        )

                    deuda_rows = obs_alert_base[obs_alert_base["Deuda a la fecha"].abs() > 0].copy()
                    if not deuda_rows.empty:
                        deuda_rows["Deuda_abs"] = deuda_rows["Deuda a la fecha"].abs()
                        deuda_top = deuda_rows.sort_values("Deuda_abs", ascending=False).iloc[0]
                        deuda_total = float(deuda_rows["Deuda_abs"].sum())
                        deuda_share = float(deuda_top["Deuda_abs"] / deuda_total) if deuda_total else 0.0
                        deuda_badge = "Crítico" if deuda_share >= 0.55 else ("Presión" if deuda_share >= 0.35 else "Atención")
                        detalle_alertas.append(
                            (
                                deuda_badge,
                                f"OBS con mayor deuda: {escape(detalle_short(deuda_top['OBS']))}, por {fmt_clp_largo(float(deuda_top['Deuda_abs']))}.",
                            )
                        )

                    mora_rows = obs_alert_base[obs_alert_base["No pagado"].abs() > 0].copy()
                    if not mora_rows.empty:
                        mora_rows["Mora_abs"] = mora_rows["No pagado"].abs()
                        mora_total = float(mora_rows["Mora_abs"].sum())
                        mora_count = int(len(mora_rows))
                        mora_badge = "Crítico" if mora_count >= 5 else ("Presión" if mora_count >= 3 else "Atención")
                        detalle_alertas.append(
                            (
                                mora_badge,
                                f"Mora activa en {mora_count} OBS por {fmt_clp_largo(mora_total)}; priorizar recuperación por antigüedad y monto.",
                            )
                        )

                    critical_rows = obs_alert_base.assign(
                        Criticidad=obs_alert_base["Pagado"].clip(upper=0).abs() + obs_alert_base["No pagado"].abs() + obs_alert_base["Deuda a la fecha"].abs()
                    )
                    critical_rows = critical_rows[critical_rows["Criticidad"] > 0].sort_values("Criticidad", ascending=False)
                    if not critical_rows.empty:
                        top_criticas = [escape(detalle_short(v, 24)) for v in critical_rows["OBS"].head(3).tolist()]
                        crit_badge = "Crítico" if len(top_criticas) >= 3 else "Presión"
                        detalle_alertas.append(
                            (
                                crit_badge,
                                f"Categorías críticas principales: {', '.join(top_criticas)}.",
                            )
                        )

                    if not detalle_alertas:
                        detalle_alertas.append(("Saludable", "No se detectan alertas operacionales relevantes bajo los filtros actuales."))

                    detalle_alertas_html = "".join(
                        f"""
                        <div class="detalle-alert">
                            <div class="detalle-alert-badge {detalle_alert_badge_class(badge)}">{badge}</div>
                            <div class="detalle-alert-copy">{copy}</div>
                        </div>
                        """
                        for badge, copy in detalle_alertas[:4]
                    )
                    st.markdown(
                        f"""
                        <div class="detalle-alerts-card">
                            <div class="detalle-alerts-head">
                                <div>
                                    <div class="detalle-alerts-title">Alertas operacionales</div>
                                    <div class="detalle-alerts-sub">Insights automáticos según resumen OBS y filtros activos</div>
                                </div>
                            </div>
                            <div class="detalle-alerts-grid">{detalle_alertas_html}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                    chart_obs = resumen_obs.copy()
                    chart_obs["Año_chart"] = pd.to_numeric(chart_obs.get("Año_sel"), errors="coerce")
                    chart_obs["Mes_chart"] = pd.to_numeric(chart_obs.get("Mes_sel"), errors="coerce")
                    chart_obs = chart_obs.dropna(subset=["Año_chart", "Mes_chart", "Monto"])
                    if not chart_obs.empty:
                        chart_obs["Periodo_chart"] = pd.to_datetime(
                            dict(
                                year=chart_obs["Año_chart"].astype(int),
                                month=chart_obs["Mes_chart"].astype(int),
                                day=1,
                            ),
                            errors="coerce",
                        )
                        chart_obs = chart_obs.dropna(subset=["Periodo_chart"])

                    if not chart_obs.empty:
                        chart_periodo = (
                            chart_obs.groupby("Periodo_chart", as_index=False)
                            .agg(
                                Registros=("Monto", "size"),
                                Pagado=("Monto_pagado_obs", "sum"),
                                **{"No pagado": ("Monto_no_pagado_obs", "sum")},
                                Abono=("Monto_abono_obs", "sum"),
                            )
                            .sort_values("Periodo_chart")
                        )
                        chart_periodo["RESULTADO"] = chart_periodo["Pagado"] + chart_periodo["No pagado"]
                        chart_periodo["Deuda a la fecha"] = np.where(
                            (chart_periodo["Pagado"] < 0) & (chart_periodo["No pagado"] < 0),
                            chart_periodo["No pagado"] + chart_periodo["Abono"],
                            np.where(
                                chart_periodo["Pagado"] < 0,
                                chart_periodo["Pagado"] + chart_periodo["Abono"],
                                chart_periodo["No pagado"] - chart_periodo["Abono"],
                            ),
                        )
                        chart_periodo["Periodo_txt"] = chart_periodo["Periodo_chart"].dt.strftime("%Y-%m")
                        chart_periodo["Hover"] = (
                            "Registros: " + chart_periodo["Registros"].astype(int).astype(str)
                            + "<br>Pagado: $" + chart_periodo["Pagado"].map(lambda v: f"{v:,.0f}")
                            + "<br>No pagado: $" + chart_periodo["No pagado"].map(lambda v: f"{v:,.0f}")
                            + "<br>Abono: $" + chart_periodo["Abono"].map(lambda v: f"{v:,.0f}")
                            + "<br>Deuda: $" + chart_periodo["Deuda a la fecha"].map(lambda v: f"{v:,.0f}")
                        )

                        st.markdown(
                            section_heading(
                                "📊",
                                "Evolución por año y mes según OBS",
                                "Montos calculados con los mismos selectores del resumen",
                                weight_class="section-heading-title-soft",
                            ),
                            unsafe_allow_html=True,
                        )

                        grafico_obs_export = chart_periodo.copy()
                        fig_obs_periodo = build_obs_periodo_figure(chart_periodo)
                        st.plotly_chart(fig_obs_periodo, use_container_width=True)

                        temporal_insights = []
                        temporal_base = chart_periodo.copy().sort_values("Periodo_chart")
                        temporal_base["Deuda_abs"] = temporal_base["Deuda a la fecha"].abs()
                        temporal_base["Mora_abs"] = temporal_base["No pagado"].abs()

                        if not temporal_base.empty and temporal_base["Deuda_abs"].max() > 0:
                            peak_row = temporal_base.loc[temporal_base["Deuda_abs"].idxmax()]
                            temporal_insights.append(
                                (
                                    "Crítico" if float(peak_row["Deuda_abs"]) >= float(temporal_base["Deuda_abs"].quantile(0.75)) else "Presión",
                                    f"Peak histórico de deuda en {pd.to_datetime(peak_row['Periodo_chart']).strftime('%b %Y')}: {fmt_clp_largo(float(peak_row['Deuda_abs']))}.",
                                )
                            )

                        if len(temporal_base) >= 2:
                            last_row_tmp = temporal_base.iloc[-1]
                            prev_row_tmp = temporal_base.iloc[-2]
                            mora_delta = float(last_row_tmp["Mora_abs"] - prev_row_tmp["Mora_abs"])
                            deuda_delta = float(last_row_tmp["Deuda_abs"] - prev_row_tmp["Deuda_abs"])
                            if mora_delta > 0:
                                temporal_insights.append(
                                    (
                                        "Presión" if mora_delta < max(float(prev_row_tmp["Mora_abs"]) * 0.35, 1) else "Crítico",
                                        f"Mora activa aumenta {fmt_clp_largo(mora_delta)} vs período anterior; revisar recuperación inmediata.",
                                    )
                                )
                            elif mora_delta < 0:
                                temporal_insights.append(
                                    (
                                        "Saludable",
                                        f"Mejora operacional: la mora baja {fmt_clp_largo(abs(mora_delta))} frente al período anterior.",
                                    )
                                )

                            if deuda_delta < 0:
                                temporal_insights.append(
                                    (
                                        "Saludable",
                                        f"La deuda a la fecha disminuye {fmt_clp_largo(abs(deuda_delta))}; tendencia reciente favorable.",
                                    )
                                )

                        if len(temporal_base) >= 3:
                            recent_debt = temporal_base["Deuda_abs"].tail(3)
                            avg_recent_debt = float(recent_debt.mean())
                            max_recent_debt = float(recent_debt.max())
                            min_recent_debt = float(recent_debt.min())
                            if avg_recent_debt > 0 and (max_recent_debt - min_recent_debt) <= max(avg_recent_debt * 0.10, 1):
                                temporal_insights.append(
                                    (
                                        "Atención",
                                        f"Estabilización detectada: deuda se mantiene en torno a {fmt_clp_largo(avg_recent_debt)} durante los últimos 3 períodos.",
                                    )
                                )
                            if (temporal_base["Deuda_abs"].tail(3) > 0).all():
                                temporal_insights.append(
                                    (
                                        "Presión",
                                        "Presión persistente: la deuda permanece activa en los últimos 3 períodos observados.",
                                    )
                                )

                        if not temporal_insights:
                            temporal_insights.append(
                                (
                                    "Saludable",
                                    "No se detectan presiones temporales relevantes en la evolución del período filtrado.",
                                )
                            )

                        temporal_seen = set()
                        temporal_unique = []
                        for badge, copy in temporal_insights:
                            if copy in temporal_seen:
                                continue
                            temporal_seen.add(copy)
                            temporal_unique.append((badge, copy))

                        temporal_insights_html = "".join(
                            f"""
                            <div class="detalle-alert">
                                <div class="detalle-alert-badge {detalle_alert_badge_class(badge)}">{badge}</div>
                                <div class="detalle-alert-copy">{copy}</div>
                            </div>
                            """
                            for badge, copy in temporal_unique[:4]
                        )
                        st.markdown(
                            f"""
                            <div class="detalle-alerts-card">
                                <div class="detalle-alerts-head">
                                    <div>
                                        <div class="detalle-alerts-title">Insights operacionales detectados</div>
                                        <div class="detalle-alerts-sub">Lectura temporal automática sobre deuda, mora y presión operacional</div>
                                    </div>
                                </div>
                                <div class="detalle-alerts-grid">{temporal_insights_html}</div>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )
            if "Fecha" in df_det_f.columns:
                df_det_f["Fecha"] = pd.to_datetime(df_det_f["Fecha"], errors="coerce")
                df_det_f = df_det_f.sort_values(["Fecha", "Año", "Mes"], ascending=[False, False, False], na_position="last")
            else:
                df_det_f = df_det_f.sort_values(["Año", "Mes"], ascending=[False, False], na_position="last")

            cols_det = [c for c in ["Fecha", "Año", "Mes", "Esp", "Responsable", "CC", "CC1", "Obs", "Sit", "Monto"] if c in df_det_f.columns]
            df_det_view = df_det_f[cols_det].copy().reset_index(drop=True)

            st.caption(f"Universo operacional bajo análisis: {len(df_det_view):,} movimientos.".replace(",", "."))
            if df_det_view.empty:
                st.info("No se registran movimientos relevantes para el universo filtrado.")
            else:
                df_det_show = df_det_view.copy()
                if "Fecha" in df_det_show.columns:
                    df_det_show["Fecha"] = (
                        pd.to_datetime(df_det_show["Fecha"], errors="coerce")
                        .dt.strftime("%Y-%m-%d")
                        .fillna("Sin fecha")
                    )

                filtros_det_export = {
                    "Búsqueda": busqueda_det if busqueda_det else "Sin búsqueda",
                    "CC1": sel_cc1_det,
                    "OBS": sel_obs_det,
                    "Responsable": sel_resp_det,
                    "Año": sel_anio_det,
                    "Mes": sel_mes_det,
                    "Registros": str(len(df_det_view)),
                }
                kpis_det_export = {
                    "Total tabla (filtro)": monto_total_det,
                    "Monto por pagar": monto_por_pagar_det,
                    "Monto pagado": monto_pagado_det,
                    "Abonos": monto_abonos_det,
                }
                pdf_detalle = build_detalle_movimientos_pdf(
                    df_det_view,
                    filtros_det_export,
                    kpis_det_export,
                    resumen_obs_export,
                    grafico_obs_export,
                )
                excel_detalle = build_detalle_movimientos_excel(
                    df_det_view,
                    filtros_det_export,
                    kpis_det_export,
                    resumen_obs_export,
                    grafico_obs_export,
                )

                with st.container():
                    st.markdown('<div class="detalle-export-toolbar-anchor"></div>', unsafe_allow_html=True)
                    exp_title, exp_pdf, exp_excel = st.columns([4.6, 0.92, 1.02], gap="small")
                    with exp_title:
                        st.markdown(
                            f"""
                            <div>
                                <div class="detalle-export-title">Tabla operacional</div>
                                <div class="detalle-export-sub">{len(df_det_view):,} movimientos bajo análisis · visualización operativa con navegación interna</div>
                            </div>
                            """.replace(",", "."),
                            unsafe_allow_html=True,
                        )
                    with exp_pdf:
                        st.download_button(
                            "PDF",
                            data=pdf_detalle,
                            file_name="detalle_movimientos_filtrado.pdf",
                            mime="application/pdf",
                            key="download_detalle_movimientos_pdf",
                            use_container_width=True,
                        )
                    with exp_excel:
                        st.download_button(
                            "Excel",
                            data=excel_detalle,
                            file_name="detalle_movimientos_filtrado.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_detalle_movimientos_excel",
                            use_container_width=True,
                        )

                visible_rows = 12
                row_height_px = 28
                header_px = 32
                table_height = header_px + (visible_rows * row_height_px)
                left_cols_det = [c for c in ["Responsable", "CC1", "Obs"] if c in df_det_show.columns]
                responsable_cols_det = [c for c in ["Responsable"] if c in df_det_show.columns]
                sit_cols_det = [c for c in ["Sit"] if c in df_det_show.columns]
                cc_cols_det = [c for c in ["CC"] if c in df_det_show.columns]
                monto_cols_det = [c for c in ["Monto"] if c in df_det_show.columns]
                critical_cols_det = [c for c in ["Responsable", "Sit", "Monto"] if c in df_det_show.columns]
                secondary_cols_det = [c for c in df_det_show.columns if c not in set(critical_cols_det)]

                if "Fecha" in df_det_show.columns:
                    group_keys_det = df_det_show["Fecha"].astype(str).str.slice(0, 7)
                elif {"Año", "Mes"}.issubset(df_det_show.columns):
                    group_keys_det = (
                        df_det_show["Año"].astype(str).str.strip()
                        + "-"
                        + df_det_show["Mes"].astype(str).str.strip().str.zfill(2)
                    )
                else:
                    group_keys_det = pd.Series("", index=df_det_show.index)
                group_break_rows_det = set(group_keys_det[group_keys_det.ne(group_keys_det.shift())].index.tolist())

                def _enterprise_row_style(row):
                    styles = [""] * len(row)
                    if row.name == 0:
                        styles = ["background-color:#F8FBFF; font-weight:720;"] * len(row)
                    elif row.name in group_break_rows_det:
                        styles = ["border-top:1px solid rgba(148,163,184,0.20);"] * len(row)
                    return styles

                def _style_sit(v):
                    s = str(v).strip().upper()
                    if s == "NO PAGADO":
                        return "background-color:#FFF1F1; color:#B42318; font-weight:900; border-radius:6px;"
                    if s == "PAGADO":
                        return "background-color:#F0FDF4; color:#027A48; font-weight:900; border-radius:6px;"
                    return ""

                def _style_cc(v):
                    s = str(v).strip().upper()
                    if s == "EGRESO":
                        return "color:#B42318; font-weight:700;"
                    if s == "INGRESO":
                        return "color:#027A48; font-weight:700;"
                    return ""

                def _style_monto(v):
                    try:
                        n = float(v)
                        if n < 0:
                            return "background-color:#FFF1F1; color:#B42318; font-weight:950; font-variant-numeric:tabular-nums; letter-spacing:-0.01em;"
                        if n > 0:
                            return "background-color:#F0FDF4; color:#027A48; font-weight:950; font-variant-numeric:tabular-nums; letter-spacing:-0.01em;"
                        return "background-color:#F8FAFC; color:#344054; font-weight:850; font-variant-numeric:tabular-nums;"
                    except Exception:
                        pass
                    return ""

                styler_det = (
                    df_det_show.style
                    .format({"Monto": "${:,.0f}"})
                    .set_table_styles([
                        {
                            "selector": "thead th",
                            "props": [
                                ("position", "sticky"),
                                ("top", "0"),
                                ("z-index", "2"),
                                ("background", "linear-gradient(180deg,#F8FAFC 0%,#EEF3F8 100%)"),
                                ("color", "#475569"),
                                ("font-weight", "780"),
                                ("font-size", "10.8px"),
                                ("border-bottom", "1px solid rgba(148,163,184,0.24)"),
                                ("text-align", "center"),
                                ("padding", "5px 7px"),
                                ("letter-spacing", ".02em"),
                            ],
                        },
                        {
                            "selector": "tbody td",
                            "props": [
                                ("font-size", "11px"),
                                ("padding", "4px 7px"),
                                ("border-bottom", "1px solid rgba(226,232,240,0.36)"),
                                ("vertical-align", "middle"),
                            ],
                        },
                        {
                            "selector": "tbody tr:nth-child(even)",
                            "props": [("background-color", "#FCFDFF")],
                        },
                        {
                            "selector": "tbody tr:nth-child(odd)",
                            "props": [("background-color", "#FFFFFF")],
                        },
                        {
                            "selector": "tbody tr:hover",
                            "props": [
                                ("background-color", "#F5F9FF"),
                                ("box-shadow", "inset 2px 0 0 rgba(37,99,235,0.55)"),
                            ],
                        },
                    ])
                    .set_properties(subset=left_cols_det, **{"text-align": "left"})
                    .set_properties(
                        subset=monto_cols_det,
                        **{
                            "font-weight": "950",
                            "text-align": "right",
                            "font-variant-numeric": "tabular-nums",
                            "font-family": "ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace",
                        },
                    )
                    .apply(_enterprise_row_style, axis=1)
                )
                if secondary_cols_det:
                    styler_det = styler_det.set_properties(
                        subset=secondary_cols_det,
                        **{"color": "#64748B", "font-weight": "560"},
                    )
                if responsable_cols_det:
                    styler_det = styler_det.set_properties(
                        subset=responsable_cols_det,
                        **{"color": "#0F172A", "font-weight": "850"},
                    )
                if sit_cols_det:
                    styler_det = styler_det.map(_style_sit, subset=sit_cols_det)
                if cc_cols_det:
                    styler_det = styler_det.map(_style_cc, subset=cc_cols_det)
                if monto_cols_det:
                    styler_det = styler_det.map(_style_monto, subset=monto_cols_det)
                st.caption("Detalle transaccional secundario: foco visual en responsable, estado operativo y monto.")
                st.dataframe(
                    styler_det,
                    use_container_width=True,
                    height=table_height,
                )

        render_detalle_filtrable_movimientos()

    render_flujo_operacional()

# =========================================================
# 🏗️ TAB 6: CAPEX
# =========================================================
if active_section == "🏗️ Capex":
    import plotly.graph_objects as go

    capex_view = capex_df.copy()
    if capex_view.empty:
        st.warning("La fuente Capex no tiene datos válidos para mostrar.")
        st.stop()

    def fmt_m(v: float) -> str:
        return f"${float(v) / 1_000_000:.1f}M"

    total_capex_view = float(capex_view["Monto"].sum())
    costo_total = float(capex_view.loc[capex_view["Estado_norm"].eq("COSTO"), "Monto"].sum())
    gasto_total = float(capex_view.loc[capex_view["Estado_norm"].eq("GASTO"), "Monto"].sum())
    cost_share = costo_total / total_capex_view if total_capex_view else 0.0
    gasto_share = gasto_total / total_capex_view if total_capex_view else 0.0

    top_cat = capex_view.groupby("CCCC", as_index=False)["Monto"].sum().sort_values("Monto", ascending=False)
    top_cat_name = str(top_cat.iloc[0]["CCCC"]) if not top_cat.empty else "Sin datos"
    top_cat_value = float(top_cat.iloc[0]["Monto"]) if not top_cat.empty else 0.0
    top_share = top_cat_value / total_capex_view if total_capex_view else 0.0
    top3_sum = float(top_cat.head(3)["Monto"].sum()) if not top_cat.empty else 0.0
    top5_sum = float(top_cat.head(5)["Monto"].sum()) if not top_cat.empty else 0.0

    capex_export = capex_view[
        ["Año", "Periodo", "Situación", "CCCC", "Estado", "Monto", "Periodo_ref"]
    ].sort_values(["Periodo_ref", "CCCC", "Situación"]).copy()
    capex_export["Periodo_ref"] = capex_export["Periodo_ref"].dt.strftime("%Y-%m")
    csv_b64 = base64.b64encode(capex_export.to_csv(index=False).encode("utf-8")).decode()

    monthly = (
        capex_view.groupby(["Periodo_ref", "Estado"], as_index=False)["Monto"].sum()
        .sort_values("Periodo_ref")
    )
    monthly_pivot = (
        monthly.pivot_table(index="Periodo_ref", columns="Estado", values="Monto", aggfunc="sum", fill_value=0)
        .sort_index()
    )
    for col in ["Costo", "Gasto"]:
        if col not in monthly_pivot.columns:
            monthly_pivot[col] = 0
    monthly_pivot["Total"] = monthly_pivot[["Costo", "Gasto"]].sum(axis=1)
    monthly_pivot["Acumulado"] = monthly_pivot["Total"].cumsum()
    monthly_plot = monthly_pivot.reset_index()
    first_month_total = float(monthly_plot["Total"].iloc[0]) if not monthly_plot.empty else 0.0
    last_month_total = float(monthly_plot["Total"].iloc[-1]) if not monthly_plot.empty else 0.0
    period_delta = (last_month_total / first_month_total - 1) if first_month_total else 0.0

    fig_month = go.Figure()
    fig_month.add_trace(go.Bar(
        x=monthly_plot["Periodo_ref"],
        y=monthly_plot["Costo"],
        name="Costo (Capitalizable)",
        marker_color="#20C978",
        hovertemplate="<b>%{x|%b %Y}</b><br>Costo: $%{y:,.0f}<extra></extra>",
    ))
    fig_month.add_trace(go.Bar(
        x=monthly_plot["Periodo_ref"],
        y=monthly_plot["Gasto"],
        name="Gasto (Operacional)",
        marker_color="#FF343E",
        hovertemplate="<b>%{x|%b %Y}</b><br>Gasto: $%{y:,.0f}<extra></extra>",
    ))
    fig_month.add_trace(go.Scatter(
        x=monthly_plot["Periodo_ref"],
        y=monthly_plot["Acumulado"],
        yaxis="y2",
        mode="lines+markers",
        name="Total CAPEX (Tendencia)",
        line=dict(color="#001A5A", width=2.5),
        marker=dict(size=7, color="#001A5A", line=dict(color="#FFFFFF", width=1.2)),
        hovertemplate="<b>%{x|%b %Y}</b><br>Acumulado: $%{y:,.0f}<extra></extra>",
    ))
    fig_month.update_layout(
        height=215,
        barmode="stack",
        margin=dict(l=8, r=8, t=4, b=18),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#FFFFFF",
        legend=dict(orientation="h", y=1.18, x=0, font=dict(size=10, color="#14244A")),
        xaxis=dict(title="", tickformat="%b %Y", showgrid=False, color="#22345F", tickfont=dict(size=10)),
        yaxis=dict(title="", tickprefix="$", separatethousands=True, gridcolor="#E5EAF2", color="#22345F", tickfont=dict(size=10)),
        yaxis2=dict(
            title="",
            overlaying="y",
            side="right",
            tickprefix="$",
            separatethousands=True,
            gridcolor="rgba(0,0,0,0)",
            color="#22345F",
            tickfont=dict(size=10),
        ),
    )

    donut_values = [top3_sum, max(top5_sum - top3_sum, 0), max(total_capex_view - top5_sum, 0)]
    fig_donut = go.Figure(data=[go.Pie(
        values=donut_values,
        labels=["Top 3 categorías", "Top 5 categorías", "Resto de categorías"],
        hole=0.62,
        marker=dict(colors=["#20C978", "#001A5A", "#9B5CF6"], line=dict(color="#FFFFFF", width=2)),
        textinfo="none",
        hovertemplate="%{label}<br>$%{value:,.0f}<extra></extra>",
    )])
    fig_donut.update_layout(
        height=128,
        showlegend=False,
        margin=dict(l=0, r=0, t=0, b=0),
        paper_bgcolor="rgba(0,0,0,0)",
        annotations=[dict(
            text=f"<b>{top3_sum / total_capex_view:.0%}</b><br><span style='font-size:9px'>Top 3</span>",
            x=0.5,
            y=0.5,
            font=dict(size=18, color="#001133"),
            showarrow=False,
        )],
    )

    top_rows = []
    top_max = float(top_cat["Monto"].max()) if not top_cat.empty else 1.0
    for i, row in enumerate(top_cat.head(10).itertuples(index=False), start=1):
        width = (float(row.Monto) / top_max) * 100 if top_max else 0
        share = float(row.Monto) / total_capex_view if total_capex_view else 0
        top_rows.append(
            f"""
            <div class="capex-rank-row">
                <div class="capex-rank-num">{i}</div>
                <div class="capex-rank-name">{escape(str(row.CCCC))}</div>
                <div class="capex-rank-track"><div style="width:{width:.1f}%"></div></div>
                <div class="capex-rank-value">{fmt_m(float(row.Monto))} ({share:.1%})</div>
            </div>
            """
        )

    heat = capex_view.groupby(["Año", "Mes_num"], as_index=False)["Monto"].sum().dropna(subset=["Año", "Mes_num"])
    heat_lookup = {(int(r["Año"]), int(r["Mes_num"])): float(r["Monto"]) for _, r in heat.iterrows()}
    heat_max = max(heat_lookup.values()) if heat_lookup else 1.0
    month_labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    heat_rows = []
    for year in sorted(capex_view["Año"].dropna().astype(int).unique()):
        cells = []
        for month in range(1, 13):
            val = heat_lookup.get((year, month), 0.0)
            alpha = 0.08 + 0.82 * (val / heat_max if heat_max else 0)
            cells.append(f'<div class="capex-heat-cell" title="{year}-{month:02d}: {fmt_clp_largo(val)}" style="background:rgba(32,201,120,{alpha:.2f});"></div>')
        heat_rows.append(f'<div class="capex-heat-year">{year}</div>{"".join(cells)}')
    heat_months = "".join([f"<div>{m}</div>" for m in month_labels])
    last_update = pd.Timestamp.now().strftime("%d de %B de %Y, %H:%M")

    st.markdown(
        """
        <style>
        .capex-page {
            color:#001133;
            font-family: inherit;
            padding:0 2px 8px 2px;
        }
        .capex-hero {
            display:flex;
            align-items:flex-start;
            justify-content:space-between;
            gap:12px;
            margin:-68px 0 8px 0;
        }
        .capex-title {
            font-size:28px;
            line-height:1;
            font-weight:950;
            color:#001133;
            letter-spacing:-0.025em;
        }
        .capex-subtitle {
            color:#65719A;
            font-size:12px;
            font-weight:650;
            margin-top:6px;
        }
        .capex-actions {
            display:flex;
            gap:7px;
            align-items:center;
            flex-wrap:wrap;
            justify-content:flex-end;
        }
        .capex-action {
            height:34px;
            min-width:34px;
            display:flex;
            align-items:center;
            justify-content:center;
            border:1px solid #E1E7F0;
            border-radius:8px;
            background:#FFFFFF;
            color:#001A5A;
            box-shadow:0 8px 22px rgba(15,23,42,0.04);
            font-size:13px;
            font-weight:900;
            text-decoration:none;
        }
        .capex-action-share {
            padding:0 12px;
            gap:6px;
            font-size:12px;
        }
        .capex-action-primary {
            min-width:142px;
            padding:0 12px;
            gap:6px;
            border-color:#0A55F7;
            background:#0B5AF4;
            color:#FFFFFF !important;
            font-size:12px;
            box-shadow:0 14px 28px rgba(11,90,244,0.22);
        }
        .capex-kpi-grid {
            display:grid;
            grid-template-columns:repeat(4, minmax(0, 1fr));
            gap:8px;
            margin-bottom:8px;
        }
        .capex-card {
            border:1px solid #E1E7F0;
            border-radius:9px;
            background:#FFFFFF;
            box-shadow:0 8px 18px rgba(15,23,42,0.04);
        }
        .capex-kpi {
            min-height:100px;
            display:grid;
            grid-template-columns:40px minmax(0,1fr);
            gap:10px;
            padding:14px 15px;
            box-sizing:border-box;
        }
        .capex-kpi-icon {
            width:38px;
            height:38px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:var(--halo);
            color:var(--accent);
            font-size:19px;
            font-weight:950;
        }
        .capex-kpi-label {
            color:#001133;
            font-size:10.5px;
            font-weight:950;
            letter-spacing:.02em;
            text-transform:uppercase;
            margin-bottom:7px;
        }
        .capex-kpi-value {
            color:var(--accent);
            font-size:25px;
            line-height:1;
            font-weight:950;
            letter-spacing:-0.035em;
            white-space:normal;
            overflow-wrap:anywhere;
        }
        .capex-kpi-copy {
            color:#334675;
            font-size:11px;
            font-weight:650;
            margin-top:8px;
        }
        .capex-pill {
            display:inline-flex;
            align-items:center;
            gap:5px;
            height:22px;
            border-radius:999px;
            padding:0 9px;
            margin-top:8px;
            background:#EEF2F7;
            color:#001133;
            font-size:10px;
            font-weight:800;
        }
        .capex-progress {
            height:8px;
            width:100%;
            border-radius:999px;
            background:#E2E7EF;
            overflow:hidden;
            margin-top:13px;
        }
        .capex-progress > div {
            height:100%;
            border-radius:999px;
            background:var(--accent);
            box-shadow:inset 0 0 0 2px rgba(0,0,0,0.04);
        }
        .capex-insights {
            padding:9px 12px 11px 12px;
            margin-bottom:8px;
        }
        .capex-section-title {
            color:#001133;
            font-size:14px;
            font-weight:950;
            margin-bottom:8px;
        }
        .capex-insight-grid {
            display:grid;
            grid-template-columns:repeat(3, minmax(0, 1fr));
            gap:9px;
        }
        .capex-insight {
            min-height:74px;
            display:grid;
            grid-template-columns:38px minmax(0,1fr);
            gap:10px;
            padding:12px;
            border:1px solid var(--border);
            border-radius:8px;
            background:linear-gradient(135deg, #FFFFFF 0%, var(--soft) 100%);
        }
        .capex-insight-icon {
            width:36px;
            height:36px;
            border-radius:999px;
            background:var(--halo);
            color:var(--accent);
            display:flex;
            align-items:center;
            justify-content:center;
            font-size:16px;
            font-weight:950;
        }
        .capex-insight-title {
            color:var(--accent);
            font-size:13px;
            font-weight:950;
            margin-bottom:5px;
        }
        .capex-insight-copy {
            color:#061948;
            font-size:11px;
            line-height:1.32;
            font-weight:620;
        }
        .capex-chart-frame {
            border:1px solid #E1E7F0;
            border-radius:9px;
            background:#FFFFFF;
            box-shadow:0 8px 18px rgba(15,23,42,0.04);
            padding:10px 13px 5px 13px;
            min-height:276px;
            margin-bottom:8px;
        }
        .capex-card-head {
            display:flex;
            justify-content:space-between;
            gap:10px;
            align-items:center;
            color:#001133;
            font-size:14px;
            font-weight:950;
            margin-bottom:5px;
        }
        .capex-dots {
            color:#001A5A;
            letter-spacing:3px;
            font-size:14px;
            font-weight:950;
        }
        .capex-rank-list {
            display:flex;
            flex-direction:column;
            gap:4px;
            padding:8px 4px 2px 4px;
        }
        .capex-rank-row {
            display:grid;
            grid-template-columns:20px 148px minmax(90px,1fr) 105px;
            gap:8px;
            align-items:center;
            color:#061948;
            font-size:10.5px;
            font-weight:750;
        }
        .capex-rank-num {
            width:18px;
            height:18px;
            border-radius:999px;
            background:#EFF5F7;
            border:1px solid #D8E5EC;
            display:flex;
            align-items:center;
            justify-content:center;
            color:#52627D;
            font-size:10px;
            font-weight:900;
        }
        .capex-rank-track {
            height:12px;
            border-radius:4px;
            background:transparent;
        }
        .capex-rank-track div {
            height:100%;
            border-radius:4px;
            background:#20C978;
            box-shadow:inset 0 0 0 2px rgba(0,0,0,0.04);
        }
        .capex-rank-value {
            color:#24375F;
            white-space:nowrap;
            font-weight:800;
        }
        .capex-small-card {
            min-height:192px;
            padding:10px 13px;
            box-sizing:border-box;
        }
        .capex-donut-grid {
            display:grid;
            grid-template-columns:125px 1fr;
            gap:5px;
            align-items:center;
        }
        .capex-legend {
            display:flex;
            flex-direction:column;
            gap:8px;
            color:#24375F;
            font-size:10.5px;
            font-weight:750;
        }
        .capex-legend-row {
            display:grid;
            grid-template-columns:10px 1fr auto;
            gap:6px;
            align-items:start;
        }
        .capex-dot {
            width:8px;
            height:8px;
            border-radius:999px;
            margin-top:4px;
            background:var(--dot);
        }
        .capex-legend-val {
            color:#001133;
            font-size:12px;
            font-weight:950;
            text-align:right;
        }
        .capex-legend-val span {
            display:block;
            color:#65719A;
            font-size:9px;
            font-weight:750;
            margin-top:1px;
        }
        .capex-callout {
            color:#0B5AF4;
            font-size:10px;
            font-weight:850;
            margin-top:4px;
        }
        .capex-heat {
            margin-top:13px;
        }
        .capex-heat-grid {
            display:grid;
            grid-template-columns:38px repeat(12, 1fr);
            gap:2px;
            align-items:center;
        }
        .capex-heat-year {
            color:#334675;
            font-size:10px;
            font-weight:800;
            text-align:right;
            padding-right:6px;
        }
        .capex-heat-cell {
            height:18px;
            border-radius:1px;
            border:1px solid rgba(255,255,255,0.88);
        }
        .capex-heat-months {
            display:grid;
            grid-template-columns:38px repeat(12, 1fr);
            gap:2px;
            margin-top:5px;
            color:#334675;
            font-size:9px;
            font-weight:750;
            text-align:center;
        }
        .capex-heat-scale {
            display:flex;
            align-items:center;
            justify-content:center;
            gap:8px;
            margin-top:14px;
            color:#334675;
            font-size:9.5px;
            font-weight:750;
        }
        .capex-gradient {
            width:145px;
            height:10px;
            border-radius:3px;
            background:linear-gradient(90deg, rgba(32,201,120,.08), rgba(32,201,120,.22), rgba(32,201,120,.45), rgba(32,201,120,.7), rgba(32,201,120,.95));
        }
        .capex-summary-grid {
            display:grid;
            grid-template-columns:repeat(3, 1fr);
            gap:0;
            margin-top:12px;
        }
        .capex-summary-item {
            padding:0 10px;
            border-right:1px solid #E2E7EF;
        }
        .capex-summary-item:last-child {
            border-right:0;
        }
        .capex-summary-icon {
            width:30px;
            height:30px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            color:var(--accent);
            background:var(--halo);
            font-weight:950;
            font-size:14px;
            margin-bottom:9px;
        }
        .capex-summary-label {
            color:#001133;
            font-size:10.5px;
            font-weight:950;
            margin-bottom:7px;
        }
        .capex-summary-value {
            color:var(--accent);
            font-size:20px;
            line-height:1;
            font-weight:950;
            letter-spacing:-0.025em;
        }
        .capex-summary-sub {
            color:#65719A;
            font-size:10px;
            font-weight:650;
            margin-top:7px;
        }
        .capex-info {
            margin-top:13px;
            border:1px solid #CFE0FF;
            background:#F2F7FF;
            color:#0B5AF4;
            border-radius:6px;
            padding:8px 10px;
            font-size:10.5px;
            font-weight:850;
        }
        .capex-footer {
            display:flex;
            justify-content:center;
            align-items:center;
            gap:10px;
            color:#65719A;
            font-size:10.5px;
            font-weight:750;
            margin:8px 0 0 0;
        }
        @media (max-width: 1320px) {
            .capex-kpi-grid, .capex-insight-grid { grid-template-columns:1fr; }
            .capex-hero { margin-top:-42px; flex-direction:column; }
            .capex-rank-row { grid-template-columns:26px 150px minmax(120px,1fr) 116px; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        <div class="capex-page">
            <div class="capex-hero">
                <div>
                    <div class="capex-title">CAPEX Intelligence</div>
                    <div class="capex-subtitle">Análisis y control de inversiones en infraestructura</div>
                </div>
                <div class="capex-actions">
                    <div class="capex-action capex-action-share">⌯ Compartir</div>
                    <div class="capex-action">☆</div>
                    <div class="capex-action">...</div>
                    <a class="capex-action capex-action-primary" href="data:text/csv;base64,{csv_b64}" download="capex_filtrado_bodegas2025.csv">⇩ Descargar reporte</a>
                </div>
            </div>
            <div class="capex-kpi-grid">
                <div class="capex-card capex-kpi" style="--accent:#0B5AF4;--halo:#EAF2FF;">
                    <div class="capex-kpi-icon">$</div>
                    <div>
                        <div class="capex-kpi-label">CAPEX TOTAL FILTRADO</div>
                        <div class="capex-kpi-value">{fmt_m(total_capex_view)}</div>
                        <div class="capex-kpi-copy">CAPEX consolidado filtrado</div>
                        <div class="capex-pill" style="color:{'#06A861' if period_delta >= 0 else '#DC2626'};">↗ {period_delta:+.1%} vs período anterior</div>
                    </div>
                </div>
                <div class="capex-card capex-kpi" style="--accent:#14B86A;--halo:#E9F8F0;">
                    <div class="capex-kpi-icon">♣</div>
                    <div>
                        <div class="capex-kpi-label">COSTO CAPITALIZABLE</div>
                        <div class="capex-kpi-value">{cost_share:.1%}</div>
                        <div class="capex-kpi-copy">Activo depreciable</div>
                        <div class="capex-progress"><div style="width:{min(cost_share * 100, 100):.1f}%"></div></div>
                    </div>
                </div>
                <div class="capex-card capex-kpi" style="--accent:#F01822;--halo:#FFF0F1;">
                    <div class="capex-kpi-icon">●</div>
                    <div>
                        <div class="capex-kpi-label">GASTO OPERACIONAL</div>
                        <div class="capex-kpi-value">{gasto_share:.1%}</div>
                        <div class="capex-kpi-copy">Impacto inmediato en resultado</div>
                        <div class="capex-progress"><div style="width:{min(gasto_share * 100, 100):.1f}%"></div></div>
                    </div>
                </div>
                <div class="capex-card capex-kpi" style="--accent:#7C2CF4;--halo:#F2E9FF;">
                    <div class="capex-kpi-icon">↑</div>
                    <div>
                        <div class="capex-kpi-label">PRINCIPAL DRIVER</div>
                        <div class="capex-kpi-value" style="font-size:22px;line-height:1.05;">{escape(top_cat_name)}</div>
                        <div class="capex-kpi-copy">{top_share:.1%} del CAPEX total</div>
                        <div class="capex-progress"><div style="width:{min(top_share * 100, 100):.1f}%"></div></div>
                    </div>
                </div>
            </div>
            <div class="capex-card capex-insights">
                <div class="capex-section-title">Insights ejecutivos</div>
                <div class="capex-insight-grid">
                    <div class="capex-insight" style="--accent:#0B5AF4;--halo:#EAF2FF;--soft:#FAFCFF;--border:#DCEAFF;">
                        <div class="capex-insight-icon">◌</div>
                        <div><div class="capex-insight-title">Lectura técnica</div><div class="capex-insight-copy">Las 3 categorías principales concentran {top3_sum / total_capex_view:.1%} del CAPEX, evidenciando una estructura focalizada en activos estructurales.</div></div>
                    </div>
                    <div class="capex-insight" style="--accent:#14B86A;--halo:#E9F8F0;--soft:#FBFFFD;--border:#DCEFE7;">
                        <div class="capex-insight-icon">▤</div>
                        <div><div class="capex-insight-title">Clasificación contable</div><div class="capex-insight-copy">La separación entre costo y gasto permite distinguir partidas capitalizables de impacto operacional inmediato y optimizar la capitalización.</div></div>
                    </div>
                    <div class="capex-insight" style="--accent:#7C2CF4;--halo:#F2E9FF;--soft:#FEFBFF;--border:#E9DDFC;">
                        <div class="capex-insight-icon">◇</div>
                        <div><div class="capex-insight-title">Calidad de datos</div><div class="capex-insight-copy">El dataset presenta validación automática de períodos, montos y normalización de registros para garantizar integridad y trazabilidad.</div></div>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns([0.95, 1.05])
    with c1:
        with st.container(border=True):
            st.markdown(
                '<div class="capex-card-head"><span>Evolución mensual CAPEX por clasificación</span><span class="capex-dots">⋮</span></div>',
                unsafe_allow_html=True,
            )
            st.plotly_chart(fig_month, use_container_width=True, config={"displayModeBar": False})
    with c2:
        st.markdown(
            f"""
            <div class="capex-chart-frame">
                <div class="capex-card-head"><span>Top 15 categorías CCCC por inversión</span><span class="capex-dots">⋮</span></div>
                <div class="capex-rank-list">{''.join(top_rows)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    b1, b2, b3 = st.columns([0.9, 0.9, 1.05])
    with b1:
        with st.container(border=True):
            st.markdown(
                '<div class="capex-card-head"><span>Concentración de inversión</span><span class="capex-dots">⋮</span></div>',
                unsafe_allow_html=True,
            )
            donut_col, legend_col = st.columns([0.95, 1])
            with donut_col:
                st.plotly_chart(fig_donut, use_container_width=True, config={"displayModeBar": False})
            with legend_col:
                st.markdown(
                    f"""
                    <div class="capex-legend">
                        <div class="capex-legend-row"><div class="capex-dot" style="--dot:#20C978;"></div><div>Top 3 categorías</div><div class="capex-legend-val">{top3_sum / total_capex_view:.0%}<span>{fmt_m(top3_sum)}</span></div></div>
                        <div class="capex-legend-row"><div class="capex-dot" style="--dot:#001A5A;"></div><div>Top 5 categorías</div><div class="capex-legend-val">{top5_sum / total_capex_view:.0%}<span>{fmt_m(top5_sum)}</span></div></div>
                        <div class="capex-legend-row"><div class="capex-dot" style="--dot:#9B5CF6;"></div><div>Resto de categorías</div><div class="capex-legend-val">{(total_capex_view - top5_sum) / total_capex_view:.0%}<span>{fmt_m(total_capex_view - top5_sum)}</span></div></div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            st.markdown('<div class="capex-callout">ⓘ Alta concentración en categorías estratégicas.</div>', unsafe_allow_html=True)
    with b2:
        st.markdown(
            f"""
            <div class="capex-card capex-small-card">
                <div class="capex-card-head"><span>Timeline de inversión (intensidad mensual)</span><span class="capex-dots">⋮</span></div>
                <div class="capex-heat">
                    <div class="capex-heat-grid">{''.join(heat_rows)}</div>
                    <div class="capex-heat-months"><div></div>{heat_months}</div>
                    <div class="capex-heat-scale"><span>Baja inversión</span><div class="capex-gradient"></div><span>Alta inversión</span></div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with b3:
        st.markdown(
            f"""
            <div class="capex-card capex-small-card">
                <div class="capex-card-head"><span>Resumen del período</span></div>
                <div class="capex-summary-grid">
                    <div class="capex-summary-item" style="--accent:#001133;--halo:#E9F8F0;">
                        <div class="capex-summary-icon">↟</div>
                        <div class="capex-summary-label">Total filtrado</div>
                        <div class="capex-summary-value">{fmt_m(total_capex_view)}</div>
                        <div class="capex-summary-sub">100% del total</div>
                    </div>
                    <div class="capex-summary-item" style="--accent:#001133;--halo:#FFF7E8;">
                        <div class="capex-summary-icon" style="color:#F3A417;">◎</div>
                        <div class="capex-summary-label">Costo capitalizable</div>
                        <div class="capex-summary-value">{fmt_m(costo_total)}</div>
                        <div class="capex-summary-sub">{cost_share:.1%} del total</div>
                    </div>
                    <div class="capex-summary-item" style="--accent:#F01822;--halo:#FFF0F1;">
                        <div class="capex-summary-icon">♢</div>
                        <div class="capex-summary-label">Gasto operacional</div>
                        <div class="capex-summary-value">{fmt_m(gasto_total)}</div>
                        <div class="capex-summary-sub">{gasto_share:.1%} del total</div>
                    </div>
                </div>
                <div class="capex-info">ⓘ El costo capitalizable representa el {cost_share:.1%} del CAPEX total filtrado.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown(f'<div class="capex-footer">↻ Última actualización: {last_update}</div>', unsafe_allow_html=True)
    with st.expander("Ver detalle de movimientos CAPEX"):
        capex_table = capex_export.rename(columns={"Periodo_ref": "Periodo normalizado"})
        st.dataframe(capex_table.style.format({"Monto": "${:,.0f}"}), use_container_width=True, height=430)
    st.stop()

    st.markdown(
        tab_header("Capex", "Inversión histórica del activo desde Google Sheets", show_download=False),
        unsafe_allow_html=True,
    )

    capex_base = capex_df.copy()
    st.markdown(
        """
        <style>
        .capex-filter-note {
            margin:-2px 0 10px 0;
            color:#475569;
            font-size:12px;
            font-weight:650;
        }
        .capex-kpi-grid {
            display:grid;
            grid-template-columns:repeat(5, minmax(0, 1fr));
            gap:9px;
            margin:8px 0 12px 0;
        }
        .capex-kpi {
            min-height:92px;
            border:1px solid var(--border);
            border-radius:10px;
            background:linear-gradient(135deg, #ffffff 0%, var(--soft) 100%);
            padding:11px 12px;
            box-shadow:0 10px 24px rgba(15,23,42,0.04);
        }
        .capex-kpi-label {
            color:#475569;
            font-size:10px;
            font-weight:900;
            letter-spacing:.08em;
            text-transform:uppercase;
            margin-bottom:8px;
        }
        .capex-kpi-value {
            color:var(--accent);
            font-size:21px;
            line-height:1.05;
            font-weight:950;
            letter-spacing:-.025em;
            font-variant-numeric:tabular-nums;
        }
        .capex-kpi-sub {
            margin-top:6px;
            color:#64748b;
            font-size:10.5px;
            line-height:1.22;
            font-weight:720;
        }
        .capex-analysis-grid {
            display:grid;
            grid-template-columns:repeat(3, minmax(0, 1fr));
            gap:9px;
            margin:8px 0 13px 0;
        }
        .capex-analysis {
            border:1px solid #dbe3ee;
            border-radius:10px;
            background:#ffffff;
            padding:12px 13px;
            min-height:106px;
            box-shadow:0 10px 24px rgba(15,23,42,0.035);
        }
        .capex-analysis-title {
            color:#081735;
            font-size:12px;
            line-height:1.18;
            font-weight:950;
            margin-bottom:7px;
        }
        .capex-analysis-body {
            color:#475569;
            font-size:11px;
            line-height:1.32;
            font-weight:650;
        }
        @media (max-width: 1280px) {
            .capex-kpi-grid { grid-template-columns:repeat(2, minmax(0, 1fr)); }
            .capex-analysis-grid { grid-template-columns:1fr; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    if capex_base.empty:
        st.warning("La fuente Capex no tiene datos válidos para mostrar.")
        st.stop()

    years = sorted([int(y) for y in capex_base["Año"].dropna().unique()])
    estados = sorted([e for e in capex_base["Estado"].dropna().unique() if str(e).strip()])
    categorias = sorted([c for c in capex_base["CCCC"].dropna().unique() if str(c).strip()])

    f1, f2, f3 = st.columns([1, 1, 2])
    with f1:
        sel_years = st.multiselect("Año", years, default=years, key="capex_years")
    with f2:
        sel_estados = st.multiselect("Estado", estados, default=estados, key="capex_estados")
    with f3:
        sel_categorias = st.multiselect("Categoría CCCC", categorias, default=categorias, key="capex_categorias")

    capex_view = capex_base[
        capex_base["Año"].isin(sel_years)
        & capex_base["Estado"].isin(sel_estados)
        & capex_base["CCCC"].isin(sel_categorias)
    ].copy()

    st.markdown(
        f'<div class="capex-filter-note">Fuente viva: Google Sheets CSV · registros filtrados: {len(capex_view):,}</div>',
        unsafe_allow_html=True,
    )

    total_capex_view = float(capex_view["Monto"].sum()) if not capex_view.empty else 0.0
    costo_total = float(capex_view.loc[capex_view["Estado_norm"].eq("COSTO"), "Monto"].sum()) if not capex_view.empty else 0.0
    gasto_total = float(capex_view.loc[capex_view["Estado_norm"].eq("GASTO"), "Monto"].sum()) if not capex_view.empty else 0.0
    top_cat = (
        capex_view.groupby("CCCC", as_index=False)["Monto"].sum().sort_values("Monto", ascending=False)
        if not capex_view.empty else pd.DataFrame(columns=["CCCC", "Monto"])
    )
    top_cat_name = str(top_cat.iloc[0]["CCCC"]) if not top_cat.empty else "Sin datos"
    top_cat_value = float(top_cat.iloc[0]["Monto"]) if not top_cat.empty else 0.0
    capex_declared = 127_742_570
    capex_delta = total_capex_view - capex_declared

    st.markdown(
        f"""
        <div class="capex-kpi-grid">
            <div class="capex-kpi" style="--accent:#0B3A86;--soft:#f6f9ff;--border:#d4e1f6;">
                <div class="capex-kpi-label">CAPEX filtrado</div>
                <div class="capex-kpi-value">{fmt_clp_largo(total_capex_view)}</div>
                <div class="capex-kpi-sub">Suma normalizada de la columna Monto.</div>
            </div>
            <div class="capex-kpi" style="--accent:#B7791F;--soft:#fffaf0;--border:#eadfbd;">
                <div class="capex-kpi-label">Costo</div>
                <div class="capex-kpi-value">{fmt_clp_largo(costo_total)}</div>
                <div class="capex-kpi-sub">{(costo_total / total_capex_view if total_capex_view else 0):.1%} del total filtrado.</div>
            </div>
            <div class="capex-kpi" style="--accent:#DC2626;--soft:#fff7f7;--border:#f1caca;">
                <div class="capex-kpi-label">Gasto</div>
                <div class="capex-kpi-value">{fmt_clp_largo(gasto_total)}</div>
                <div class="capex-kpi-sub">{(gasto_total / total_capex_view if total_capex_view else 0):.1%} del total filtrado.</div>
            </div>
            <div class="capex-kpi" style="--accent:#047857;--soft:#f6fffb;--border:#cfe9de;">
                <div class="capex-kpi-label">Principal CCCC</div>
                <div class="capex-kpi-value">{fmt_short(top_cat_value)}</div>
                <div class="capex-kpi-sub">{escape(top_cat_name)} · {(top_cat_value / total_capex_view if total_capex_view else 0):.1%}.</div>
            </div>
            <div class="capex-kpi" style="--accent:#6D28D9;--soft:#fbf8ff;--border:#e0d3f5;">
                <div class="capex-kpi-label">Control vs hoja</div>
                <div class="capex-kpi-value">{fmt_clp_largo(capex_delta)}</div>
                <div class="capex-kpi-sub">Diferencia contra total publicado {fmt_clp_largo(capex_declared)}.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if capex_view.empty:
        st.info("No hay movimientos Capex para los filtros seleccionados.")
        st.stop()

    top3_sum = float(top_cat.head(3)["Monto"].sum()) if not top_cat.empty else 0.0
    cost_share = costo_total / total_capex_view if total_capex_view else 0.0
    first_period = capex_view["Periodo_ref"].min()
    last_period = capex_view["Periodo_ref"].max()
    period_txt = (
        f"{first_period.strftime('%b %Y')} a {last_period.strftime('%b %Y')}"
        if pd.notna(first_period) and pd.notna(last_period)
        else "sin período válido"
    )
    st.markdown(
        f"""
        <div class="capex-analysis-grid">
            <div class="capex-analysis">
                <div class="capex-analysis-title">Lectura técnica de concentración</div>
                <div class="capex-analysis-body">Las 3 categorías principales explican {top3_sum / total_capex_view:.1%} del Capex filtrado. Esto permite auditar primero los rubros estructurales y luego revisar partidas menores por excepción.</div>
            </div>
            <div class="capex-analysis">
                <div class="capex-analysis-title">Clasificación contable</div>
                <div class="capex-analysis-body">La fuente separa Costo y Gasto. En la vista actual, Costo representa {cost_share:.1%}; conviene mantener esa taxonomía porque afecta capitalización, depreciación y lectura tributaria.</div>
            </div>
            <div class="capex-analysis">
                <div class="capex-analysis-title">Calidad de dato</div>
                <div class="capex-analysis-body">El CSV trae columnas residuales vacías desde Google Sheets. La carga las elimina y valida período, monto y año antes de graficar. Cobertura temporal: {period_txt}.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    monthly = (
        capex_view.groupby(["Periodo_ref", "Estado"], as_index=False)["Monto"].sum()
        .sort_values("Periodo_ref")
    )
    fig_month = px.bar(
        monthly,
        x="Periodo_ref",
        y="Monto",
        color="Estado",
        color_discrete_map={"Costo": CHART_TEAL, "Gasto": CHART_RED},
        title="Evolución mensual Capex por clasificación",
    )
    fig_month.update_layout(
        height=380,
        barmode="stack",
        margin=dict(l=20, r=20, t=52, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#F8FAFC",
        legend=dict(orientation="h", y=1.04, x=0),
        xaxis_title="Periodo",
        yaxis_title="Monto (CLP)",
        yaxis_tickprefix="$",
        yaxis_separatethousands=True,
    )

    top_categories = top_cat.head(15).sort_values("Monto", ascending=True)
    fig_cat = px.bar(
        top_categories,
        x="Monto",
        y="CCCC",
        orientation="h",
        title="Top 15 categorías CCCC por inversión",
        color_discrete_sequence=["#0B3A86"],
    )
    fig_cat.update_layout(
        height=380,
        margin=dict(l=20, r=20, t=52, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#F8FAFC",
        xaxis_title="Monto (CLP)",
        yaxis_title="",
        xaxis_tickprefix="$",
        xaxis_separatethousands=True,
    )

    c1, c2 = st.columns([1.15, 1])
    with c1:
        st.plotly_chart(fig_month, use_container_width=True)
    with c2:
        st.plotly_chart(fig_cat, use_container_width=True)

    pareto = top_cat.copy()
    pareto["Participación"] = pareto["Monto"] / total_capex_view if total_capex_view else 0.0
    pareto["Acumulado"] = pareto["Participación"].cumsum()
    pareto_top = pareto.head(12)
    fig_pareto = go.Figure()
    fig_pareto.add_trace(
        go.Bar(
            x=pareto_top["CCCC"],
            y=pareto_top["Monto"],
            name="Monto",
            marker_color="#7FA6A2",
            hovertemplate="<b>%{x}</b><br>Monto: $%{y:,.0f}<extra></extra>",
        )
    )
    fig_pareto.add_trace(
        go.Scatter(
            x=pareto_top["CCCC"],
            y=pareto_top["Acumulado"],
            name="Acumulado",
            yaxis="y2",
            mode="lines+markers",
            line=dict(color="#4B5563", width=3),
            marker=dict(size=7),
            hovertemplate="<b>%{x}</b><br>Acumulado: %{y:.1%}<extra></extra>",
        )
    )
    fig_pareto.update_layout(
        title="Pareto Capex por CCCC",
        height=420,
        margin=dict(l=20, r=30, t=52, b=92),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#F8FAFC",
        xaxis=dict(tickangle=-28),
        yaxis=dict(title="Monto (CLP)", tickprefix="$", separatethousands=True),
        yaxis2=dict(title="Acumulado", overlaying="y", side="right", tickformat=".0%", range=[0, 1.05]),
        legend=dict(orientation="h", y=1.04, x=0),
    )
    st.plotly_chart(fig_pareto, use_container_width=True)

    capex_export = capex_view[
        ["Año", "Periodo", "Situación", "CCCC", "Estado", "Monto", "Periodo_ref"]
    ].sort_values(["Periodo_ref", "CCCC", "Situación"]).copy()
    capex_export["Periodo_ref"] = capex_export["Periodo_ref"].dt.strftime("%Y-%m")
    csv_capex = capex_export.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Descargar CSV Capex filtrado",
        data=csv_capex,
        file_name="capex_filtrado_bodegas2025.csv",
        mime="text/csv",
        key="download_capex_csv",
    )

    capex_table = capex_export.rename(columns={"Periodo_ref": "Periodo normalizado"})
    st.dataframe(
        capex_table.style.format({"Monto": "${:,.0f}"}),
        use_container_width=True,
        height=430,
    )

# =========================================================
# ⚡ TAB 7: ELECTRICIDAD (Excel por pestaña)
# =========================================================
if active_section == "⚡ Consumos Energéticos":
    title_col, btn_col = st.columns([6, 1])
    with title_col:
        st.markdown(
            tab_header("Consumos Energéticos", "Liquidación por bodega e inputs de facturación"),
            unsafe_allow_html=True,
        )
        st.caption("Vista idéntica al Excel: inputs generales, boleta CGE, inputs por bodega y liquidación.")
    with btn_col:
        st.markdown("")
        st.markdown("")
        # Placeholder for PDF button (se setea más abajo cuando tengamos los datos)
        pdf_btn_placeholder = st.empty()

    try:
        parsed_sheets = load_electricidad_parsed(ELECTRICIDAD_XLSX)
    except ImportError:
        parsed_sheets = {}
        st.error(
            "Falta la dependencia `openpyxl` para leer archivos .xlsx. "
            "Instálala en tu entorno con: `pip install openpyxl`"
        )
    except Exception as e:
        parsed_sheets = {}
        st.error(f"No se pudo cargar el Excel de electricidad: {e}")
    if not parsed_sheets:
        st.warning(
            "No se encontraron hojas válidas de electricidad para mostrar. "
            "La fuente puede venir vacía, con una estructura distinta o haber fallado durante el parseo."
        )
        st.stop()

    # Preferir hojas tipo MES-AÑO (ej. FEB-2026)
    month_sheets = [s for s in parsed_sheets.keys() if "-" in s]
    sel_months = st.multiselect(
        "Meses",
        month_sheets or list(parsed_sheets.keys()),
        default=[month_sheets[0]] if month_sheets else list(parsed_sheets.keys())[:1],
        key="elec_months",
    )
    if not sel_months:
        st.info("Selecciona al menos un mes.")
        st.stop()

    # Selector de bodega
    first_parsed = parsed_sheets[sel_months[0]]
    bodega_col = first_parsed["inputs_bodega"].columns[0]
    bodegas = (
        first_parsed["inputs_bodega"][bodega_col]
        .dropna()
        .astype(str)
        .tolist()
    )
    bodegas = [b for b in bodegas if b and b.upper() != "TOTAL"]
    sel_bodega = st.selectbox("Bodega", ["Todas"] + bodegas, index=0, key="elec_bodega")

    # Parse selected months
    parsed_by_month = {m: parsed_sheets[m] for m in sel_months}

    # Encabezado estilo Excel
    st.markdown(
        """
        <div style="background:#1f4e78;color:white;padding:8px 12px;border-radius:6px;font-weight:700;">
        Liquidación Eléctrica por Bodega
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("")

    col_left, col_right = st.columns([1, 1])
    with col_left:
        st.markdown(
            section_heading("📥", "Inputs generales", weight_class="section-heading-title-soft"),
            unsafe_allow_html=True,
        )
        _render_table(first_parsed["inputs_generales"], header_bg="#1f4e78", header_fg="white", row_alt="#fbf3d6")
    with col_right:
        st.markdown(
            section_heading("🧾", "Boleta CGE (Inputs de facturación)", weight_class="section-heading-title-soft"),
            unsafe_allow_html=True,
        )
        # Promedio por concepto según meses seleccionados
        boletas = []
        for m, p in parsed_by_month.items():
            df_b = p["boleta_cge"].copy()
            df_b["Monto $"] = pd.to_numeric(df_b["Monto $"], errors="coerce")
            boletas.append(df_b)
        if boletas:
            boleta_all = pd.concat(boletas, ignore_index=True)
            boleta_avg = (
                boleta_all
                .groupby("Concepto", as_index=False)["Monto $"]
                .mean()
            )
            # mantener orden original de la primera hoja
            orden = first_parsed["boleta_cge"]["Concepto"].tolist()
            boleta_avg["__ord"] = boleta_avg["Concepto"].apply(
                lambda x: orden.index(x) if x in orden else 999
            )
            boleta_avg = boleta_avg.sort_values("__ord").drop(columns="__ord")
            _render_table(
                boleta_avg,
                header_bg="#1f4e78",
                header_fg="white",
                row_alt="#fbf3d6",
            )
        else:
            boleta_avg = first_parsed["boleta_cge"].copy()
            _render_table(boleta_avg, header_bg="#1f4e78", header_fg="white", row_alt="#fbf3d6")

    st.markdown("")
    st.markdown(
        section_heading("🏢", "Inputs por bodega (Remarcador + horario efectivo)", weight_class="section-heading-title-soft"),
        unsafe_allow_html=True,
    )
    inputs_bodega = []
    for m, p in parsed_by_month.items():
        df_in = p["inputs_bodega"].copy()
        if sel_bodega != "Todas":
            df_in = df_in[df_in[bodega_col].astype(str) == sel_bodega]
        df_in.insert(0, "Mes", m)
        inputs_bodega.append(df_in)
    inputs_bodega = pd.concat(inputs_bodega, ignore_index=True) if inputs_bodega else pd.DataFrame()
    _render_table(
        inputs_bodega,
        header_bg="#1f4e78",
        header_fg="white",
        row_alt="#fff7e6",
        int_cols={"kWh post-18 (calc)", "kWh día (calc)"},
    )

    st.markdown("")
    st.markdown(
        section_heading("⚙️", "Liquidación por bodega (Asignación de costos de boleta + criterio horario)", weight_class="section-heading-title-soft"),
        unsafe_allow_html=True,
    )
    liquidacion = []
    for m, p in parsed_by_month.items():
        df_liq = p["liquidacion"].copy()
        if sel_bodega != "Todas":
            df_liq = df_liq[df_liq[df_liq.columns[0]].astype(str) == sel_bodega]
        df_liq.insert(0, "Mes", m)
        liquidacion.append(df_liq)
    liquidacion = pd.concat(liquidacion, ignore_index=True) if liquidacion else pd.DataFrame()
    _render_table(
        liquidacion,
        header_bg="#1f4e78",
        header_fg="white",
        row_alt="#eef8ee",
        int_cols={"kWh post-18"},
        cur_cols_extra={"IVA"},
    )

    # Gráfico profesional de Liquidación por Bodega
    st.markdown(
        section_heading("📈", "Liquidación por Bodega — Distribución de costos", weight_class="section-heading-title-soft"),
        unsafe_allow_html=True,
    )
    liq_chart = liquidacion.copy()
    if not liq_chart.empty:
        liq_chart = liq_chart[liq_chart[liq_chart.columns[1]].astype(str).str.upper() != "TOTAL"]

    if liq_chart.empty:
        st.info("No hay datos suficientes para el gráfico.")
    else:
        # Columnas esperadas
        col_mes = "Mes"
        col_bodega = liq_chart.columns[1]
        cols_costos = ["$ Energía", "$ Punta", "$ Reactiva", "$ Cargos Fijos", "$ Interés"]
        col_total = "TOTAL c/IVA $"

        # Normalizar numéricos
        for c in cols_costos + [col_total]:
            if c in liq_chart.columns:
                liq_chart[c] = pd.to_numeric(liq_chart[c], errors="coerce")
            else:
                liq_chart[c] = 0

        import plotly.graph_objects as go

        palette = {
            "$ Energía": "#7FA6A2",       # verde agua
            "$ Punta": "#4B5563",         # gris pizarra
            "$ Reactiva": "#DCAA67",      # mostaza
            "$ Cargos Fijos": "#D85E5D",  # coral
            "$ Interés": "#A8A8A8",       # gris neutro
        }

        def build_liq_fig(df_plot: pd.DataFrame, x_axis, x_title, height=520, show_legend=True):
            fig = go.Figure()
            for c in cols_costos:
                fig.add_trace(
                    go.Bar(
                        x=x_axis,
                        y=df_plot[c],
                        name=c,
                        marker=dict(color=palette.get(c, "#94a3b8")),
                        text=df_plot[c],
                        texttemplate="%{text:,.0f}",
                        textposition="inside",
                        hovertemplate=f"<b>%{{x}}</b><br>{c}: $%{{y:,.0f}}<extra></extra>",
                    )
                )
            if col_total in df_plot.columns:
                fig.add_trace(
                    go.Scatter(
                        x=x_axis,
                        y=df_plot[col_total],
                        name="TOTAL c/IVA $",
                        mode="lines+markers",
                        line=dict(color="#4B5563", width=3),
                        marker=dict(size=7, color="#4B5563"),
                        hovertemplate="<b>%{x}</b><br>Total c/IVA: $%{y:,.0f}<extra></extra>",
                        yaxis="y2",
                    )
                )
            fig.update_layout(
                template="plotly_white",
                barmode="stack",
                height=height,
                margin=dict(l=20, r=40, t=40, b=20),
                legend=dict(
                    orientation="h",
                    y=1.2,
                    x=0.5,
                    xanchor="center",
                    yanchor="bottom",
                    traceorder="normal",
                    font=dict(size=11),
                    entrywidth=120,
                    entrywidthmode="pixels",
                ) if show_legend else None,
                xaxis=dict(title=x_title, showgrid=False),
                yaxis=dict(title="Costo (CLP)", tickformat=",.0f", gridcolor="rgba(148,163,184,0.25)"),
                yaxis2=dict(
                    title="Total c/IVA (CLP)",
                    overlaying="y",
                    side="right",
                    tickformat=",.0f",
                    showgrid=False,
                ),
                hovermode="x unified",
            )
            return fig

        def build_liq_single_fig(df_plot: pd.DataFrame, height=420):
            agg_vals = df_plot[cols_costos].sum(numeric_only=True)
            vals = [float(agg_vals.get(c, 0) or 0) for c in cols_costos]
            labels = cols_costos
            total_val = float(pd.to_numeric(df_plot[col_total], errors="coerce").sum())

            fig = go.Figure(
                data=[
                    go.Pie(
                        labels=labels,
                        values=vals,
                        hole=0.58,
                        sort=False,
                        marker=dict(colors=[palette.get(c, "#94a3b8") for c in labels]),
                        texttemplate="%{percent}",
                        textposition="inside",
                        hovertemplate="%{label}<br>$%{value:,.0f} (%{percent})<extra></extra>",
                    )
                ]
            )

            bodega_label = str(df_plot[col_bodega].iloc[0]) if not df_plot.empty else ""
            mes_label = str(df_plot[col_mes].iloc[0]) if not df_plot.empty else ""
            center_text = f"<b>{bodega_label}</b><br>{mes_label}<br><span style='font-size:14px'>Total c/IVA</span><br><b>${total_val:,.0f}</b>"

            fig.update_layout(
                template="plotly_white",
                height=height,
                margin=dict(l=20, r=20, t=20, b=20),
                legend=dict(
                    orientation="h",
                    y=1.08,
                    x=0.5,
                    xanchor="center",
                    yanchor="bottom",
                    font=dict(size=11),
                ),
                annotations=[
                    dict(
                        text=center_text,
                        x=0.5,
                        y=0.5,
                        xref="paper",
                        yref="paper",
                        showarrow=False,
                        align="center",
                        font=dict(size=15, color="#111827"),
                    )
                ],
            )
            return fig

        single_period = len(sel_months) == 1
        charts_for_pdf = []
        single_month_single_bodega = len(sel_months) == 1 and sel_bodega != "Todas"
        if sel_bodega == "Todas" and len(sel_months) > 1:
            bodegas_plot = liq_chart[col_bodega].dropna().astype(str).unique().tolist()
            cols = st.columns(2)
            for i, b in enumerate(bodegas_plot):
                df_b = liq_chart[liq_chart[col_bodega].astype(str) == b]
                fig = build_liq_fig(df_b, df_b[col_mes], "Mes", height=360, show_legend=True)
                with cols[i % 2]:
                    st.markdown(f"**{b}**")
                    st.plotly_chart(fig, use_container_width=True, config={"displaylogo": False})
                charts_for_pdf.append({
                    "df": df_b,
                    "x_col": col_mes,
                    "title": f"{b}",
                    "cols_costos": cols_costos,
                    "col_total": col_total,
                    "palette": palette,
                    "chart_type": "stacked",
                })
        elif single_month_single_bodega:
            fig = build_liq_single_fig(liq_chart, height=430)
            st.plotly_chart(fig, use_container_width=True, config={"displaylogo": False})
            charts_for_pdf.append({
                "df": liq_chart,
                "x_col": col_bodega if single_period else col_mes,
                "title": "",
                "cols_costos": cols_costos,
                "col_total": col_total,
                "palette": palette,
                "chart_type": "donut",
            })
        else:
            x_axis = liq_chart[col_bodega] if single_period else liq_chart[col_mes]
            x_title = "Bodega" if single_period else "Mes"
            fig = build_liq_fig(liq_chart, x_axis, x_title, height=520, show_legend=True)
            st.plotly_chart(fig, use_container_width=True, config={"displaylogo": False})
            charts_for_pdf.append({
                "df": liq_chart,
                "x_col": col_bodega if single_period else col_mes,
                "title": "",
                "cols_costos": cols_costos,
                "col_total": col_total,
                "palette": palette,
                "chart_type": "stacked",
            })

        # PDF download button (ubicado en el header)
        try:
            pdf_bytes = build_electricidad_pdf(
                title="Liquidación Eléctrica por Bodega",
                sel_months=sel_months,
                sel_bodega=sel_bodega,
                inputs_generales=first_parsed["inputs_generales"],
                boleta_avg=boleta_avg,
                inputs_bodega=inputs_bodega,
                liquidacion=liquidacion,
                charts=charts_for_pdf,
            )
            pdf_btn_placeholder.download_button(
                "⬇️ Descargar PDF de Liquidación Eléctrica",
                data=pdf_bytes,
                file_name="liquidacion_electrica_bodega.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except Exception as e:
            pdf_btn_placeholder.error(f"No se pudo generar el PDF: {e}")
